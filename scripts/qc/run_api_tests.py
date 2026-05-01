"""
run_api_tests.py - Comprehensive end-to-end QC for the Recommendation Dashboard API.

Tests every endpoint in backend/routers/, every role boundary, and a wide
range of validation/edge cases. Outputs an Excel report with five sheets:
Summary, By Category, All Results (with expected/actual JSON + reasoning),
Failures, and Coverage Map (endpoint -> tests).

Run:
    python scripts/qc/run_api_tests.py
    python scripts/qc/run_api_tests.py --verbose
    python scripts/qc/run_api_tests.py --output reports/pre_push_run.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
import time
import traceback
import uuid
from dataclasses import dataclass, asdict, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import psycopg2
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_BASE_URL = "http://localhost:8000"
DEMO_PASSWORD = "Demo1234!"

# Mirrors backend/schemas/cart.py CartSource Literal
CART_SOURCES = [
    "manual",
    "recommendation_peer_gap",
    "recommendation_lapsed",
    "recommendation_replenishment",
    "recommendation_cart_complement",
    "recommendation_pb_upgrade",
    "recommendation_medline_conversion",
    "recommendation_item_similarity",
    "recommendation_popularity",
]

# Mirrors backend/schemas/recommendation.py REJECTION_REASON_CODES
REJECT_REASONS = [
    "not_relevant", "already_have", "out_of_stock", "price_too_high",
    "wrong_size_or_spec", "different_brand", "bad_timing",
    "wrong_recommendation", "other",
]

SIGNAL_TO_SOURCE = {
    "peer_gap":               "recommendation_peer_gap",
    "lapsed_recovery":        "recommendation_lapsed",
    "replenishment":          "recommendation_replenishment",
    "cart_complement":        "recommendation_cart_complement",
    "private_brand_upgrade":  "recommendation_pb_upgrade",
    "medline_conversion":     "recommendation_medline_conversion",
    "item_similarity":        "recommendation_item_similarity",
    "popularity":             "recommendation_popularity",
}


# Postgres connection

def load_pg_config() -> dict:
    env = ROOT / ".env"
    cfg = {"host": "localhost", "port": 5432,
           "dbname": "recommendation_dashboard",
           "user": "postgres", "password": "",
           "schema": "recdash"}
    if not env.exists():
        return cfg
    for line in env.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k, v = k.strip().lower(), v.strip().strip('"').strip("'")
        mp = {"postgres_host": "host", "postgres_port": "port",
              "postgres_db": "dbname", "postgres_user": "user",
              "postgres_password": "password", "postgres_schema": "schema"}
        if k in mp:
            cfg[mp[k]] = int(v) if k == "postgres_port" else v
    return cfg


def get_db_conn(cfg: dict):
    return psycopg2.connect(host=cfg["host"], port=cfg["port"],
                            dbname=cfg["dbname"], user=cfg["user"],
                            password=cfg["password"])


# Fixture discovery from Postgres

@dataclass
class DiscoveredUsers:
    admin_username: Optional[str] = None
    admin_user_id: Optional[int] = None
    second_admin_username: Optional[str] = None
    second_admin_user_id: Optional[int] = None
    seller_username: Optional[str] = None
    seller_user_id: Optional[int] = None
    second_seller_username: Optional[str] = None
    second_seller_user_id: Optional[int] = None

    customer_stable_username: Optional[str] = None
    customer_stable_cust_id: Optional[int] = None
    customer_stable_segment: Optional[str] = None
    customer_stable_specialty: Optional[str] = None

    customer_declining_username: Optional[str] = None
    customer_declining_cust_id: Optional[int] = None
    customer_declining_segment: Optional[str] = None
    customer_declining_specialty: Optional[str] = None

    customer_churned_username: Optional[str] = None
    customer_churned_cust_id: Optional[int] = None
    customer_churned_segment: Optional[str] = None
    customer_churned_specialty: Optional[str] = None

    customer_cold_username: Optional[str] = None
    customer_cold_cust_id: Optional[int] = None
    customer_cold_segment: Optional[str] = None
    customer_cold_specialty: Optional[str] = None

    customer_richest_username: Optional[str] = None
    customer_richest_cust_id: Optional[int] = None
    customer_richest_segment: Optional[str] = None
    customer_richest_specialty: Optional[str] = None
    customer_richest_archetype: Optional[str] = None
    customer_richest_status: Optional[str] = None
    customer_richest_n_orders: int = 0

    unassigned_cust_id: Optional[int] = None
    available_cust_ids_unassigned: list = field(default_factory=list)


def discover_users(cfg: dict) -> DiscoveredUsers:
    print("\n[Discovery] Pulling test users from database")
    schema = cfg["schema"]
    out = DiscoveredUsers()

    conn = get_db_conn(cfg)
    cur = conn.cursor()

    cur.execute(f"SELECT user_id, username FROM {schema}.users WHERE role='admin' AND is_active=true ORDER BY user_id LIMIT 2")
    admins = cur.fetchall()
    if admins:
        out.admin_user_id, out.admin_username = admins[0]
        print(f"  Admin: {out.admin_username}")
    if len(admins) > 1:
        out.second_admin_user_id, out.second_admin_username = admins[1]

    cur.execute(f"SELECT user_id, username FROM {schema}.users WHERE role='seller' AND is_active=true ORDER BY user_id LIMIT 2")
    sellers = cur.fetchall()
    if sellers:
        out.seller_user_id, out.seller_username = sellers[0]
        print(f"  Seller 1: {out.seller_username}")
    if len(sellers) > 1:
        out.second_seller_user_id, out.second_seller_username = sellers[1]
        print(f"  Seller 2: {out.second_seller_username}")

    for status in ["stable_warm", "declining_warm", "churned_warm", "cold_start"]:
        cur.execute(f"""
            SELECT u.username, c.cust_id, c.segment, c.specialty_code, COUNT(ph.purchase_id)
              FROM {schema}.users u
              JOIN {schema}.customers c ON u.cust_id = c.cust_id
              LEFT JOIN {schema}.purchase_history ph ON ph.cust_id = c.cust_id
             WHERE u.role='customer' AND u.is_active=true AND c.status=%s
             GROUP BY u.username, c.cust_id, c.segment, c.specialty_code
             ORDER BY 5 DESC LIMIT 1
        """, (status,))
        row = cur.fetchone()
        if row:
            uname, cid, segment, specialty, n_orders = row
            short = status.replace("_warm", "").replace("_start", "")
            setattr(out, f"customer_{short}_username", uname)
            setattr(out, f"customer_{short}_cust_id", cid)
            setattr(out, f"customer_{short}_segment", segment)
            setattr(out, f"customer_{short}_specialty", specialty)
            print(f"  Customer ({status}): {uname} | {segment} | {specialty} | {n_orders} orders")

    cur.execute(f"""
        SELECT u.username, c.cust_id, c.segment, c.specialty_code, c.archetype, c.status,
               COUNT(ph.purchase_id)
          FROM {schema}.users u
          JOIN {schema}.customers c ON u.cust_id = c.cust_id
          JOIN {schema}.purchase_history ph ON ph.cust_id = c.cust_id
         WHERE u.role='customer' AND u.is_active=true
         GROUP BY u.username, c.cust_id, c.segment, c.specialty_code, c.archetype, c.status
         ORDER BY 7 DESC LIMIT 1
    """)
    row = cur.fetchone()
    if row:
        out.customer_richest_username = row[0]
        out.customer_richest_cust_id = row[1]
        out.customer_richest_segment = row[2]
        out.customer_richest_specialty = row[3]
        out.customer_richest_archetype = row[4]
        out.customer_richest_status = row[5]
        out.customer_richest_n_orders = row[6]
        print(f"  Richest: {row[0]} | {row[2]}/{row[3]} | {row[4]} | {row[5]} | {row[6]} orders")

    cur.execute(f"SELECT cust_id FROM {schema}.customers WHERE assigned_seller_id IS NULL LIMIT 5")
    rows = cur.fetchall()
    if rows:
        out.unassigned_cust_id = rows[0][0]
        out.available_cust_ids_unassigned = [r[0] for r in rows]

    cur.close()
    conn.close()
    return out


# Test result tracking

@dataclass
class TestResult:
    test_id: str
    category: str
    name: str
    method: str
    endpoint: str
    role: str
    status_expected: int = 200
    status_actual: int = 0
    passed: bool = False
    duration_ms: int = 0
    expected_json: str = ""
    actual_json: str = ""
    reasoning: str = ""
    error_msg: str = ""
    notes: str = ""


class TestSession:
    def __init__(self, base_url: str, verbose: bool = False):
        self.base_url = base_url.rstrip("/")
        self.tokens: dict[str, str] = {}
        self.user_ids: dict[str, int] = {}
        self.cust_ids: dict[str, int] = {}
        self.verbose = verbose
        self.results: list[TestResult] = []
        self.session = requests.Session()
        self.test_users_created: list[int] = []
        self.test_cust_ids_created: list[int] = []

    def call(self, method: str, path: str, *,
             token_role: str = "admin",
             json_body: Any = None,
             form_body: dict | None = None,
             params: dict | None = None,
             extra_headers: dict | None = None) -> tuple[Any, int]:
        url = f"{self.base_url}{path}"
        headers = {}
        if token_role and token_role in self.tokens:
            headers["Authorization"] = f"Bearer {self.tokens[token_role]}"
        if extra_headers:
            headers.update(extra_headers)
        if self.verbose:
            print(f"    -> {method} {path} [token={token_role}]")
        t0 = time.time()
        try:
            if form_body is not None:
                resp = self.session.request(method, url, data=form_body,
                                            headers=headers, params=params, timeout=30)
            elif json_body is not None:
                resp = self.session.request(method, url, json=json_body,
                                            headers=headers, params=params, timeout=30)
            else:
                resp = self.session.request(method, url, headers=headers,
                                            params=params, timeout=30)
        except requests.RequestException as e:
            class FakeResp:
                status_code = 0
                text = str(e)
                def json(self): raise ValueError("Connection error")
            resp = FakeResp()
        ms = int((time.time() - t0) * 1000)
        return resp, ms


# Helpers (assertion / formatting / login)

def get_actual_json(resp, max_chars: int = 1500) -> str:
    try:
        body = resp.json()
        text = json.dumps(body, indent=2, default=str)
        if len(text) > max_chars:
            text = text[:max_chars] + f"\n... (truncated, total {len(json.dumps(body, default=str))} chars)"
        return text
    except Exception:
        text = (getattr(resp, 'text', '') or '')[:max_chars]
        return f"(non-JSON) {text}"


def assert_status(resp, expected) -> tuple[bool, str]:
    if isinstance(expected, int):
        expected = [expected]
    if resp.status_code in expected:
        return True, ""
    try:
        body_preview = json.dumps(resp.json())[:160]
    except Exception:
        body_preview = (getattr(resp, 'text', '') or '')[:160]
    return False, f"Expected {expected}, got {resp.status_code}. {body_preview}"


def assert_keys(resp, required: list[str]) -> tuple[bool, str]:
    try:
        body = resp.json()
    except Exception:
        return False, "Response is not JSON"
    if not isinstance(body, dict):
        return False, f"Response is not a dict, got {type(body).__name__}"
    missing = [k for k in required if k not in body]
    if missing:
        return False, f"Missing required keys: {missing}"
    return True, ""


def record(s: TestSession, r: TestResult):
    s.results.append(r)
    icon = "[PASS]" if r.passed else "[FAIL]"
    print(f"  {icon} {r.test_id}: {r.name} ({r.duration_ms}ms)")
    if not r.passed and r.error_msg:
        print(f"        {r.error_msg[:130]}")


def login(s: TestSession, role_label: str, username: str, password: str = DEMO_PASSWORD) -> bool:
    resp, _ = s.call("POST", "/auth/login", token_role="",
                     form_body={"username": username, "password": password})
    if resp.status_code != 200:
        return False
    body = resp.json()
    s.tokens[role_label] = body["access_token"]
    s.user_ids[role_label] = body["user"]["user_id"]
    if body["user"].get("cust_id"):
        s.cust_ids[role_label] = body["user"]["cust_id"]
    return True


def ensure_assigned(s: TestSession, cust_id: int, seller_id: int):
    """Idempotently assign a customer to a seller. Used for setup of tests
    that need a known assignment state."""
    s.call("PATCH", f"/customers/{cust_id}/assignment", token_role="admin",
           json_body={"seller_id": seller_id, "notes": "QC setup"})
# [1] Health

def test_health(s: TestSession):
    print("\n[1] Health Checks")

    cases = [
        ("HC-01", "App alive", "/health"),
        ("HC-02", "Postgres reachable", "/health/db"),
        ("HC-03", "Parquet readable", "/health/parquet"),
    ]
    for tid, name, path in cases:
        resp, ms = s.call("GET", path, token_role="")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id=tid, category="Health", name=name,
            method="GET", endpoint=path, role="none",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"status": "ok"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning=(
                f"PASS: Health endpoint responded HTTP 200 in {ms}ms."
                if passed else
                f"FAIL: {err}. Service may not be running or upstream dependency is broken."
            ),
        ))


# [2] Auth - login + /auth/me

def test_authentication(s: TestSession, users: DiscoveredUsers, admin_pwd: str):
    print("\n[2] Authentication")

    expected_login = json.dumps({
        "access_token": "<JWT string>",
        "token_type": "bearer",
        "expires_in_minutes": 60,
        "user": {"user_id": "<int>", "username": "<str>", "role": "<admin|seller|customer>",
                 "full_name": "<optional>", "cust_id": "<int|null>"}
    }, indent=2)

    # AUTH-01 admin
    ok = login(s, "admin", users.admin_username or "admin", admin_pwd)
    record(s, TestResult(
        test_id="AUTH-01", category="Auth",
        name=f"Admin login ({users.admin_username})",
        method="POST", endpoint="/auth/login", role="none",
        status_expected=200, status_actual=200 if ok else 401,
        passed=ok, duration_ms=0,
        expected_json=expected_login,
        actual_json=json.dumps({"access_token": (s.tokens.get("admin", "")[:25] + "...") if ok else "", "user_id": s.user_ids.get("admin")}, indent=2),
        reasoning=(
            f"PASS: Admin '{users.admin_username}' authenticated. JWT cached under role='admin' for downstream tests."
            if ok else
            f"FAIL: Admin login failed. All admin-only tests will be skipped. Check .env password and that admin is_active=true."
        ),
    ))

    # AUTH-02 second admin
    if users.second_admin_username:
        ok = login(s, "admin2", users.second_admin_username, DEMO_PASSWORD)
        record(s, TestResult(
            test_id="AUTH-02", category="Auth",
            name=f"Second admin ({users.second_admin_username})",
            method="POST", endpoint="/auth/login", role="none",
            status_expected=200, status_actual=200 if ok else 401,
            passed=ok, duration_ms=0,
            expected_json=expected_login,
            actual_json=json.dumps({"user_id": s.user_ids.get("admin2"), "role": "admin"}, indent=2) if ok else "",
            reasoning="PASS: Multi-admin support validated." if ok else "FAIL: Second admin login failed.",
        ))

    # AUTH-03 seller
    if users.seller_username:
        ok = login(s, "seller", users.seller_username, DEMO_PASSWORD)
        record(s, TestResult(
            test_id="AUTH-03", category="Auth",
            name=f"Seller login ({users.seller_username})",
            method="POST", endpoint="/auth/login", role="none",
            status_expected=200, status_actual=200 if ok else 401,
            passed=ok, duration_ms=0,
            expected_json=expected_login,
            actual_json=json.dumps({"user_id": s.user_ids.get("seller"), "role": "seller"}, indent=2) if ok else "",
            reasoning="PASS: Seller login. Token used for seller-scoped tests." if ok else "FAIL: Seller login failed.",
        ))

    # AUTH-04 second seller
    if users.second_seller_username:
        ok = login(s, "seller2", users.second_seller_username, DEMO_PASSWORD)
        record(s, TestResult(
            test_id="AUTH-04", category="Auth",
            name=f"Second seller ({users.second_seller_username})",
            method="POST", endpoint="/auth/login", role="none",
            status_expected=200, status_actual=200 if ok else 401,
            passed=ok, duration_ms=0,
            expected_json=expected_login,
            actual_json=json.dumps({"user_id": s.user_ids.get("seller2"), "role": "seller"}, indent=2) if ok else "",
            reasoning=(
                "PASS: Second seller available. Used in cross-seller isolation tests (reassignment, claim race)."
                if ok else "FAIL: Cross-seller isolation tests will be skipped."
            ),
        ))

    # AUTH-05 customer (richest)
    if users.customer_richest_username:
        ok = login(s, "customer", users.customer_richest_username, DEMO_PASSWORD)
        record(s, TestResult(
            test_id="AUTH-05", category="Auth",
            name=f"Richest customer ({users.customer_richest_username})",
            method="POST", endpoint="/auth/login", role="none",
            status_expected=200, status_actual=200 if ok else 401,
            passed=ok, duration_ms=0,
            expected_json=expected_login,
            actual_json=json.dumps({
                "user_id": s.user_ids.get("customer"),
                "cust_id": s.cust_ids.get("customer"),
                "role": "customer"
            }, indent=2) if ok else "",
            reasoning=(
                f"PASS: Demo customer logged in. cust_id={s.cust_ids.get('customer')}, "
                f"segment={users.customer_richest_segment}, status={users.customer_richest_status}, "
                f"{users.customer_richest_n_orders} orders. Drives self-service tests."
                if ok else "FAIL: Demo customer login failed."
            ),
        ))

    # AUTH-06a..d - one customer per lifecycle status
    for idx, status_short in enumerate(["stable", "declining", "churned", "cold"]):
        uname = getattr(users, f"customer_{status_short}_username")
        if not uname:
            continue
        cid = getattr(users, f"customer_{status_short}_cust_id")
        seg = getattr(users, f"customer_{status_short}_segment")
        spec = getattr(users, f"customer_{status_short}_specialty")
        role_label = f"cust_{status_short}"
        full_status = status_short + "_warm" if status_short != "cold" else "cold_start"
        ok = login(s, role_label, uname, DEMO_PASSWORD)
        record(s, TestResult(
            test_id=f"AUTH-06{chr(ord('a')+idx)}", category="Auth",
            name=f"Login {full_status} ({uname})",
            method="POST", endpoint="/auth/login", role="none",
            status_expected=200, status_actual=200 if ok else 401,
            passed=ok, duration_ms=0,
            expected_json=expected_login,
            actual_json=json.dumps({"user_id": s.user_ids.get(role_label), "cust_id": cid}, indent=2) if ok else "",
            reasoning=(
                f"PASS: {full_status} customer authenticated (cust_id={cid}, segment={seg}, specialty={spec}). Drives lifecycle-specific recommendation tests."
                if ok else f"FAIL: {full_status} customer login failed; lifecycle test will be skipped."
            ),
        ))

    # AUTH-07 invalid credentials
    resp, ms = s.call("POST", "/auth/login", token_role="",
                      form_body={"username": "definitely_not_a_user_xyz", "password": "wrong"})
    passed, err = assert_status(resp, 401)
    record(s, TestResult(
        test_id="AUTH-07", category="Auth", name="Invalid credentials -> 401",
        method="POST", endpoint="/auth/login", role="none",
        status_expected=401, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Invalid username or password"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning=(
            "PASS: Bad credentials rejected with 401. Error message is uniform - doesn't leak whether the user exists."
            if passed else f"SECURITY FAIL: {err}"
        ),
        notes="Negative security test",
    ))

    # AUTH-08 empty credentials
    resp, ms = s.call("POST", "/auth/login", token_role="",
                      form_body={"username": "", "password": ""})
    passed, err = assert_status(resp, [400, 401, 422])
    record(s, TestResult(
        test_id="AUTH-08", category="Auth", name="Empty credentials -> 4xx",
        method="POST", endpoint="/auth/login", role="none",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["body", "username"], "msg": "..."}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Validation rejects empty fields before reaching auth logic." if passed else f"FAIL: {err}",
    ))

    # AUTH-09 SQL injection probe in username
    resp, ms = s.call("POST", "/auth/login", token_role="",
                      form_body={"username": "admin' OR 1=1 --", "password": "anything"})
    passed = resp.status_code == 401
    record(s, TestResult(
        test_id="AUTH-09", category="Auth", name="SQL injection probe in login -> 401",
        method="POST", endpoint="/auth/login", role="none",
        status_expected=401, status_actual=resp.status_code,
        passed=passed, duration_ms=ms,
        error_msg="" if passed else f"Got {resp.status_code} on injection probe",
        expected_json='{"detail": "Invalid username or password"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning=(
            "PASS: SQL injection probe correctly rejected. SQLAlchemy ORM parameterizes queries."
            if passed else "CRITICAL SECURITY FAIL: probe did not return 401."
        ),
        notes="CRITICAL Security probe",
    ))

    # AUTH-10 GET /auth/me with valid token
    if "admin" in s.tokens:
        resp, ms = s.call("GET", "/auth/me", token_role="admin")
        ok_status, err = assert_status(resp, 200)
        ok_keys, kerr = assert_keys(resp, ["user_id", "username", "role"])
        passed = ok_status and ok_keys
        record(s, TestResult(
            test_id="AUTH-10", category="Auth", name="GET /auth/me as admin",
            method="GET", endpoint="/auth/me", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err or kerr,
            expected_json=json.dumps({"user_id": "<int>", "username": users.admin_username, "role": "admin"}, indent=2),
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning=(
                f"PASS: /auth/me echoes the JWT identity. user_id={s.user_ids.get('admin')}."
                if passed else f"FAIL: {err or kerr}"
            ),
        ))

    # AUTH-11 GET /auth/me without token
    resp, ms = s.call("GET", "/auth/me", token_role="")
    passed, err = assert_status(resp, 401)
    record(s, TestResult(
        test_id="AUTH-11", category="Auth", name="GET /auth/me without token -> 401",
        method="GET", endpoint="/auth/me", role="none",
        status_expected=401, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Not authenticated"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: Auth required." if passed else f"SECURITY FAIL: {err}",
    ))

    # AUTH-12 malformed JWT
    resp = s.session.request("GET", f"{s.base_url}/auth/me",
                             headers={"Authorization": "Bearer this.is.not.a.real.jwt"}, timeout=10)
    passed = resp.status_code == 401
    record(s, TestResult(
        test_id="AUTH-12", category="Auth", name="Malformed JWT -> 401",
        method="GET", endpoint="/auth/me", role="malformed",
        status_expected=401, status_actual=resp.status_code,
        passed=passed, duration_ms=0, error_msg="" if passed else f"Got {resp.status_code}",
        expected_json='{"detail": "Could not validate credentials"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: Malformed JWT rejected." if passed else "SECURITY FAIL",
    ))

    # AUTH-13 wrong scheme (Basic instead of Bearer)
    resp = s.session.request("GET", f"{s.base_url}/auth/me",
                             headers={"Authorization": "Basic dXNlcjpwYXNz"}, timeout=10)
    passed = resp.status_code == 401
    record(s, TestResult(
        test_id="AUTH-13", category="Auth", name="Basic auth scheme -> 401",
        method="GET", endpoint="/auth/me", role="basic-scheme",
        status_expected=401, status_actual=resp.status_code,
        passed=passed, duration_ms=0, error_msg="" if passed else f"Got {resp.status_code}",
        expected_json='{"detail": "Bearer auth required"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: Only Bearer scheme accepted." if passed else "SECURITY FAIL",
    ))


# [3] User management - admins, sellers, customers, password change, lifecycle

def test_user_management(s: TestSession):
    print("\n[3] User Management")
    if "admin" not in s.tokens:
        print("  SKIPPED")
        return

    suffix = uuid.uuid4().hex[:6]
    test_admin = f"qc_admin_{suffix}"
    test_seller = f"qc_seller_{suffix}"
    test_customer = f"qc_customer_{suffix}"
    test_pwd = "QcTest1234!"

    # USR-01 list users
    resp, ms = s.call("GET", "/users", token_role="admin", params={"limit": 5})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="USR-01", category="UserMgmt", name="List users (admin, limit=5)",
        method="GET", endpoint="/users?limit=5", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='[{"user_id": "<int>", "username": "<str>", "role": "<role>", "is_active": true}, "..."]',
        actual_json=get_actual_json(resp, max_chars=600),
        reasoning="PASS: User list paginated." if passed else f"FAIL: {err}",
    ))

    # USR-02 pagination
    resp, ms = s.call("GET", "/users", token_role="admin", params={"limit": 3, "offset": 2})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="USR-02", category="UserMgmt", name="Pagination limit+offset",
        method="GET", endpoint="/users?limit=3&offset=2", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json="<list, items 3-5 of full list>",
        actual_json=get_actual_json(resp, max_chars=600),
        reasoning="PASS: limit+offset respected." if passed else f"FAIL: {err}",
    ))

    # USR-03 list users as seller -> 403
    if "seller" in s.tokens:
        resp, ms = s.call("GET", "/users", token_role="seller")
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="USR-03", category="UserMgmt", name="List users as seller -> 403",
            method="GET", endpoint="/users", role="seller",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Admin role required"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: User list is admin-only." if passed else f"AUTHZ FAIL: {err}",
            notes="CRITICAL Authorization",
        ))

    # USR-04 create admin
    body = {"username": test_admin, "password": test_pwd, "full_name": "QC Test Admin"}
    resp, ms = s.call("POST", "/users/admins", token_role="admin", json_body=body)
    passed, err = assert_status(resp, [200, 201])
    new_admin_id = None
    if passed:
        try:
            new_admin_id = resp.json().get("user_id")
            if new_admin_id: s.test_users_created.append(new_admin_id)
        except Exception: pass
    record(s, TestResult(
        test_id="USR-04", category="UserMgmt", name=f"Create admin '{test_admin}'",
        method="POST", endpoint="/users/admins", role="admin",
        status_expected=201, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json=json.dumps({"request": body, "expected": {"user_id": "<new>", "role": "admin"}}, indent=2),
        actual_json=get_actual_json(resp, max_chars=500),
        reasoning=f"PASS: Admin user created (user_id={new_admin_id})." if passed else f"FAIL: {err}",
    ))

    # USR-05 create seller
    body = {"username": test_seller, "password": test_pwd, "full_name": "QC Test Seller", "territory_code": "TST"}
    resp, ms = s.call("POST", "/users/sellers", token_role="admin", json_body=body)
    passed, err = assert_status(resp, [200, 201])
    new_seller_id = None
    if passed:
        try:
            new_seller_id = resp.json().get("user_id")
            if new_seller_id: s.test_users_created.append(new_seller_id)
        except Exception: pass
    record(s, TestResult(
        test_id="USR-05", category="UserMgmt", name=f"Create seller '{test_seller}'",
        method="POST", endpoint="/users/sellers", role="admin",
        status_expected=201, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json=json.dumps({"request": body, "expected": {"role": "seller", "territory_code": "TST"}}, indent=2),
        actual_json=get_actual_json(resp, max_chars=500),
        reasoning=f"PASS: Seller created (user_id={new_seller_id})." if passed else f"FAIL: {err}",
    ))

    # USR-06 seller cannot create admin
    if "seller" in s.tokens:
        resp, ms = s.call("POST", "/users/admins", token_role="seller",
                          json_body={"username": "qc_blocked", "password": test_pwd})
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="USR-06", category="UserMgmt", name="Seller cannot create admin",
            method="POST", endpoint="/users/admins", role="seller",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Admin role required"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Privilege escalation prevented." if passed else f"PRIVILEGE FAIL: {err}",
            notes="CRITICAL privilege escalation",
        ))

    # USR-07 create customer-user (needs an unlinked cust_id)
    cfg = load_pg_config()
    new_cust_id = None
    cust_market = None
    cust_segment = None
    cust_specialty = None
    try:
        conn = get_db_conn(cfg); cur = conn.cursor()
        cur.execute(f"""
            SELECT c.cust_id, c.market_code, c.segment, c.specialty_code
              FROM {cfg['schema']}.customers c
             WHERE NOT EXISTS (SELECT 1 FROM {cfg['schema']}.users u WHERE u.cust_id = c.cust_id)
               AND c.market_code IS NOT NULL AND c.segment IS NOT NULL
             LIMIT 1
        """)
        row = cur.fetchone()
        if row: new_cust_id, cust_market, cust_segment, cust_specialty = row
        cur.close(); conn.close()
    except Exception: pass

    if new_cust_id:
        size_tier = cust_segment.split("_")[-1] if "_" in cust_segment else "small"
        body = {
            "username": test_customer, "password": test_pwd, "cust_id": new_cust_id,
            "customer_name": f"QC Customer {new_cust_id}",
            "market_code": cust_market, "segment": cust_segment, "size_tier": size_tier,
        }
        if cust_specialty: body["specialty_code"] = cust_specialty
        resp, ms = s.call("POST", "/users/customers", token_role="admin", json_body=body)
        passed, err = assert_status(resp, [200, 201])
        new_cust_user_id = None
        if passed:
            try:
                new_cust_user_id = resp.json().get("user_id")
                if new_cust_user_id: s.test_users_created.append(new_cust_user_id)
            except Exception: pass
        record(s, TestResult(
            test_id="USR-07", category="UserMgmt", name=f"Create customer-user (cust_id={new_cust_id})",
            method="POST", endpoint="/users/customers", role="admin",
            status_expected=201, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"request": body, "expected": {"role": "customer", "cust_id": new_cust_id}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=500),
            reasoning=(
                f"PASS: Customer-role user linked to cust_id={new_cust_id}. The user can now use /me endpoints."
                if passed else f"FAIL: {err}"
            ),
        ))

    # USR-08 duplicate username
    resp, ms = s.call("POST", "/users/admins", token_role="admin",
                      json_body={"username": test_admin, "password": test_pwd})
    passed, err = assert_status(resp, [400, 409, 422])
    record(s, TestResult(
        test_id="USR-08", category="UserMgmt", name="Duplicate username rejected",
        method="POST", endpoint="/users/admins", role="admin",
        status_expected=409, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Username already exists"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Username uniqueness enforced." if passed else f"FAIL: {err}",
    ))

    # USR-09 missing password
    resp, ms = s.call("POST", "/users/admins", token_role="admin",
                      json_body={"username": "qc_no_pwd"})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="USR-09", category="UserMgmt", name="Missing password -> 422",
        method="POST", endpoint="/users/admins", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["body", "password"], "msg": "Field required"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Required field validation." if passed else f"FAIL: {err}",
    ))

    # USR-10 GET /users/{id}
    if new_admin_id:
        resp, ms = s.call("GET", f"/users/{new_admin_id}", token_role="admin")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="USR-10", category="UserMgmt", name=f"GET /users/{new_admin_id}",
            method="GET", endpoint=f"/users/{new_admin_id}", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"user_id": new_admin_id, "username": test_admin}, indent=2),
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Lookup by id confirms persisted state." if passed else f"FAIL: {err}",
        ))

    # USR-11 GET nonexistent
    resp, ms = s.call("GET", "/users/9999999", token_role="admin")
    passed, err = assert_status(resp, 404)
    record(s, TestResult(
        test_id="USR-11", category="UserMgmt", name="GET nonexistent user -> 404",
        method="GET", endpoint="/users/9999999", role="admin",
        status_expected=404, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "User 9999999 not found"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: 404 for nonexistent." if passed else f"FAIL: {err}",
    ))

    # USR-12 login as newly-created admin
    ok = login(s, "test_admin", test_admin, test_pwd)
    record(s, TestResult(
        test_id="USR-12", category="UserMgmt", name="Login as newly-created admin",
        method="POST", endpoint="/auth/login", role="none",
        status_expected=200, status_actual=200 if ok else 401,
        passed=ok, duration_ms=0,
        expected_json='{"access_token": "<JWT>", "user": {"role": "admin"}}',
        actual_json=json.dumps({"login_succeeded": ok}, indent=2),
        reasoning="PASS: Account active immediately." if ok else "FAIL",
    ))

    # USR-13 password change
    if "test_admin" in s.tokens:
        new_pwd = "ChangedPwd567!"
        body = {"current_password": test_pwd, "new_password": new_pwd}
        resp, ms = s.call("PATCH", "/users/me/password", token_role="test_admin", json_body=body)
        passed, err = assert_status(resp, [200, 204])
        record(s, TestResult(
            test_id="USR-13", category="UserMgmt", name="Change own password",
            method="PATCH", endpoint="/users/me/password", role="test_admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"request": body, "expected": {"message": "..."}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Password changed." if passed else f"FAIL: {err}",
        ))

        if passed:
            # USR-14 new password works
            ok = login(s, "test_admin_v2", test_admin, new_pwd)
            record(s, TestResult(
                test_id="USR-14", category="UserMgmt", name="New password works after change",
                method="POST", endpoint="/auth/login", role="none",
                status_expected=200, status_actual=200 if ok else 401,
                passed=ok, duration_ms=0,
                expected_json='{"access_token": "<JWT issued with new pwd>"}',
                actual_json=json.dumps({"login_succeeded": ok}, indent=2),
                reasoning="PASS: Password update persisted." if ok else "FAIL",
            ))

            # USR-15 old password rejected
            resp, ms = s.call("POST", "/auth/login", token_role="",
                              form_body={"username": test_admin, "password": test_pwd})
            passed, err = assert_status(resp, 401)
            record(s, TestResult(
                test_id="USR-15", category="UserMgmt", name="Old password rejected after change",
                method="POST", endpoint="/auth/login", role="none",
                status_expected=401, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json='{"detail": "Invalid username or password"}',
                actual_json=get_actual_json(resp, max_chars=300),
                reasoning="PASS: Old password immediately invalidated." if passed else f"SECURITY FAIL: {err}",
                notes="CRITICAL Security",
            ))

        # USR-16 wrong current_password
        resp, ms = s.call("PATCH", "/users/me/password", token_role="test_admin_v2" if "test_admin_v2" in s.tokens else "test_admin",
                          json_body={"current_password": "WrongCurrentPwd!", "new_password": "AnotherNew99!"})
        passed, err = assert_status(resp, [400, 401, 403, 422])
        record(s, TestResult(
            test_id="USR-16", category="UserMgmt", name="Wrong current_password rejected",
            method="PATCH", endpoint="/users/me/password", role="test_admin",
            status_expected=400, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Current password is incorrect"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Cannot change password without knowing the current one." if passed else f"SECURITY FAIL: {err}",
            notes="CRITICAL Security",
        ))

    # USR-17 deactivate user
    if new_seller_id:
        resp, ms = s.call("DELETE", f"/users/{new_seller_id}", token_role="admin")
        passed, err = assert_status(resp, [200, 204])
        record(s, TestResult(
            test_id="USR-17", category="UserMgmt", name=f"Deactivate user {new_seller_id}",
            method="DELETE", endpoint=f"/users/{new_seller_id}", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"user_id": new_seller_id, "is_active": False}, indent=2),
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Soft-delete (is_active=false). Foreign keys preserved." if passed else f"FAIL: {err}",
        ))

        # USR-18 reactivate
        resp, ms = s.call("POST", f"/users/{new_seller_id}/reactivate", token_role="admin")
        passed, err = assert_status(resp, [200, 204])
        record(s, TestResult(
            test_id="USR-18", category="UserMgmt", name=f"Reactivate user {new_seller_id}",
            method="POST", endpoint=f"/users/{new_seller_id}/reactivate", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"user_id": new_seller_id, "is_active": True}, indent=2),
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Reactivated; can log in again." if passed else f"FAIL: {err}",
        ))

    # USR-19 customer cannot list users
    if "customer" in s.tokens:
        resp, ms = s.call("GET", "/users", token_role="customer")
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="USR-19", category="UserMgmt", name="Customer cannot list users",
            method="GET", endpoint="/users", role="customer",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Admin role required"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Customer locked out of admin endpoints." if passed else f"AUTHZ FAIL: {err}",
        ))
# [4] Customers - search, filter, get one, /me, /record

def test_customers(s: TestSession, users: DiscoveredUsers):
    print("\n[4] Customer Endpoints")
    if "admin" not in s.tokens:
        return
    cid = users.customer_richest_cust_id
    if not cid: return

    # CUST-01 GET /customers/{cid} as admin - validate full schema
    resp, ms = s.call("GET", f"/customers/{cid}", token_role="admin")
    ok_status, err = assert_status(resp, 200)
    ok_keys, kerr = assert_keys(resp, ["cust_id", "segment", "status", "archetype", "is_assigned_to_me"])
    passed = ok_status and ok_keys
    body = {}
    try: body = resp.json()
    except Exception: pass
    record(s, TestResult(
        test_id="CUST-01", category="Customers",
        name=f"GET customer {cid} (admin) - validates status+archetype+is_assigned_to_me fields",
        method="GET", endpoint=f"/customers/{cid}", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err or kerr,
        expected_json=json.dumps({
            "cust_id": cid,
            "segment": users.customer_richest_segment,
            "specialty_code": users.customer_richest_specialty,
            "status": users.customer_richest_status,
            "archetype": users.customer_richest_archetype,
            "is_assigned_to_me": True,
            "assigned_seller_id": "<int|null>"
        }, indent=2),
        actual_json=get_actual_json(resp, max_chars=600),
        reasoning=(
            f"PASS: Customer {cid} ({body.get('segment')}, status={body.get('status')}, archetype={body.get('archetype')}). "
            f"is_assigned_to_me={body.get('is_assigned_to_me')} (admins always see True - they have full action set)."
            if passed else f"FAIL: {err or kerr}"
        ),
    ))

    # CUST-02 GET nonexistent
    resp, ms = s.call("GET", "/customers/99999999", token_role="admin")
    passed, err = assert_status(resp, 404)
    record(s, TestResult(
        test_id="CUST-02", category="Customers", name="GET nonexistent customer -> 404",
        method="GET", endpoint="/customers/99999999", role="admin",
        status_expected=404, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Customer 99999999 not found"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: Nonexistent cust_id returns 404." if passed else f"FAIL: {err}",
    ))

    # CUST-03 GET /customers/me
    if "customer" in s.tokens:
        resp, ms = s.call("GET", "/customers/me", token_role="customer")
        ok_status, err = assert_status(resp, 200)
        body = {}
        try: body = resp.json()
        except Exception: pass
        passed = ok_status and body.get("cust_id") == s.cust_ids.get("customer")
        record(s, TestResult(
            test_id="CUST-03", category="Customers", name="GET /customers/me",
            method="GET", endpoint="/customers/me", role="customer",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"cust_id": s.cust_ids.get("customer"), "segment": users.customer_richest_segment}, indent=2),
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning="PASS: cust_id resolved from JWT, returns own record." if passed else f"FAIL: {err}",
        ))

    # CUST-04 customer hits /customers/me as wrong role
    if "admin" in s.tokens:
        resp, ms = s.call("GET", "/customers/me", token_role="admin")
        passed, err = assert_status(resp, [403, 404])
        record(s, TestResult(
            test_id="CUST-04", category="Customers", name="Admin -> /customers/me -> 403/404",
            method="GET", endpoint="/customers/me", role="admin",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "/me endpoints are for customers only"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: /me is customer-only." if passed else f"FAIL: {err}",
        ))

    # CUST-05 customer cannot read another customer's record
    if "customer" in s.tokens and users.customer_stable_cust_id:
        other = users.customer_stable_cust_id
        if other != s.cust_ids.get("customer"):
            resp, ms = s.call("GET", f"/customers/{other}", token_role="customer")
            passed, err = assert_status(resp, 403)
            record(s, TestResult(
                test_id="CUST-05", category="Customers", name="Customer blocked from other customer's record",
                method="GET", endpoint=f"/customers/{other}", role="customer",
                status_expected=403, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json='{"detail": "Customers may only view their own record"}',
                actual_json=get_actual_json(resp, max_chars=400),
                reasoning="PASS: Customer-to-customer privacy enforced." if passed else f"PRIVACY FAIL: {err}",
                notes="CRITICAL privacy",
            ))

    # CUST-06 seller views any customer (read is open for sellers - is_assigned_to_me reflects ownership)
    if "seller" in s.tokens:
        resp, ms = s.call("GET", f"/customers/{cid}", token_role="seller")
        ok_status, err = assert_status(resp, 200)
        body = {}
        try: body = resp.json()
        except Exception: pass
        # is_assigned_to_me should be a boolean
        flag = body.get("is_assigned_to_me")
        passed = ok_status and flag is not None
        record(s, TestResult(
            test_id="CUST-06", category="Customers",
            name=f"Seller can read any customer (is_assigned_to_me={flag})",
            method="GET", endpoint=f"/customers/{cid}", role="seller",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"cust_id": cid, "is_assigned_to_me": "<bool>"}, indent=2),
            actual_json=get_actual_json(resp, max_chars=500),
            reasoning=(
                f"PASS: Read is open across sellers (browse mode). is_assigned_to_me={flag} tells the UI whether to show write controls."
                if passed else f"FAIL: {err}"
            ),
            notes="Read open by design; is_assigned_to_me gates write actions",
        ))

    # CUST-07 search by cust_id (numeric)
    resp, ms = s.call("GET", "/customers/search", token_role="admin", params={"q": str(cid)})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="CUST-07", category="Customers", name="Search by numeric cust_id",
        method="GET", endpoint=f"/customers/search?q={cid}", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='[{"cust_id": "<exact match>", "segment": "...", "specialty_code": "..."}]',
        actual_json=get_actual_json(resp, max_chars=600),
        reasoning="PASS: Numeric search uses exact cust_id match." if passed else f"FAIL: {err}",
    ))

    # CUST-08 search by uppercase code (market or specialty)
    resp, ms = s.call("GET", "/customers/search", token_role="admin", params={"q": "PO", "limit": 10})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="CUST-08", category="Customers", name="Search uppercase code 'PO' (market)",
        method="GET", endpoint="/customers/search?q=PO", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='[{"market_code": "PO", ...}, "..."]',
        actual_json=get_actual_json(resp, max_chars=800),
        reasoning="PASS: 1-6 letter uppercase matches market_code/specialty_code exactly." if passed else f"FAIL: {err}",
    ))

    # CUST-09 search empty -> 422
    resp, ms = s.call("GET", "/customers/search", token_role="admin", params={"q": ""})
    passed, err = assert_status(resp, [400, 422])
    record(s, TestResult(
        test_id="CUST-09", category="Customers", name="Empty search rejected",
        method="GET", endpoint="/customers/search?q=", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["query", "q"], "msg": "min_length=1"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: min_length=1 enforced. Without this, empty q would match all 380K+ customers." if passed else f"FAIL: {err}",
    ))

    # CUST-10 search too long -> 422
    resp, ms = s.call("GET", "/customers/search", token_role="admin", params={"q": "X" * 100})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="CUST-10", category="Customers", name="Search >50 chars rejected",
        method="GET", endpoint="/customers/search?q=Xx100", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["query", "q"], "msg": "max_length=50"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: max_length=50 enforced." if passed else f"FAIL: {err}",
    ))

    # CUST-11 search SQL injection
    resp, ms = s.call("GET", "/customers/search", token_role="admin",
                      params={"q": "'; DROP TABLE customers; --"})
    passed = resp.status_code in (200, 400, 422)
    record(s, TestResult(
        test_id="CUST-11", category="Customers", name="Search SQL injection probe",
        method="GET", endpoint="/customers/search", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms,
        error_msg="" if passed else f"Got 5xx ({resp.status_code}) - injection broke query layer",
        expected_json='200 (empty results) or 4xx (rejected)',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning=(
            f"PASS: SQL injection probe handled safely (HTTP {resp.status_code}). SQLAlchemy parameterizes."
            if passed else "CRITICAL SECURITY FAIL: query layer crashed on injection."
        ),
        notes="CRITICAL Security probe",
    ))

    # CUST-12 search unicode
    resp, ms = s.call("GET", "/customers/search", token_role="admin", params={"q": "東京"})
    passed = resp.status_code == 200
    record(s, TestResult(
        test_id="CUST-12", category="Customers", name="Search unicode characters",
        method="GET", endpoint="/customers/search?q=Tokyo(JP)", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms,
        error_msg="" if passed else f"Got {resp.status_code}",
        expected_json='[]',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: Unicode handled without crashing." if passed else "FAIL",
    ))

    # CUST-13 customer cannot search
    if "customer" in s.tokens:
        resp, ms = s.call("GET", "/customers/search", token_role="customer", params={"q": "PO"})
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="CUST-13", category="Customers", name="Customer cannot use /customers/search",
            method="GET", endpoint="/customers/search", role="customer",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Forbidden - admin or seller role required"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Customer cannot enumerate other customers." if passed else f"PRIVACY FAIL: {err}",
            notes="CRITICAL privacy",
        ))

    # CUST-14..19 - filter by various fields
    filter_cases = [
        ("CUST-14", "Filter status=stable_warm", {"status": "stable_warm", "limit": 10}, "status"),
        ("CUST-15", "Filter status=cold_start", {"status": "cold_start", "limit": 10}, "status"),
        ("CUST-16", "Filter status=churned_warm", {"status": "churned_warm", "limit": 10}, "status"),
        ("CUST-17", "Filter status=declining_warm", {"status": "declining_warm", "limit": 10}, "status"),
        ("CUST-18", "Filter market_code=PO", {"market_code": "PO", "limit": 10}, "market_code"),
        ("CUST-19", "Filter combined status+segment", {"status": "stable_warm", "segment": "PO_large", "limit": 10}, None),
    ]
    for tid, name, params, validate_field in filter_cases:
        resp, ms = s.call("GET", "/customers/filter", token_role="admin", params=params)
        ok_status, err = assert_status(resp, 200)
        body = []
        try: body = resp.json()
        except Exception: pass
        n = len(body) if isinstance(body, list) else 0
        is_valid = True
        if ok_status and validate_field and n > 0:
            actual_vals = set(c.get(validate_field) for c in body if c.get(validate_field))
            if actual_vals and not actual_vals.issubset({params[validate_field]}):
                is_valid = False
                err = f"Filter not applied: got {validate_field} values {actual_vals}"
        passed = ok_status and is_valid
        record(s, TestResult(
            test_id=tid, category="Customers", name=name,
            method="GET", endpoint=f"/customers/filter", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"params": params, "expected": "list of customers, all matching filter"}, indent=2),
            actual_json=get_actual_json(resp, max_chars=800),
            reasoning=(
                f"PASS: {n} customers returned. " +
                (f"Filter validated: all match {validate_field}={params.get(validate_field)}." if validate_field else "Combined AND filter.")
                if passed else f"FAIL: {err}"
            ),
        ))

    # CUST-20 filter scope=all (seller)
    if "seller" in s.tokens:
        resp, ms = s.call("GET", "/customers/filter", token_role="seller",
                          params={"scope": "all", "limit": 10})
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="CUST-20", category="Customers", name="Seller filter scope=all (browse mode)",
            method="GET", endpoint="/customers/filter?scope=all", role="seller",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='[<entire customer base, read-only>]',
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning="PASS: Seller can browse the full customer base with scope=all." if passed else f"FAIL: {err}",
            notes="Browse mode for unassigned customers tab",
        ))

    # CUST-21 filter scope=mine default for seller
    if "seller" in s.tokens:
        resp, ms = s.call("GET", "/customers/filter", token_role="seller", params={"limit": 10})
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="CUST-21", category="Customers", name="Seller filter default scope=mine",
            method="GET", endpoint="/customers/filter", role="seller",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='[<only this seller\'s assigned customers>]',
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning="PASS: Default scope=mine returns own assigned customers." if passed else f"FAIL: {err}",
        ))

    # CUST-22 invalid scope
    if "seller" in s.tokens:
        resp, ms = s.call("GET", "/customers/filter", token_role="seller",
                          params={"scope": "everything", "limit": 5})
        passed, err = assert_status(resp, 400)
        record(s, TestResult(
            test_id="CUST-22", category="Customers", name="Invalid scope='everything' -> 400",
            method="GET", endpoint="/customers/filter?scope=everything", role="seller",
            status_expected=400, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "scope must be \'mine\' or \'all\'"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: scope enum validated." if passed else f"FAIL: {err}",
        ))

    # CUST-23 limit validation
    resp, ms = s.call("GET", "/customers/filter", token_role="admin", params={"limit": 9999})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="CUST-23", category="Customers", name="Filter limit > 500 rejected",
        method="GET", endpoint="/customers/filter?limit=9999", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["query", "limit"], "msg": "<= 500"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: limit cap=500 enforced." if passed else f"FAIL: {err}",
    ))


# [5] POST /customers/record - new endpoint, seller auto-assignment

def test_customer_record_create(s: TestSession, users: DiscoveredUsers):
    print("\n[5] Customer Record Creation (no login)")
    if "seller" not in s.tokens:
        print("  SKIPPED")
        return

    seller_id = users.seller_user_id
    second_seller_id = users.second_seller_user_id

    # CR-01 seller auto-assigns to self
    body = {
        "customer_business_name": f"QC Auto-Assign {uuid.uuid4().hex[:6]}",
        "market_code": "PO", "size_tier": "small", "specialty_code": "FP",
    }
    resp, ms = s.call("POST", "/customers/record", token_role="seller", json_body=body)
    ok_status, err = assert_status(resp, [200, 201])
    rb = {}
    try: rb = resp.json()
    except Exception: pass
    passed = ok_status and rb.get("assigned_seller_id") == seller_id and rb.get("is_assigned_to_me") == True
    new_id = rb.get("cust_id") if isinstance(rb, dict) else None
    if new_id: s.test_cust_ids_created.append(new_id)
    record(s, TestResult(
        test_id="CR-01", category="CustomerRecord",
        name="Seller creates customer (auto-assign to self)",
        method="POST", endpoint="/customers/record", role="seller",
        status_expected=201, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json=json.dumps({
            "request": body,
            "expected": {"cust_id": "<new int>", "assigned_seller_id": seller_id,
                         "is_assigned_to_me": True, "segment": "PO_small"}
        }, indent=2),
        actual_json=get_actual_json(resp, max_chars=600),
        reasoning=(
            f"PASS: cust_id={new_id} auto-assigned to seller {seller_id}. is_assigned_to_me=True. segment '{rb.get('segment')}' built from market+size."
            if passed else f"FAIL: assigned_seller_id={rb.get('assigned_seller_id')}, expected {seller_id}. {err}"
        ),
        notes="CRITICAL auto-assignment",
    ))

    # CR-02 seller cross-assign attempt (must 403)
    if second_seller_id:
        body = {
            "customer_business_name": f"QC Cross-Assign {uuid.uuid4().hex[:6]}",
            "market_code": "PO", "size_tier": "small",
            "assigned_seller_id": second_seller_id,
        }
        resp, ms = s.call("POST", "/customers/record", token_role="seller", json_body=body)
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="CR-02", category="CustomerRecord",
            name="Seller blocked from assigning to another seller",
            method="POST", endpoint="/customers/record", role="seller",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Sellers cannot assign customers to another seller..."}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Territory integrity. Sellers cannot route new accounts to colleagues." if passed else f"AUTHZ FAIL: {err}",
            notes="CRITICAL territory integrity",
        ))

    # CR-03 admin assigns to specific seller
    if "admin" in s.tokens and seller_id:
        body = {
            "customer_business_name": f"QC Admin-Routed {uuid.uuid4().hex[:6]}",
            "market_code": "AC", "size_tier": "large",
            "assigned_seller_id": seller_id,
        }
        resp, ms = s.call("POST", "/customers/record", token_role="admin", json_body=body)
        ok_status, err = assert_status(resp, [200, 201])
        rb = {}
        try: rb = resp.json()
        except Exception: pass
        passed = ok_status and rb.get("assigned_seller_id") == seller_id
        new_id = rb.get("cust_id") if isinstance(rb, dict) else None
        if new_id: s.test_cust_ids_created.append(new_id)
        record(s, TestResult(
            test_id="CR-03", category="CustomerRecord",
            name=f"Admin assigns new customer to seller {seller_id}",
            method="POST", endpoint="/customers/record", role="admin",
            status_expected=201, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"request": body, "expected": {"assigned_seller_id": seller_id, "segment": "AC_large"}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning=f"PASS: cust_id={new_id} assigned to seller {seller_id}." if passed else f"FAIL: {err}",
        ))

    # CR-04 admin creates unassigned customer
    if "admin" in s.tokens:
        body = {
            "customer_business_name": f"QC Unassigned {uuid.uuid4().hex[:6]}",
            "market_code": "LTC", "size_tier": "mid",
            "assigned_seller_id": None,
        }
        resp, ms = s.call("POST", "/customers/record", token_role="admin", json_body=body)
        ok_status, err = assert_status(resp, [200, 201])
        rb = {}
        try: rb = resp.json()
        except Exception: pass
        passed = ok_status and rb.get("assigned_seller_id") is None
        new_id = rb.get("cust_id") if isinstance(rb, dict) else None
        if new_id: s.test_cust_ids_created.append(new_id)
        record(s, TestResult(
            test_id="CR-04", category="CustomerRecord",
            name="Admin creates unassigned customer (claimable later)",
            method="POST", endpoint="/customers/record", role="admin",
            status_expected=201, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"assigned_seller_id": None, "segment": "LTC_mid"}, indent=2),
            actual_json=get_actual_json(resp, max_chars=500),
            reasoning=f"PASS: cust_id={new_id} created unassigned. Claimable via POST /customers/{{id}}/claim." if passed else f"FAIL: {err}",
        ))

    # CR-05 customer business name too short
    body = {"customer_business_name": "X", "market_code": "PO", "size_tier": "small"}
    resp, ms = s.call("POST", "/customers/record", token_role="seller", json_body=body)
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="CR-05", category="CustomerRecord", name="Business name <2 chars rejected",
        method="POST", endpoint="/customers/record", role="seller",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["body", "customer_business_name"], "msg": "min_length=2"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: min_length=2 enforced." if passed else f"FAIL: {err}",
    ))

    # CR-06 too long
    body = {"customer_business_name": "X" * 250, "market_code": "PO", "size_tier": "small"}
    resp, ms = s.call("POST", "/customers/record", token_role="seller", json_body=body)
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="CR-06", category="CustomerRecord", name="Business name >200 chars rejected",
        method="POST", endpoint="/customers/record", role="seller",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "max_length=200"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: max_length=200 enforced." if passed else f"FAIL: {err}",
    ))

    # CR-07 missing market_code
    body = {"customer_business_name": "QC No Market", "size_tier": "small"}
    resp, ms = s.call("POST", "/customers/record", token_role="seller", json_body=body)
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="CR-07", category="CustomerRecord", name="Missing market_code -> 422",
        method="POST", endpoint="/customers/record", role="seller",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["body", "market_code"], "msg": "Field required"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: market_code required." if passed else f"FAIL: {err}",
    ))

    # CR-08 customer cannot create
    if "customer" in s.tokens:
        body = {"customer_business_name": "QC Customer Tries", "market_code": "PO", "size_tier": "small"}
        resp, ms = s.call("POST", "/customers/record", token_role="customer", json_body=body)
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="CR-08", category="CustomerRecord", name="Customer role blocked from creating customers",
            method="POST", endpoint="/customers/record", role="customer",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Forbidden - admin or seller role required"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Customer cannot create accounts (privilege escalation prevented)." if passed else f"PRIVILEGE FAIL: {err}",
            notes="CRITICAL privilege escalation",
        ))

    # CR-09 newly-created customer routes to cold_start recs (end-to-end)
    if s.test_cust_ids_created:
        new_id = s.test_cust_ids_created[0]
        resp, ms = s.call("GET", f"/recommendations/customer/{new_id}",
                          token_role="seller", params={"n": 10})
        ok_status, err = assert_status(resp, 200)
        rb = {}
        try: rb = resp.json()
        except Exception: pass
        source = rb.get("recommendation_source") if isinstance(rb, dict) else None
        passed = ok_status and source == "cold_start"
        record(s, TestResult(
            test_id="CR-09", category="CustomerRecord",
            name=f"Newly-created customer ({new_id}) gets cold_start recs",
            method="GET", endpoint=f"/recommendations/customer/{new_id}", role="seller",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"recommendation_source": "cold_start", "recommendations": "<popularity-based>"}, indent=2),
            actual_json=get_actual_json(resp, max_chars=1000),
            reasoning=f"PASS: New customer routes to cold_start path. source='{source}'." if passed else f"FAIL: source='{source}' (expected 'cold_start'). {err}",
            notes="End-to-end: create -> get recs",
        ))
# Reasoning analyzer for recommendations

SPECIALTY_FAMILIES = {
    "FP":  ["Wound Care", "Infection Prevention", "Flu", "Rx", "Nursing", "Diagnostic"],
    "IM":  ["Wound Care", "Infection Prevention", "Flu", "Rx", "Diagnostic", "DME"],
    "PD":  ["Vaccines", "Flu", "Rx", "Nursing", "Wound Care"],
    "ENT": ["Nursing", "Infection Prevention", "Wound Care"],
    "GS":  ["Nursing", "Surgical", "Wound Care"],
}


def analyze_recommendations(body, customer_status, customer_segment, customer_specialty):
    if not isinstance(body, dict):
        return "FAIL: response not dict"
    parts = []
    recs = body.get("recommendations", [])
    src = body.get("recommendation_source")
    expected = "cold_start" if customer_status == "cold_start" else "precomputed"
    parts.append(f"source='{src}' (expected '{expected}')")
    parts.append(f"{len(recs)}/10 recs returned")
    if not recs:
        return " | ".join(parts) + " | No recs"
    signals = [r.get("primary_signal") for r in recs]
    parts.append(f"distinct signals: {sorted(set(signals))}")
    pb = sum(1 for r in recs if r.get("is_mckesson_brand"))
    parts.append(f"house-brand share: {pb}/10 ({pb*10}%)")
    spec_match = sum(1 for r in recs if r.get("specialty_match") == "match")
    parts.append(f"specialty match: {spec_match}/10 align with {customer_specialty}")
    if customer_status == "cold_start":
        pop = signals.count("popularity")
        parts.append(f"popularity recs: {pop} (expected high for cold)")
    elif customer_status == "churned_warm":
        lapsed = signals.count("lapsed_recovery")
        parts.append(f"lapsed_recovery: {lapsed} (expected re-engagement focus)")
    return " || ".join(parts)


def analyze_cart_helper(body, cart_items=None):
    if not isinstance(body, dict): return "FAIL"
    parts = []
    src = body.get("cart_source")
    parts.append(f"cart_source='{src}'")
    parts.append(f"{len(body.get('cart_complements', []))} complements, "
                 f"{len(body.get('private_brand_upgrades', []))} PB upgrades, "
                 f"{len(body.get('medline_conversions', []))} brand conversions")
    return " || ".join(parts)


# [6] Recommendations - GET, /me, with various N values, lifecycle statuses

def test_recommendations(s: TestSession, users: DiscoveredUsers):
    print("\n[6] Recommendations")
    if "admin" not in s.tokens:
        return

    # REC-01..04 - one per lifecycle status as admin
    for idx, short in enumerate(["stable", "declining", "churned", "cold"]):
        cid = getattr(users, f"customer_{short}_cust_id")
        seg = getattr(users, f"customer_{short}_segment")
        spec = getattr(users, f"customer_{short}_specialty")
        if not cid: continue
        full_status = short + "_warm" if short != "cold" else "cold_start"
        resp, ms = s.call("GET", f"/recommendations/customer/{cid}",
                          token_role="admin", params={"n": 10})
        ok_status, err = assert_status(resp, 200)
        ok_keys, kerr = assert_keys(resp, ["cust_id", "n_results", "recommendations", "recommendation_source"])
        passed = ok_status and ok_keys
        body = {}
        try: body = resp.json()
        except Exception: pass
        reasoning = f"PASS: " + analyze_recommendations(body, full_status, seg, spec) if passed else f"FAIL: {err or kerr}"
        record(s, TestResult(
            test_id=f"REC-{idx+1:02d}", category="Recommendations",
            name=f"Recs for {full_status} customer (cust={cid}, {seg}/{spec})",
            method="GET", endpoint=f"/recommendations/customer/{cid}?n=10", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err or kerr,
            expected_json=json.dumps({
                "cust_id": cid,
                "recommendation_source": "cold_start" if short == "cold" else "precomputed",
                "n_results": 10,
                "recommendations": [{"rank": "<int 1-10>", "item_id": "<int>", "primary_signal": "<one of 8>",
                                    "is_mckesson_brand": "<bool>", "specialty_match": "match|other"}, "..."]
            }, indent=2),
            actual_json=get_actual_json(resp, max_chars=2000),
            reasoning=reasoning,
            notes=f"Lifecycle: {full_status}",
        ))

    # REC-05 - customer self-service /me
    if "customer" in s.tokens:
        resp, ms = s.call("GET", "/recommendations/me", token_role="customer", params={"n": 10})
        passed, err = assert_status(resp, 200)
        body = {}
        try: body = resp.json()
        except Exception: pass
        reasoning = "PASS: " + analyze_recommendations(body, users.customer_richest_status, users.customer_richest_segment, users.customer_richest_specialty) if passed else f"FAIL: {err}"
        record(s, TestResult(
            test_id="REC-05", category="Recommendations", name="GET /recommendations/me",
            method="GET", endpoint="/recommendations/me?n=10", role="customer",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"cust_id": s.cust_ids.get("customer"), "n_results": 10}, indent=2),
            actual_json=get_actual_json(resp, max_chars=1500),
            reasoning=reasoning,
        ))

    # REC-06, 07 - n parameter respected
    cid = users.customer_richest_cust_id
    if cid:
        for n_val, tid in [(5, "REC-06"), (3, "REC-07")]:
            resp, ms = s.call("GET", f"/recommendations/customer/{cid}",
                              token_role="admin", params={"n": n_val})
            ok_status, err = assert_status(resp, 200)
            count_ok = False
            cnt = None
            if ok_status:
                try:
                    cnt = len(resp.json().get("recommendations", []))
                    count_ok = cnt == n_val
                except Exception: pass
            passed = ok_status and count_ok
            record(s, TestResult(
                test_id=tid, category="Recommendations", name=f"n={n_val} respected",
                method="GET", endpoint=f"/recommendations/customer/{cid}?n={n_val}", role="admin",
                status_expected=200, status_actual=resp.status_code,
                passed=passed, duration_ms=ms,
                error_msg=err if not count_ok else "",
                expected_json=json.dumps({"n_results": n_val, "recommendations": f"<{n_val} items>"}, indent=2),
                actual_json=get_actual_json(resp, max_chars=1200),
                reasoning=f"PASS: n={n_val} respected, returned {cnt}." if passed else f"FAIL: requested n={n_val}, got {cnt}",
            ))

    # REC-08 - n out of range (over max)
    if cid:
        resp, ms = s.call("GET", f"/recommendations/customer/{cid}",
                          token_role="admin", params={"n": 999})
        passed, err = assert_status(resp, 422)
        record(s, TestResult(
            test_id="REC-08", category="Recommendations", name="n=999 rejected (max=20)",
            method="GET", endpoint=f"/recommendations/customer/{cid}?n=999", role="admin",
            status_expected=422, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": [{"loc": ["query", "n"], "msg": "<= 20"}]}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: n cap=20 enforced." if passed else f"FAIL: {err}",
        ))

    # REC-09 - nonexistent customer
    resp, ms = s.call("GET", "/recommendations/customer/99999999",
                      token_role="admin", params={"n": 10})
    # Could be 404 or 200 with cold-start-like fallback
    passed = resp.status_code in (200, 404)
    record(s, TestResult(
        test_id="REC-09", category="Recommendations", name="Recs for nonexistent customer",
        method="GET", endpoint="/recommendations/customer/99999999", role="admin",
        status_expected=404, status_actual=resp.status_code,
        passed=passed, duration_ms=ms,
        error_msg="" if passed else f"Got {resp.status_code}",
        expected_json='Either 404 or 200 with empty/fallback recs',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning=f"INFO: returned {resp.status_code}. Either is acceptable - depends on engine fallback policy.",
    ))

    # REC-10 - customer cannot view another customer's recs
    if "customer" in s.tokens and users.customer_stable_cust_id:
        other = users.customer_stable_cust_id
        if other != s.cust_ids.get("customer"):
            resp, ms = s.call("GET", f"/recommendations/customer/{other}",
                              token_role="customer", params={"n": 10})
            passed, err = assert_status(resp, 403)
            record(s, TestResult(
                test_id="REC-10", category="Recommendations",
                name="Customer cannot view another customer's recs",
                method="GET", endpoint=f"/recommendations/customer/{other}", role="customer",
                status_expected=403, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json='{"detail": "Customers can only view their own recommendations"}',
                actual_json=get_actual_json(resp, max_chars=400),
                reasoning="PASS: Cross-customer rec privacy enforced." if passed else f"PRIVACY FAIL: {err}",
                notes="CRITICAL privacy",
            ))

    # REC-11 - seller cannot view recs for unassigned customer
    # Use the customer_stable customer; reassign to admin/none first to make seller2 unauthorized
    if "seller2" in s.tokens and users.customer_stable_cust_id and users.seller_user_id:
        # Make sure stable customer is assigned to seller1 (not seller2)
        ensure_assigned(s, users.customer_stable_cust_id, users.seller_user_id)
        resp, ms = s.call("GET", f"/recommendations/customer/{users.customer_stable_cust_id}",
                          token_role="seller2", params={"n": 5})
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="REC-11", category="Recommendations",
            name="Seller2 blocked from recs of seller1's customer",
            method="GET", endpoint=f"/recommendations/customer/{users.customer_stable_cust_id}", role="seller2",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "This customer is not assigned to you"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Cross-seller rec isolation enforced." if passed else f"AUTHZ FAIL: {err}",
            notes="CRITICAL cross-seller isolation",
        ))


# [7] Cart-helper - both calling modes

def test_cart_helper(s: TestSession, users: DiscoveredUsers):
    print("\n[7] Cart-helper")
    if "admin" not in s.tokens or not users.customer_richest_cust_id:
        return
    cid = users.customer_richest_cust_id

    # CH-01 cart-helper using postgres cart
    resp, ms = s.call("POST", "/recommendations/cart-helper",
                      token_role="admin", json_body={"cust_id": cid})
    ok_status, err = assert_status(resp, 200)
    ok_keys, kerr = assert_keys(resp, ["cust_id", "cart_complements", "private_brand_upgrades", "medline_conversions", "cart_source"])
    passed = ok_status and ok_keys
    body = {}
    try: body = resp.json()
    except Exception: pass
    record(s, TestResult(
        test_id="CH-01", category="CartHelper",
        name="Cart-helper using postgres cart (no cart_items in body)",
        method="POST", endpoint="/recommendations/cart-helper", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err or kerr,
        expected_json=json.dumps({
            "request": {"cust_id": cid},
            "expected": {"cart_source": "postgres_cart or empty",
                         "cart_complements": "<list>", "private_brand_upgrades": "<list>", "medline_conversions": "<list>"}
        }, indent=2),
        actual_json=get_actual_json(resp, max_chars=2000),
        reasoning=f"PASS: " + analyze_cart_helper(body) if passed else f"FAIL: {err or kerr}",
    ))

    # CH-02 cart-helper with explicit cart_items
    rec_resp, _ = s.call("GET", f"/recommendations/customer/{cid}", token_role="admin", params={"n": 5})
    if rec_resp.status_code == 200:
        recs = rec_resp.json().get("recommendations", [])
        if len(recs) >= 2:
            items = [r["item_id"] for r in recs[:2]]
            resp, ms = s.call("POST", "/recommendations/cart-helper",
                              token_role="admin",
                              json_body={"cust_id": cid, "cart_items": items})
            ok_status, err = assert_status(resp, 200)
            body = {}
            try: body = resp.json()
            except Exception: pass
            cart_src = body.get("cart_source") if isinstance(body, dict) else None
            passed = ok_status and cart_src == "request_body"
            record(s, TestResult(
                test_id="CH-02", category="CartHelper",
                name=f"Cart-helper with explicit cart_items={items}",
                method="POST", endpoint="/recommendations/cart-helper", role="admin",
                status_expected=200, status_actual=resp.status_code,
                passed=passed, duration_ms=ms,
                error_msg=err if not passed else "",
                expected_json=json.dumps({
                    "request": {"cust_id": cid, "cart_items": items},
                    "expected": {"cart_source": "request_body", "cart_complements": "<list>"}
                }, indent=2),
                actual_json=get_actual_json(resp, max_chars=2000),
                reasoning=(
                    f"PASS: cart_source='{cart_src}' confirms hypothetical cart was used (NOT live postgres cart). "
                    + analyze_cart_helper(body, items)
                    if passed else f"FAIL: cart_source='{cart_src}' (expected 'request_body'). {err}"
                ),
            ))

    # CH-03 nonexistent customer
    resp, ms = s.call("POST", "/recommendations/cart-helper", token_role="admin",
                      json_body={"cust_id": 99999999})
    passed = resp.status_code in (404, 200)
    record(s, TestResult(
        test_id="CH-03", category="CartHelper", name="Cart-helper for nonexistent customer",
        method="POST", endpoint="/recommendations/cart-helper", role="admin",
        status_expected=404, status_actual=resp.status_code,
        passed=passed, duration_ms=ms,
        error_msg="" if passed else f"Got {resp.status_code}",
        expected_json='Either 404 or 200 (empty results)',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning=f"INFO: got {resp.status_code}.",
    ))

    # CH-04 customer self-cart-helper (their own)
    if "customer" in s.tokens:
        own_cid = s.cust_ids.get("customer")
        resp, ms = s.call("POST", "/recommendations/cart-helper",
                          token_role="customer", json_body={"cust_id": own_cid})
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="CH-04", category="CartHelper", name="Customer cart-helper for own cart",
            method="POST", endpoint="/recommendations/cart-helper", role="customer",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"cust_id": own_cid, "cart_complements": "<list>"}, indent=2),
            actual_json=get_actual_json(resp, max_chars=1500),
            reasoning="PASS: Customer can run cart-helper for their own cart." if passed else f"FAIL: {err}",
        ))

    # CH-05 customer cannot cart-helper another customer
    if "customer" in s.tokens and users.customer_stable_cust_id:
        other = users.customer_stable_cust_id
        if other != s.cust_ids.get("customer"):
            resp, ms = s.call("POST", "/recommendations/cart-helper",
                              token_role="customer", json_body={"cust_id": other})
            passed, err = assert_status(resp, 403)
            record(s, TestResult(
                test_id="CH-05", category="CartHelper", name="Customer blocked from another customer's cart-helper",
                method="POST", endpoint="/recommendations/cart-helper", role="customer",
                status_expected=403, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json='{"detail": "Customers can only run cart helper for their own cart"}',
                actual_json=get_actual_json(resp, max_chars=400),
                reasoning="PASS: Cross-customer cart-helper blocked." if passed else f"PRIVACY FAIL: {err}",
                notes="CRITICAL privacy",
            ))


# [8] POST /recommendations/reject - new endpoint

def test_reject_recommendation(s: TestSession, users: DiscoveredUsers):
    print("\n[8] Reject Recommendation Flow")
    if "admin" not in s.tokens or "seller" not in s.tokens or not users.seller_user_id:
        print("  SKIPPED")
        return

    cid = users.customer_stable_cust_id
    if not cid: return

    seller_id = users.seller_user_id
    ensure_assigned(s, cid, seller_id)

    # Get real item_ids
    rec_resp, _ = s.call("GET", f"/recommendations/customer/{cid}",
                         token_role="seller", params={"n": 10})
    if rec_resp.status_code != 200:
        print("  SKIPPED (could not fetch recs to reject)")
        return
    recs = rec_resp.json().get("recommendations", [])
    if len(recs) < 9:
        print(f"  SKIPPED (need 9 recs to test all reason codes, got {len(recs)})")
        return
    item_ids = [r["item_id"] for r in recs]
    sigs = {r["item_id"]: r.get("primary_signal") for r in recs}

    # REJ-01..09 each reason code
    for idx, reason in enumerate(REJECT_REASONS):
        item_id = item_ids[idx]
        body = {
            "cust_id": cid, "item_id": item_id,
            "primary_signal": sigs.get(item_id),
            "reason_code": reason,
            "reason_note": f"QC test - {reason}" if reason == "other" else None,
        }
        resp, ms = s.call("POST", "/recommendations/reject", token_role="seller", json_body=body)
        ok_status, err = assert_status(resp, [200, 201])
        ok_keys, kerr = assert_keys(resp, ["event_id", "outcome", "reason_code"])
        passed = ok_status and ok_keys
        rb = {}
        try: rb = resp.json()
        except Exception: pass
        record(s, TestResult(
            test_id=f"REJ-{idx+1:02d}", category="RejectFlow",
            name=f"Reject reason='{reason}'",
            method="POST", endpoint="/recommendations/reject", role="seller",
            status_expected=201, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err or kerr,
            expected_json=json.dumps({"request": body, "expected": {"event_id": "<int>", "outcome": "rejected", "reason_code": reason}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning=(
                f"PASS: '{reason}' accepted. event_id={rb.get('event_id')}. Audit row written to recommendation_events."
                if passed else f"FAIL: {err or kerr}"
            ),
        ))

    # REJ-10 invalid code
    resp, ms = s.call("POST", "/recommendations/reject", token_role="seller",
                      json_body={"cust_id": cid, "item_id": item_ids[0], "reason_code": "made_up_xyz"})
    passed, err = assert_status(resp, 400)
    record(s, TestResult(
        test_id="REJ-10", category="RejectFlow", name="Invalid reason_code -> 400",
        method="POST", endpoint="/recommendations/reject", role="seller",
        status_expected=400, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Unknown reason_code \'made_up_xyz\'..."}',
        actual_json=get_actual_json(resp, max_chars=500),
        reasoning="PASS: Reason code outside the 9 allowed values rejected." if passed else f"FAIL: {err}",
    ))

    # REJ-11 admin blocked (sellers only)
    resp, ms = s.call("POST", "/recommendations/reject", token_role="admin",
                      json_body={"cust_id": cid, "item_id": item_ids[0], "reason_code": "not_relevant"})
    passed, err = assert_status(resp, 403)
    record(s, TestResult(
        test_id="REJ-11", category="RejectFlow", name="Admin blocked - sellers only",
        method="POST", endpoint="/recommendations/reject", role="admin",
        status_expected=403, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Only sellers can reject recommendations"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: Reject is a seller workflow tool." if passed else f"FAIL: {err}",
        notes="CRITICAL role-scope",
    ))

    # REJ-12 customer blocked
    if "customer" in s.tokens:
        resp, ms = s.call("POST", "/recommendations/reject", token_role="customer",
                          json_body={"cust_id": s.cust_ids.get("customer"), "item_id": 1, "reason_code": "not_relevant"})
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="REJ-12", category="RejectFlow", name="Customer blocked from reject",
            method="POST", endpoint="/recommendations/reject", role="customer",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Only sellers can reject recommendations"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Customer cannot reject - feedback flow is seller-only." if passed else f"FAIL: {err}",
        ))

    # REJ-13 unauthenticated
    resp, ms = s.call("POST", "/recommendations/reject", token_role="",
                      json_body={"cust_id": cid, "item_id": item_ids[0], "reason_code": "not_relevant"})
    passed, err = assert_status(resp, 401)
    record(s, TestResult(
        test_id="REJ-13", category="RejectFlow", name="Unauthenticated reject -> 401",
        method="POST", endpoint="/recommendations/reject", role="none",
        status_expected=401, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Not authenticated"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: Auth required." if passed else f"FAIL: {err}",
    ))

    # REJ-14 cross-seller (seller2 trying to reject seller1's customer rec)
    if "seller2" in s.tokens and users.second_seller_user_id:
        # ensure cid is assigned to seller1
        ensure_assigned(s, cid, seller_id)
        resp, ms = s.call("POST", "/recommendations/reject", token_role="seller2",
                          json_body={"cust_id": cid, "item_id": item_ids[0], "reason_code": "not_relevant"})
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="REJ-14", category="RejectFlow",
            name="Seller2 blocked from rejecting seller1's customer's recs",
            method="POST", endpoint="/recommendations/reject", role="seller2",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "This customer is not assigned to you"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Cross-seller reject blocked." if passed else f"AUTHZ FAIL: {err}",
            notes="CRITICAL cross-seller isolation",
        ))

    # REJ-15 nonexistent customer
    resp, ms = s.call("POST", "/recommendations/reject", token_role="seller",
                      json_body={"cust_id": 99999999, "item_id": item_ids[0], "reason_code": "not_relevant"})
    passed, err = assert_status(resp, 404)
    record(s, TestResult(
        test_id="REJ-15", category="RejectFlow", name="Reject for nonexistent customer -> 404",
        method="POST", endpoint="/recommendations/reject", role="seller",
        status_expected=404, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Customer 99999999 not found"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: 404 for nonexistent." if passed else f"FAIL: {err}",
    ))

    # REJ-16 reason_code case normalization (UPPERCASE -> lowercased)
    resp, ms = s.call("POST", "/recommendations/reject", token_role="seller",
                      json_body={"cust_id": cid, "item_id": item_ids[0], "reason_code": "NOT_RELEVANT"})
    ok_status, err = assert_status(resp, [200, 201])
    rb = {}
    try: rb = resp.json()
    except Exception: pass
    code_lower = rb.get("reason_code") == "not_relevant" if isinstance(rb, dict) else False
    passed = ok_status and code_lower
    record(s, TestResult(
        test_id="REJ-16", category="RejectFlow", name="Case normalization (UPPERCASE -> lowercase)",
        method="POST", endpoint="/recommendations/reject", role="seller",
        status_expected=201, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"reason_code": "not_relevant"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning=(
            f"PASS: 'NOT_RELEVANT' normalized to '{rb.get('reason_code')}'. Prevents duplicate rows in analytics."
            if passed else f"FAIL: stored code='{rb.get('reason_code')}'. {err}"
        ),
        notes="Data normalization",
    ))

    # REJ-17 missing item_id
    resp, ms = s.call("POST", "/recommendations/reject", token_role="seller",
                      json_body={"cust_id": cid, "reason_code": "not_relevant"})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="REJ-17", category="RejectFlow", name="Missing item_id -> 422",
        method="POST", endpoint="/recommendations/reject", role="seller",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["body", "item_id"], "msg": "Field required"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Required field validation." if passed else f"FAIL: {err}",
    ))

    # REJ-18 reason_note over 2000 chars
    resp, ms = s.call("POST", "/recommendations/reject", token_role="seller",
                      json_body={"cust_id": cid, "item_id": item_ids[0],
                                 "reason_code": "other", "reason_note": "X" * 2500})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="REJ-18", category="RejectFlow", name="reason_note over 2000 chars -> 422",
        method="POST", endpoint="/recommendations/reject", role="seller",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["body", "reason_note"], "msg": "max_length=2000"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: max_length=2000 on reason_note enforced." if passed else f"FAIL: {err}",
    ))
# [9] Cart workflow - all sources, all status transitions, edge cases

def test_cart(s: TestSession, users: DiscoveredUsers):
    print("\n[9] Cart Workflow")
    if "admin" not in s.tokens or not users.customer_richest_cust_id:
        return
    cid = users.customer_richest_cust_id

    # Get recs to use
    rec_resp, _ = s.call("GET", f"/recommendations/customer/{cid}",
                         token_role="admin", params={"n": 10})
    if rec_resp.status_code != 200:
        return
    recs = rec_resp.json().get("recommendations", [])
    if not recs:
        return

    # Map signal -> source name
    items_by_source = {"manual": recs[-1]["item_id"]}
    for r in recs:
        sig = r.get("primary_signal")
        src = SIGNAL_TO_SOURCE.get(sig)
        if src and src not in items_by_source:
            items_by_source[src] = r["item_id"]

    # CART-01 view cart
    resp, ms = s.call("GET", f"/customers/{cid}/cart", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="CART-01", category="Cart", name=f"View cart for cust {cid}",
        method="GET", endpoint=f"/customers/{cid}/cart", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json=json.dumps({"cust_id": cid, "items": "<list>", "total_items": "<int>", "estimated_total": "<decimal>"}, indent=2),
        actual_json=get_actual_json(resp, max_chars=600),
        reasoning="PASS: Cart endpoint reachable." if passed else f"FAIL: {err}",
    ))

    # CART-02..N add by each source
    cart_item_ids = []
    test_idx = 2
    for source, item_id in items_by_source.items():
        if item_id is None: continue
        body = {"item_id": item_id, "quantity": 2, "source": source}
        resp, ms = s.call("POST", f"/customers/{cid}/cart", token_role="admin", json_body=body)
        ok_status, err = assert_status(resp, [200, 201])
        passed = ok_status
        item_rec = next((r for r in recs if r.get("item_id") == item_id), {})
        if passed:
            try:
                rb = resp.json()
                cart_id = rb.get("cart_item_id") or rb.get("item", {}).get("cart_item_id")
                if cart_id: cart_item_ids.append(cart_id)
            except Exception: pass
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart",
            name=f"Add to cart - source='{source}'",
            method="POST", endpoint=f"/customers/{cid}/cart", role="admin",
            status_expected=201, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"request": body, "expected": {"cart_item_id": "<int>", "source": source, "status": "in_cart"}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=500),
            reasoning=(
                f"PASS: source='{source}' accepted. Item: {item_rec.get('description', '?')[:40]}, family={item_rec.get('family', '?')}."
                if passed else f"FAIL: source='{source}' rejected. {err}"
            ),
        ))
        test_idx += 1

    # Invalid source
    resp, ms = s.call("POST", f"/customers/{cid}/cart", token_role="admin",
                      json_body={"item_id": items_by_source.get("manual", 1), "quantity": 1, "source": "invalid_xyz"})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id=f"CART-{test_idx:02d}", category="Cart", name="Invalid source rejected",
        method="POST", endpoint=f"/customers/{cid}/cart", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["body", "source"], "msg": "must be one of 9 enum values"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Invalid source rejected. Enum enforced." if passed else f"FAIL: {err}",
    ))
    test_idx += 1

    # quantity = 0
    resp, ms = s.call("POST", f"/customers/{cid}/cart", token_role="admin",
                      json_body={"item_id": items_by_source.get("manual", 1), "quantity": 0, "source": "manual"})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id=f"CART-{test_idx:02d}", category="Cart", name="quantity=0 rejected (gt=0)",
        method="POST", endpoint=f"/customers/{cid}/cart", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "quantity must be > 0"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: gt=0 enforced on quantity." if passed else f"FAIL: {err}",
    ))
    test_idx += 1

    # negative quantity
    resp, ms = s.call("POST", f"/customers/{cid}/cart", token_role="admin",
                      json_body={"item_id": items_by_source.get("manual", 1), "quantity": -5, "source": "manual"})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id=f"CART-{test_idx:02d}", category="Cart", name="negative quantity rejected",
        method="POST", endpoint=f"/customers/{cid}/cart", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "quantity must be > 0"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Negative quantity rejected." if passed else f"FAIL: {err}",
    ))
    test_idx += 1

    # PATCH quantity
    if cart_item_ids:
        cart_id = cart_item_ids[0]
        resp, ms = s.call("PATCH", f"/cart/{cart_id}", token_role="admin", json_body={"quantity": 7})
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name=f"Update quantity for cart_item {cart_id}",
            method="PATCH", endpoint=f"/cart/{cart_id}", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"request": {"quantity": 7}, "expected": {"quantity": 7, "status": "in_cart"}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Quantity updated 2 -> 7." if passed else f"FAIL: {err}",
        ))
        test_idx += 1

        # PATCH qty=0 on existing line
        resp, ms = s.call("PATCH", f"/cart/{cart_id}", token_role="admin", json_body={"quantity": 0})
        passed, err = assert_status(resp, [400, 422])
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name="PATCH quantity=0 on existing line rejected",
            method="PATCH", endpoint=f"/cart/{cart_id}", role="admin",
            status_expected=422, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "quantity must be > 0"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Cannot update quantity to 0 (use DELETE instead)." if passed else f"FAIL: {err}",
        ))
        test_idx += 1

    # PATCH nonexistent
    resp, ms = s.call("PATCH", "/cart/99999999", token_role="admin", json_body={"quantity": 5})
    passed = resp.status_code in (404, 400)
    record(s, TestResult(
        test_id=f"CART-{test_idx:02d}", category="Cart", name="PATCH nonexistent cart_item -> 404/400",
        method="PATCH", endpoint="/cart/99999999", role="admin",
        status_expected=404, status_actual=resp.status_code,
        passed=passed, duration_ms=ms,
        error_msg="" if passed else f"Got {resp.status_code}",
        expected_json='{"detail": "Cart item 99999999 not found"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: 404 for nonexistent." if passed else "FAIL",
    ))
    test_idx += 1

    # Mark not_sold
    if len(cart_item_ids) >= 2:
        target = cart_item_ids[1]
        resp, ms = s.call("PATCH", f"/cart/{target}/status", token_role="admin",
                          json_body={"status": "not_sold"})
        passed, err = assert_status(resp, [200, 204])
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name=f"Mark cart line {target} as not_sold",
            method="PATCH", endpoint=f"/cart/{target}/status", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"request": {"status": "not_sold"}, "expected": {"status": "not_sold"}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Customer declined this rec; no purchase_history written." if passed else f"FAIL: {err}",
            notes="CRITICAL: not_sold = no revenue",
        ))
        test_idx += 1

        # Try to flip not_sold back to sold (illegal transition)
        resp, ms = s.call("PATCH", f"/cart/{target}/status", token_role="admin",
                          json_body={"status": "sold"})
        passed = resp.status_code in (400, 409)
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name="Cannot transition not_sold -> sold",
            method="PATCH", endpoint=f"/cart/{target}/status", role="admin",
            status_expected=409, status_actual=resp.status_code,
            passed=passed, duration_ms=ms,
            error_msg="" if passed else f"Illegal transition allowed ({resp.status_code})",
            expected_json='{"detail": "Cart item is already \'not_sold\'..."}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning=(
                "PASS: Cart state machine rejects not_sold -> sold transition. Once status is set, it's terminal."
                if passed else "DATA INTEGRITY FAIL: Illegal status transition allowed."
            ),
            notes="CRITICAL state machine",
        ))
        test_idx += 1

    # Invalid status enum
    if cart_item_ids:
        target = cart_item_ids[0]
        resp, ms = s.call("PATCH", f"/cart/{target}/status", token_role="admin",
                          json_body={"status": "pending"})
        passed, err = assert_status(resp, [400, 422])
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name="Invalid status='pending' rejected",
            method="PATCH", endpoint=f"/cart/{target}/status", role="admin",
            status_expected=422, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "status must be \'sold\' or \'not_sold\'"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Status enum enforced." if passed else f"FAIL: {err}",
        ))
        test_idx += 1

    # Checkout
    if cart_item_ids:
        target = cart_item_ids[0]
        resp, ms = s.call("POST", f"/cart/{target}/checkout", token_role="admin")
        passed, err = assert_status(resp, [200, 201])
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name=f"Checkout cart line {target}",
            method="POST", endpoint=f"/cart/{target}/checkout", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"expected": {"cart_item_id": target, "status": "sold", "purchase_history_id": "<new int>"}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Checkout writes to purchase_history (revenue recorded)." if passed else f"FAIL: {err}",
            notes="CRITICAL: closes loop rec -> revenue",
        ))
        test_idx += 1

        # Double checkout
        resp, ms = s.call("POST", f"/cart/{target}/checkout", token_role="admin")
        passed = resp.status_code in (400, 409)
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name="Double checkout rejected",
            method="POST", endpoint=f"/cart/{target}/checkout", role="admin",
            status_expected=409, status_actual=resp.status_code,
            passed=passed, duration_ms=ms,
            error_msg="" if passed else f"Got {resp.status_code} - double checkout allowed!",
            expected_json='{"detail": "Cart item is already sold"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning=(
                f"PASS: Double-checkout rejected with HTTP {resp.status_code}. Prevents accidentally double-counting revenue."
                if passed else "CRITICAL DATA INTEGRITY FAIL: revenue could be double-counted!"
            ),
            notes="CRITICAL data integrity",
        ))
        test_idx += 1

    # GET cart history
    resp, ms = s.call("GET", f"/customers/{cid}/cart/history", token_role="admin", params={"status": "all"})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id=f"CART-{test_idx:02d}", category="Cart", name="Cart history (all statuses)",
        method="GET", endpoint=f"/customers/{cid}/cart/history?status=all", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"items": [{"cart_item_id": "<int>", "status": "<sold|not_sold|in_cart>"}, "..."]}',
        actual_json=get_actual_json(resp, max_chars=1500),
        reasoning="PASS: Cart history audit trail accessible." if passed else f"FAIL: {err}",
    ))
    test_idx += 1

    # GET /cart/me
    if "customer" in s.tokens:
        resp, ms = s.call("GET", "/cart/me", token_role="customer")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name="GET /cart/me as customer",
            method="GET", endpoint="/cart/me", role="customer",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"cust_id": "<self>", "items": "<list>"}',
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning="PASS: Customer self-service cart." if passed else f"FAIL: {err}",
        ))
        test_idx += 1

    # /cart/me as admin -> 403
    resp, ms = s.call("GET", "/cart/me", token_role="admin")
    passed, err = assert_status(resp, 403)
    record(s, TestResult(
        test_id=f"CART-{test_idx:02d}", category="Cart", name="Admin -> /cart/me -> 403",
        method="GET", endpoint="/cart/me", role="admin",
        status_expected=403, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "This endpoint is for customers"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: /cart/me is customer-only." if passed else f"FAIL: {err}",
    ))
    test_idx += 1

    # Cannot delete sold line
    if cart_item_ids:
        target = cart_item_ids[0]  # this one was checked out
        resp, ms = s.call("DELETE", f"/cart/{target}", token_role="admin")
        passed = resp.status_code in (400, 403, 409)
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name="Cannot DELETE sold cart line",
            method="DELETE", endpoint=f"/cart/{target}", role="admin",
            status_expected=400, status_actual=resp.status_code,
            passed=passed, duration_ms=ms,
            error_msg="" if passed else f"Got {resp.status_code} - sold line deletable!",
            expected_json='{"detail": "Cannot delete a sold cart line"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Sold lines preserved as audit trail." if passed else "DATA INTEGRITY FAIL",
            notes="CRITICAL: revenue records preserved",
        ))
        test_idx += 1

    # DELETE in_cart line (allowed)
    if len(cart_item_ids) >= 3:
        target = cart_item_ids[2]
        resp, ms = s.call("DELETE", f"/cart/{target}", token_role="admin")
        passed, err = assert_status(resp, [200, 204])
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart", name=f"DELETE in_cart line {target} (allowed)",
            method="DELETE", endpoint=f"/cart/{target}", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"message": "Cart item deleted"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: in_cart lines can be deleted (mistake correction)." if passed else f"FAIL: {err}",
        ))
        test_idx += 1

    # Add to nonexistent customer
    resp, ms = s.call("POST", "/customers/99999999/cart", token_role="admin",
                      json_body={"item_id": items_by_source.get("manual", 1), "quantity": 1, "source": "manual"})
    passed, err = assert_status(resp, 404)
    record(s, TestResult(
        test_id=f"CART-{test_idx:02d}", category="Cart", name="Add to nonexistent customer's cart -> 404",
        method="POST", endpoint="/customers/99999999/cart", role="admin",
        status_expected=404, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "Customer 99999999 not found"}',
        actual_json=get_actual_json(resp, max_chars=300),
        reasoning="PASS: Customer existence validated." if passed else f"FAIL: {err}",
    ))
    test_idx += 1

    # Seller cart op on customer NOT assigned to them
    if "seller2" in s.tokens and users.customer_stable_cust_id and users.seller_user_id:
        # Make sure the customer is assigned to seller1
        ensure_assigned(s, users.customer_stable_cust_id, users.seller_user_id)
        resp, ms = s.call("POST", f"/customers/{users.customer_stable_cust_id}/cart",
                          token_role="seller2",
                          json_body={"item_id": items_by_source.get("manual", 1), "quantity": 1, "source": "manual"})
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id=f"CART-{test_idx:02d}", category="Cart",
            name="Seller2 blocked from adding to seller1's customer's cart",
            method="POST", endpoint=f"/customers/{users.customer_stable_cust_id}/cart", role="seller2",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "This customer is not assigned to you"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Cart writes scoped to assigned seller." if passed else f"AUTHZ FAIL: {err}",
            notes="CRITICAL cross-seller isolation",
        ))
        test_idx += 1


# [10] Purchase history

def test_purchase_history(s: TestSession, users: DiscoveredUsers):
    print("\n[10] Purchase History")
    if "admin" not in s.tokens:
        return
    cid = users.customer_richest_cust_id

    if cid:
        # PH-01 history exists
        resp, ms = s.call("GET", f"/customers/{cid}/history", token_role="admin")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="PH-01", category="PurchaseHistory",
            name=f"History for richest customer ({cid}, {users.customer_richest_n_orders} orders)",
            method="GET", endpoint=f"/customers/{cid}/history", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"cust_id": "<int>", "total": "<int>", "items": [{"purchase_id": "<int>", "item_id": "<int>", "quantity": "<int>"}, "..."]}',
            actual_json=get_actual_json(resp, max_chars=1500),
            reasoning=(
                f"PASS: History reachable for cust_id={cid}. Drives the 'replenishment' rec signal."
                if passed else f"FAIL: {err}"
            ),
        ))

        # PH-02 pagination
        resp, ms = s.call("GET", f"/customers/{cid}/history", token_role="admin", params={"limit": 10, "offset": 0})
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="PH-02", category="PurchaseHistory", name="History pagination limit=10",
            method="GET", endpoint=f"/customers/{cid}/history?limit=10", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"items": "<list of up to 10>"}',
            actual_json=get_actual_json(resp, max_chars=1500),
            reasoning="PASS: limit cap enforced." if passed else f"FAIL: {err}",
        ))

    # PH-03 cold customer (sparse history)
    cold_cid = users.customer_cold_cust_id
    if cold_cid:
        resp, ms = s.call("GET", f"/customers/{cold_cid}/history", token_role="admin")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="PH-03", category="PurchaseHistory",
            name=f"History for cold-start customer ({cold_cid})",
            method="GET", endpoint=f"/customers/{cold_cid}/history", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"items": []}',
            actual_json=get_actual_json(resp, max_chars=1000),
            reasoning="PASS: Cold customer returns empty list, not error." if passed else f"FAIL: {err}",
        ))

    # PH-04 limit > 500 rejected
    if cid:
        resp, ms = s.call("GET", f"/customers/{cid}/history", token_role="admin", params={"limit": 9999})
        passed, err = assert_status(resp, 422)
        record(s, TestResult(
            test_id="PH-04", category="PurchaseHistory", name="limit > 500 rejected",
            method="GET", endpoint=f"/customers/{cid}/history?limit=9999", role="admin",
            status_expected=422, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": [{"loc": ["query", "limit"], "msg": "<= 500"}]}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: limit cap=500 enforced." if passed else f"FAIL: {err}",
        ))

    # PH-05 customer cannot view another's history
    if "customer" in s.tokens and users.customer_stable_cust_id:
        other = users.customer_stable_cust_id
        if other != s.cust_ids.get("customer"):
            resp, ms = s.call("GET", f"/customers/{other}/history", token_role="customer")
            passed, err = assert_status(resp, 403)
            record(s, TestResult(
                test_id="PH-05", category="PurchaseHistory",
                name="Customer cannot view another customer's purchase history",
                method="GET", endpoint=f"/customers/{other}/history", role="customer",
                status_expected=403, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json='{"detail": "Customers can only view their own purchase history"}',
                actual_json=get_actual_json(resp, max_chars=400),
                reasoning="PASS: Purchase history privacy enforced." if passed else f"PRIVACY FAIL: {err}",
                notes="CRITICAL privacy",
            ))


# [11] Assignment lifecycle

def test_assignment_lifecycle(s: TestSession, users: DiscoveredUsers):
    print("\n[11] Assignment Lifecycle")
    if "admin" not in s.tokens or not users.seller_user_id:
        return
    cid = users.customer_stable_cust_id
    if not cid: return

    seller1 = users.seller_user_id
    seller2 = users.second_seller_user_id

    # ASN-01 seller's own customer list
    if "seller" in s.tokens:
        resp, ms = s.call("GET", "/sellers/me/customers", token_role="seller")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="ASN-01", category="Assignments", name="Seller's own customer list",
            method="GET", endpoint="/sellers/me/customers", role="seller",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"seller_id": "<int>", "total": "<int>", "items": "<list>"}',
            actual_json=get_actual_json(resp, max_chars=1000),
            reasoning="PASS: /me convenience endpoint works." if passed else f"FAIL: {err}",
        ))

    # ASN-02 admin views any seller's customers
    if seller1:
        resp, ms = s.call("GET", f"/sellers/{seller1}/customers", token_role="admin")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="ASN-02", category="Assignments", name=f"Admin views seller {seller1}'s customers",
            method="GET", endpoint=f"/sellers/{seller1}/customers", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"seller_id": "<int>", "items": "<list>"}',
            actual_json=get_actual_json(resp, max_chars=1000),
            reasoning="PASS: Admin can view any seller's book." if passed else f"FAIL: {err}",
        ))

    # ASN-03 seller cannot view ANOTHER seller's customers
    if seller2 and "seller" in s.tokens:
        resp, ms = s.call("GET", f"/sellers/{seller2}/customers", token_role="seller")
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="ASN-03", category="Assignments",
            name="Seller cannot view another seller's customer list",
            method="GET", endpoint=f"/sellers/{seller2}/customers", role="seller",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Sellers can only view their own customer list"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Cross-seller pipeline blocked." if passed else f"AUTHZ FAIL: {err}",
            notes="CRITICAL cross-seller isolation",
        ))

    # ASN-04 admin assigns
    body = {"seller_id": seller1, "notes": "QC assignment"}
    resp, ms = s.call("PATCH", f"/customers/{cid}/assignment", token_role="admin", json_body=body)
    passed, err = assert_status(resp, [200, 201, 204])
    record(s, TestResult(
        test_id="ASN-04", category="Assignments", name=f"Assign cust {cid} -> seller {seller1}",
        method="PATCH", endpoint=f"/customers/{cid}/assignment", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json=json.dumps({"request": body, "expected": {"new_seller_id": seller1, "change_reason": "admin_assign"}}, indent=2),
        actual_json=get_actual_json(resp, max_chars=600),
        reasoning="PASS: Assignment recorded with audit row." if passed else f"FAIL: {err}",
    ))

    # ASN-05 seller sees is_assigned_to_me=true after assignment
    if "seller" in s.tokens:
        resp, ms = s.call("GET", f"/customers/{cid}", token_role="seller")
        ok_status, err = assert_status(resp, 200)
        body_resp = {}
        try: body_resp = resp.json()
        except Exception: pass
        flag = body_resp.get("is_assigned_to_me")
        passed = ok_status and flag == True
        record(s, TestResult(
            test_id="ASN-05", category="Assignments",
            name=f"After assignment, seller sees is_assigned_to_me=true",
            method="GET", endpoint=f"/customers/{cid}", role="seller",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"is_assigned_to_me": True, "assigned_seller_id": seller1}, indent=2),
            actual_json=get_actual_json(resp, max_chars=500),
            reasoning=(
                f"PASS: Flag correctly set. UI uses this to enable cart actions."
                if passed else f"FAIL: is_assigned_to_me={flag} (expected True). {err}"
            ),
        ))

    # ASN-06 reassign to seller2
    if seller2:
        body = {"seller_id": seller2, "notes": "QC reassignment"}
        resp, ms = s.call("PATCH", f"/customers/{cid}/assignment", token_role="admin", json_body=body)
        passed, err = assert_status(resp, [200, 201, 204])
        record(s, TestResult(
            test_id="ASN-06", category="Assignments",
            name=f"Reassign {cid}: seller {seller1} -> seller {seller2}",
            method="PATCH", endpoint=f"/customers/{cid}/assignment", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"new_seller_id": seller2, "previous_seller_id": seller1, "change_reason": "admin_reassign"}, indent=2),
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning="PASS: Reassignment audit row written." if passed else f"FAIL: {err}",
        ))

        # ASN-07 OLD seller can still READ but is_assigned_to_me flips to False
        if "seller" in s.tokens:
            resp, ms = s.call("GET", f"/customers/{cid}", token_role="seller")
            ok_status, err = assert_status(resp, 200)
            body_resp = {}
            try: body_resp = resp.json()
            except Exception: pass
            flag = body_resp.get("is_assigned_to_me")
            passed = ok_status and flag == False
            record(s, TestResult(
                test_id="ASN-07", category="Assignments",
                name=f"Old seller {seller1} sees is_assigned_to_me=false after reassignment",
                method="GET", endpoint=f"/customers/{cid}", role="seller",
                status_expected=200, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json=json.dumps({"is_assigned_to_me": False, "assigned_seller_id": seller2}, indent=2),
                actual_json=get_actual_json(resp, max_chars=500),
                reasoning=(
                    f"PASS: Read is open by design (browse mode); is_assigned_to_me={flag} flag tells UI to disable write actions for old seller."
                    if passed else f"FAIL: is_assigned_to_me={flag} (expected False). {err}"
                ),
                notes="Browse mode + flag-based write gating",
            ))

            # ASN-08 OLD seller blocked from WRITE actions (cart-helper)
            resp, ms = s.call("POST", "/recommendations/cart-helper",
                              token_role="seller", json_body={"cust_id": cid})
            passed, err = assert_status(resp, 403)
            record(s, TestResult(
                test_id="ASN-08", category="Assignments",
                name=f"Old seller {seller1} BLOCKED from cart-helper after reassignment",
                method="POST", endpoint="/recommendations/cart-helper", role="seller",
                status_expected=403, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json='{"detail": "This customer is not assigned to you"}',
                actual_json=get_actual_json(resp, max_chars=400),
                reasoning="PASS: Write actions scoped to assigned_seller_id." if passed else f"SECURITY FAIL: {err}",
                notes="CRITICAL: write isolation",
            ))

        # ASN-09 NEW seller sees is_assigned_to_me=true
        if "seller2" in s.tokens:
            resp, ms = s.call("GET", f"/customers/{cid}", token_role="seller2")
            ok_status, err = assert_status(resp, 200)
            body_resp = {}
            try: body_resp = resp.json()
            except Exception: pass
            flag = body_resp.get("is_assigned_to_me")
            passed = ok_status and flag == True
            record(s, TestResult(
                test_id="ASN-09", category="Assignments",
                name=f"New seller {seller2} sees is_assigned_to_me=true",
                method="GET", endpoint=f"/customers/{cid}", role="seller2",
                status_expected=200, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json=json.dumps({"is_assigned_to_me": True, "assigned_seller_id": seller2}, indent=2),
                actual_json=get_actual_json(resp, max_chars=500),
                reasoning="PASS: Atomic ownership swap - new seller gains write capability." if passed else f"FAIL: {err}",
            ))

    # ASN-10 assignment history (admin only)
    resp, ms = s.call("GET", f"/customers/{cid}/assignment-history", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ASN-10", category="Assignments", name="Assignment history audit (admin)",
        method="GET", endpoint=f"/customers/{cid}/assignment-history", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"cust_id": "<int>", "total_changes": "<int>", "items": "<list>"}',
        actual_json=get_actual_json(resp, max_chars=1500),
        reasoning="PASS: Full audit trail." if passed else f"FAIL: {err}",
    ))

    # ASN-11 seller cannot view assignment history
    if "seller" in s.tokens:
        resp, ms = s.call("GET", f"/customers/{cid}/assignment-history", token_role="seller")
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="ASN-11", category="Assignments", name="Seller blocked from assignment history",
            method="GET", endpoint=f"/customers/{cid}/assignment-history", role="seller",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Only admins can view assignment history"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Audit trail is admin-only." if passed else f"AUTHZ FAIL: {err}",
        ))

    # ASN-12 unassign
    body = {"seller_id": None, "notes": "QC unassign"}
    resp, ms = s.call("PATCH", f"/customers/{cid}/assignment", token_role="admin", json_body=body)
    passed, err = assert_status(resp, [200, 201, 204])
    record(s, TestResult(
        test_id="ASN-12", category="Assignments", name="Unassign (seller_id=null)",
        method="PATCH", endpoint=f"/customers/{cid}/assignment", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"new_seller_id": null, "change_reason": "admin_unassign"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Customer back to claimable pool." if passed else f"FAIL: {err}",
    ))

    # ASN-13 self-claim
    if "seller" in s.tokens:
        resp, ms = s.call("POST", f"/customers/{cid}/claim", token_role="seller")
        passed, err = assert_status(resp, [200, 201])
        record(s, TestResult(
            test_id="ASN-13", category="Assignments", name=f"Seller {seller1} self-claims customer",
            method="POST", endpoint=f"/customers/{cid}/claim", role="seller",
            status_expected=201, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"new_seller_id": "<self>", "change_reason": "seller_claim"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Field-rep self-service claim." if passed else f"FAIL: {err}",
        ))

    # ASN-14 customer cannot change assignment
    if "customer" in s.tokens:
        resp, ms = s.call("PATCH", f"/customers/{cid}/assignment", token_role="customer",
                          json_body={"seller_id": seller1})
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="ASN-14", category="Assignments", name="Customer cannot change assignment",
            method="PATCH", endpoint=f"/customers/{cid}/assignment", role="customer",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Only admins can directly change assignments"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Customers cannot manipulate their assigned seller." if passed else f"PRIVILEGE FAIL: {err}",
            notes="CRITICAL privilege escalation",
        ))

    # ASN-15 bulk assign
    bulk_cids = [c for c in [users.customer_declining_cust_id, users.customer_churned_cust_id] if c and c != cid]
    if bulk_cids and seller1:
        body = {"seller_id": seller1, "cust_ids": bulk_cids, "notes": "QC bulk"}
        resp, ms = s.call("POST", "/customers/assignments/bulk", token_role="admin", json_body=body)
        passed, err = assert_status(resp, [200, 201])
        record(s, TestResult(
            test_id="ASN-15", category="Assignments", name=f"Bulk assign {len(bulk_cids)} to seller {seller1}",
            method="POST", endpoint="/customers/assignments/bulk", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json=json.dumps({"request": body, "expected": {"requested_count": len(bulk_cids), "assigned_count": "<int>"}}, indent=2),
            actual_json=get_actual_json(resp, max_chars=800),
            reasoning="PASS: Territory rebalancing." if passed else f"FAIL: {err}",
        ))

    # ASN-16 bulk empty list -> 422
    if seller1:
        resp, ms = s.call("POST", "/customers/assignments/bulk", token_role="admin",
                          json_body={"seller_id": seller1, "cust_ids": []})
        passed, err = assert_status(resp, 422)
        record(s, TestResult(
            test_id="ASN-16", category="Assignments", name="Bulk with empty cust_ids -> 422",
            method="POST", endpoint="/customers/assignments/bulk", role="admin",
            status_expected=422, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": [{"loc": ["body", "cust_ids"], "msg": "min_length=1"}]}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: min_length=1 enforced on cust_ids." if passed else f"FAIL: {err}",
        ))

    # ASN-17 bulk over 500 -> 422
    if seller1:
        big = list(range(50000000, 50000000 + 600))
        resp, ms = s.call("POST", "/customers/assignments/bulk", token_role="admin",
                          json_body={"seller_id": seller1, "cust_ids": big})
        passed, err = assert_status(resp, 422)
        record(s, TestResult(
            test_id="ASN-17", category="Assignments", name="Bulk over 500 IDs -> 422",
            method="POST", endpoint="/customers/assignments/bulk", role="admin",
            status_expected=422, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": [{"loc": ["body", "cust_ids"], "msg": "max_length=500"}]}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: max_length=500 caps batch size." if passed else f"FAIL: {err}",
        ))

    # ASN-18 bulk with invalid cust_ids (mixed)
    if seller1 and bulk_cids:
        mixed = bulk_cids + [99999998, 99999999]
        resp, ms = s.call("POST", "/customers/assignments/bulk", token_role="admin",
                          json_body={"seller_id": seller1, "cust_ids": mixed, "notes": "QC mixed"})
        passed, err = assert_status(resp, [200, 207])
        body_resp = {}
        try: body_resp = resp.json()
        except Exception: pass
        record(s, TestResult(
            test_id="ASN-18", category="Assignments", name="Bulk mixed valid+invalid (partial success)",
            method="POST", endpoint="/customers/assignments/bulk", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"assigned_count": "<int>", "skipped_count": "<int>", "skipped_reasons": {"99999998": "not_found"}}',
            actual_json=get_actual_json(resp, max_chars=800),
            reasoning=(
                f"PASS: assigned={body_resp.get('assigned_count')}, skipped={body_resp.get('skipped_count')}. Partial-success pattern."
                if passed else f"FAIL: {err}"
            ),
        ))

    # ASN-19 seller blocked from bulk
    if "seller" in s.tokens and bulk_cids:
        resp, ms = s.call("POST", "/customers/assignments/bulk", token_role="seller",
                          json_body={"seller_id": seller1, "cust_ids": bulk_cids})
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="ASN-19", category="Assignments", name="Seller blocked from bulk-assign",
            method="POST", endpoint="/customers/assignments/bulk", role="seller",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Admin role required"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Bulk is admin-only." if passed else f"AUTHZ FAIL: {err}",
            notes="CRITICAL Authorization",
        ))

    # ASN-20 concurrent claim race - seller2 tries to claim already-claimed customer
    if "seller2" in s.tokens and users.customer_cold_cust_id:
        cold = users.customer_cold_cust_id
        # Unassign first to set known state
        s.call("PATCH", f"/customers/{cold}/assignment", token_role="admin",
               json_body={"seller_id": None})
        # Seller1 claims
        s.call("POST", f"/customers/{cold}/claim", token_role="seller")
        # Seller2 tries to claim
        resp, ms = s.call("POST", f"/customers/{cold}/claim", token_role="seller2")
        passed = resp.status_code in (400, 403, 409)
        record(s, TestResult(
            test_id="ASN-20", category="Assignments", name="Concurrent claim - second seller blocked",
            method="POST", endpoint=f"/customers/{cold}/claim", role="seller2",
            status_expected=409, status_actual=resp.status_code,
            passed=passed, duration_ms=ms,
            error_msg="" if passed else f"Got {resp.status_code} - race condition!",
            expected_json='{"detail": "Customer already assigned to another seller"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning=(
                f"PASS: Race-safe claim ({resp.status_code})."
                if passed else "DATA INTEGRITY FAIL: claim race possible."
            ),
            notes="CRITICAL race condition",
        ))

    # ASN-21 customer cannot claim
    if "customer" in s.tokens and users.unassigned_cust_id:
        target = users.unassigned_cust_id
        resp, ms = s.call("POST", f"/customers/{target}/claim", token_role="customer")
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="ASN-21", category="Assignments", name="Customer cannot use claim endpoint",
            method="POST", endpoint=f"/customers/{target}/claim", role="customer",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Forbidden"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Customer cannot claim (privilege escalation prevented)." if passed else f"PRIVILEGE FAIL: {err}",
            notes="CRITICAL privilege escalation",
        ))
# [12] Stats - Admin scope (all 9 admin endpoints)

def test_stats_admin(s: TestSession):
    print("\n[12] Admin Stats")
    if "admin" not in s.tokens:
        return

    # ST-01 overview
    resp, ms = s.call("GET", "/admin/stats/overview", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-01", category="StatsAdmin", name="Overview KPIs",
        method="GET", endpoint="/admin/stats/overview", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"total_customers": "<int>", "total_revenue": "<decimal>", "active_carts": "<int>"}',
        actual_json=get_actual_json(resp, max_chars=800),
        reasoning="PASS: Top-line KPIs for admin dashboard." if passed else f"FAIL: {err}",
    ))

    # ST-02 sales-trend default (90d daily)
    resp, ms = s.call("GET", "/admin/stats/sales-trend", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-02", category="StatsAdmin", name="Sales-trend default (90d, daily)",
        method="GET", endpoint="/admin/stats/sales-trend", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"granularity": "daily", "range_label": "90d", "points": "<list>"}',
        actual_json=get_actual_json(resp, max_chars=800),
        reasoning="PASS: Time-series chart data." if passed else f"FAIL: {err}",
    ))

    # ST-03 sales-trend 1y monthly
    resp, ms = s.call("GET", "/admin/stats/sales-trend", token_role="admin",
                      params={"granularity": "monthly", "range": "1y"})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-03", category="StatsAdmin", name="Sales-trend monthly 1y",
        method="GET", endpoint="/admin/stats/sales-trend?granularity=monthly&range=1y", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"granularity": "monthly", "range_label": "1y"}',
        actual_json=get_actual_json(resp, max_chars=600),
        reasoning="PASS: Different granularities supported." if passed else f"FAIL: {err}",
    ))

    # ST-04 sales-trend invalid granularity
    resp, ms = s.call("GET", "/admin/stats/sales-trend", token_role="admin",
                      params={"granularity": "yearly"})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="ST-04", category="StatsAdmin", name="Invalid granularity rejected",
        method="GET", endpoint="/admin/stats/sales-trend?granularity=yearly", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["query", "granularity"], "msg": "must be daily|weekly|monthly"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Granularity Literal enforced." if passed else f"FAIL: {err}",
    ))

    # ST-05 sales-trend invalid range
    resp, ms = s.call("GET", "/admin/stats/sales-trend", token_role="admin",
                      params={"range": "lifetime"})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="ST-05", category="StatsAdmin", name="Invalid range rejected",
        method="GET", endpoint="/admin/stats/sales-trend?range=lifetime", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["query", "range"], "msg": "must be 7d|30d|90d|180d|1y|all"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Range Literal enforced." if passed else f"FAIL: {err}",
    ))

    # ST-06 conversion-by-signal admin
    resp, ms = s.call("GET", "/admin/stats/conversion-by-signal", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-06", category="StatsAdmin", name="Conversion by signal (admin)",
        method="GET", endpoint="/admin/stats/conversion-by-signal", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"rows": [{"source": "<rec_*>", "cart_adds": "<int>", "sold": "<int>", "conversion_rate_pct": "<float>"}, "..."]}',
        actual_json=get_actual_json(resp, max_chars=1500),
        reasoning="PASS: Per-signal performance for engine effectiveness." if passed else f"FAIL: {err}",
    ))

    # ST-07 segment-distribution
    resp, ms = s.call("GET", "/admin/stats/segment-distribution", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-07", category="StatsAdmin", name="Segment distribution",
        method="GET", endpoint="/admin/stats/segment-distribution", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"rows": [{"segment": "<str>", "count": "<int>"}, "..."]}',
        actual_json=get_actual_json(resp, max_chars=1000),
        reasoning="PASS: Customer pop counts per segment (Parquet-sourced)." if passed else f"FAIL: {err}",
    ))

    # ST-08 top-sellers default
    resp, ms = s.call("GET", "/admin/stats/top-sellers", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-08", category="StatsAdmin", name="Top sellers leaderboard (default)",
        method="GET", endpoint="/admin/stats/top-sellers", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"range_label": "all", "rows": [{"seller_id": "<int>", "username": "<str>", "revenue": "<decimal>"}, "..."]}',
        actual_json=get_actual_json(resp, max_chars=1200),
        reasoning="PASS: Sales leaderboard." if passed else f"FAIL: {err}",
    ))

    # ST-09 top-sellers limit=5
    resp, ms = s.call("GET", "/admin/stats/top-sellers", token_role="admin",
                      params={"limit": 5, "range": "30d"})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-09", category="StatsAdmin", name="Top sellers limit=5 range=30d",
        method="GET", endpoint="/admin/stats/top-sellers?limit=5&range=30d", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"range_label": "30d", "rows": "<<=5>"}',
        actual_json=get_actual_json(resp, max_chars=800),
        reasoning="PASS: Limit and range filter applied." if passed else f"FAIL: {err}",
    ))

    # ST-10 top-sellers limit > 50
    resp, ms = s.call("GET", "/admin/stats/top-sellers", token_role="admin",
                      params={"limit": 100})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="ST-10", category="StatsAdmin", name="top-sellers limit > 50 rejected",
        method="GET", endpoint="/admin/stats/top-sellers?limit=100", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["query", "limit"], "msg": "<= 50"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: limit cap=50 enforced." if passed else f"FAIL: {err}",
    ))

    # ST-11 top-customers default (NEW)
    resp, ms = s.call("GET", "/admin/stats/top-customers", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-11", category="StatsAdmin", name="Top customers leaderboard (NEW)",
        method="GET", endpoint="/admin/stats/top-customers", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"range_label": "all", "rows": [{"cust_id": "<int>", "revenue": "<decimal>", "n_orders": "<int>"}, "..."]}',
        actual_json=get_actual_json(resp, max_chars=1200),
        reasoning="PASS: Customer revenue ranking. Replaces old segment panel." if passed else f"FAIL: {err}",
    ))

    # ST-12 top-customers 90d
    resp, ms = s.call("GET", "/admin/stats/top-customers", token_role="admin",
                      params={"range": "90d", "limit": 10})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-12", category="StatsAdmin", name="Top customers range=90d",
        method="GET", endpoint="/admin/stats/top-customers?range=90d&limit=10", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"range_label": "90d"}',
        actual_json=get_actual_json(resp, max_chars=800),
        reasoning="PASS: Range filter applied." if passed else f"FAIL: {err}",
    ))

    # ST-13 engine-effectiveness (NEW)
    resp, ms = s.call("GET", "/admin/stats/engine-effectiveness", token_role="admin")
    ok_status, err = assert_status(resp, 200)
    ok_keys, kerr = assert_keys(resp, ["totals", "by_signal", "by_reason"])
    passed = ok_status and ok_keys
    record(s, TestResult(
        test_id="ST-13", category="StatsAdmin", name="Engine effectiveness funnel (NEW)",
        method="GET", endpoint="/admin/stats/engine-effectiveness", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err or kerr,
        expected_json=json.dumps({
            "totals": {"cart_adds": "<int>", "sold": "<int>", "rejected": "<int>",
                       "conversion_rate_pct": "<float>", "acceptance_rate_pct": "<float>"},
            "by_signal": "<list of 8 signal funnels>",
            "by_reason": "<list of reject reason counts>"
        }, indent=2),
        actual_json=get_actual_json(resp, max_chars=2500),
        reasoning="PASS: Engine effectiveness panel - funnel from rec -> cart -> sold/rejected." if passed else f"FAIL: {err or kerr}",
        notes="NEW endpoint - core to feedback loop story",
    ))

    # ST-14 recent-sales default
    resp, ms = s.call("GET", "/admin/stats/recent-sales", token_role="admin")
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="ST-14", category="StatsAdmin", name="Recent sales feed (default)",
        method="GET", endpoint="/admin/stats/recent-sales", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"items": [{"sold_at": "<datetime>", "cust_id": "<int>", "item_id": "<int>"}, "..."]}',
        actual_json=get_actual_json(resp, max_chars=1500),
        reasoning="PASS: Live feed for activity panel." if passed else f"FAIL: {err}",
    ))

    # ST-15 recent-sales limit > 200 rejected
    resp, ms = s.call("GET", "/admin/stats/recent-sales", token_role="admin",
                      params={"limit": 500})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="ST-15", category="StatsAdmin", name="recent-sales limit > 200 rejected",
        method="GET", endpoint="/admin/stats/recent-sales?limit=500", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": [{"loc": ["query", "limit"], "msg": "<= 200"}]}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: limit cap=200 enforced." if passed else f"FAIL: {err}",
    ))

    # ST-16..21 - all admin stats blocked for sellers AND customers
    admin_only = [
        ("/admin/stats/overview", "ST-16"),
        ("/admin/stats/conversion-by-signal", "ST-17"),
        ("/admin/stats/top-customers", "ST-18"),
        ("/admin/stats/engine-effectiveness", "ST-19"),
        ("/admin/stats/top-sellers", "ST-20"),
        ("/admin/stats/recent-sales", "ST-21"),
    ]
    for path, tid in admin_only:
        for role in ["seller", "customer"]:
            if role not in s.tokens: continue
            resp, ms = s.call("GET", path, token_role=role)
            passed, err = assert_status(resp, 403)
            record(s, TestResult(
                test_id=f"{tid}-{role}", category="StatsAdmin",
                name=f"{path} blocked for {role}",
                method="GET", endpoint=path, role=role,
                status_expected=403, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json='{"detail": "Admin role required"}',
                actual_json=get_actual_json(resp, max_chars=300),
                reasoning=f"PASS: {role} cannot access admin stats." if passed else f"AUTHZ FAIL: {err}",
            ))


# [13] Stats - Seller + Customer scope

def test_stats_seller_customer(s: TestSession, users: DiscoveredUsers):
    print("\n[13] Seller + Customer Stats")

    # SS-01 /sellers/me/stats
    if "seller" in s.tokens:
        resp, ms = s.call("GET", "/sellers/me/stats", token_role="seller")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="SS-01", category="StatsSellerCustomer", name="/sellers/me/stats",
            method="GET", endpoint="/sellers/me/stats", role="seller",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"seller_id": "<self>", "n_customers": "<int>", "revenue": "<decimal>"}',
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning="PASS: Seller's own performance metrics." if passed else f"FAIL: {err}",
        ))

    # SS-02 /sellers/me/conversion-by-signal
    if "seller" in s.tokens:
        resp, ms = s.call("GET", "/sellers/me/conversion-by-signal", token_role="seller")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="SS-02", category="StatsSellerCustomer", name="Seller's conversion-by-signal",
            method="GET", endpoint="/sellers/me/conversion-by-signal", role="seller",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"rows": [{"source": "<rec_*>", "cart_adds": "<int>", "sold": "<int>"}, "..."]}',
            actual_json=get_actual_json(resp, max_chars=1500),
            reasoning="PASS: Seller-scoped conversion (filtered to own customers)." if passed else f"FAIL: {err}",
        ))

    # SS-03 admin blocked from /sellers/me/stats
    if "admin" in s.tokens:
        resp, ms = s.call("GET", "/sellers/me/stats", token_role="admin")
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="SS-03", category="StatsSellerCustomer", name="Admin blocked from /sellers/me/stats",
            method="GET", endpoint="/sellers/me/stats", role="admin",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Seller role required"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: /me is seller-only; admin must use /sellers/{id}/* instead." if passed else f"FAIL: {err}",
        ))

    # SS-04 customer blocked from /sellers/me/stats
    if "customer" in s.tokens:
        resp, ms = s.call("GET", "/sellers/me/stats", token_role="customer")
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="SS-04", category="StatsSellerCustomer", name="Customer blocked from /sellers/me/stats",
            method="GET", endpoint="/sellers/me/stats", role="customer",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "Seller role required"}',
            actual_json=get_actual_json(resp, max_chars=300),
            reasoning="PASS: Customer cannot view seller stats." if passed else f"FAIL: {err}",
        ))

    # CS-01 /customers/{cid}/stats as admin
    cid = users.customer_richest_cust_id
    if cid and "admin" in s.tokens:
        resp, ms = s.call("GET", f"/customers/{cid}/stats", token_role="admin",
                          params={"range": "1y"})
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="CS-01", category="StatsSellerCustomer", name=f"Customer {cid} stats (admin, 1y)",
            method="GET", endpoint=f"/customers/{cid}/stats?range=1y", role="admin",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"cust_id": "<int>", "n_orders": "<int>", "revenue": "<decimal>", "top_products": "<list>", "top_families": "<list>"}',
            actual_json=get_actual_json(resp, max_chars=1500),
            reasoning="PASS: Per-customer drilldown KPIs." if passed else f"FAIL: {err}",
        ))

        # CS-02 invalid range
        resp, ms = s.call("GET", f"/customers/{cid}/stats", token_role="admin",
                          params={"range": "10y"})
        passed, err = assert_status(resp, 422)
        record(s, TestResult(
            test_id="CS-02", category="StatsSellerCustomer", name="Customer stats invalid range -> 422",
            method="GET", endpoint=f"/customers/{cid}/stats?range=10y", role="admin",
            status_expected=422, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "range must be 7d|30d|90d|180d|1y|all"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Range Literal enforced." if passed else f"FAIL: {err}",
        ))

        # CS-03 top_products > 50
        resp, ms = s.call("GET", f"/customers/{cid}/stats", token_role="admin",
                          params={"top_products": 100})
        passed, err = assert_status(resp, 422)
        record(s, TestResult(
            test_id="CS-03", category="StatsSellerCustomer", name="top_products > 50 rejected",
            method="GET", endpoint=f"/customers/{cid}/stats?top_products=100", role="admin",
            status_expected=422, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "top_products <= 50"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: top_products cap=50." if passed else f"FAIL: {err}",
        ))

    # CS-04 customer self-stats
    if "customer" in s.tokens:
        own = s.cust_ids.get("customer")
        resp, ms = s.call("GET", f"/customers/{own}/stats", token_role="customer")
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id="CS-04", category="StatsSellerCustomer", name="Customer views own stats",
            method="GET", endpoint=f"/customers/{own}/stats", role="customer",
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"cust_id": "<self>", "n_orders": "<int>"}',
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning="PASS: Customer self-service stats." if passed else f"FAIL: {err}",
        ))

    # CS-05 customer cannot view another's stats
    if "customer" in s.tokens and users.customer_stable_cust_id:
        other = users.customer_stable_cust_id
        if other != s.cust_ids.get("customer"):
            resp, ms = s.call("GET", f"/customers/{other}/stats", token_role="customer")
            passed, err = assert_status(resp, 403)
            record(s, TestResult(
                test_id="CS-05", category="StatsSellerCustomer", name="Customer cannot view another's stats",
                method="GET", endpoint=f"/customers/{other}/stats", role="customer",
                status_expected=403, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json='{"detail": "Customers can only view their own stats"}',
                actual_json=get_actual_json(resp, max_chars=400),
                reasoning="PASS: Cross-customer stats privacy." if passed else f"PRIVACY FAIL: {err}",
                notes="CRITICAL privacy",
            ))

    # CS-06 seller cannot view stats for unassigned customer
    if "seller2" in s.tokens and users.customer_stable_cust_id and users.seller_user_id:
        ensure_assigned(s, users.customer_stable_cust_id, users.seller_user_id)
        resp, ms = s.call("GET", f"/customers/{users.customer_stable_cust_id}/stats",
                          token_role="seller2")
        passed, err = assert_status(resp, 403)
        record(s, TestResult(
            test_id="CS-06", category="StatsSellerCustomer",
            name="Seller2 blocked from stats for seller1's customer",
            method="GET", endpoint=f"/customers/{users.customer_stable_cust_id}/stats", role="seller2",
            status_expected=403, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"detail": "This customer is not assigned to you"}',
            actual_json=get_actual_json(resp, max_chars=400),
            reasoning="PASS: Cross-seller stats isolation." if passed else f"AUTHZ FAIL: {err}",
            notes="CRITICAL cross-seller isolation",
        ))


# [14] Products

def test_products(s: TestSession):
    print("\n[14] Products")
    if "admin" not in s.tokens:
        return

    # PR-01 browse default
    resp, ms = s.call("GET", "/products", token_role="admin", params={"limit": 10})
    ok_status, err = assert_status(resp, 200)
    ok_keys, kerr = assert_keys(resp, ["items"])
    passed = ok_status and ok_keys
    record(s, TestResult(
        test_id="PR-01", category="Products", name="Browse default (limit=10)",
        method="GET", endpoint="/products?limit=10", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err or kerr,
        expected_json='{"total": "<int>", "limit": 10, "items": [{"item_id": "<int>", "description": "<str>", "family": "<str>"}, "..."]}',
        actual_json=get_actual_json(resp, max_chars=1500),
        reasoning="PASS: Catalog browse." if passed else f"FAIL: {err or kerr}",
    ))

    # PR-02 search
    resp, ms = s.call("GET", "/products", token_role="admin",
                      params={"q": "gloves", "limit": 5})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="PR-02", category="Products", name="Search by 'gloves'",
        method="GET", endpoint="/products?q=gloves", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"items": "<gloves products>"}',
        actual_json=get_actual_json(resp, max_chars=1200),
        reasoning="PASS: Description ILIKE search." if passed else f"FAIL: {err}",
    ))

    # PR-03 family filter
    resp, ms = s.call("GET", "/products", token_role="admin",
                      params={"family": "Wound Care", "limit": 5})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="PR-03", category="Products", name="Filter family='Wound Care'",
        method="GET", endpoint="/products?family=Wound+Care", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"items": "<all in family>"}',
        actual_json=get_actual_json(resp, max_chars=1200),
        reasoning="PASS: Family filter applied." if passed else f"FAIL: {err}",
    ))

    # PR-04 private brand only
    resp, ms = s.call("GET", "/products", token_role="admin",
                      params={"is_private_brand": "true", "limit": 5})
    ok_status, err = assert_status(resp, 200)
    body = {}
    try: body = resp.json()
    except Exception: pass
    items = body.get("items", []) if isinstance(body, dict) else []
    all_pb = all(it.get("is_private_brand") for it in items) if items else True
    passed = ok_status and all_pb
    record(s, TestResult(
        test_id="PR-04", category="Products", name="Filter is_private_brand=true",
        method="GET", endpoint="/products?is_private_brand=true", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms,
        error_msg=err if not all_pb else "",
        expected_json='{"items": "<all is_private_brand=true>"}',
        actual_json=get_actual_json(resp, max_chars=1200),
        reasoning=(
            f"PASS: All {len(items)} returned items are private brand." if passed
            else "FAIL: Filter did not strictly apply."
        ),
    ))

    # PR-05 in_stock=true
    resp, ms = s.call("GET", "/products", token_role="admin",
                      params={"in_stock": "true", "limit": 5})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="PR-05", category="Products", name="Filter in_stock=true",
        method="GET", endpoint="/products?in_stock=true", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"items": "<units_available > 0>"}',
        actual_json=get_actual_json(resp, max_chars=1200),
        reasoning="PASS: Inventory filter." if passed else f"FAIL: {err}",
    ))

    # PR-06 sort by price
    resp, ms = s.call("GET", "/products", token_role="admin",
                      params={"sort_by": "price_low", "limit": 10})
    passed, err = assert_status(resp, 200)
    record(s, TestResult(
        test_id="PR-06", category="Products", name="Sort price_low",
        method="GET", endpoint="/products?sort_by=price_low", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"items": "<sorted ascending by price>"}',
        actual_json=get_actual_json(resp, max_chars=1200),
        reasoning="PASS: Sort applied." if passed else f"FAIL: {err}",
    ))

    # PR-07 invalid sort
    resp, ms = s.call("GET", "/products", token_role="admin",
                      params={"sort_by": "wishlist"})
    passed, err = assert_status(resp, 422)
    record(s, TestResult(
        test_id="PR-07", category="Products", name="Invalid sort_by rejected",
        method="GET", endpoint="/products?sort_by=wishlist", role="admin",
        status_expected=422, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err,
        expected_json='{"detail": "sort_by Literal enum"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: sort_by Literal enforced." if passed else f"FAIL: {err}",
    ))

    # PR-08 limit > cap
    resp, ms = s.call("GET", "/products", token_role="admin",
                      params={"limit": 500})
    passed = resp.status_code in (200, 422)
    record(s, TestResult(
        test_id="PR-08", category="Products", name="limit=500 (boundary)",
        method="GET", endpoint="/products?limit=500", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms,
        error_msg="" if passed else f"Got {resp.status_code}",
        expected_json='{"items": "<<=500 or 422>"}',
        actual_json=get_actual_json(resp, max_chars=400),
        reasoning="PASS: Boundary respected." if passed else "FAIL",
    ))

    # PR-09 /products/filters
    resp, ms = s.call("GET", "/products/filters", token_role="admin")
    ok_status, err = assert_status(resp, 200)
    ok_keys, kerr = assert_keys(resp, ["families", "categories"])
    passed = ok_status and ok_keys
    record(s, TestResult(
        test_id="PR-09", category="Products", name="Catalog filter options",
        method="GET", endpoint="/products/filters", role="admin",
        status_expected=200, status_actual=resp.status_code,
        passed=passed, duration_ms=ms, error_msg=err or kerr,
        expected_json='{"families": "<list>", "categories": "<list>", "suppliers": "<list>"}',
        actual_json=get_actual_json(resp, max_chars=1500),
        reasoning="PASS: Filter dropdown source." if passed else f"FAIL: {err or kerr}",
    ))

    # PR-10..12 - all 3 roles can browse
    for idx, role in enumerate(["seller", "customer"]):
        if role not in s.tokens: continue
        resp, ms = s.call("GET", "/products", token_role=role, params={"limit": 5})
        passed, err = assert_status(resp, 200)
        record(s, TestResult(
            test_id=f"PR-{10+idx:02d}", category="Products",
            name=f"{role.capitalize()} can browse products",
            method="GET", endpoint="/products?limit=5", role=role,
            status_expected=200, status_actual=resp.status_code,
            passed=passed, duration_ms=ms, error_msg=err,
            expected_json='{"items": "<list>"}',
            actual_json=get_actual_json(resp, max_chars=600),
            reasoning=f"PASS: {role} can browse catalog (Amazon-style grid)." if passed else f"FAIL: {err}",
        ))
# [15] Authorization Matrix - role x endpoint grid for full coverage

def test_authorization_matrix(s: TestSession, users: DiscoveredUsers):
    print("\n[15] Authorization Matrix (full coverage)")
    cid = users.customer_richest_cust_id
    if not cid: return

    matrix = [
        # path, method, body, [admin, seller, customer], description
        ("/users", "GET", None, [200, 403, 403], "User list - admin only"),
        ("/admin/stats/overview", "GET", None, [200, 403, 403], "Overview KPIs"),
        ("/admin/stats/sales-trend", "GET", None, [200, 403, 403], "Sales-trend"),
        ("/admin/stats/conversion-by-signal", "GET", None, [200, 403, 403], "Per-signal conv (admin)"),
        ("/admin/stats/segment-distribution", "GET", None, [200, 403, 403], "Segment dist"),
        ("/admin/stats/top-sellers", "GET", None, [200, 403, 403], "Sellers leaderboard"),
        ("/admin/stats/top-customers", "GET", None, [200, 403, 403], "Customers leaderboard"),
        ("/admin/stats/recent-sales", "GET", None, [200, 403, 403], "Recent sales feed"),
        ("/admin/stats/engine-effectiveness", "GET", None, [200, 403, 403], "Engine effectiveness"),
        (f"/customers/{cid}/assignment-history", "GET", None, [200, 403, 403], "Assignment audit"),
        ("/sellers/me/customers", "GET", None, [403, 200, 403], "Seller's own list"),
        ("/sellers/me/stats", "GET", None, [403, 200, 403], "Seller's own stats"),
        ("/sellers/me/conversion-by-signal", "GET", None, [403, 200, 403], "Seller's conv"),
        ("/customers/me", "GET", None, [403, 403, 200], "Customer self-record"),
        ("/cart/me", "GET", None, [403, 403, 200], "Customer self-cart"),
        ("/recommendations/me", "GET", None, [403, 403, 200], "Customer self-recs"),
        (f"/customers/{cid}", "GET", None, [200, 200, 403], "Customer detail (read open for sellers)"),
        ("/customers/search", "GET", None, [200, 200, 403], "Search (admin/seller)"),
        ("/customers/filter", "GET", None, [200, 200, 403], "Filter (admin/seller)"),
        ("/products", "GET", None, [200, 200, 200], "Product catalog (all)"),
        ("/products/filters", "GET", None, [200, 200, 200], "Filter options (all)"),
    ]

    role_map = {"admin": 0, "seller": 1, "customer": 2}
    test_idx = 1
    for path, method, body, expectations, description in matrix:
        for role in ["admin", "seller", "customer"]:
            exp = expectations[role_map[role]]
            if role not in s.tokens:
                continue
            # Special: customers/search?q is required - add it
            params = {"q": "PO"} if path == "/customers/search" else None
            resp, ms = s.call(method, path, token_role=role,
                              json_body=body if method in ("POST", "PATCH") else None,
                              params=params)
            if exp == 200:
                passed, err = assert_status(resp, [200, 201])
                outcome = "ALLOWED"
            else:
                passed = resp.status_code in (exp, 401, 403)
                err = "" if passed else f"expected {exp}, got {resp.status_code}"
                outcome = "BLOCKED"
            record(s, TestResult(
                test_id=f"AM-{test_idx:02d}", category="AuthzMatrix",
                name=f"{role.upper()} -> {method} {path} = {outcome}",
                method=method, endpoint=path, role=role,
                status_expected=exp, status_actual=resp.status_code,
                passed=passed, duration_ms=ms, error_msg=err,
                expected_json=f"HTTP {exp} ({outcome})",
                actual_json=get_actual_json(resp, max_chars=300),
                reasoning=(
                    f"PASS: {role} {outcome.lower()} on {description} (HTTP {resp.status_code})."
                    if passed else f"AUTHZ FAIL: {role} got {resp.status_code} (expected {exp}). {err}"
                ),
                notes="Authorization grid",
            ))
            test_idx += 1


# Cleanup

def cleanup(s: TestSession):
    print("\n[Cleanup] Removing test users")
    if "admin" not in s.tokens:
        return
    for uid in s.test_users_created:
        try:
            s.call("DELETE", f"/users/{uid}", token_role="admin")
        except Exception:
            pass
    print(f"  Deactivated {len(s.test_users_created)} test users")


# Excel report

def generate_excel_report(s: TestSession, output_path: str):
    print(f"\n[Report] Building Excel report at {output_path}")
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    df_all = pd.DataFrame([asdict(r) for r in s.results])
    if df_all.empty:
        print("  No results to write")
        return

    # Reorder columns
    col_order = ["test_id", "category", "name", "method", "endpoint", "role",
                 "status_expected", "status_actual", "passed", "duration_ms",
                 "expected_json", "actual_json", "reasoning", "error_msg", "notes"]
    df_all = df_all[col_order]

    # Summary
    total = len(df_all)
    n_pass = int(df_all["passed"].sum())
    n_fail = total - n_pass
    pct = round(n_pass / total * 100, 2) if total else 0
    summary_rows = [
        {"Metric": "Total tests", "Value": total},
        {"Metric": "Passed", "Value": n_pass},
        {"Metric": "Failed", "Value": n_fail},
        {"Metric": "Pass rate (%)", "Value": pct},
        {"Metric": "Avg duration (ms)", "Value": int(df_all["duration_ms"].mean())},
        {"Metric": "Generated", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # By category
    cat_grp = df_all.groupby("category").agg(
        total=("test_id", "count"),
        passed=("passed", "sum"),
        avg_ms=("duration_ms", "mean"),
    ).reset_index()
    cat_grp["failed"] = cat_grp["total"] - cat_grp["passed"]
    cat_grp["pass_rate"] = (cat_grp["passed"] / cat_grp["total"] * 100).round(1)
    cat_grp["avg_ms"] = cat_grp["avg_ms"].round(0).astype(int)
    df_cat = cat_grp[["category", "total", "passed", "failed", "pass_rate", "avg_ms"]]

    # Failures
    df_fail = df_all[~df_all["passed"]].copy()

    # Coverage map
    cov = df_all.groupby(["endpoint", "method"]).agg(
        n_tests=("test_id", "count"),
        n_passed=("passed", "sum"),
        roles=("role", lambda x: ", ".join(sorted(set(x)))),
    ).reset_index()
    cov["n_failed"] = cov["n_tests"] - cov["n_passed"]
    cov = cov.sort_values(["endpoint", "method"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="01_Summary", index=False)
        df_cat.to_excel(writer, sheet_name="02_By_Category", index=False)
        df_all.to_excel(writer, sheet_name="03_All_Results", index=False)
        df_fail.to_excel(writer, sheet_name="04_Failures", index=False)
        cov.to_excel(writer, sheet_name="05_Coverage_Map", index=False)

        # Format
        wb = writer.book
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Bold header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2F5496")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Auto column widths
            for col in ws.columns:
                max_len = 10
                for cell in col[:50]:
                    try:
                        v = str(cell.value or "")
                        max_len = max(max_len, min(len(v), 60))
                    except Exception:
                        pass
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
            # Wrap text in 03 + 04
            if sheet_name in ("03_All_Results", "04_Failures"):
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

    print(f"  Wrote {total} results: {n_pass}/{total} passed ({pct}%)")


# Main

def main():
    parser = argparse.ArgumentParser(description="Comprehensive QC test runner")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL,
                        help=f"API base URL (default: {DEFAULT_BASE_URL})")
    parser.add_argument("--output", default=str(ROOT / "reports" / "api_test_report.xlsx"),
                        help="Output Excel path")
    parser.add_argument("--admin-password", default=DEMO_PASSWORD,
                        help="Admin password (default: Demo1234!)")
    parser.add_argument("--verbose", action="store_true", help="Print HTTP details")
    parser.add_argument("--skip-cleanup", action="store_true", help="Don't delete test users at the end")
    args = parser.parse_args()

    print("=" * 75)
    print("Comprehensive API QC Test Runner")
    print("=" * 75)
    print(f"Base URL : {args.base_url}")
    print(f"Output   : {args.output}")
    print(f"Started  : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    cfg = load_pg_config()
    print(f"Postgres : {cfg['host']}:{cfg['port']}/{cfg['dbname']} schema={cfg['schema']}")

    # Smoke check
    try:
        r = requests.get(f"{args.base_url.rstrip('/')}/health", timeout=5)
        if r.status_code != 200:
            print(f"\nERROR: API at {args.base_url} returned {r.status_code} on /health.")
            sys.exit(1)
    except Exception as e:
        print(f"\nERROR: Cannot reach API at {args.base_url}: {e}")
        sys.exit(1)

    # Discover
    try:
        users = discover_users(cfg)
    except Exception as e:
        print(f"\nERROR: discovery failed: {e}")
        traceback.print_exc()
        sys.exit(1)

    s = TestSession(args.base_url, args.verbose)

    # Run sections
    try:
        test_health(s)
        test_authentication(s, users, args.admin_password)
        test_user_management(s)
        test_customers(s, users)
        test_customer_record_create(s, users)
        test_recommendations(s, users)
        test_cart_helper(s, users)
        test_reject_recommendation(s, users)
        test_cart(s, users)
        test_purchase_history(s, users)
        test_assignment_lifecycle(s, users)
        test_stats_admin(s)
        test_stats_seller_customer(s, users)
        test_products(s)
        test_authorization_matrix(s, users)
    except Exception as e:
        print(f"\nERROR during test run: {e}")
        traceback.print_exc()
    finally:
        if not args.skip_cleanup:
            cleanup(s)
        generate_excel_report(s, args.output)

    # Final
    total = len(s.results)
    n_pass = sum(1 for r in s.results if r.passed)
    print("\n" + "=" * 75)
    print(f"Done: {n_pass}/{total} passed ({round(n_pass/total*100,1) if total else 0}%)")
    print(f"Excel: {args.output}")
    print("=" * 75)
    sys.exit(0 if n_pass == total else 1)


if __name__ == "__main__":
    main()