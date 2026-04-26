from __future__ import annotations

import sys
import time
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# Paths

ROOT         = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN   = ROOT / "data_clean"
FEATURE_FILE = DATA_CLEAN / "features"  / "customer_features.parquet"
PRECOMP_DIR  = DATA_CLEAN / "serving"   / "precomputed"
OUT_ANALYSIS = DATA_CLEAN / "analysis"
OUT_CHARTS   = OUT_ANALYSIS / "charts"


# Configuration

# Market type display labels
MKT_CD_LABELS = {
    "PO":  "Physician Office",
    "LTC": "Long Term Care",
    "SC":  "Surgery Center / Specialty",
    "LC":  "Lab / Clinical",
    "HC":  "Home Care",
    "AC":  "Acute Care",
}
MKT_CD_OTHER = "OTHER"

# Size tier ordering for consistent display
SIZE_TIER_ORDER = ["new", "small", "mid", "large", "enterprise"]

# Size tier display labels
SIZE_TIER_LABELS = {
    "new":        "New Customer",
    "small":      "Small",
    "mid":        "Mid",
    "large":      "Large",
    "enterprise": "Enterprise",
}

# Product family preferences per market type.
# Used for the segmentation_report.xlsx strategy sheet only — informational,
# not for scoring. Scoring logic lives exclusively in recommendation_factors.py.
PRIMARY_FAMILIES = {
    "PO":  ["Nursing and Surgical Supplies", "Infection Prevention",
            "Wound Care & Skin Care", "Rx", "Lab-Waived Lab",
            "Office and Facility Supplies"],
    "LTC": ["Incontinence", "Nursing and Surgical Supplies",
            "Nutrition and Feeding Supplies", "Wound Care & Skin Care",
            "Patient Therapy/Personal Care", "Infection Prevention"],
    "SC":  ["Wound Care & Skin Care", "Infection Prevention",
            "Nursing and Surgical Supplies", "Equipment & Equip Disposables",
            "Office and Facility Supplies", "Rx"],
    "LC":  ["Lab-Waived Lab", "Lab-Non-Waived Lab",
            "Lab-Ancillary Lab Products", "Nursing and Surgical Supplies",
            "Infection Prevention"],
    "HC":  ["Patient Therapy/Personal Care", "Nursing and Surgical Supplies",
            "Incontinence", "Wound Care & Skin Care",
            "Nutrition and Feeding Supplies"],
    "AC":  ["Nursing and Surgical Supplies", "Infection Prevention",
            "Wound Care & Skin Care", "Rx", "Equipment & Equip Disposables",
            "Lab-Waived Lab", "Respiratory Products"],
}

# Scoring weights per size tier.
# Larger customers have more budget for cross-sell, smaller customers rely
# more on staying on file with lapsed reorders.
# These MUST match the values in recommendation_factors.py SEGMENT_WEIGHTS.
SCORING_REFERENCE = {
    "_new": {
        "peer_gap": 2.0, "lapsed": 1.0,
        "rationale": "New customer — stick to safe high-adoption items, no aggressive pitches"
    },
    "_small": {
        "peer_gap": 2.0, "lapsed": 3.0,
        "rationale": "Small customer — limited budget, prioritize staying on file via lapsed reorders"
    },
    "_mid": {
        "peer_gap": 2.5, "lapsed": 2.5,
        "rationale": "Mid-size customer — balanced cross-sell and reorders"
    },
    "_large": {
        "peer_gap": 3.0, "lapsed": 2.0,
        "rationale": "Large customer — budget available, cross-sell new categories"
    },
    "_enterprise": {
        "peer_gap": 3.5, "lapsed": 1.5,
        "rationale": "Enterprise customer — aggressive cross-sell, expansion focus"
    },
}
PRIVATE_BRAND_BOOST = 0.5

# Minimum segment size below which we flag as too small for reliable peer stats.
# Segments below this still exist but the recommendation engine will produce
# fewer recommendations for them (MIN_PEER_SIZE in recommendation_factors.py).
MIN_SEGMENT_SIZE = 5


# Logging

def _s(title):
    print(f"\n{'-' * 64}\n  {title}\n{'-' * 64}", flush=True)


def _log(msg):
    print(f"  {msg}", flush=True)


# Excel helpers

def _style(ws, df, hc="1F4E79"):
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, 12), 55)


# Step 1: Load features

def load_features():
    _s("Step 1: Loading customer features")
    if not FEATURE_FILE.exists():
        print(f"\nFATAL: customer_features.parquet not found", file=sys.stderr)
        sys.exit(1)
    df = pd.read_parquet(FEATURE_FILE)
    _log(f"Loaded : {len(df):,} customers  |  {df.shape[1]} columns")
    required = [
        "DIM_CUST_CURR_ID", "MKT_CD", "size_tier",
        "monetary", "recency_days", "frequency",
        "avg_revenue_per_order", "n_categories_bought",
        "category_hhi", "cycle_regularity",
        "median_monthly_spend", "active_months_last_12",
        "affordability_ceiling",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"\nFATAL: Missing columns: {missing}", file=sys.stderr)
        print("Run clean_data.py first to generate size_tier and affordability_ceiling.",
              file=sys.stderr)
        sys.exit(1)
    return df


# Step 2: Assign segments

def assign_segments(df):
    # Build segment label as MKT_CD_sizeTier (e.g. PO_small, LTC_enterprise).
    # MKT_CD values not in the known list fall back to OTHER.
    # size_tier is already computed in clean_data.py Step 6c.

    _s("Step 2: Assigning segments — MKT_CD x size tier")
    df = df.copy()
    df["mkt_cd_clean"] = df["MKT_CD"].fillna(MKT_CD_OTHER).str.strip().str.upper()
    df["mkt_cd_clean"] = df["mkt_cd_clean"].apply(
        lambda x: x if x in MKT_CD_LABELS else MKT_CD_OTHER
    )

    # size_tier already exists from clean_data.py — use as-is
    df["segment"] = df["mkt_cd_clean"] + "_" + df["size_tier"]

    desc_map = {
        "new":        "New customer — safe high-adoption items only",
        "small":      "Small customer — prioritize lapsed reorders",
        "mid":        "Mid-size customer — balanced cross-sell and reorders",
        "large":      "Large customer — cross-sell new categories",
        "enterprise": "Enterprise customer — aggressive cross-sell and expansion",
    }
    df["segment_description"] = df["size_tier"].map(desc_map)

    seg_counts = (
        df.groupby(["mkt_cd_clean", "size_tier"])
        .size()
        .reset_index(name="n")
    )
    seg_counts["segment"] = seg_counts["mkt_cd_clean"] + "_" + seg_counts["size_tier"]
    seg_counts = seg_counts.sort_values("n", ascending=False)

    _log(f"\n  {'Segment':<30} {'Customers':>10} {'% Portfolio':>12}")
    _log(f"  {'-' * 30} {'-' * 10} {'-' * 12}")
    for _, row in seg_counts.iterrows():
        pct = row["n"] / len(df) * 100
        warn = "  <-- small" if row["n"] < MIN_SEGMENT_SIZE else ""
        _log(f"  {row['segment']:<30} {row['n']:>10,} {pct:>11.1f}%{warn}")

    n_small_segs = (seg_counts["n"] < MIN_SEGMENT_SIZE).sum()
    if n_small_segs > 0:
        _log(f"\n  Warning: {n_small_segs} segment(s) have fewer than {MIN_SEGMENT_SIZE} customers.")
        _log(f"  These will produce fewer recommendations due to MIN_PEER_SIZE threshold.")

    if "churn_label" in df.columns:
        churn_by_seg = (
            df[df["churn_label"].isin([0, 1])]
            .groupby("segment")["churn_label"]
            .mean().mul(100).round(1)
        )
        _log(f"\n  Churn rate by segment (top 10 highest risk):")
        for seg, rate in churn_by_seg.sort_values(ascending=False).head(10).items():
            _log(f"    {seg:<30} {rate:.1f}%")
    return df


# Step 3: Build segment profiles

def build_segment_profiles(df):
    # Aggregate descriptive statistics per segment for the team report.
    # This is read-only analysis — no scoring happens here.

    _s("Step 3: Building segment profiles")
    agg = (
        df.groupby("segment")
        .agg(
            mkt_cd                = ("mkt_cd_clean",          "first"),
            size_tier             = ("size_tier",             "first"),
            n_customers           = ("DIM_CUST_CURR_ID",      "count"),
            median_monetary       = ("monetary",              "median"),
            median_monthly_spend  = ("median_monthly_spend",  "median"),
            median_recency        = ("recency_days",          "median"),
            median_frequency      = ("frequency",             "median"),
            median_avg_order      = ("avg_revenue_per_order", "median"),
            median_afford_ceiling = ("affordability_ceiling", "median"),
            median_active_months  = ("active_months_last_12", "median"),
            median_n_cats         = ("n_categories_bought",   "median"),
            median_hhi            = ("category_hhi",          "median"),
            median_cycle          = ("cycle_regularity",      "median"),
        )
        .reset_index()
    )
    if "churn_label" in df.columns:
        churn = (
            df[df["churn_label"].isin([0, 1])]
            .groupby("segment")["churn_label"]
            .mean().mul(100).round(1).reset_index()
            .rename(columns={"churn_label": "churn_rate_pct"})
        )
        agg = agg.merge(churn, on="segment", how="left")

    # Attach scoring weights from the reference table (size-tier-based)
    def _tier_weights(tier, side):
        key = "_" + tier
        return SCORING_REFERENCE.get(key, SCORING_REFERENCE["_mid"])[side]

    agg["peer_gap_weight"] = agg["size_tier"].apply(lambda t: _tier_weights(t, "peer_gap"))
    agg["lapsed_weight"]   = agg["size_tier"].apply(lambda t: _tier_weights(t, "lapsed"))
    agg["primary_signal"]  = agg.apply(
        lambda r: "peer_gap" if r["peer_gap_weight"] >= r["lapsed_weight"] else "lapsed",
        axis=1
    )
    agg["mkt_label"] = agg["mkt_cd"].map(lambda x: MKT_CD_LABELS.get(x, "Other"))

    # Sort for display: market then size tier in defined order
    agg["_tier_order"] = agg["size_tier"].map({t: i for i, t in enumerate(SIZE_TIER_ORDER)})
    agg = agg.sort_values(["mkt_cd", "_tier_order"]).drop(columns=["_tier_order"]).reset_index(drop=True)

    _log(f"\n  {'Segment':<30} {'N':>8} {'Churn%':>7} {'Median$/mo':>11} {'$/order':>9} {'Freq':>6}")
    _log(f"  {'-' * 30} {'-' * 8} {'-' * 7} {'-' * 11} {'-' * 9} {'-' * 6}")
    for _, r in agg.iterrows():
        churn = f"{r.get('churn_rate_pct', 0):.1f}%" if pd.notna(r.get("churn_rate_pct")) else "n/a"
        _log(f"  {r['segment']:<30} {r['n_customers']:>8,} {churn:>7} "
             f"${r['median_monthly_spend']:>10,.0f} ${r['median_avg_order']:>8,.0f} "
             f"{r['median_frequency']:>6.0f}")
    return agg


# Step 4: Save outputs

def save_outputs(df, profiles):
    # Write the segment parquet (consumed by recommendation_factors.py) and
    # the team-facing Excel report. No scoring happens here.

    _s("Step 4: Saving outputs")
    PRECOMP_DIR.mkdir(parents=True, exist_ok=True)
    OUT_ANALYSIS.mkdir(parents=True, exist_ok=True)

    # customer_segments.parquet — one row per customer, minimal columns
    seg_out = df[[
        "DIM_CUST_CURR_ID", "segment", "mkt_cd_clean", "size_tier"
    ]].copy()

    # Type safety for downstream joins
    seg_out["DIM_CUST_CURR_ID"] = seg_out["DIM_CUST_CURR_ID"].astype("int64")
    for c in ["segment", "mkt_cd_clean", "size_tier"]:
        seg_out[c] = seg_out[c].fillna("").astype(str)

    seg_path = PRECOMP_DIR / "customer_segments.parquet"
    seg_out.to_parquet(seg_path, index=False)
    _log(f"Saved: {seg_path.relative_to(ROOT)}")

    # Strategy-by-segment reference table (size tier x market, informational)
    strat_rows = []
    for mkt_cd in list(MKT_CD_LABELS.keys()) + [MKT_CD_OTHER]:
        label = MKT_CD_LABELS.get(mkt_cd, "Other")
        fams = PRIMARY_FAMILIES.get(mkt_cd, [])
        for tier in SIZE_TIER_ORDER:
            key = "_" + tier
            w = SCORING_REFERENCE[key]
            strat_rows.append({
                "segment":           f"{mkt_cd}_{tier}",
                "customer_type":     label,
                "size_tier":         tier,
                "size_tier_label":   SIZE_TIER_LABELS[tier],
                "peer_gap_weight":   w["peer_gap"],
                "lapsed_weight":     w["lapsed"],
                "primary_signal":    "peer_gap" if w["peer_gap"] >= w["lapsed"] else "lapsed",
                "scoring_rationale": w["rationale"],
                "primary_families":  ", ".join(fams[:4]),
                "medline_rule":      "Never recommend Medline — substitution framing for medline_only customers",
            })
    strat_df = pd.DataFrame(strat_rows)

    xlsx_path = OUT_ANALYSIS / "segmentation_report.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        profiles.to_excel(writer, sheet_name="01_segment_profiles", index=False)
        strat_df.to_excel(writer, sheet_name="02_strategy_by_segment", index=False)
        wb = writer.book
        for name, color, df_out in [
            ("01_segment_profiles",   "1F4E79", profiles),
            ("02_strategy_by_segment", "833C00", strat_df),
        ]:
            if name in wb.sheetnames:
                wb[name].sheet_properties.tabColor = color
                _style(writer.sheets[name], df_out, hc=color)
    _log(f"Saved: {xlsx_path.relative_to(ROOT)}")


# Step 5: Save PNG charts

def save_charts(df, profiles):
    # Writes 4 PNG charts to data_clean/analysis/charts/.
    # Chart 1: size_tier distribution
    # Chart 2: market code distribution
    # Chart 3: market x size heatmap
    # Chart 4: median monthly spend per size tier

    _s("Step 5: Saving PNG charts")
    OUT_CHARTS.mkdir(parents=True, exist_ok=True)

    # Chart 1: size_tier distribution
    tier_counts = (
        df["size_tier"].value_counts()
        .reindex(SIZE_TIER_ORDER, fill_value=0)
    )
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = ["#888888", "#375623", "#1F4E79", "#7030A0", "#C00000"]
    bars = ax.bar(
        [SIZE_TIER_LABELS[t] for t in tier_counts.index],
        tier_counts.values,
        color=colors
    )
    for bar, count in zip(bars, tier_counts.values):
        pct = count / tier_counts.sum() * 100
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            f"{count:,}\n({pct:.1f}%)",
            ha="center", va="bottom", fontsize=9
        )
    ax.set_title("Customer count by size tier", fontsize=13, fontweight="bold")
    ax.set_ylabel("Number of customers")
    ax.set_ylim(0, tier_counts.max() * 1.15)
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    path = OUT_CHARTS / "01_size_tier_distribution.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    _log(f"Saved: {path.relative_to(ROOT)}")

    # Chart 2: market code distribution
    mkt_counts = df["mkt_cd_clean"].value_counts()
    mkt_labels = [
        f"{code}\n{MKT_CD_LABELS.get(code, 'Other')}"
        for code in mkt_counts.index
    ]
    fig, ax = plt.subplots(figsize=(12, 6))
    bars = ax.bar(mkt_labels, mkt_counts.values, color="#1F4E79")
    for bar, count in zip(bars, mkt_counts.values):
        pct = count / mkt_counts.sum() * 100
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            f"{count:,}\n({pct:.1f}%)",
            ha="center", va="bottom", fontsize=9
        )
    ax.set_title("Customer count by market code", fontsize=13, fontweight="bold")
    ax.set_ylabel("Number of customers")
    ax.set_ylim(0, mkt_counts.max() * 1.15)
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    path = OUT_CHARTS / "02_market_distribution.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    _log(f"Saved: {path.relative_to(ROOT)}")

    # Chart 3: market x size heatmap
    pivot = (
        df.groupby(["mkt_cd_clean", "size_tier"])
        .size()
        .unstack(fill_value=0)
        .reindex(columns=SIZE_TIER_ORDER, fill_value=0)
    )

    # Sort markets by total customer count for readability
    pivot = pivot.loc[pivot.sum(axis=1).sort_values(ascending=False).index]

    fig, ax = plt.subplots(figsize=(10, 7))
    im = ax.imshow(pivot.values, aspect="auto", cmap="Blues")

    # Annotations
    for i in range(pivot.shape[0]):
        for j in range(pivot.shape[1]):
            val = pivot.values[i, j]
            text_color = "white" if val > pivot.values.max() * 0.5 else "black"
            ax.text(
                j, i, f"{val:,}",
                ha="center", va="center",
                color=text_color, fontsize=9
            )

    ax.set_xticks(range(len(SIZE_TIER_ORDER)))
    ax.set_xticklabels([SIZE_TIER_LABELS[t] for t in SIZE_TIER_ORDER], rotation=0)
    ax.set_yticks(range(len(pivot.index)))
    ax.set_yticklabels(pivot.index)
    ax.set_title("Customer count by market code x size tier",
                 fontsize=13, fontweight="bold")
    ax.set_xlabel("Size tier")
    ax.set_ylabel("Market code")

    cbar = plt.colorbar(im, ax=ax)
    cbar.set_label("Number of customers")

    plt.tight_layout()
    path = OUT_CHARTS / "03_market_x_size_heatmap.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    _log(f"Saved: {path.relative_to(ROOT)}")

    # Chart 4: median monthly spend per size tier
    spend_by_tier = (
        df.groupby("size_tier")["median_monthly_spend"]
        .median()
        .reindex(SIZE_TIER_ORDER)
    )
    fig, ax = plt.subplots(figsize=(10, 6))
    bars = ax.bar(
        [SIZE_TIER_LABELS[t] for t in spend_by_tier.index],
        spend_by_tier.values,
        color=colors
    )
    for bar, val in zip(bars, spend_by_tier.values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            f"${val:,.0f}",
            ha="center", va="bottom", fontsize=10, fontweight="bold"
        )
    ax.set_title(
        "Median monthly spend by size tier (validates tier definitions)",
        fontsize=13, fontweight="bold"
    )
    ax.set_ylabel("Median monthly spend ($)")
    ax.set_yscale("log")
    ax.grid(axis="y", alpha=0.3, which="both")
    plt.tight_layout()
    path = OUT_CHARTS / "04_median_spend_by_tier.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    _log(f"Saved: {path.relative_to(ROOT)}")


# Main

def main():
    print()
    print("=" * 64)
    print("  CUSTOMER SEGMENTATION — MKT_CD x SIZE TIER")
    print("=" * 64)
    start = time.time()
    PRECOMP_DIR.mkdir(parents=True, exist_ok=True)
    OUT_ANALYSIS.mkdir(parents=True, exist_ok=True)
    OUT_CHARTS.mkdir(parents=True, exist_ok=True)

    features = load_features()
    df       = assign_segments(features)
    profiles = build_segment_profiles(df)
    save_outputs(df, profiles)
    save_charts(df, profiles)

    _s("Complete")
    _log(f"Total time: {round(time.time() - start, 1)}s")
    _log("")
    _log("Segment logic: MKT_CD (PO/LTC/SC/LC/HC/AC/OTHER) x size_tier (new/small/mid/large/enterprise)")
    _log("Scoring weights: defined in recommendation_factors.py SEGMENT_WEIGHTS")
    _log("")
    _log("Key outputs:")
    _log("  customer_segments.parquet        segment per customer (read by recommendation_factors.py)")
    _log("  segmentation_report.xlsx         profiles and strategy per segment")
    _log("  charts/01_size_tier_distribution.png")
    _log("  charts/02_market_distribution.png")
    _log("  charts/03_market_x_size_heatmap.png")
    _log("  charts/04_median_spend_by_tier.png")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise