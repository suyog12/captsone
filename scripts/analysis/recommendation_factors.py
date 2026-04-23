from __future__ import annotations

import json
import sys
import time
import warnings
from pathlib import Path

import duckdb
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import roc_auc_score, classification_report
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore")


# Paths

ROOT         = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN   = ROOT / "data_clean"
MERGED_FILE  = DATA_CLEAN / "serving"  / "merged_dataset.parquet"
PROD_FILE    = DATA_CLEAN / "product"  / "products_clean.parquet"
CUST_FILE    = DATA_CLEAN / "customer" / "customers_clean.parquet"
FEATURE_FILE = DATA_CLEAN / "features" / "customer_features.parquet"
RFM_FILE     = DATA_CLEAN / "features" / "customer_rfm.parquet"
OUT_PRECOMP  = DATA_CLEAN / "serving"  / "precomputed"
OUT_ANALYSIS = DATA_CLEAN / "analysis"

# DuckDB spill directory — used when memory runs out during the big aggregation
SPILL_DIR = DATA_CLEAN / "_duckdb_spill"


# Configuration

ITEM_ID_COL       = "DIM_ITEM_E1_CURR_ID"
ITEM_DSC_COL      = "ITEM_DSC"
PROD_FAMILY_COL   = "PROD_FMLY_LVL1_DSC"
PROD_CATEG_COL    = "PROD_CTGRY_LVL2_DSC"
CUST_ID_COL       = "DIM_CUST_CURR_ID"
REVENUE_COL       = "UNIT_SLS_AMT"
SPCLTY_COL        = "SPCLTY_CD"
PRIVATE_BRAND_COL = "is_private_brand"

FISCAL_YEARS     = ("FY2425", "FY2526")

EXCLUDED_SUPPLIERS = ("MEDLINE", "MEDLINE INDUSTRIES")
EXCLUDED_FAMILIES  = {"Fee", "Unknown", "NaN", "nan", ""}

MIN_PEER_SIZE      = 5
MIN_ADOPTION_PCT   = 0.15
MIN_PEER_PURCHASES = 3

LAPSE_MULTIPLIER = 2.5

TOP_N_RECS = 10

# Scoring weights per segment tier.
#
# Formula for peer_gap score:  2.0 * adoption_rate * peer_w
# Formula for lapsed   score:  lapsed_w
# Private brand bonus:         +0.5 added after the base score
#
# Math validation at peer_gap adoption_rate = 0.50 (typical top-ranked peer product):
#   _high tier: 2 * 0.50 * 3.5 = 3.50 vs lapsed 1.0 -> peer_gap dominates
#   _mid  tier: 2 * 0.50 * 2.5 = 2.50 vs lapsed 2.5 -> balanced (tie at 0.5)
#   _low  tier: 2 * 0.50 * 1.0 = 1.00 vs lapsed 3.5 -> lapsed dominates
#
# For mid-tier customers, products with adoption > 0.50 win as peer_gap,
# products with adoption < 0.50 lose to lapsed. Result: roughly 50/50 split.

SEGMENT_WEIGHTS: dict[str, tuple[float, float]] = {
    "PO_high":  (3.5, 1.0),
    "LTC_high": (3.5, 1.0),
    "SC_high":  (3.5, 1.0),
    "HC_high":  (3.5, 1.0),
    "LC_high":  (3.5, 1.0),
    "AC_high":  (3.5, 1.0),
    "PO_mid":   (2.5, 2.5),
    "LTC_mid":  (2.5, 2.5),
    "SC_mid":   (2.5, 2.5),
    "HC_mid":   (2.5, 2.5),
    "LC_mid":   (2.5, 2.5),
    "AC_mid":   (2.5, 2.5),
    "PO_low":   (1.0, 3.5),
    "LTC_low":  (1.0, 3.5),
    "SC_low":   (1.0, 3.5),
    "HC_low":   (1.0, 3.5),
    "LC_low":   (1.0, 3.5),
    "AC_low":   (1.0, 3.5),
    "OTHER_low": (1.0, 3.5),
    "unknown":   (3.0, 2.0),
}
DEFAULT_WEIGHTS = (3.0, 2.0)
PRIVATE_BRAND_BONUS = 0.5


# Logging

def _s(title: str) -> None:
    print(f"\n{'-'*64}\n  {title}\n{'-'*64}", flush=True)

def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


# Excel helpers

def _style(ws, df: pd.DataFrame, hc: str = "1F4E79") -> None:
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w+2, 12), 55)


def _chart(ws, n, lc, vc, anchor, title, xtitle, color="1F4E79") -> None:
    ch = BarChart()
    ch.type = "bar"; ch.grouping = "clustered"; ch.title = title
    ch.x_axis.title = xtitle; ch.legend = None
    ch.width = 26; ch.height = 16; ch.style = 2
    ch.add_data(Reference(ws, min_col=vc, min_row=1, max_row=n+1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=lc, min_row=2, max_row=n+1))
    ch.series[0].graphicalProperties.solidFill = color
    ch.series[0].graphicalProperties.line.solidFill = color
    ws.add_chart(ch, anchor)


# DuckDB connection with memory-safe settings

def _memory_safe_duckdb() -> duckdb.DuckDBPyConnection:
    # Open DuckDB with conservative memory settings.
    # - 4 GB hard limit (leaves room for pandas + Python + OS)
    # - 1 thread (each thread has its own working memory)
    # - preserve_insertion_order = false (reduces memory overhead)
    # - spill to disk when memory runs out

    SPILL_DIR.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(":memory:")
    con.execute("SET memory_limit = '4GB'")
    con.execute("SET threads = 1")
    con.execute("SET preserve_insertion_order = false")
    con.execute(f"SET temp_directory = '{SPILL_DIR.as_posix()}'")
    return con


# Step 1: Load features and build peer groups

def load_features_and_segments() -> pd.DataFrame:
    # Pull customer features and attach peer_group. Returns ~389k rows.

    _s("Step 1: Loading customer features and building peer groups")

    feat_cols_all = set(
        duckdb.connect().execute(
            f"DESCRIBE SELECT * FROM read_parquet('{FEATURE_FILE.as_posix()}') LIMIT 0"
        ).fetchdf()["column_name"].tolist()
    )
    wanted = [
        "DIM_CUST_CURR_ID", "SPCLTY_CD", "R_score", "F_score",
        "monetary", "recency_days", "frequency",
        "avg_order_gap_days", "specialty_tier", "CUST_TYPE_CD", "MKT_CD",
        "supplier_profile",
        "pct_of_total_revenue", "specialty_revenue_trend_pct",
        "avg_revenue_per_order", "n_categories_bought",
        "category_hhi", "cycle_regularity", "M_score",
    ]
    cols = [c for c in wanted if c in feat_cols_all]
    features = pd.read_parquet(FEATURE_FILE, columns=cols)
    _log(f"Loaded: {len(features):,} customers  |  {len(cols)} columns")

    if "supplier_profile" not in features.columns:
        _log("Warning: supplier_profile missing — defaulting all to 'mixed'")
        features["supplier_profile"] = "mixed"
    else:
        features["supplier_profile"] = features["supplier_profile"].fillna("mixed")

    def rfm_tier(r):
        if r["R_score"] >= 4 and r["F_score"] >= 4:
            return "high"
        elif r["R_score"] <= 2 or r["F_score"] <= 2:
            return "low"
        return "mid"

    features["rfm_tier"]   = features.apply(rfm_tier, axis=1)
    features["peer_group"] = (
        features["SPCLTY_CD"].fillna("UNKNOWN") + "|" + features["rfm_tier"]
    )
    tier3 = features["specialty_tier"] == 3
    features.loc[tier3, "peer_group"] = "TIER3|" + features.loc[tier3, "rfm_tier"]

    peer_sizes = features.groupby("peer_group").size()
    _log(f"Total peer groups       : {features['peer_group'].nunique():,}")
    _log(f"Groups >= {MIN_PEER_SIZE} customers  : {(peer_sizes >= MIN_PEER_SIZE).sum():,}")

    return features


# Step 2: Aggregate customer-product history in DuckDB

def load_customer_product_history(features: pd.DataFrame) -> pd.DataFrame:
    # Two-step strategy for memory safety on 8-16 GB machines:
    #   Step 2a: Write filtered+aggregated transactions to a disk parquet.
    #            Only keep the numeric/ID columns we need for aggregation.
    #            DuckDB streams this and spills to disk when RAM fills up.
    #   Step 2b: Join product metadata (item_dsc, prod_family, etc.)
    #            and customer peer_group from smaller tables.

    _s("Step 2: Aggregating customer-product purchase history (DuckDB, two-pass)")

    # Step 2a: numeric-only aggregation (streams through DuckDB, spills to disk)

    tmp_agg = SPILL_DIR / "_history_agg.parquet"
    SPILL_DIR.mkdir(parents=True, exist_ok=True)

    con = _memory_safe_duckdb()

    desc = con.execute(
        f"DESCRIBE SELECT * FROM read_parquet('{MERGED_FILE.as_posix()}') LIMIT 0"
    ).fetchdf()
    cols = set(desc["column_name"].tolist())
    required = {ITEM_ID_COL, PROD_FAMILY_COL, CUST_ID_COL,
                REVENUE_COL, PRIVATE_BRAND_COL}
    missing = required - cols
    if missing:
        raise ValueError(f"Missing columns in merged_dataset: {missing}")

    supplier_col = next(
        (c for c in ["SUPLR_ROLLUP_DSC", "SUPLR_DSC"] if c in cols), None
    )
    if supplier_col:
        excl_sql = ", ".join(f"'{s}'" for s in EXCLUDED_SUPPLIERS)
        supplier_filter = (
            f"AND UPPER(COALESCE({supplier_col}, '')) NOT IN ({excl_sql})"
        )
        _log(f"Supplier exclusion  : {list(EXCLUDED_SUPPLIERS)} via {supplier_col}")
    else:
        supplier_filter = ""
        _log("Supplier exclusion  : supplier column absent")

    family_excl_sql = ", ".join(f"'{f}'" for f in sorted(EXCLUDED_FAMILIES))
    _log(f"Family exclusion    : {sorted(EXCLUDED_FAMILIES)}")

    fy_sql = ", ".join(f"'{fy}'" for fy in FISCAL_YEARS)
    _log(f"Fiscal years        : {list(FISCAL_YEARS)}")

    t0 = time.time()
    _log("Step 2a: Streaming aggregation to disk (~3-6 minutes)...")

    # COPY to parquet rather than returning a pandas DF.
    # DuckDB streams this and can spill to temp_directory as needed.
    # We aggregate only numeric columns — no ANY_VALUE() on strings yet.
    con.execute(f"""
        COPY (
            SELECT
                CAST({CUST_ID_COL} AS BIGINT)  AS cust_id,
                CAST({ITEM_ID_COL} AS BIGINT)  AS item_id,
                SUM({REVENUE_COL})              AS total_spend,
                COUNT(DISTINCT ORDR_NUM)        AS order_count,
                MAX(DIM_ORDR_DT_ID)             AS last_order_dt
            FROM read_parquet('{MERGED_FILE.as_posix()}')
            WHERE {REVENUE_COL} > 0
              AND fiscal_year IN ({fy_sql})
              AND {ITEM_ID_COL} IS NOT NULL
              AND {CUST_ID_COL} IS NOT NULL
              AND COALESCE({PROD_FAMILY_COL}, '') NOT IN ({family_excl_sql})
              {supplier_filter}
            GROUP BY
                CAST({CUST_ID_COL} AS BIGINT),
                CAST({ITEM_ID_COL} AS BIGINT)
        ) TO '{tmp_agg.as_posix()}' (FORMAT PARQUET, COMPRESSION SNAPPY)
    """)
    _log(f"Step 2a complete in {time.time()-t0:.1f}s")

    # Step 2b: join product and customer metadata back in

    _log("Step 2b: Joining product metadata and customer peer_group...")

    # Register features subset for the join
    con.register("features_df", features[[
        "DIM_CUST_CURR_ID", "peer_group", "avg_order_gap_days"
    ]])

    # Build product lookup from products_clean (much smaller than merged_dataset)
    has_categ = PROD_CATEG_COL in cols
    categ_col = PROD_CATEG_COL if has_categ else "'Unknown'"

    t1 = time.time()
    history = con.execute(f"""
        SELECT
            h.cust_id,
            h.item_id,
            COALESCE(p.{ITEM_DSC_COL}, '')                      AS item_dsc,
            COALESCE(p.{PROD_FAMILY_COL}, 'Unknown')            AS prod_family,
            COALESCE({("p." + PROD_CATEG_COL) if has_categ else "'Unknown'"}, 'Unknown') AS prod_category,
            COALESCE(p.{PRIVATE_BRAND_COL}, 0)                  AS is_private_brand,
            h.total_spend,
            h.order_count,
            h.last_order_dt,
            f.peer_group,
            f.avg_order_gap_days
        FROM read_parquet('{tmp_agg.as_posix()}') h
        LEFT JOIN read_parquet('{PROD_FILE.as_posix()}') p
            ON h.item_id = CAST(p.{ITEM_ID_COL} AS BIGINT)
        LEFT JOIN features_df f
            ON h.cust_id = CAST(f.DIM_CUST_CURR_ID AS BIGINT)
    """).fetchdf()
    con.close()

    _log(f"Step 2b complete in {time.time()-t1:.1f}s")
    _log(f"Loaded {len(history):,} customer-product pairs total")
    _log(f"Unique customers : {history['cust_id'].nunique():,}")
    _log(f"Unique products  : {history['item_id'].nunique():,}")
    _log(f"Unique families  : {history['prod_family'].nunique():,}")

    # Cleanup temp parquet
    try:
        tmp_agg.unlink()
    except Exception:
        pass

    # Filter: rows must have peer_group AND not be in excluded families
    # (excluded families could leak in if a product was reclassified between
    # merged_dataset.parquet and products_clean.parquet — extremely unlikely
    # but costs nothing to guard against)
    before = len(history)
    history = history[~history["prod_family"].isin(EXCLUDED_FAMILIES)]
    history = history.dropna(subset=["peer_group"])
    dropped = before - len(history)
    if dropped > 0:
        _log(f"Dropped {dropped:,} rows (no peer_group or excluded family)")

    return history


# Step 3: Peer adoption

def compute_peer_adoption(history: pd.DataFrame) -> pd.DataFrame:
    _s("Step 3: Computing product adoption rates within peer groups")

    peer_sizes = (
        history[["cust_id", "peer_group"]]
        .drop_duplicates()
        .groupby("peer_group")
        .size()
        .reset_index(name="peer_group_size")
    )

    adoption = (
        history.groupby(["peer_group", "item_id", "item_dsc",
                          "prod_family", "prod_category", "is_private_brand"])
        .agg(
            buyer_count  = ("cust_id",    "nunique"),
            total_orders = ("order_count", "sum"),
            median_spend = ("total_spend", "median"),
            total_spend  = ("total_spend", "sum"),
        )
        .reset_index()
    )

    adoption = adoption.merge(peer_sizes, on="peer_group", how="left")
    adoption["adoption_rate"] = (
        adoption["buyer_count"] / adoption["peer_group_size"]
    ).round(4)

    adoption = adoption[
        (adoption["buyer_count"]    >= MIN_PEER_PURCHASES) &
        (adoption["peer_group_size"] >= MIN_PEER_SIZE)     &
        (adoption["adoption_rate"]   >= MIN_ADOPTION_PCT)
    ].copy()

    adoption = adoption.sort_values(
        ["peer_group", "adoption_rate"], ascending=[True, False]
    ).reset_index(drop=True)

    _log(f"Product-peer_group combos after filtering: {len(adoption):,}")
    _log(f"Unique products in adoption table        : {adoption['item_id'].nunique():,}")

    top_products = (
        adoption.groupby(["item_id", "item_dsc", "prod_family"])
        ["adoption_rate"].mean()
        .sort_values(ascending=False)
        .head(20)
        .reset_index()
    )
    _log("\n  Top 10 products by avg peer adoption rate:")
    for _, r in top_products.head(10).iterrows():
        dsc = str(r["item_dsc"])[:48]
        fam = str(r["prod_family"])[:28]
        _log(f"  {dsc:<50} {fam:<30} {r['adoption_rate']:>10.1%}")

    return adoption


# Step 4: Lapsed products

def compute_lapsed_products(history: pd.DataFrame) -> pd.DataFrame:
    _s("Step 4: Identifying lapsed products (due for reorder)")

    ref_date = int(history["last_order_dt"].max())

    def yyyymmdd_to_days_ago(d: int) -> int:
        from datetime import datetime
        ref = datetime.strptime(str(ref_date), "%Y%m%d")
        try:
            dt = datetime.strptime(str(int(d)), "%Y%m%d")
            return max(0, (ref - dt).days)
        except Exception:
            return 999

    history = history.copy()
    history["days_since_last_order"] = history["last_order_dt"].apply(
        yyyymmdd_to_days_ago
    )

    lapse_threshold = (
        history["avg_order_gap_days"].fillna(30) * LAPSE_MULTIPLIER
    )
    history["is_lapsed"] = (
        history["days_since_last_order"] > lapse_threshold
    ).astype(int)

    lapsed = history[history["is_lapsed"] == 1].copy()
    lapsed["lapse_reason"] = lapsed.apply(
        lambda r: (
            f"Last ordered {r['days_since_last_order']:.0f} days ago "
            f"(avg gap {r['avg_order_gap_days']:.0f}d — overdue by "
            f"{r['days_since_last_order'] - r['avg_order_gap_days']:.0f}d)"
        ), axis=1
    )

    _log(f"Total customer-product pairs   : {len(history):,}")
    _log(f"Lapsed pairs (overdue reorder) : {len(lapsed):,} "
         f"({len(lapsed)/len(history)*100:.1f}%)")

    return lapsed[["cust_id", "item_id", "item_dsc", "prod_family",
                   "prod_category", "is_private_brand", "total_spend",
                   "days_since_last_order", "lapse_reason"]]


# Step 5: Build recommendations

def build_recommendations(
    adoption:  pd.DataFrame,
    history:   pd.DataFrame,
    lapsed:    pd.DataFrame,
    features:  pd.DataFrame,
) -> pd.DataFrame:
    _s("Step 5: Building per-customer recommendation list")

    seq_context: dict[str, dict[str, tuple]] = {}
    seq_path = OUT_PRECOMP / "segment_sequences.parquet"
    if seq_path.exists():
        seq_df = pd.read_parquet(seq_path)
        for _, row in seq_df.iterrows():
            seg  = row["segment"]
            from_cat = row["from_category"]
            to_cat   = row["to_category"]
            prob     = row["transition_prob"]
            if seg not in seq_context:
                seq_context[seg] = {}
            if from_cat not in seq_context[seg]:
                seq_context[seg][from_cat] = (to_cat, prob)
        _log(f"Segment sequences loaded: {len(seq_context)} segments")
    else:
        _log("segment_sequences.parquet not found — pitch messages will omit sequence context")

    seg_path = OUT_PRECOMP / "customer_segments.parquet"
    cust_segment: dict[int, str] = {}
    if seg_path.exists():
        seg_labels = pd.read_parquet(seg_path, columns=["DIM_CUST_CURR_ID", "segment"])
        cust_segment = dict(zip(
            seg_labels["DIM_CUST_CURR_ID"].astype(int),
            seg_labels["segment"]
        ))
        _log(f"Customer segments loaded: {len(cust_segment):,} customers")
    else:
        _log("customer_segments.parquet not found — recommendations will use 'unknown' segment")

    cust_supplier_profile: dict[int, str] = dict(zip(
        features["DIM_CUST_CURR_ID"].astype(int),
        features["supplier_profile"].fillna("mixed")
    ))

    cust_products = (
        history.groupby("cust_id")["item_id"]
        .apply(set)
        .reset_index()
        .rename(columns={"item_id": "bought_set"})
    )

    cust_peers = features[["DIM_CUST_CURR_ID", "peer_group"]].copy()
    cust_peers = cust_peers.rename(columns={"DIM_CUST_CURR_ID": "cust_id"})
    cust_df = cust_products.merge(cust_peers, on="cust_id", how="inner")

    _log(f"Customers to score: {len(cust_df):,}")

    lapsed_idx = (
        lapsed.groupby("cust_id")[["item_id", "item_dsc", "prod_family",
                                    "prod_category", "is_private_brand",
                                    "lapse_reason"]]
        .apply(lambda x: x.to_dict("records"))
        .to_dict()
    )

    adoption_by_peer = dict(iter(adoption.groupby("peer_group")))

    CHUNK = 10_000
    all_recs = []

    t0 = time.time()
    for start in range(0, len(cust_df), CHUNK):
        chunk = cust_df.iloc[start:start+CHUNK]

        for _, crow in chunk.iterrows():
            cust_id    = int(crow["cust_id"])
            bought_set = crow["bought_set"]
            peer_group = crow["peer_group"]

            seg_label = cust_segment.get(cust_id, "unknown")
            peer_w, lapsed_w = SEGMENT_WEIGHTS.get(seg_label, DEFAULT_WEIGHTS)

            recs = {}

            # Signal 1: peer_gap — products peers buy that this customer doesn't.
            # Score formula: 2.0 * adoption_rate * peer_w
            # At adoption 0.5 with mid-tier weight 2.5, score = 2.5 which exactly
            # matches the lapsed flat score. Products with adoption > 0.5 tip toward
            # peer_gap; products with adoption < 0.5 tip toward lapsed.
            # This produces balanced 50/50 rank-1 mix in mid tiers while keeping
            # peer_gap dominant in _high and lapsed dominant in _low tiers.
            peer_products = adoption_by_peer.get(peer_group)
            if peer_products is not None:
                for _, prod in peer_products.iterrows():
                    iid = int(prod["item_id"])
                    if iid in bought_set:
                        continue
                    score = 2.0 * float(prod["adoption_rate"]) * peer_w
                    if prod["is_private_brand"]:
                        score += PRIVATE_BRAND_BONUS
                    recs[iid] = {
                        "item_id":          iid,
                        "item_dsc":         prod["item_dsc"],
                        "prod_family":      prod["prod_family"],
                        "prod_category":    prod["prod_category"],
                        "is_private_brand": int(prod["is_private_brand"]),
                        "score":            score,
                        "adoption_rate":    float(prod["adoption_rate"]),
                        "peer_group_size":  int(prod["peer_group_size"]),
                        "buyer_count":      int(prod["buyer_count"]),
                        "median_peer_spend": float(prod["median_spend"]),
                        "signal":           "peer_gap",
                    }

            # Signal 2: lapsed — products this customer used to buy but stopped.
            # Score formula: lapsed_w (flat)
            for lp in lapsed_idx.get(cust_id, []):
                iid = int(lp["item_id"])
                score = lapsed_w
                if lp["is_private_brand"]:
                    score += PRIVATE_BRAND_BONUS
                if iid not in recs or recs[iid]["score"] < score:
                    recs[iid] = {
                        "item_id":          iid,
                        "item_dsc":         lp["item_dsc"],
                        "prod_family":      lp["prod_family"],
                        "prod_category":    lp["prod_category"],
                        "is_private_brand": int(lp["is_private_brand"]),
                        "score":            score,
                        "adoption_rate":    0.0,
                        "peer_group_size":  0,
                        "buyer_count":      0,
                        "median_peer_spend": 0.0,
                        "signal":           "lapsed",
                    }

            ranked = sorted(recs.values(), key=lambda x: x["score"], reverse=True)
            supplier_profile = cust_supplier_profile.get(cust_id, "mixed")

            for rank, r in enumerate(ranked[:TOP_N_RECS], 1):
                r["cust_id"]          = cust_id
                r["segment"]          = seg_label
                r["supplier_profile"] = supplier_profile
                r["rank"]             = rank
                r["pitch_message"]    = _make_pitch_message(
                    r, seg_label, supplier_profile,
                    seq_context.get(seg_label, {})
                )
                all_recs.append(r)

        if (start // CHUNK) % 5 == 0:
            elapsed = time.time() - t0
            pct = min(100, (start + CHUNK) / len(cust_df) * 100)
            _log(f"  Progress: {pct:.0f}%  |  {elapsed:.0f}s elapsed  |  {len(all_recs):,} recs so far")

    recs_df = pd.DataFrame(all_recs)
    _log(f"Total recommendation rows : {len(recs_df):,}")
    _log(f"Customers with recs       : {recs_df['cust_id'].nunique():,}")

    sig_counts = recs_df["signal"].value_counts()
    _log(f"\n  Recommendation signal breakdown:")
    for sig, cnt in sig_counts.items():
        _log(f"    {sig:<20} {cnt:>8,} rows  ({cnt/len(recs_df)*100:.1f}%)")

    return recs_df


def _make_pitch_message(
    r: dict,
    segment: str,
    supplier_profile: str,
    seq_context: dict,
) -> str:
    signal   = r["signal"]
    dsc      = r["item_dsc"][:60] if r["item_dsc"] else "this product"
    fam      = r.get("prod_family", "")
    pb_label = " (McKesson private brand)" if r["is_private_brand"] else ""
    seg_display = segment.replace("_", " ") if segment != "unknown" else "peer"

    seq_suffix = ""
    if fam and fam in seq_context:
        next_cat, prob = seq_context[fam]
        seq_suffix = (
            f" Customers in this segment next buy {next_cat[:30]} "
            f"{prob:.0f}% of the time."
        )

    if supplier_profile == "medline_only":
        prefix = "Medline substitution" if signal == "peer_gap" else "Reorder swap"
    elif signal == "peer_gap":
        if "_high" in segment:
            prefix = "Cross-sell opportunity"
        elif "_low" in segment:
            prefix = "Category expansion"
        else:
            prefix = "Peer-adopted product"
    elif signal == "lapsed":
        if "_low" in segment or "_mid" in segment:
            prefix = "Re-engagement priority"
        else:
            prefix = "Reorder reminder"
    else:
        prefix = "Recommended product"

    if signal == "peer_gap":
        pct    = r["adoption_rate"] * 100
        buyers = r["buyer_count"]
        size   = r["peer_group_size"]
        spend  = r["median_peer_spend"]
        if supplier_profile == "medline_only":
            return (
                f"{prefix} — {dsc}{pb_label}: "
                f"instead of Medline, {pct:.0f}% of similar {seg_display} customers "
                f"buy this McKesson alternative ({buyers}/{size}). "
                f"Peers spend ~${spend:,.0f} per order.{seq_suffix}"
            )
        return (
            f"{prefix} — {dsc}{pb_label}: "
            f"{pct:.0f}% of similar {seg_display} customers ({buyers}/{size}) "
            f"buy this. Peers spend ~${spend:,.0f} per order.{seq_suffix}"
        )
    elif signal == "lapsed":
        if supplier_profile == "medline_only":
            return (
                f"{prefix} — {dsc}{pb_label}: "
                f"previously ordered, not reordered recently. "
                f"Propose McKesson reorder in place of Medline.{seq_suffix}"
            )
        return (
            f"{prefix} — {dsc}{pb_label}: "
            f"previously ordered, not reordered recently. "
            f"Easy re-engagement — they already know the product.{seq_suffix}"
        )
    return f"{prefix} — {dsc}{pb_label}: pitch opportunity.{seq_suffix}"


# Step 6: Propensity model

def compute_recommendation_factors(
    history:   pd.DataFrame,
    adoption:  pd.DataFrame,
    features:  pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, float]:
    _s("Step 6: Building purchase propensity dataset for recommendation model")

    bought_pairs = set(zip(
        history["cust_id"].astype(int),
        history["item_id"].astype(int)
    ))
    _log(f"Unique customer-product pairs in history : {len(bought_pairs):,}")

    cust_family_spend = (
        history.groupby(["cust_id", "prod_family"])["total_spend"]
        .sum()
        .reset_index()
        .rename(columns={"total_spend": "spend_in_family"})
    )

    family_counts = (
        history.groupby("cust_id")["prod_family"]
        .nunique()
        .reset_index()
        .rename(columns={"prod_family": "n_families_bought"})
    )

    _log("Building customer-product candidate pairs...")

    top_products_per_group = (
        adoption
        .sort_values(["peer_group", "adoption_rate"], ascending=[True, False])
        .groupby("peer_group")
        .head(50)
        [["peer_group", "item_id", "item_dsc", "prod_family",
          "adoption_rate", "is_private_brand",
          "buyer_count", "median_spend"]]
        .copy()
    )

    cust_peer = (
        features[["DIM_CUST_CURR_ID", "peer_group"]]
        .rename(columns={"DIM_CUST_CURR_ID": "cust_id"})
        .dropna()
    )

    pairs = cust_peer.merge(top_products_per_group, on="peer_group", how="inner")
    _log(f"Candidate pairs (customer x peer product): {len(pairs):,}")

    pairs["bought"] = pairs.apply(
        lambda r: 1 if (int(r["cust_id"]), int(r["item_id"])) in bought_pairs else 0,
        axis=1
    )
    pos_rate = pairs["bought"].mean() * 100
    _log(f"Positive rate (customer already buys it): {pos_rate:.1f}%")
    _log(f"Negatives (pitch candidates): {(pairs['bought']==0).sum():,}")

    _cust_wanted = [
        "DIM_CUST_CURR_ID", "specialty_tier",
        "R_score", "F_score", "M_score",
        "recency_days", "frequency", "monetary",
        "avg_order_gap_days",
        "pct_of_total_revenue",
        "specialty_revenue_trend_pct",
        "avg_revenue_per_order",
        "n_categories_bought",
        "category_hhi",
        "cycle_regularity",
    ]
    _cust_available = [c for c in _cust_wanted if c in features.columns]
    _cust_missing   = [c for c in _cust_wanted if c not in features.columns]
    if _cust_missing:
        _log(f"Note: {len(_cust_missing)} customer columns absent: {_cust_missing}")
    cust_feats = features[_cust_available].rename(columns={"DIM_CUST_CURR_ID": "cust_id"})

    fam_enc = {f: i for i, f in enumerate(pairs["prod_family"].dropna().unique())}
    pairs["prod_family_encoded"] = pairs["prod_family"].map(fam_enc).fillna(-1)

    item_group_count = (
        adoption.groupby("item_id")["peer_group"]
        .nunique()
        .reset_index()
        .rename(columns={"peer_group": "n_peer_groups_carrying"})
    )

    df = (
        pairs
        .merge(cust_feats,       on="cust_id",  how="left")
        .merge(family_counts,    on="cust_id",  how="left")
        .merge(item_group_count, on="item_id",  how="left")
        .merge(
            cust_family_spend,
            on=["cust_id", "prod_family"],
            how="left"
        )
    )
    df["n_families_bought"]      = df["n_families_bought"].fillna(1)
    df["n_peer_groups_carrying"] = df["n_peer_groups_carrying"].fillna(0)
    df["spend_in_family"]        = df["spend_in_family"].fillna(0)

    df["already_buys_this_family"] = (df["spend_in_family"] > 0).astype(int)
    df["adoption_x_recency"] = (
        df["adoption_rate"] * df["R_score"].fillna(1)
    )

    X_COLS_ALL = [
        "adoption_rate",
        "n_peer_groups_carrying",
        "is_private_brand",
        "prod_family_encoded",
        "recency_days",
        "frequency",
        "monetary",
        "avg_order_gap_days",
        "R_score",
        "F_score",
        "M_score",
        "specialty_tier",
        "n_families_bought",
        "already_buys_this_family",
        "adoption_x_recency",
        "spend_in_family",
        "pct_of_total_revenue",
        "specialty_revenue_trend_pct",
    ]
    X_cols = [c for c in X_COLS_ALL if c in df.columns]

    product_feats     = {"adoption_rate", "n_peer_groups_carrying",
                         "is_private_brand", "prod_family_encoded"}
    customer_feats    = {"recency_days", "frequency", "monetary", "avg_order_gap_days",
                         "R_score", "F_score", "M_score", "specialty_tier",
                         "n_families_bought", "pct_of_total_revenue",
                         "specialty_revenue_trend_pct"}
    interaction_feats = {"already_buys_this_family", "adoption_x_recency",
                         "spend_in_family"}

    def get_group(f: str) -> str:
        if f in product_feats:     return "PRODUCT"
        if f in customer_feats:    return "CUSTOMER"
        if f in interaction_feats: return "INTERACTION"
        return "OTHER"

    sample = df.sample(min(500_000, len(df)), random_state=42)
    X = sample[X_cols].fillna(0)
    y = sample["bought"]

    _log(f"Training dataset: {X.shape[0]:,} rows x {X.shape[1]} features")
    _log(f"Positive class (bought=1): {y.sum():,} ({y.mean()*100:.1f}%)")

    if y.sum() < 100:
        _log("Not enough positives to train — returning feature list only")
        factor_df = pd.DataFrame({
            "feature":     X_cols,
            "group":       [get_group(c) for c in X_cols],
            "description": [_desc(c) for c in X_cols],
        })
        return pd.DataFrame(), factor_df, 0.0

    X_tr, X_te, y_tr, y_te = train_test_split(
        X, y, test_size=0.20, random_state=42, stratify=y
    )
    model = RandomForestClassifier(
        n_estimators=100, max_depth=10, n_jobs=-1,
        class_weight="balanced", random_state=42
    )
    _log("Training purchase propensity model...")
    t0 = time.time()
    model.fit(X_tr, y_tr)
    _log(f"Fit in {time.time()-t0:.1f}s")

    auc = roc_auc_score(y_te, model.predict_proba(X_te)[:, 1])
    _log(f"AUC-ROC: {auc:.4f}")

    report = classification_report(
        y_te, model.predict(X_te),
        target_names=["Will not buy", "Will buy"]
    )
    for line in report.strip().split("\n"):
        _log(f"  {line}")

    imp = pd.DataFrame({
        "feature":    X_cols,
        "importance": model.feature_importances_,
    }).sort_values("importance", ascending=False).reset_index(drop=True)
    imp["rank"]         = range(1, len(imp)+1)
    imp["importance"]   = imp["importance"].round(6)
    imp["pct_of_total"] = (imp["importance"] / imp["importance"].sum() * 100).round(2)
    imp["group"]        = imp["feature"].apply(get_group)

    df_lbl = X.copy()
    df_lbl["bought"] = y.values

    rows = []
    for _, r in imp.iterrows():
        feat    = r["feature"]
        col     = df_lbl[feat]
        med     = col.median()
        hi_rate = df_lbl.loc[col >= med, "bought"].mean() * 100
        lo_rate = df_lbl.loc[col <  med, "bought"].mean() * 100
        hi_rate = 0.0 if hi_rate != hi_rate else round(hi_rate, 1)
        lo_rate = 0.0 if lo_rate != lo_rate else round(lo_rate, 1)
        diff    = round(hi_rate - lo_rate, 1)
        diff    = 0.0 if diff != diff else diff

        rows.append({
            "rank":          int(r["rank"]),
            "group":         r["group"],
            "feature":       feat,
            "description":   _desc(feat),
            "gini_pct":      r["pct_of_total"],
            "buy_rate_high": hi_rate,
            "buy_rate_low":  lo_rate,
            "difference_pp": diff,
            "what_it_means": _meaning(feat, hi_rate, lo_rate, diff, med),
        })

    factor_df = pd.DataFrame(rows)

    _log(f"\n  {'RNK':<4} {'GRP':<12} {'FEATURE':<30} {'GINI%':>6} {'HI%':>6} {'LO%':>6} {'DIFF':>6}")
    for _, r in factor_df.iterrows():
        _log(
            f"  {r['rank']:<4} {r['group']:<12} {r['feature']:<30} "
            f"{r['gini_pct']:>6.1f} {r['buy_rate_high']:>6.1f} "
            f"{r['buy_rate_low']:>6.1f} {r['difference_pp']:>+6.1f}"
        )

    _log("\n  Buy rate by feature group:")
    for grp in ["PRODUCT", "INTERACTION", "CUSTOMER"]:
        grp_rows = factor_df[factor_df["group"] == grp]
        avg_imp = grp_rows["gini_pct"].sum()
        _log(f"    {grp:<14} {avg_imp:>6.1f}% of total model importance")

    return imp, factor_df, auc


def _desc(feat: str) -> str:
    d = {
        "adoption_rate":             "% of peer group who buy this product",
        "n_peer_groups_carrying":    "How many peer groups carry this product",
        "is_private_brand":          "McKesson private brand flag",
        "prod_family_encoded":       "Product family (encoded)",
        "recency_days":              "Days since customer last ordered anything",
        "frequency":                 "Customer total order count",
        "monetary":                  "Customer total spend",
        "avg_order_gap_days":        "Average days between customer orders",
        "R_score":                   "Recency quintile 1-5",
        "F_score":                   "Frequency quintile 1-5",
        "M_score":                   "Monetary quintile 1-5",
        "specialty_tier":            "Specialty tier 1-3",
        "n_families_bought":         "Product families customer already buys from",
        "already_buys_this_family":  "Customer already buys from this product family",
        "adoption_x_recency":        "Adoption rate x recency score interaction",
        "spend_in_family":           "Customer existing spend in this product family",
        "pct_of_total_revenue":      "Specialty share of McKesson total revenue",
        "specialty_revenue_trend_pct": "Specialty revenue trend FY2425->FY2526",
    }
    return d.get(feat, feat)


def _meaning(feat: str, hi_rate: float, lo_rate: float,
             diff: float, med: float) -> str:
    if feat == "adoption_rate":
        return (
            f"Products bought by >= {med:.0%} of peers: {hi_rate:.0f}% buy rate. "
            f"Below {med:.0%}: {lo_rate:.0f}% buy rate."
        )
    if feat == "already_buys_this_family":
        return (
            f"Buys this family: {hi_rate:.0f}% buy rate. "
            f"Does not: {lo_rate:.0f}% buy rate."
        )
    if feat == "spend_in_family":
        return (
            f"Spend in family >= ${med:,.0f}: {hi_rate:.0f}% buy rate. "
            f"No spend: {lo_rate:.0f}% buy rate."
        )
    if feat == "adoption_x_recency":
        return (
            f"High adoption x high recency: {hi_rate:.0f}% buy rate. "
            f"Best pitch window."
        )
    direction = "above" if diff > 0 else "below"
    return (
        f"Customers {direction} median {feat}: {max(hi_rate, lo_rate):.0f}% buy rate. "
        f"Difference of {abs(diff):.1f} percentage points."
    )


# Step 7: Save outputs

def save_outputs(
    recs_df:        pd.DataFrame,
    adoption:       pd.DataFrame,
    imp:            pd.DataFrame,
    factor_signals: pd.DataFrame,
    auc:            float,
) -> None:
    _s("Step 7: Saving outputs")

    OUT_PRECOMP.mkdir(parents=True, exist_ok=True)
    OUT_ANALYSIS.mkdir(parents=True, exist_ok=True)

    recs_df = _normalise_recs_schema(recs_df)
    recs_df = recs_df.sort_values(["cust_id", "rank"]).reset_index(drop=True)

    recs_path = OUT_PRECOMP / "customer_recommendations.parquet"
    recs_df.to_parquet(recs_path, index=False)
    _log(f"Saved: {recs_path.relative_to(ROOT)}")

    _write_recs_schema(recs_path.with_suffix(".schema.json"), recs_df)

    adoption = _normalise_adoption_schema(adoption)
    adop_path = OUT_PRECOMP / "product_adoption_rates.parquet"
    adoption.to_parquet(adop_path, index=False)
    _log(f"Saved: {adop_path.relative_to(ROOT)}")

    if len(factor_signals) > 0:
        fac_path = OUT_PRECOMP / "recommendation_factors.parquet"
        factor_signals.to_parquet(fac_path, index=False)
        _log(f"Saved: {fac_path.relative_to(ROOT)}")

    CHART_N   = 20
    xlsx_path = OUT_ANALYSIS / "recommendation_factors.xlsx"

    family_adoption = (
        adoption.groupby("prod_family")
        .agg(
            unique_products     = ("item_id",       "nunique"),
            avg_adoption_rate   = ("adoption_rate", "mean"),
            max_adoption_rate   = ("adoption_rate", "max"),
            total_peer_buyers   = ("buyer_count",   "sum"),
        )
        .reset_index()
        .sort_values("avg_adoption_rate", ascending=False)
        .reset_index(drop=True)
    )
    family_adoption["avg_adoption_rate"] = family_adoption["avg_adoption_rate"].round(4)
    family_adoption["max_adoption_rate"] = family_adoption["max_adoption_rate"].round(4)

    top_products = (
        adoption.groupby(["item_id", "item_dsc", "prod_family",
                           "prod_category", "is_private_brand"])
        .agg(
            avg_adoption    = ("adoption_rate", "mean"),
            peer_groups     = ("peer_group",    "nunique"),
            total_buyers    = ("buyer_count",   "sum"),
            median_spend    = ("median_spend",  "median"),
        )
        .reset_index()
        .sort_values("avg_adoption", ascending=False)
        .head(200)
        .reset_index(drop=True)
    )
    top_products["avg_adoption"] = top_products["avg_adoption"].round(4)
    top_products["median_spend"] = top_products["median_spend"].round(2)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        family_adoption.to_excel(writer, sheet_name="01_family_adoption", index=False)
        top_products.to_excel(   writer, sheet_name="02_top_products",    index=False)
        if len(imp)            > 0: imp.to_excel(writer, sheet_name="03_factor_importance", index=False)
        if len(factor_signals) > 0: factor_signals.to_excel(writer, sheet_name="04_factor_signals", index=False)
        recs_df.head(5000).to_excel(writer, sheet_name="05_sample_recs", index=False)

        wb = writer.book
        colors = {
            "01_family_adoption":   "1F4E79",
            "02_top_products":      "375623",
            "03_factor_importance": "7030A0",
            "04_factor_signals":    "833C00",
            "05_sample_recs":       "C00000",
        }
        dfs = {
            "01_family_adoption":   family_adoption,
            "02_top_products":      top_products,
            "03_factor_importance": imp if len(imp) > 0 else pd.DataFrame(),
            "04_factor_signals":    factor_signals if len(factor_signals) > 0 else pd.DataFrame(),
            "05_sample_recs":       recs_df.head(5000),
        }
        for name, color in colors.items():
            if name in wb.sheetnames and len(dfs[name]) > 0:
                wb[name].sheet_properties.tabColor = color
                _style(writer.sheets[name], dfs[name], hc=color)

        if "01_family_adoption" in wb.sheetnames:
            _chart(wb["01_family_adoption"], n=min(CHART_N, len(family_adoption)),
                   lc=1, vc=3, anchor="I2",
                   title="Avg product adoption rate by family",
                   xtitle="Average adoption rate within peer groups",
                   color="1F4E79")
        if "02_top_products" in wb.sheetnames:
            _chart(wb["02_top_products"], n=min(CHART_N, len(top_products)),
                   lc=2, vc=6, anchor="L2",
                   title=f"Top {CHART_N} products by avg adoption rate",
                   xtitle="Avg adoption rate across peer groups",
                   color="375623")
        if "03_factor_importance" in wb.sheetnames and len(imp) > 0:
            _chart(wb["03_factor_importance"], n=min(CHART_N, len(imp)),
                   lc=1, vc=2, anchor="I2",
                   title="Factors driving recommendation confidence",
                   xtitle="Gini importance score", color="7030A0")

    _log(f"Saved: {xlsx_path.relative_to(ROOT)}")

    _save_pngs(family_adoption, imp, recs_df)


def _normalise_recs_schema(recs_df: pd.DataFrame) -> pd.DataFrame:
    df = recs_df.copy()

    int_cols   = ["cust_id", "item_id", "is_private_brand", "rank",
                  "peer_group_size", "buyer_count"]
    float_cols = ["score", "adoption_rate", "median_peer_spend"]
    text_cols  = ["item_dsc", "prod_family", "prod_category", "signal",
                  "segment", "supplier_profile", "pitch_message"]

    for c in int_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype("int64")
    for c in float_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype("float64")
    for c in text_cols:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)

    col_order = [
        "cust_id", "rank", "segment", "supplier_profile",
        "item_id", "item_dsc", "prod_family", "prod_category",
        "is_private_brand", "signal", "score",
        "adoption_rate", "peer_group_size", "buyer_count", "median_peer_spend",
        "pitch_message",
    ]
    ordered = [c for c in col_order if c in df.columns]
    remaining = [c for c in df.columns if c not in ordered]
    df = df[ordered + remaining]

    return df


def _normalise_adoption_schema(adoption: pd.DataFrame) -> pd.DataFrame:
    df = adoption.copy()
    for c in ["item_id", "buyer_count", "total_orders", "peer_group_size",
              "is_private_brand"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype("int64")
    for c in ["median_spend", "total_spend", "adoption_rate"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype("float64")
    for c in ["peer_group", "item_dsc", "prod_family", "prod_category"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)
    return df


def _write_recs_schema(path: Path, df: pd.DataFrame) -> None:
    schema = {
        "file":    "customer_recommendations.parquet",
        "purpose": "Precomputed top-N recommendations per customer. Read by the API.",
        "sort_order": ["cust_id", "rank"],
        "row_count":  int(len(df)),
        "columns": [],
    }
    col_meta = {
        "cust_id":          ("int64",  "Customer ID — joins to customers_clean"),
        "rank":             ("int64",  "1-indexed rank within customer. 1 = top pick."),
        "segment":          ("string", "MKT_CD_tier label (e.g. PO_high). 'unknown' if missing."),
        "supplier_profile": ("string", "One of: medline_only, mckesson_primary, mixed."),
        "item_id":          ("int64",  "Product ID — joins to products_clean"),
        "item_dsc":         ("string", "Human-readable product name"),
        "prod_family":      ("string", "PROD_FMLY_LVL1_DSC"),
        "prod_category":    ("string", "PROD_CTGRY_LVL2_DSC"),
        "is_private_brand": ("int64",  "1 if McKesson private brand, else 0"),
        "signal":           ("string", "'peer_gap' or 'lapsed'"),
        "score":            ("float64","Final scoring value — higher ranks first"),
        "adoption_rate":    ("float64","Peer adoption rate for peer_gap signal (0 for lapsed)"),
        "peer_group_size":  ("int64",  "Size of the peer group the rate was computed over"),
        "buyer_count":      ("int64",  "Number of peers who bought this item"),
        "median_peer_spend":("float64","Median spend per order among peers who buy"),
        "pitch_message":    ("string", "Seller-facing pitch text."),
    }
    for c in df.columns:
        dtype, desc = col_meta.get(c, (str(df[c].dtype), ""))
        schema["columns"].append({"name": c, "type": dtype, "description": desc})

    path.write_text(json.dumps(schema, indent=2), encoding="utf-8")
    _log(f"Saved: {path.relative_to(ROOT)}")


def _save_pngs(
    family_adoption: pd.DataFrame,
    imp:             pd.DataFrame,
    recs_df:         pd.DataFrame,
) -> None:
    if len(family_adoption) > 0:
        top = family_adoption.head(15).iloc[::-1]
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.barh(top["prod_family"].astype(str).str[:35],
                top["avg_adoption_rate"] * 100, color="#1F4E79")
        ax.set_xlabel("Average adoption rate within peer groups (%)")
        ax.set_title("Product family adoption rates (top 15)")
        plt.tight_layout()
        out = OUT_ANALYSIS / "recommendation_family_adoption.png"
        plt.savefig(out, dpi=150, bbox_inches="tight")
        plt.close(fig)
        _log(f"Saved: {out.relative_to(ROOT)}")

    if len(imp) > 0:
        top = imp.head(15).iloc[::-1]
        colors = {"PRODUCT": "#1F4E79", "CUSTOMER": "#375623",
                  "INTERACTION": "#7030A0", "OTHER": "#888888"}
        fig, ax = plt.subplots(figsize=(10, 6))
        bar_colors = [colors.get(g, "#888888") for g in top["group"]]
        ax.barh(top["feature"], top["pct_of_total"], color=bar_colors)
        ax.set_xlabel("% of total model importance (Gini)")
        ax.set_title("Top 15 factors driving recommendation confidence")
        import matplotlib.patches as mpatches
        legend_handles = [mpatches.Patch(color=c, label=g)
                           for g, c in colors.items() if g != "OTHER"]
        ax.legend(handles=legend_handles, loc="lower right")
        plt.tight_layout()
        out = OUT_ANALYSIS / "recommendation_factor_importance.png"
        plt.savefig(out, dpi=150, bbox_inches="tight")
        plt.close(fig)
        _log(f"Saved: {out.relative_to(ROOT)}")

    if len(recs_df) > 0 and "segment" in recs_df.columns:
        rank1 = recs_df[recs_df["rank"] == 1]
        if len(rank1) > 0:
            mix = (
                rank1.groupby(["segment", "signal"])
                .size()
                .unstack(fill_value=0)
            )
            if "peer_gap" not in mix.columns: mix["peer_gap"] = 0
            if "lapsed"   not in mix.columns: mix["lapsed"]   = 0
            mix["total"] = mix["peer_gap"] + mix["lapsed"]
            mix = mix[mix["total"] >= 100].copy()
            if len(mix) > 0:
                mix["peer_gap_pct"] = mix["peer_gap"] / mix["total"] * 100
                mix["lapsed_pct"]   = mix["lapsed"]   / mix["total"] * 100
                mix = mix.sort_values("peer_gap_pct", ascending=True)

                fig, ax = plt.subplots(figsize=(10, 7))
                y = range(len(mix))
                ax.barh(y, mix["peer_gap_pct"], color="#1F4E79", label="peer_gap")
                ax.barh(y, mix["lapsed_pct"], left=mix["peer_gap_pct"],
                        color="#C00000", label="lapsed")
                ax.set_yticks(y)
                ax.set_yticklabels(mix.index)
                ax.set_xlabel("% of rank-1 recommendations")
                ax.set_xlim(0, 100)
                ax.set_title("Signal mix by segment at rank 1")
                ax.legend(loc="lower right")
                plt.tight_layout()
                out = OUT_ANALYSIS / "recommendation_signal_mix_by_segment.png"
                plt.savefig(out, dpi=150, bbox_inches="tight")
                plt.close(fig)
                _log(f"Saved: {out.relative_to(ROOT)}")


# Main

def main() -> None:
    print()
    print("="*64)
    print("  PRODUCT RECOMMENDATION FACTOR ANALYSIS")
    print("="*64)
    start = time.time()

    OUT_PRECOMP.mkdir(parents=True, exist_ok=True)
    OUT_ANALYSIS.mkdir(parents=True, exist_ok=True)
    SPILL_DIR.mkdir(parents=True, exist_ok=True)

    for f, lbl in [
        (MERGED_FILE,  "merged_dataset.parquet"),
        (PROD_FILE,    "products_clean.parquet"),
        (FEATURE_FILE, "customer_features.parquet"),
    ]:
        if not f.exists():
            print(f"\nFATAL: {lbl} not found. Run clean_data.py first.",
                  file=sys.stderr)
            sys.exit(1)

    try:
        features              = load_features_and_segments()
        history               = load_customer_product_history(features)
        adoption              = compute_peer_adoption(history)
        lapsed                = compute_lapsed_products(history)
        recs_df               = build_recommendations(adoption, history, lapsed, features)
        imp, factor_signals, auc = compute_recommendation_factors(history, adoption, features)

        save_outputs(recs_df, adoption, imp, factor_signals, auc)
    finally:
        # Clean up the spill directory on success or failure
        try:
            import shutil
            if SPILL_DIR.exists():
                shutil.rmtree(SPILL_DIR, ignore_errors=True)
        except Exception:
            pass

    _s("Complete")
    _log(f"Total time: {time.time()-start:.1f}s")
    _log("")
    _log("Key outputs:")
    _log("  customer_recommendations.parquet        — top-N per customer (API)")
    _log("  customer_recommendations.schema.json    — schema sidecar")
    _log("  product_adoption_rates.parquet          — peer adoption per product")
    _log("  recommendation_factors.parquet          — model feature signals")
    _log("  recommendation_factors.xlsx             — team analysis report")
    _log("  recommendation_family_adoption.png")
    _log("  recommendation_factor_importance.png")
    _log("  recommendation_signal_mix_by_segment.png")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"\nFATAL ERROR: {e}", file=sys.stderr)
        raise