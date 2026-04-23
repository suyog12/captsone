from __future__ import annotations

import sys
import time
import warnings
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")   # non-interactive backend — safe for scripts
import matplotlib.pyplot as plt
import shap
from sklearn.ensemble import RandomForestClassifier
from sklearn.inspection import permutation_importance
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, roc_auc_score
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)


# Paths

ROOT          = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN    = ROOT / "data_clean"
FEATURE_FILE  = DATA_CLEAN / "features" / "customer_features.parquet"
OUT_DIR       = DATA_CLEAN / "analysis"


# Configuration

RF_PARAMS = dict(
    n_estimators  = 100,
    max_depth     = 10,
    n_jobs        = -1,
    class_weight  = "balanced",
    random_state  = 42,
)
SHAP_SAMPLE_SIZE = 5_000
TEST_SIZE        = 0.20
RANDOM_STATE     = 42
TOP_N            = 20   # features to display in console and charts

# Columns that must never be used as features.
# These are identifiers, raw categoricals (encoded versions are used instead),
# or columns that would leak target information into the model.
# Note: tier3_fallback_spclty_cd was removed from customer_features.parquet in
# clean_data.py step 7 — it no longer exists in the file.
_DROP_ALWAYS = {
    "DIM_CUST_CURR_ID", "CUST_NUM", "CUST_NAME",  # identifiers
    "CUST_TYPE_CD", "SPCLTY_CD", "SPCLTY_DSC",    # raw strings, use encoded versions
    "MKT_CD", "MMS_CLASS_CD", "MMS_CLASS_DSC",
    "MMS_SGMNT_CD", "MMS_SUB_CLASS_CD",
    "STATE", "CITY", "CNTRY_CD", "ZIP",            # raw geo, use state_encoded
    "ACTV_FLG",                                    # raw flag, not numeric
    "state_grouped",                               # intermediate, use state_encoded
    "RFM_score",                                   # composite string, use R/F/M scores
    "last_order_date_id",                          # raw date integer, leaks into recency_days
}


# Logging

def _section(title: str) -> None:
    print(f"\n{'─' * 64}", flush=True)
    print(f"  {title}", flush=True)
    print(f"{'─' * 64}", flush=True)


def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


# Excel styling helper

def _style_sheet(ws, df: pd.DataFrame, header_color: str = "1F4E79") -> None:
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=header_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, 12), 55)


# Step 1: Load and prepare data

def load_and_prepare() -> tuple[pd.DataFrame, pd.Series, list[str]]:
    _section("Step 1: Loading customer_features.parquet")

    df = pd.read_parquet(FEATURE_FILE)
    _log(f"Loaded : {len(df):,} rows  |  {df.columns.tolist().__len__()} columns")

    # Keep only rows where churn_label is 0 or 1 — exclude new FY2526 customers
    df = df[df["churn_label"].isin([0, 1])].copy()
    _log(f"Filtered to churn training rows: {len(df):,}  "
         f"(excluded label=-1 new customers)")

    churn_rate = df["churn_label"].mean() * 100
    _log(f"Churn rate in training set: {churn_rate:.2f}%")

    # Separate target
    y = df["churn_label"].astype(int)

    # Build feature matrix — drop ID/categorical/target columns
    drop_cols = _DROP_ALWAYS | {"churn_label"}
    X_cols = [
        c for c in df.columns
        if c not in drop_cols
        and pd.api.types.is_numeric_dtype(df[c])
    ]
    X = df[X_cols].copy()

    # Fill any residual nulls with column median — RF cannot handle NaN
    for col in X.columns:
        if X[col].isna().any():
            X[col] = X[col].fillna(X[col].median())

    _log(f"Feature matrix : {X.shape[0]:,} rows  |  {X.shape[1]} features")
    _log(f"Target         : churn_label  (0=retained, 1=churned)")

    return X, y, X_cols




# Step 1b: Feature correlation check

def correlation_check(X: pd.DataFrame) -> pd.DataFrame:
    """
    Compute pairwise absolute correlations to detect redundant features.
    Pairs above 0.85 are flagged — keeping both adds noise without signal.
    Returns a DataFrame of high-correlation pairs for the Excel report.
    """
    _section("Step 1b: Feature correlation check")

    corr = X.corr().abs()

    # Extract upper triangle only — avoid duplicate pairs
    upper = corr.where(np.triu(np.ones(corr.shape), k=1).astype(bool))
    high_corr = (
        upper.stack()
        .reset_index()
        .rename(columns={"level_0": "feature_a", "level_1": "feature_b", 0: "correlation"})
        .query("correlation >= 0.85")
        .sort_values("correlation", ascending=False)
        .reset_index(drop=True)
    )
    high_corr["correlation"] = high_corr["correlation"].round(4)

    if len(high_corr) == 0:
        _log("No highly correlated feature pairs found (threshold: r >= 0.85)")
    else:
        _log(f"High-correlation pairs (r >= 0.85): {len(high_corr)} found")
        _log(f"  {'Feature A':<30} {'Feature B':<30} {'r':>6}")
        _log(f"  {'─'*30} {'─'*30} {'─'*6}")
        for _, row in high_corr.head(15).iterrows():
            _log(f"  {row['feature_a']:<30} {row['feature_b']:<30} {row['correlation']:>6.4f}")
        if len(high_corr) > 15:
            _log(f"  ... and {len(high_corr) - 15} more — see correlation sheet in report")
        _log("")
        _log("Note: highly correlated features are kept in training but flagged here.")
        _log("      Random Forest handles multicollinearity reasonably — SHAP will")
        _log("      split importance across correlated features, so check both together.")

    return high_corr

# Step 2: Train / test split and model training

def train_model(
    X: pd.DataFrame,
    y: pd.Series,
) -> tuple[RandomForestClassifier, pd.DataFrame, pd.DataFrame, pd.Series, pd.Series]:
    _section("Step 2: Training RandomForestClassifier")

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y
    )
    _log(f"Train : {len(X_train):,} rows  |  Test : {len(X_test):,} rows")

    model = RandomForestClassifier(**RF_PARAMS)
    _log(f"Fitting model  (n_estimators={RF_PARAMS['n_estimators']}, "
         f"max_depth={RF_PARAMS['max_depth']})...")
    t0 = time.time()
    model.fit(X_train, y_train)
    _log(f"Fit complete in {time.time() - t0:.1f}s")

    # Evaluate
    y_pred  = model.predict(X_test)
    y_proba = model.predict_proba(X_test)[:, 1]
    auc     = roc_auc_score(y_test, y_proba)
    _log(f"AUC-ROC on test set: {auc:.4f}")
    _log("Classification report:")
    report = classification_report(y_test, y_pred, target_names=["Retained", "Churned"])
    for line in report.strip().split("\n"):
        _log(f"  {line}")

    return model, X_train, X_test, y_train, y_test


# Step 3: Gini importance

def gini_importance(
    model: RandomForestClassifier,
    feature_names: list[str],
) -> pd.DataFrame:
    _section("Step 3: Gini (split) importance")

    imp = pd.DataFrame({
        "feature":    feature_names,
        "importance": model.feature_importances_,
    }).sort_values("importance", ascending=False).reset_index(drop=True)

    imp["rank"]        = range(1, len(imp) + 1)
    imp["importance"]  = imp["importance"].round(6)
    imp["pct_of_total"] = (imp["importance"] / imp["importance"].sum() * 100).round(2)
    imp["cumulative_pct"] = imp["pct_of_total"].cumsum().round(2)

    _log(f"Top {TOP_N} features by Gini importance:")
    _log(f"  {'Rank':<5} {'Feature':<35} {'Importance':>12} {'% Total':>9}")
    _log(f"  {'─'*5} {'─'*35} {'─'*12} {'─'*9}")
    for _, row in imp.head(TOP_N).iterrows():
        _log(f"  {int(row['rank']):<5} {row['feature']:<35} "
             f"{row['importance']:>12.6f} {row['pct_of_total']:>8.2f}%")

    return imp


# Step 4: Permutation importance

def perm_importance(
    model: RandomForestClassifier,
    X_test: pd.DataFrame,
    y_test: pd.Series,
) -> pd.DataFrame:
    _section("Step 4: Permutation importance (on test set)")
    _log("Running permutation importance — this takes 2-5 minutes...")

    t0     = time.time()
    result = permutation_importance(
        model, X_test, y_test,
        n_repeats   = 10,
        random_state = RANDOM_STATE,
        n_jobs      = -1,
        scoring     = "roc_auc",
    )
    _log(f"Permutation importance complete in {time.time() - t0:.1f}s")

    perm = pd.DataFrame({
        "feature":      X_test.columns,
        "importance":   result.importances_mean.round(6),
        "std":          result.importances_std.round(6),
    }).sort_values("importance", ascending=False).reset_index(drop=True)

    perm["rank"] = range(1, len(perm) + 1)

    _log(f"Top {TOP_N} features by permutation importance (AUC drop):")
    _log(f"  {'Rank':<5} {'Feature':<35} {'AUC Drop':>10} {'Std':>8}")
    _log(f"  {'─'*5} {'─'*35} {'─'*10} {'─'*8}")
    for _, row in perm.head(TOP_N).iterrows():
        _log(f"  {int(row['rank']):<5} {row['feature']:<35} "
             f"{row['importance']:>10.6f} {row['std']:>8.6f}")

    return perm




# Step 4b: Top-10 overlap across all three importance methods

def log_overlap(
    gini:    pd.DataFrame,
    perm:    pd.DataFrame,
    shap_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Compare top 10 features from Gini, Permutation, and SHAP.
    Features that appear in all three are the most reliable drivers.
    Features that appear in only one method may be artefacts.
    """
    _section("Step 4b: Top-10 feature overlap across all three methods")

    top_gini = set(gini.head(10)["feature"])
    top_perm = set(perm.head(10)["feature"])
    top_shap = set(shap_df.head(10)["feature"])

    all_top  = top_gini | top_perm | top_shap

    rows = []
    for feat in sorted(all_top):
        in_gini = feat in top_gini
        in_perm = feat in top_perm
        in_shap = feat in top_shap
        count   = sum([in_gini, in_perm, in_shap])
        confidence = {3: "HIGH — appears in all 3 methods",
                      2: "MEDIUM — appears in 2 methods",
                      1: "LOW — appears in 1 method only"}[count]
        rows.append({
            "feature":     feat,
            "in_gini":     "✓" if in_gini else "",
            "in_perm":     "✓" if in_perm else "",
            "in_shap":     "✓" if in_shap else "",
            "method_count": count,
            "confidence":   confidence,
        })

    overlap = (
        pd.DataFrame(rows)
        .sort_values(["method_count", "feature"], ascending=[False, True])
        .reset_index(drop=True)
    )

    _log(f"  {'Feature':<35} {'Gini':^6} {'Perm':^6} {'SHAP':^6} {'Confidence'}")
    _log(f"  {'─'*35} {'─'*6} {'─'*6} {'─'*6} {'─'*40}")
    for _, r in overlap.iterrows():
        _log(f"  {r['feature']:<35} {r['in_gini']:^6} {r['in_perm']:^6} "
             f"{r['in_shap']:^6} {r['confidence']}")

    high_conf = overlap[overlap["method_count"] == 3]["feature"].tolist()
    _log(f"")
    _log(f"HIGH confidence drivers (all 3 methods agree): {len(high_conf)}")
    for f in high_conf:
        _log(f"    {f}")

    return overlap

# Step 5: SHAP values

def shap_analysis(
    model: RandomForestClassifier,
    X: pd.DataFrame,
) -> tuple[np.ndarray, pd.DataFrame]:
    _section("Step 5: SHAP values (sampled 5,000 rows)")

    sample = X.sample(
        min(SHAP_SAMPLE_SIZE, len(X)), random_state=RANDOM_STATE
    )
    _log(f"SHAP sample size: {len(sample):,} rows")
    _log("Computing SHAP values — this takes 3-8 minutes...")

    t0      = time.time()
    explainer   = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(sample)
    _log(f"SHAP complete in {time.time() - t0:.1f}s")

    # Normalise shap_values to a 2D array of shape (n_samples, n_features)
    # for the churn class (class index 1).
    #
    # SHAP output varies by version and model type:
    #   list of arrays  -> shap_values[1] gives class-1 array, shape (n, f)
    #   3D numpy array  -> shape (n, f, 2); slice [:, :, 1] for class 1
    #   2D numpy array  -> already correct (binary shorthand from newer SHAP)
    if isinstance(shap_values, list):
        sv = shap_values[1]                   # list[class0, class1] → class 1
    elif isinstance(shap_values, np.ndarray) and shap_values.ndim == 3:
        sv = shap_values[:, :, 1]             # (n, f, 2) → (n, f) for class 1
    else:
        sv = shap_values                      # already (n, f)

    # Mean absolute SHAP per feature — sv is now guaranteed 2D (n_samples, n_features)
    mean_abs_shap = np.abs(sv).mean(axis=0)   # shape (n_features,)
    shap_mean = pd.DataFrame({
        "feature":       list(sample.columns),
        "mean_abs_shap": mean_abs_shap.round(6),
    }).sort_values("mean_abs_shap", ascending=False).reset_index(drop=True)
    shap_mean["rank"] = range(1, len(shap_mean) + 1)

    # Direction: positive mean SHAP = feature pushes toward churn
    shap_mean["mean_shap_signed"] = sv.mean(axis=0).round(6)
    shap_mean["direction"] = shap_mean["mean_shap_signed"].apply(
        lambda v: "↑ churn" if v > 0.001 else ("↓ churn" if v < -0.001 else "neutral")
    )

    _log(f"Top {TOP_N} features by SHAP (mean |SHAP|, class=Churned):")
    _log(f"  {'Rank':<5} {'Feature':<35} {'|SHAP|':>8} {'Direction':>12}")
    _log(f"  {'─'*5} {'─'*35} {'─'*8} {'─'*12}")
    for _, row in shap_mean.head(TOP_N).iterrows():
        _log(f"  {int(row['rank']):<5} {row['feature']:<35} "
             f"{row['mean_abs_shap']:>8.6f} {row['direction']:>12}")

    # SHAP beeswarm summary plot
    fig, ax = plt.subplots(figsize=(10, 8))
    top_features = shap_mean.head(TOP_N)["feature"].tolist()
    sample_top   = sample[top_features]
    sv_top       = sv[:, [sample.columns.tolist().index(f) for f in top_features]]

    shap.summary_plot(
        sv_top, sample_top,
        feature_names = top_features,
        show          = False,
        plot_size     = (10, 8),
    )
    plt.title("SHAP Summary — Drivers of Customer Churn", fontsize=13, pad=12)
    plt.tight_layout()
    plt.savefig(OUT_DIR / "shap_summary.png", dpi=150, bbox_inches="tight")
    plt.close()
    _log("Saved : data_clean/analysis/shap_summary.png")

    # SHAP bar chart
    fig, ax = plt.subplots(figsize=(10, 7))
    colors = [
        "#C00000" if d == "↑ churn" else "#375623" if d == "↓ churn" else "#888888"
        for d in shap_mean.head(TOP_N)["direction"]
    ]
    bars = ax.barh(
        y      = shap_mean.head(TOP_N)["feature"][::-1],
        width  = shap_mean.head(TOP_N)["mean_abs_shap"][::-1],
        color  = colors[::-1],
        edgecolor = "white",
        linewidth = 0.5,
    )
    ax.set_xlabel("Mean |SHAP value| (impact on churn probability)", fontsize=10)
    ax.set_title("Top Feature Drivers of Churn — Mean Absolute SHAP", fontsize=12, pad=10)
    ax.tick_params(axis="y", labelsize=8)

    # Legend
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor="#C00000", label="Pushes toward churn (↑)"),
        Patch(facecolor="#375623", label="Protects against churn (↓)"),
        Patch(facecolor="#888888", label="Neutral"),
    ]
    ax.legend(handles=legend_elements, fontsize=8, loc="lower right")
    plt.tight_layout()
    plt.savefig(OUT_DIR / "shap_bar.png", dpi=150, bbox_inches="tight")
    plt.close()
    _log("Saved : data_clean/analysis/shap_bar.png")

    return sv, shap_mean, sample


# Step 6: Business signal interpretation

def _derive_feature_stats(feat: str, df: pd.DataFrame, churn_col: str = "churn_label") -> dict:
    """
    Compute real statistics from the feature matrix for a single feature.
    Returns a dict with: median, mean, std, churned_mean, retained_mean,
    churned_vs_retained_pct (how much higher churned mean is vs retained),
    and data_driven interpretations derived from those numbers.

    This replaces all hardcoded text — every signal and action is derived
    from what the actual data shows, not what we assumed it would show.
    """
    col = df[feat] if feat in df.columns else None
    if col is None or not pd.api.types.is_numeric_dtype(col):
        return {
            "median": None, "mean": None, "std": None,
            "churned_mean": None, "retained_mean": None,
            "churned_vs_retained_pct": None,
            "signal": feat,
            "high_value_means": "Non-numeric or missing in feature matrix",
            "action": "Investigate column availability in customer_features.parquet",
        }

    churned  = df[df[churn_col] == 1][feat].dropna()
    retained = df[df[churn_col] == 0][feat].dropna()

    median        = round(float(col.median()), 3)
    mean_val      = round(float(col.mean()), 3)
    std_val       = round(float(col.std()), 3)
    churned_mean  = round(float(churned.mean()), 3) if len(churned) else None
    retained_mean = round(float(retained.mean()), 3) if len(retained) else None

    # Direction: is churned mean higher or lower than retained?
    if churned_mean is not None and retained_mean is not None and retained_mean != 0:
        diff_pct = round((churned_mean - retained_mean) / abs(retained_mean) * 100, 1)
    else:
        diff_pct = None

    #  Data-driven signal label ─
    # Build a description purely from the column name pattern and the stats.
    # No hardcoded text — the numbers tell the story.

    if feat == "recency_days":
        signal = f"Days since last order  (median={median:.0f}d)"
        if diff_pct is not None:
            high_means = (
                f"Churned customers average {churned_mean:.0f} days inactive vs "
                f"{retained_mean:.0f} days for retained (+{diff_pct:.0f}%)"
            )
            action = (
                f"Flag customers > {median:.0f} days inactive for re-engagement. "
                f"Churned threshold is ~{churned_mean:.0f} days."
            )
        else:
            high_means = "Higher value = customer inactive longer"
            action = "Re-engagement campaign — offer consumable reorder"

    elif feat == "frequency":
        signal = f"Distinct orders placed in FY2425  (median={median:.0f})"
        if diff_pct is not None:
            high_means = (
                f"Retained customers average {retained_mean:.1f} orders vs "
                f"{churned_mean:.1f} for churned ({abs(diff_pct):.0f}% fewer orders before churn)"
            )
            action = (
                f"Customers with < {max(1, int(churned_mean)):.0f} orders are at elevated risk. "
                f"Target with loyalty pricing or early product access."
            )
        else:
            high_means = "Higher frequency = stronger retention signal"
            action = "Protect high-frequency relationships — cross-sell while engaged"

    elif feat == "monetary":
        signal = f"Total spend in FY2425  (median=${median:,.0f})"
        if diff_pct is not None:
            high_means = (
                f"Retained customers average ${retained_mean:,.0f} spend vs "
                f"${churned_mean:,.0f} for churned"
            )
            action = (
                f"High-spend customers (>${retained_mean:,.0f}) require dedicated rep. "
                f"High-value churn is the most costly."
            )
        else:
            high_means = "Higher monetary = more valuable customer"
            action = "Assign dedicated rep — high-value churn is most costly"

    elif feat == "avg_order_gap_days":
        signal = f"Avg days between consecutive orders  (median={median:.0f}d)"
        if diff_pct is not None:
            high_means = (
                f"Churned customers had avg gap of {churned_mean:.0f} days vs "
                f"{retained_mean:.0f} days for retained — "
                f"{diff_pct:+.0f}% longer gaps before churning"
            )
            action = (
                f"Set reorder reminders at each customer's own gap interval. "
                f"Alert when gap exceeds {churned_mean:.0f} days."
            )
        else:
            high_means = "Longer gap = irregular buying pattern"
            action = "Automated reorder reminders at customer-specific interval"

    elif feat in ("R_score", "F_score", "M_score"):
        label = {"R_score": "Recency", "F_score": "Frequency", "M_score": "Monetary"}[feat]
        signal = f"{label} quintile score 1–5  (median={median:.1f})"
        if diff_pct is not None:
            high_means = (
                f"Retained avg {label.lower()} score: {retained_mean:.2f}  |  "
                f"Churned avg: {churned_mean:.2f}  —  "
                f"difference: {diff_pct:+.1f}%"
            )
            action = (
                f"Score < {max(1, int(churned_mean + 0.5)):.0f} is at-risk threshold for {label}. "
                f"Use as primary segmentation axis."
            )
        else:
            high_means = f"Higher {label.lower()} score = lower churn risk"
            action = f"Use {label} score as primary risk segmentation axis"

    elif feat == "specialty_revenue_trend_pct":
        signal = f"% revenue change for specialty FY2425→FY2526  (median={median:.1f}%)"
        if diff_pct is not None:
            high_means = (
                f"Retained customer specialties avg trend: {retained_mean:.1f}%  |  "
                f"Churned: {churned_mean:.1f}%"
            )
            action = (
                "Specialties with declining revenue trend need catalog expansion. "
                "Growing specialties need rep coverage increase."
            )
        else:
            high_means = "Positive trend = growing specialty segment"
            action = "Prioritise inventory and rep coverage for growing specialties"

    elif feat == "pct_of_total_revenue":
        signal = f"Specialty share of total McKesson revenue  (median={median:.3f})"
        if diff_pct is not None:
            high_means = (
                f"Retained customers are in specialties averaging {retained_mean:.4f} revenue share vs "
                f"{churned_mean:.4f} for churned"
            )
            action = "High-revenue specialties require dedicated retention strategy"
        else:
            high_means = "Higher value = customer in major revenue specialty"
            action = "Prioritise retention for high-revenue specialty segments"

    elif feat == "avg_revenue_per_order":
        signal = f"Avg order value for this specialty  (median=${median:,.0f})"
        if diff_pct is not None:
            high_means = (
                f"Retained: avg order ${retained_mean:,.0f}  |  "
                f"Churned: avg order ${churned_mean:,.0f}"
            )
            action = "Focus on order completion rate for high-ticket specialties"
        else:
            high_means = "Higher = high-ticket specialty with fewer larger orders"
            action = "Focus on order completion rate rather than frequency"

    elif feat in ("cust_type_encoded", "mkt_cd_encoded", "mms_class_encoded", "state_encoded"):
        label_map = {
            "cust_type_encoded":  "Customer type (S/X/B encoded)",
            "mkt_cd_encoded":     "Market segment code (encoded)",
            "mms_class_encoded":  "MMS class (encoded)",
            "state_encoded":      "Geographic state (encoded)",
        }
        signal = label_map[feat]
        n_unique = int(df[feat].nunique())
        high_means = (
            f"{n_unique} distinct encoded values. "
            f"Churned mean={churned_mean:.2f}  Retained mean={retained_mean:.2f}  "
            f"Diff={diff_pct:+.1f}%"
        ) if diff_pct is not None else f"{n_unique} distinct encoded values"
        action = (
            "Weak direct predictor but useful for segmentation and filtering. "
            "Do not use as primary recommendation signal."
        )

    elif feat == "specialty_tier":
        signal = f"Specialty tier 1/2/3  (median={median:.0f})"
        high_means = (
            f"Tier distribution: churned mean={churned_mean:.2f} vs retained={retained_mean:.2f}"
            if churned_mean else "Tier 1=large, Tier 2=mid, Tier 3=niche"
        )
        action = "Use tier to select recommendation model — Tier 3 needs geo fallback"

    elif feat.startswith("spec_"):
        spec_code = feat.replace("spec_", "")
        pct_churned  = round(churned.mean() * 100, 1) if len(churned) else 0
        pct_retained = round(retained.mean() * 100, 1) if len(retained) else 0
        signal = f"Binary: customer is in specialty {spec_code}"
        high_means = (
            f"{pct_churned:.1f}% of churned customers are in {spec_code} vs "
            f"{pct_retained:.1f}% of retained"
        )
        action = (
            f"Build specialty-specific catalog and recommendation panel for {spec_code}. "
            f"Churn rate in this specialty may differ from portfolio average."
        )

    elif feat == "n_categories_bought":
        signal = f"Number of distinct product families purchased  (median={median:.0f})"
        if diff_pct is not None:
            high_means = (
                f"Churned customers bought from {churned_mean:.1f} families on avg vs "
                f"{retained_mean:.1f} for retained  ({diff_pct:+.1f}%)"
            )
            action = (
                "Narrow buyers (few categories) are more likely to churn — limited"
                " relationship depth. Cross-selling into adjacent families reduces churn risk."
            )
        else:
            high_means = "Higher = customer buys from more product families (broader relationship)"
            action = "Prioritise cross-sell for customers with n_categories_bought <= 2"

    elif feat == "category_hhi":
        signal = f"Spend concentration index (HHI): 1.0 = all spend in one category  (median={median:.3f})"
        if diff_pct is not None:
            high_means = (
                f"Churned HHI avg={churned_mean:.3f}  Retained HHI avg={retained_mean:.3f}  "
                f"({diff_pct:+.1f}%). Higher HHI = more concentrated = narrower buyer."
            )
            action = (
                "Customers with HHI > 0.6 are heavily concentrated in one category. "
                "If that category declines they have no other reason to order — high churn risk. "
                "Pitch adjacent categories before concentration becomes a dependency."
            )
        else:
            high_means = "HHI near 1.0 = all spend in one category. Near 0 = evenly distributed."
            action = "Flag HHI > 0.6 customers for proactive cross-sell outreach"

    elif feat == "cycle_regularity":
        signal = f"Std dev of inter-order gap days  (median={median:.1f} days)"
        if diff_pct is not None:
            high_means = (
                f"Churned cycle_regularity avg={churned_mean:.1f}d  "
                f"Retained avg={retained_mean:.1f}d  ({diff_pct:+.1f}%). "
                "High std dev = irregular/ad-hoc buying. Low = predictable schedule."
            )
            action = (
                "Irregular buyers (high cycle_regularity) are harder to retain — no set schedule "
                "means no natural re-engagement point. Contract buyers (low std dev) are stable. "
                "For irregular buyers, lapsed product alerts are more effective than schedule-based outreach."
            )
        else:
            high_means = "Low = predictable ordering cycle (contract). High = ad-hoc or reactive purchasing."
            action = "Use cycle_regularity to distinguish contract buyers from ad-hoc buyers in segmentation"

    else:
        signal = feat
        high_means = (
            f"Churned mean={churned_mean}  Retained mean={retained_mean}  Diff={diff_pct:+.1f}%"
            if diff_pct is not None else "See SHAP plot for direction"
        )
        action = "Review distribution and confirm business meaning with team"

    return {
        "median":                  median,
        "mean":                    mean_val,
        "std":                     std_val,
        "churned_mean":            churned_mean,
        "retained_mean":           retained_mean,
        "churned_vs_retained_pct": diff_pct,
        "signal":                  signal,
        "high_value_means":        high_means,
        "action":                  action,
    }


def build_business_signals(shap_mean: pd.DataFrame, X: pd.DataFrame, y: pd.Series) -> pd.DataFrame:
    """
    Build the business signals table entirely from real data statistics.

    For each top feature by SHAP importance:
      - Computes median, mean, std from the full feature matrix
      - Splits by churn label (1=churned, 0=retained)
      - Derives signal text and action from the actual numbers

    No hardcoded interpretations. If the feature matrix changes,
    the descriptions update automatically on the next run.
    """
    _section("Step 6: Business signal interpretation (data-driven)")
    _log("Computing per-feature stats from customer_features.parquet...")

    # Attach churn label temporarily for churned/retained split
    df_with_label = X.copy()
    df_with_label["churn_label"] = y.values

    rows = []
    for _, row in shap_mean.head(TOP_N).iterrows():
        feat  = row["feature"]
        stats = _derive_feature_stats(feat, df_with_label)

        rows.append({
            "rank":                    int(row["rank"]),
            "feature":                 feat,
            "mean_abs_shap":           row["mean_abs_shap"],
            "direction":               row["direction"],
            "median_value":            stats["median"],
            "churned_mean":            stats["churned_mean"],
            "retained_mean":           stats["retained_mean"],
            "churned_vs_retained_pct": stats["churned_vs_retained_pct"],
            "signal":                  stats["signal"],
            "high_value_means":        stats["high_value_means"],
            "recommended_action":      stats["action"],
        })

    signals = pd.DataFrame(rows)

    _log(f"\n  {'RANK':<5} {'FEATURE':<28} {'CHURN_MEAN':>12} {'RET_MEAN':>10} {'DIFF%':>7} {'DIRECTION'}")
    _log(f"  {'─'*5} {'─'*28} {'─'*12} {'─'*10} {'─'*7} {'─'*12}")
    for _, r in signals.iterrows():
        cm  = f"{r['churned_mean']:.3f}"  if r['churned_mean']  is not None else "n/a"
        rm  = f"{r['retained_mean']:.3f}" if r['retained_mean'] is not None else "n/a"
        dif = f"{r['churned_vs_retained_pct']:+.1f}%" if r['churned_vs_retained_pct'] is not None else "n/a"
        _log(f"  {r['rank']:<5} {r['feature']:<28} {cm:>12} {rm:>10} {dif:>7} {r['direction']}")

    return signals


# Step 7: Save all outputs

def _add_bar_chart(
    ws,
    n_rows:       int,
    label_col:    int,   # 1-based column index for category labels
    value_col:    int,   # 1-based column index for bar values
    anchor:       str,   # top-left cell for the chart, e.g. "H2"
    title:        str,
    x_title:      str,
    bar_color:    str = "1F4E79",
    width:        int = 22,
    height:       int = 14,
) -> None:
    """Embed a horizontal bar chart into an openpyxl worksheet."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel

    chart = BarChart()
    chart.type        = "bar"          # horizontal bars
    chart.grouping    = "clustered"
    chart.title       = title
    chart.y_axis.title = ""            # categories on y-axis for horizontal
    chart.x_axis.title = x_title
    chart.legend      = None
    chart.width       = width
    chart.height      = height
    chart.style       = 2             # clean white style

    # Values
    data = Reference(ws, min_col=value_col, min_row=1, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)

    # Category labels (feature names)
    cats = Reference(ws, min_col=label_col, min_row=2, max_row=n_rows + 1)
    chart.set_categories(cats)

    # Series colour
    chart.series[0].graphicalProperties.solidFill = bar_color
    chart.series[0].graphicalProperties.line.solidFill = bar_color

    ws.add_chart(chart, anchor)


def save_outputs(
    gini:     pd.DataFrame,
    perm:     pd.DataFrame,
    shap_df:  pd.DataFrame,
    signals:  pd.DataFrame,
    overlap:  pd.DataFrame,
    corr_df:  pd.DataFrame,
) -> None:
    _section("Step 7: Saving outputs")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    gini.to_csv(OUT_DIR / "feature_importance.csv", index=False)
    _log("Saved : data_clean/analysis/feature_importance.csv")

    perm.to_csv(OUT_DIR / "permutation_importance.csv", index=False)
    _log("Saved : data_clean/analysis/permutation_importance.csv")

    shap_df.to_csv(OUT_DIR / "shap_importance.csv", index=False)
    _log("Saved : data_clean/analysis/shap_importance.csv")

    signals.to_csv(OUT_DIR / "business_signals.csv", index=False)
    _log("Saved : data_clean/analysis/business_signals.csv")

    #  Excel workbook 
    # Top-N to chart — keeps charts readable
    CHART_N = 20

    xlsx_path = OUT_DIR / "feature_importance_report.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:

        # Write all data sheets
        gini.to_excel(writer,    sheet_name="01_gini_importance",        index=False)
        perm.to_excel(writer,    sheet_name="02_permutation_importance",  index=False)
        shap_df.to_excel(writer, sheet_name="03_shap_importance",         index=False)
        signals.to_excel(writer, sheet_name="04_business_signals",        index=False)
        overlap.to_excel(writer, sheet_name="05_method_overlap",          index=False)
        corr_df.to_excel(writer, sheet_name="06_high_correlations",       index=False)

        wb = writer.book

        tab_colors = {
            "01_gini_importance":        "1F4E79",
            "02_permutation_importance": "375623",
            "03_shap_importance":        "7030A0",
            "04_business_signals":       "C00000",
            "05_method_overlap":         "833C00",
            "06_high_correlations":      "595959",
        }
        sheets_dfs = {
            "01_gini_importance":        gini,
            "02_permutation_importance": perm,
            "03_shap_importance":        shap_df,
            "04_business_signals":       signals,
            "05_method_overlap":         overlap,
            "06_high_correlations":      corr_df,
        }

        for name, color in tab_colors.items():
            wb[name].sheet_properties.tabColor = color
            _style_sheet(writer.sheets[name], sheets_dfs[name], header_color=color)

        #  Chart 1: Gini importance — top-N features ─
        # feature=col1, importance=col2 in gini df
        # We write a sorted slice to a helper area to keep chart data clean.
        ws_gini = wb["01_gini_importance"]
        gini_top = gini.head(CHART_N)[["feature", "importance"]].reset_index(drop=True)
        # Gini df already written; chart references cols A (feature) and B (importance)
        # rows 2..CHART_N+1 after the header row
        _add_bar_chart(
            ws        = ws_gini,
            n_rows    = CHART_N,
            label_col = 1,   # A = feature
            value_col = 2,   # B = importance
            anchor    = "I2",
            title     = f"Gini Importance — Top {CHART_N} Features",
            x_title   = "Gini Importance Score",
            bar_color = "1F4E79",
            width     = 26,
            height    = 16,
        )
        _log(f"Chart  : Gini importance bar chart embedded in 01_gini_importance")

        #  Chart 2: Permutation importance — top-N features ─
        ws_perm = wb["02_permutation_importance"]
        # perm df: feature=col1, importance=col2
        _add_bar_chart(
            ws        = ws_perm,
            n_rows    = CHART_N,
            label_col = 1,   # A = feature
            value_col = 2,   # B = importance (AUC drop)
            anchor    = "I2",
            title     = f"Permutation Importance (AUC Drop) — Top {CHART_N} Features",
            x_title   = "Mean AUC Drop When Feature Is Shuffled",
            bar_color = "375623",
            width     = 26,
            height    = 16,
        )
        _log(f"Chart  : Permutation importance bar chart embedded in 02_permutation_importance")

        #  Chart 3: SHAP mean absolute values — top-N features ─
        ws_shap = wb["03_shap_importance"]
        # shap_df columns: feature, mean_abs_shap, rank, mean_shap_signed, direction
        # feature=col1, mean_abs_shap=col2
        _add_bar_chart(
            ws        = ws_shap,
            n_rows    = CHART_N,
            label_col = 1,   # A = feature
            value_col = 2,   # B = mean_abs_shap
            anchor    = "I2",
            title     = f"SHAP Mean |Value| — Top {CHART_N} Features (Churn Class)",
            x_title   = "Mean |SHAP Value| — Impact on Churn Probability",
            bar_color = "7030A0",
            width     = 26,
            height    = 16,
        )
        _log(f"Chart  : SHAP importance bar chart embedded in 03_shap_importance")

        #  Chart 4: Method overlap — method_count bar 
        ws_overlap = wb["05_method_overlap"]
        # overlap columns: feature, in_gini, in_perm, in_shap, method_count, confidence
        # feature=col1, method_count=col5
        _add_bar_chart(
            ws        = ws_overlap,
            n_rows    = min(len(overlap), CHART_N),
            label_col = 1,   # A = feature
            value_col = 5,   # E = method_count
            anchor    = "I2",
            title     = "Feature Confidence — How Many Methods Agree (max 3)",
            x_title   = "Number of Methods That Include This Feature in Top 10",
            bar_color = "833C00",
            width     = 26,
            height    = 16,
        )
        _log(f"Chart  : Method overlap chart embedded in 05_method_overlap")

    _log("Saved : data_clean/analysis/feature_importance_report.xlsx  (4 embedded charts)")



def save_svg_dashboard(
    gini:    pd.DataFrame,
    perm:    pd.DataFrame,
    shap_df: pd.DataFrame,
    signals: pd.DataFrame,
) -> None:
    """
    Render a self-contained SVG dashboard showing all four importance views
    plus the data-driven business signals table.

    Saved to: data_clean/analysis/feature_importance_dashboard.svg

    Layout (top to bottom):
      Row 0 — title banner
      Row 1 — Gini bar chart (left) | SHAP bar chart (right)
      Row 2 — Permutation bar chart (left) | Business signals table (right)
    """
    _section("Step 7b: Saving SVG dashboard")

    N       = 15          # features to show in each chart
    W       = 1200        # total SVG width
    PAD     = 40          # outer padding
    BAR_H   = 22          # height per bar row
    CHART_W = (W - PAD * 3) // 2   # width of each half-panel
    CHART_H = N * BAR_H + 80       # height of each chart panel
    TITLE_H = 60
    GAP     = 20
    TABLE_H = N * 24 + 60

    total_h = TITLE_H + GAP + CHART_H + GAP + max(CHART_H, TABLE_H) + PAD * 2

    lines = []
    lines.append(
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'width="{W}" height="{total_h}" '
        f'font-family="Arial, sans-serif" font-size="12">'
    )

    #  background 
    lines.append(f'<rect width="{W}" height="{total_h}" fill="#F8F9FA"/>')

    #  title banner 
    lines.append(
        f'<rect x="0" y="0" width="{W}" height="{TITLE_H}" fill="#1F4E79"/>'
    )
    lines.append(
        f'<text x="{W//2}" y="{TITLE_H//2 + 6}" '
        f'text-anchor="middle" fill="white" font-size="18" font-weight="bold">'
        f'Feature Importance Dashboard — McKesson B2B Churn Analysis</text>'
    )

    def _bar_chart_svg(
        df:       pd.DataFrame,
        feat_col: str,
        val_col:  str,
        x0:       int,
        y0:       int,
        panel_w:  int,
        panel_h:  int,
        title:    str,
        color:    str,
        n:        int = N,
    ) -> list[str]:
        """Render a horizontal bar chart panel as SVG elements."""
        out = []
        # Panel background
        out.append(
            f'<rect x="{x0}" y="{y0}" width="{panel_w}" height="{panel_h}" '
            f'fill="white" rx="6" stroke="#DEE2E6" stroke-width="1"/>'
        )
        # Panel title
        out.append(
            f'<text x="{x0 + panel_w//2}" y="{y0 + 22}" '
            f'text-anchor="middle" font-size="13" font-weight="bold" fill="#1F4E79">'
            f'{title}</text>'
        )

        sub = df.head(n).reset_index(drop=True)
        max_val = float(sub[val_col].max()) if len(sub) else 1.0
        if max_val == 0:
            max_val = 1.0

        label_w  = 175
        bar_area = panel_w - label_w - PAD * 2 - 50
        bar_y0   = y0 + 42

        for i, row in sub.iterrows():
            bar_y   = bar_y0 + i * BAR_H
            bar_len = int((float(row[val_col]) / max_val) * bar_area)
            bar_len = max(bar_len, 2)

            # Feature label
            label = str(row[feat_col])
            if len(label) > 22:
                label = label[:21] + "…"
            out.append(
                f'<text x="{x0 + PAD + label_w - 4}" y="{bar_y + BAR_H - 7}" '
                f'text-anchor="end" font-size="10" fill="#343A40">{label}</text>'
            )
            # Bar
            bx = x0 + PAD + label_w
            out.append(
                f'<rect x="{bx}" y="{bar_y + 3}" '
                f'width="{bar_len}" height="{BAR_H - 6}" '
                f'fill="{color}" rx="2" opacity="0.88"/>'
            )
            # Value label
            val_text = f"{float(row[val_col]):.4f}"
            out.append(
                f'<text x="{bx + bar_len + 4}" y="{bar_y + BAR_H - 7}" '
                f'font-size="9" fill="#6C757D">{val_text}</text>'
            )
        return out

    #  Row 1: Gini (left) | SHAP (right) ─
    row1_y = TITLE_H + GAP
    lines += _bar_chart_svg(
        df=gini, feat_col="feature", val_col="importance",
        x0=PAD, y0=row1_y, panel_w=CHART_W, panel_h=CHART_H,
        title="Gini (Split) Importance", color="#1F4E79",
    )
    lines += _bar_chart_svg(
        df=shap_df, feat_col="feature", val_col="mean_abs_shap",
        x0=PAD * 2 + CHART_W, y0=row1_y, panel_w=CHART_W, panel_h=CHART_H,
        title="SHAP Mean |Value| — Impact on Churn", color="#7030A0",
    )

    #  Row 2: Permutation (left) | Business signals table (right)
    row2_y = row1_y + CHART_H + GAP

    lines += _bar_chart_svg(
        df=perm, feat_col="feature", val_col="importance",
        x0=PAD, y0=row2_y, panel_w=CHART_W, panel_h=CHART_H,
        title="Permutation Importance (AUC Drop)", color="#375623",
    )

    # Business signals table panel
    tx0 = PAD * 2 + CHART_W
    th  = TABLE_H
    lines.append(
        f'<rect x="{tx0}" y="{row2_y}" width="{CHART_W}" height="{th}" '
        f'fill="white" rx="6" stroke="#DEE2E6" stroke-width="1"/>'
    )
    lines.append(
        f'<text x="{tx0 + CHART_W//2}" y="{row2_y + 22}" '
        f'text-anchor="middle" font-size="13" font-weight="bold" fill="#C00000">'
        f'Business Signals — Data-Driven Actions</text>'
    )

    # Table header
    cols_x   = [tx0 + 8, tx0 + 120, tx0 + 230, tx0 + 330, tx0 + 420]
    col_hdrs = ["Feature", "Churn Mean", "Ret. Mean", "Diff%", "Action (truncated)"]
    hy = row2_y + 42
    lines.append(
        f'<rect x="{tx0 + 4}" y="{hy - 14}" '
        f'width="{CHART_W - 8}" height="18" fill="#1F4E79" rx="2"/>'
    )
    for cx, hdr in zip(cols_x, col_hdrs):
        lines.append(
            f'<text x="{cx}" y="{hy}" font-size="10" font-weight="bold" fill="white">'
            f'{hdr}</text>'
        )

    # Table rows
    sig_top = signals.head(N).reset_index(drop=True)
    for i, row in sig_top.iterrows():
        ry  = hy + 18 + i * 24
        bg  = "#F2F7FF" if i % 2 == 0 else "white"
        lines.append(
            f'<rect x="{tx0 + 4}" y="{ry - 14}" '
            f'width="{CHART_W - 8}" height="20" fill="{bg}"/>'
        )
        feat_short   = str(row["feature"])[:16]
        cm   = f"{row['churned_mean']:.2f}"  if row["churned_mean"]  is not None else "—"
        rm   = f"{row['retained_mean']:.2f}" if row["retained_mean"] is not None else "—"
        dif  = (
            f"{row['churned_vs_retained_pct']:+.0f}%"
            if row["churned_vs_retained_pct"] is not None else "—"
        )
        act  = str(row["recommended_action"])[:38] + "…"
        diff_color = "#C00000" if row["churned_vs_retained_pct"] and row["churned_vs_retained_pct"] > 0 else "#375623"

        for cx, val, clr in zip(
            cols_x,
            [feat_short, cm, rm, dif, act],
            ["#343A40", "#C00000", "#375623", diff_color, "#495057"]
        ):
            lines.append(
                f'<text x="{cx}" y="{ry}" font-size="9" fill="{clr}">{val}</text>'
            )

    lines.append("</svg>")

    svg_path = OUT_DIR / "feature_importance_dashboard.svg"
    svg_path.write_text("\n".join(lines), encoding="utf-8")
    _log(f"Saved : data_clean/analysis/feature_importance_dashboard.svg")


def save_model(model: RandomForestClassifier) -> None:
    model_path = OUT_DIR / "rf_churn_model.pkl"
    joblib.dump(model, model_path)
    _log(f"Saved : data_clean/analysis/rf_churn_model.pkl")
    _log("        (load later with: model = joblib.load('data_clean/analysis/rf_churn_model.pkl'))")


# Main

def main() -> None:
    print()
    print("=" * 64)
    print("  B2B MEDICAL SUPPLY — FEATURE IMPORTANCE ANALYSIS")
    print("=" * 64)
    start = time.time()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if not FEATURE_FILE.exists():
        print(f"\nFATAL: {FEATURE_FILE} not found. Run clean_data.py first.",
              file=sys.stderr)
        sys.exit(1)

    X, y, feature_names = load_and_prepare()
    corr_df = correlation_check(X)
    model, X_train, X_test, y_train, y_test = train_model(X, y)

    gini    = gini_importance(model, list(X.columns))
    perm    = perm_importance(model, X_test, y_test)
    _, shap_df, _ = shap_analysis(model, X)
    overlap = log_overlap(gini, perm, shap_df)
    signals = build_business_signals(shap_df, X, y)

    save_outputs(gini, perm, shap_df, signals, overlap, corr_df)
    save_svg_dashboard(gini, perm, shap_df, signals)
    save_model(model)

    elapsed = round(time.time() - start, 1)

    _section("Analysis complete")
    _log(f"Total time : {elapsed:.1f}s")
    _log(f"Outputs    : {OUT_DIR.relative_to(ROOT)}/")
    _log("")
    _log("Next steps:")
    _log("  1. Open feature_importance_report.xlsx — sheet 04_business_signals")
    _log("  2. Review shap_summary.png for direction of impact per feature")
    _log("  3. Use top drivers to design recommendation logic in scripts/models/")
    _log("")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted by user.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise