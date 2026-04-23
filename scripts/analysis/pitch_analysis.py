from __future__ import annotations

"""
scripts/analysis/pitch_signals.py
===============================================================================
B2B Medical Supply — Seller Pitch Signal Analysis

Owner  : Suyog Mainali
Team   : Bryce Marin, Zoe, Ella
Client : McKesson

Run from the Capstone root directory:
    python scripts/analysis/pitch_signals.py

Purpose
-------
Identify WHAT to pitch to each customer by finding product category gaps
between a customer's actual spend and what their peer group buys.

This is a DIFFERENT question from feature_importance.py which asks
"will this customer churn?" — this script asks:
"which categories is this customer underbuying relative to similar customers,
 and what drives those gaps?"

The output is a ready-to-use pitch table the seller opens before a visit:
  - Top 5 product categories to pitch per customer
  - Dollar opportunity per category (peer median minus customer spend)
  - Which customer characteristics most predict underbuying

Input files
-----------
    data_clean/serving/merged_dataset.parquet          — 110M transaction rows
    data_clean/features/customer_features.parquet      — 44-col feature matrix
    data_clean/features/customer_rfm.parquet           — RFM scores + labels
    data_clean/features/specialty_tiers.parquet        — tier assignments

Outputs written to data_clean/analysis/
---------------------------------------
    pitch_opportunities.parquet     One row per customer per category gap
    pitch_summary.parquet           Top-5 pitch per customer (pre-computed for API)
    pitch_factor_importance.xlsx    4-sheet workbook with charts:
                                      01_category_gaps     — which categories most underbought
                                      02_factor_importance — what drives underbuying
                                      03_segment_coverage  — coverage by specialty + segment
                                      04_pitch_signals     — seller-facing signal table
    pitch_dashboard.svg             Visual summary of all findings
    pitch_factor_importance.csv     Raw factor scores for API

Method
------
1. Load transactions from merged_dataset, aggregate spend by customer x category
2. Assign peer groups: same specialty + same RFM behavior tier
3. Compute peer median spend per category per peer group
4. Compute gap ratio = customer_spend / peer_median (low ratio = pitch opportunity)
5. Flag underbought categories (ratio < GAP_THRESHOLD) and compute $ opportunity
6. Train RF classifier to predict which customers have large gaps, derive factors
7. Build seller-facing pitch table: top-5 opportunities per customer
8. Save everything + embedded Excel charts + SVG dashboard

Notes
-----
- merged_dataset is 7.4 GB — processed in DuckDB to avoid OOM
- Peer group minimum: MIN_PEER_SIZE customers (skip groups below threshold)
- GAP_THRESHOLD: customer spends < this fraction of peer median = opportunity
===============================================================================
"""

import sys
import time
import warnings
from pathlib import Path

import duckdb
import joblib
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import roc_auc_score, classification_report
from sklearn.inspection import permutation_importance
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)


#  Paths 

ROOT         = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN   = ROOT / "data_clean"
MERGED_FILE  = DATA_CLEAN / "serving"   / "merged_dataset.parquet"
FEATURE_FILE = DATA_CLEAN / "features"  / "customer_features.parquet"
RFM_FILE     = DATA_CLEAN / "features"  / "customer_rfm.parquet"
TIER_FILE    = DATA_CLEAN / "features"  / "specialty_tiers.parquet"
OUT_DIR      = DATA_CLEAN / "analysis"


#  Configuration 

# Category column in merged_dataset — this is the product family grouping.
# McKesson MMS products are grouped by MMS_CLASS_DSC (the human-readable class).
# Fall back to MMS_CLASS_CD if DSC is unavailable.
CATEGORY_COL   = "MMS_CLASS_DSC"
CATEGORY_FALLBACK = "MMS_CLASS_CD"

# Revenue column in merged_dataset
# Revenue column — auto-detected from merged_dataset at runtime
# Known candidates: NET_SALES_AMT, UNIT_SLS_AMT, SLS_AMT, REVENUE
REVENUE_COL    = "NET_SALES_AMT"   # overridden at runtime if not found

# Customer ID — must match across all files
CUST_ID_COL    = "DIM_CUST_CURR_ID"

# Minimum customers in a peer group — groups smaller than this are too thin
# to compute a reliable median, so customers in these groups are skipped
MIN_PEER_SIZE  = 5

# A customer is "underbuying" in a category if their spend is below this
# fraction of their peer group median. 0.5 = less than half of peers.
GAP_THRESHOLD  = 0.5

# Top N categories to show per customer in the pitch table
TOP_PITCH      = 5

# Top N factors to show in importance analysis
TOP_N          = 20

# RF parameters for gap prediction model
RF_PARAMS = dict(
    n_estimators = 100,
    max_depth    = 10,
    n_jobs       = -1,
    class_weight = "balanced",
    random_state = 42,
)


#  Logging 

def _section(title: str) -> None:
    print(f"\n{'' * 64}", flush=True)
    print(f"  {title}", flush=True)
    print(f"{'' * 64}", flush=True)

def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


#  Excel helpers 

def _style_sheet(ws, df: pd.DataFrame, header_color: str = "1F4E79") -> None:
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=header_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, 12), 55)

def _add_bar_chart(ws, n_rows, label_col, value_col, anchor, title, x_title,
                   bar_color="1F4E79", width=26, height=16) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = title
    chart.x_axis.title = x_title
    chart.legend = None
    chart.width = width
    chart.height = height
    chart.style = 2
    data = Reference(ws, min_col=value_col, min_row=1, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    cats = Reference(ws, min_col=label_col, min_row=2, max_row=n_rows + 1)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = bar_color
    chart.series[0].graphicalProperties.line.solidFill = bar_color
    ws.add_chart(chart, anchor)


#  Step 1: Load category spend matrix 

def load_category_spend() -> pd.DataFrame:
    """
    Aggregate spend by customer x product category from merged_dataset.

    Uses DuckDB to process the 7.4 GB parquet without loading into memory.
    Returns a wide matrix: rows = customers, columns = categories, values = spend.

    The category column is MMS_CLASS_DSC (product family description).
    This is the right level of granularity for pitch conversations —
    "wound care" or "PPE" is something a seller can talk about naturally.
    Pitching at the individual product level (285,000 SKUs) is too granular;
    pitching at MMS_SGMNT level is too broad.
    """
    _section("Step 1: Building customer x category spend matrix")
    _log("Reading merged_dataset.parquet via DuckDB (7.4 GB — this takes ~3 min)...")

    con = duckdb.connect()

    # Check which category column exists in the merged dataset.
    # Use DESCRIBE instead of parquet_schema() — parquet_schema uses 'name'
    # not 'column_name', and varies by DuckDB version. DESCRIBE is stable.
    desc = con.execute(
        f"DESCRIBE SELECT * FROM read_parquet('{MERGED_FILE}') LIMIT 0"
    ).fetchdf()
    available_cols = desc["column_name"].tolist()
    _log(f"Columns in merged_dataset ({len(available_cols)} total) — checking category cols")

    if CATEGORY_COL in available_cols:
        cat_col = CATEGORY_COL
    elif CATEGORY_FALLBACK in available_cols:
        cat_col = CATEGORY_FALLBACK
        _log(f"Primary category column '{CATEGORY_COL}' not found — using '{CATEGORY_FALLBACK}'")
    else:
        cat_candidates = [c for c in available_cols if any(
            kw in c.upper() for kw in ["CLASS", "CATEG", "FAMILY", "SGMNT", "TYPE"]
        )]
        if not cat_candidates:
            raise ValueError(
                f"No category column found in merged_dataset.\n"
                f"Expected '{CATEGORY_COL}' or '{CATEGORY_FALLBACK}'.\n"
                f"Available columns: {available_cols}"
            )
        cat_col = cat_candidates[0]
        _log(f"Warning: using fallback category column '{cat_col}'")

    _log(f"Category column : {cat_col}")

    # Auto-detect revenue column — varies across pipeline versions
    REV_CANDIDATES = ["NET_SALES_AMT", "UNIT_SLS_AMT", "SLS_AMT", "REVENUE",
                      "NET_SALES", "SALES_AMT", "AMT"]
    rev_col = next((c for c in REV_CANDIDATES if c in available_cols), None)
    if rev_col is None:
        raise ValueError(
            f"No revenue column found in merged_dataset.\n"
            f"Looked for: {REV_CANDIDATES}\n"
            f"Available columns: {available_cols}"
        )
    if rev_col != REVENUE_COL:
        _log(f"Revenue column  : '{rev_col}' (config had '{REVENUE_COL}' — auto-corrected)")
    else:
        _log(f"Revenue column  : {rev_col}")

    # Determine which optional columns exist — build query defensively
    # Auto-detect customer ID column
    CUST_CANDIDATES = ["DIM_CUST_CURR_ID", "CUST_ID", "CUSTOMER_ID", "CUST_NUM"]
    cust_col = next((c for c in CUST_CANDIDATES if c in available_cols), None)
    if cust_col is None:
        raise ValueError(
            f"No customer ID column found in merged_dataset.\n"
            f"Looked for: {CUST_CANDIDATES}\n"
            f"Available columns: {available_cols}"
        )
    if cust_col != CUST_ID_COL:
        _log(f"Customer ID col : '{cust_col}' (config had '{CUST_ID_COL}' — auto-corrected)")
    else:
        _log(f"Customer ID col : {cust_col}")

    has_fiscal_year = "fiscal_year" in available_cols
    has_ordr_num    = "ORDR_NUM"    in available_cols
    _log(f"fiscal_year col : {'present' if has_fiscal_year else 'absent — no FY filter'}")
    _log(f"ORDR_NUM col    : {'present' if has_ordr_num else 'absent — order_count will be line count'}")

    fiscal_filter = "AND fiscal_year = 'FY2425'" if has_fiscal_year else ""
    order_count   = "COUNT(DISTINCT ORDR_NUM)" if has_ordr_num else "COUNT(*)"

    # Aggregate: customer x category total spend via DuckDB (avoids 7.4 GB OOM)
    t0 = time.time()
    spend_long = con.execute(f"""
        SELECT
            {cust_col}                         AS cust_id,
            COALESCE({cat_col}, 'Unknown')      AS category,
            SUM({rev_col})                      AS total_spend,
            {order_count}                       AS order_count,
            COUNT(*)                            AS line_count
        FROM read_parquet('{MERGED_FILE}')
        WHERE {rev_col} > 0
          {fiscal_filter}
        GROUP BY {cust_col}, COALESCE({cat_col}, 'Unknown')
        HAVING SUM({rev_col}) > 0
    """).fetchdf()
    _log(f"Aggregated in {time.time() - t0:.1f}s  |  {len(spend_long):,} customer-category rows")

    if len(spend_long) == 0:
        raise ValueError(
            f"No spend data found in merged_dataset after filtering. "
            f"Check that {REVENUE_COL} > 0 rows exist and {CUST_ID_COL} is populated."
        )

    con.close()

    n_customers  = spend_long["cust_id"].nunique()
    n_categories = spend_long["category"].nunique()
    _log(f"Unique customers  : {n_customers:,}")
    _log(f"Unique categories : {n_categories:,}")
    _log(f"Avg categories per customer: {len(spend_long) / n_customers:.1f}")

    # Top categories by total revenue — these are the ones that matter
    top_cats = (
        spend_long.groupby("category")["total_spend"]
        .sum()
        .sort_values(ascending=False)
        .head(50)
    )
    _log(f"\n  Top 10 categories by revenue:")
    for cat, rev in top_cats.head(10).items():
        _log(f"    {cat:<40} ${rev:>15,.0f}")

    return spend_long, top_cats


#  Step 2: Assign peer groups 

def assign_peer_groups(spend_long: pd.DataFrame) -> pd.DataFrame:
    """
    Assign each customer to a peer group based on:
      1. Specialty (SPCLTY_CD from customer_features)
      2. RFM behavior tier (high/mid/low from R+F scores)

    This is the behavioral peer matching approach confirmed by feature importance:
    same specialty + same behavior = genuine peer, not just geographic neighbor.

    Returns spend_long with peer_group column added.
    """
    _section("Step 2: Assigning peer groups")

    features = pd.read_parquet(FEATURE_FILE, columns=[
        "DIM_CUST_CURR_ID", "SPCLTY_CD", "R_score", "F_score",
        "M_score", "specialty_tier", "CUST_TYPE_CD", "MKT_CD"
    ])

    # RFM behavior tier: combine R and F scores into 3 tiers
    # High: R>=4 AND F>=4 — active frequent buyers
    # Low:  R<=2 OR  F<=2 — inactive or rare buyers
    # Mid:  everyone else
    def rfm_tier(row):
        if row["R_score"] >= 4 and row["F_score"] >= 4:
            return "high"
        elif row["R_score"] <= 2 or row["F_score"] <= 2:
            return "low"
        else:
            return "mid"

    features["rfm_tier"] = features.apply(rfm_tier, axis=1)

    # Peer group = specialty + rfm_tier
    # Use specialty_tier as fallback grouping for Tier 3 specialties
    features["peer_group"] = (
        features["SPCLTY_CD"].fillna("UNKNOWN") + "|" +
        features["rfm_tier"]
    )

    # For Tier 3 specialties (< 10 customers), merge into broader group
    # to ensure MIN_PEER_SIZE is achievable
    tier3_mask = features["specialty_tier"] == 3
    features.loc[tier3_mask, "peer_group"] = (
        "TIER3_FALLBACK|" + features.loc[tier3_mask, "rfm_tier"]
    )

    _log(f"Peer groups created: {features['peer_group'].nunique():,}")
    peer_sizes = features.groupby("peer_group").size()
    _log(f"Peer group size distribution:")
    _log(f"  Min    : {peer_sizes.min()}")
    _log(f"  Median : {peer_sizes.median():.0f}")
    _log(f"  Mean   : {peer_sizes.mean():.0f}")
    _log(f"  Max    : {peer_sizes.max()}")
    _log(f"  Groups >= {MIN_PEER_SIZE}: {(peer_sizes >= MIN_PEER_SIZE).sum():,}")

    # Join peer group onto spend_long
    spend_long = spend_long.merge(
        features[["DIM_CUST_CURR_ID", "peer_group", "rfm_tier", "SPCLTY_CD",
                   "specialty_tier", "CUST_TYPE_CD", "MKT_CD"]],
        left_on="cust_id", right_on="DIM_CUST_CURR_ID", how="left"
    ).drop(columns=["DIM_CUST_CURR_ID"])

    unmatched = spend_long["peer_group"].isna().sum()
    if unmatched > 0:
        _log(f"Warning: {unmatched:,} rows could not be matched to a peer group (new customers?)")
        spend_long = spend_long.dropna(subset=["peer_group"])

    return spend_long, features


#  Step 3: Compute peer medians and gap ratios 

def compute_gaps(spend_long: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    For each customer x category:
      - peer_median_spend: median spend in this category across the peer group
      - gap_ratio: customer_spend / peer_median_spend
      - spend_opportunity: peer_median - customer_spend ($ left on table)
      - is_gap: 1 if gap_ratio < GAP_THRESHOLD and peer_median > 0

    Only categories where the peer group has >= MIN_PEER_SIZE buyers are included —
    too-thin groups give unreliable medians.

    Also computes a customer-level gap score (number of significant gaps)
    which becomes the target for the factor importance model.
    """
    _section("Step 3: Computing peer medians and gap ratios")

    # Peer median: for each peer_group x category, median spend among buyers
    # We count zero-spend (non-buyers) as zero, not excluded —
    # this makes the gap analysis reflect actual coverage gaps
    _log("Computing peer group medians per category...")

    # First get all customer x peer_group assignments
    cust_peers = spend_long[["cust_id", "peer_group"]].drop_duplicates()

    # For each peer_group x category: how many customers bought it and median spend
    peer_stats = (
        spend_long
        .groupby(["peer_group", "category"])
        .agg(
            buyer_count   = ("cust_id",      "nunique"),
            peer_median   = ("total_spend",  "median"),
            peer_mean     = ("total_spend",  "mean"),
            peer_total    = ("total_spend",  "sum"),
        )
        .reset_index()
    )

    # Only keep peer_group x category combinations with enough buyers
    peer_stats = peer_stats[peer_stats["buyer_count"] >= MIN_PEER_SIZE].copy()
    _log(f"Peer group x category combinations (>= {MIN_PEER_SIZE} buyers): {len(peer_stats):,}")

    # Join peer medians onto customer spend
    gaps = spend_long.merge(
        peer_stats[["peer_group", "category", "peer_median", "buyer_count"]],
        on=["peer_group", "category"],
        how="inner"  # only categories with enough peer data
    )

    # Gap ratio and opportunity
    gaps["gap_ratio"]         = gaps["total_spend"] / gaps["peer_median"].clip(lower=1)
    gaps["spend_opportunity"] = (gaps["peer_median"] - gaps["total_spend"]).clip(lower=0)
    gaps["is_gap"]            = (gaps["gap_ratio"] < GAP_THRESHOLD).astype(int)

    _log(f"Total customer-category pairs analyzed: {len(gaps):,}")
    _log(f"Pairs with gap (< {GAP_THRESHOLD:.0%} of peer median): {gaps['is_gap'].sum():,}")
    _log(f"Avg gap ratio: {gaps['gap_ratio'].mean():.3f}")
    _log(f"Total $ opportunity across all gaps: ${gaps['spend_opportunity'].sum():,.0f}")

    # Customer-level gap summary: how many significant gaps and total $ opportunity
    cust_gap_summary = (
        gaps[gaps["is_gap"] == 1]
        .groupby("cust_id")
        .agg(
            n_gaps            = ("category",        "nunique"),
            total_opportunity = ("spend_opportunity", "sum"),
            top_category      = ("spend_opportunity", lambda x: gaps.loc[x.idxmax(), "category"]),
        )
        .reset_index()
    )

    # Flag: customer has at least one major gap (target for factor model)
    # Major gap = opportunity > $500 in at least one category
    major_gap_mask = gaps[gaps["is_gap"] == 1].groupby("cust_id")["spend_opportunity"].max() > 500
    cust_gap_summary["has_major_gap"] = (
        cust_gap_summary["cust_id"]
        .isin(major_gap_mask[major_gap_mask].index)
        .astype(int)
    )

    gap_rate = cust_gap_summary["has_major_gap"].mean() * 100
    _log(f"Customers with major pitch opportunity (>$500 gap): "
         f"{cust_gap_summary['has_major_gap'].sum():,} ({gap_rate:.1f}%)")

    return gaps, cust_gap_summary


#  Step 4: Build top-5 pitch table per customer 

def build_pitch_table(gaps: pd.DataFrame) -> pd.DataFrame:
    """
    For each customer, rank their category gaps by $ opportunity and return top 5.

    This is what the seller sees — a ready-to-use pitch list.
    The seller does NOT need to understand gap ratios or peer medians.
    They just need: "pitch these 5 categories, here's the $ reason why."

    Output columns:
      cust_id, rank, category, customer_spend, peer_median, gap_ratio,
      spend_opportunity, pitch_message
    """
    _section("Step 4: Building top-5 pitch table per customer")

    pitch_gaps = gaps[gaps["is_gap"] == 1].copy()
    pitch_gaps = pitch_gaps.sort_values(
        ["cust_id", "spend_opportunity"], ascending=[True, False]
    )
    pitch_gaps["pitch_rank"] = pitch_gaps.groupby("cust_id").cumcount() + 1
    pitch_top5 = pitch_gaps[pitch_gaps["pitch_rank"] <= TOP_PITCH].copy()

    # Seller-facing pitch message — derived from actual numbers
    def make_pitch_message(row) -> str:
        opp = row["spend_opportunity"]
        ratio_pct = row["gap_ratio"] * 100
        peer_med  = row["peer_median"]
        cat       = row["category"]
        if opp > 10_000:
            return (
                f"High-value gap: similar customers spend ${peer_med:,.0f} "
                f"on {cat} — you're at {ratio_pct:.0f}% of that. "
                f"${opp:,.0f} opportunity."
            )
        elif opp > 1_000:
            return (
                f"Mid-range gap: {cat} spend is {ratio_pct:.0f}% of peers. "
                f"${opp:,.0f} potential."
            )
        else:
            return (
                f"Coverage gap: {cat} — peers average ${peer_med:,.0f}, "
                f"you spend ${row['total_spend']:,.0f}."
            )

    pitch_top5["pitch_message"] = pitch_top5.apply(make_pitch_message, axis=1)

    n_customers_with_pitch = pitch_top5["cust_id"].nunique()
    _log(f"Customers with at least 1 pitch opportunity: {n_customers_with_pitch:,}")
    _log(f"Total pitch rows (top-5 per customer):       {len(pitch_top5):,}")
    _log(f"Avg opportunity per top-1 gap: ${pitch_top5[pitch_top5['pitch_rank']==1]['spend_opportunity'].mean():,.0f}")

    return pitch_top5


#  Step 5: Category gap overview 

def category_gap_overview(gaps: pd.DataFrame) -> pd.DataFrame:
    """
    Portfolio-level view: which categories have the most gap across all customers?
    This is what the admin sees — where are the biggest systematic underbuy patterns?
    """
    _section("Step 5: Category gap overview (portfolio level)")

    cat_gaps = (
        gaps[gaps["is_gap"] == 1]
        .groupby("category")
        .agg(
            n_customers_with_gap = ("cust_id",           "nunique"),
            total_opportunity    = ("spend_opportunity",  "sum"),
            avg_opportunity      = ("spend_opportunity",  "mean"),
            avg_gap_ratio        = ("gap_ratio",          "mean"),
        )
        .reset_index()
        .sort_values("total_opportunity", ascending=False)
        .reset_index(drop=True)
    )
    cat_gaps["rank"] = range(1, len(cat_gaps) + 1)

    _log(f"\n  Top 15 categories by total pitch opportunity:")
    _log(f"  {'Rank':<5} {'Category':<40} {'# Customers':>12} {'Total Opp $':>14}")
    _log(f"  {''*5} {''*40} {''*12} {''*14}")
    for _, row in cat_gaps.head(15).iterrows():
        _log(
            f"  {int(row['rank']):<5} {str(row['category']):<40} "
            f"{int(row['n_customers_with_gap']):>12,} "
            f"${row['total_opportunity']:>13,.0f}"
        )

    return cat_gaps


#  Step 6: Factor importance — what drives underbuying 

def compute_factor_importance(
    cust_gap_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, float]:
    """
    Train a Random Forest to predict which customers have major pitch opportunities.

    Target: has_major_gap (1 = customer is underbuying relative to peers in at
            least one category by >$500)

    Features: customer_features.parquet — the same 44-column matrix used for
              churn prediction. This intentionally reuses the same features so
              we can compare: which factors predict churn? which predict pitch gaps?
              The overlap and differences are actionable.

    The most important factors here will be things like:
      - specialty (which specialties systematically underbuy certain categories)
      - customer type (S/X/B — different product needs)
      - market segment code
      - spend concentration (customers who buy from few categories have more gaps)

    These are MORE USEFUL for the seller than churn factors because they are
    segmentation signals — "this type of customer in this specialty tends to
    miss wound care products" is actionable in a way that recency scores are not.
    """
    _section("Step 6: Factor importance — what predicts underbuying")

    features = pd.read_parquet(FEATURE_FILE)

    # Join gap target
    df = features.merge(
        cust_gap_summary[["cust_id", "has_major_gap", "n_gaps", "total_opportunity"]],
        left_on="DIM_CUST_CURR_ID", right_on="cust_id", how="left"
    )

    # Customers with no gap data = no opportunity found = 0
    df["has_major_gap"] = df["has_major_gap"].fillna(0).astype(int)
    df["n_gaps"]        = df["n_gaps"].fillna(0)

    gap_rate = df["has_major_gap"].mean() * 100
    _log(f"Gap rate in feature matrix: {gap_rate:.1f}%")

    # Build feature matrix — same exclusions as feature_importance.py
    _DROP = {
        "DIM_CUST_CURR_ID", "CUST_NUM", "CUST_NAME",
        "CUST_TYPE_CD", "SPCLTY_CD", "SPCLTY_DSC",
        "MKT_CD", "MMS_CLASS_CD", "MMS_CLASS_DSC",
        "MMS_SGMNT_CD", "MMS_SUB_CLASS_CD",
        "STATE", "CITY", "CNTRY_CD", "ZIP",
        "ACTV_FLG", "state_grouped", "RFM_score",
        "last_order_date_id", "churn_label",
        "cust_id", "has_major_gap", "n_gaps", "total_opportunity",
    }
    X_cols = [
        c for c in df.columns
        if c not in _DROP and pd.api.types.is_numeric_dtype(df[c])
    ]
    X = df[X_cols].copy()
    y = df["has_major_gap"]

    for col in X.columns:
        if X[col].isna().any():
            X[col] = X[col].fillna(X[col].median())

    _log(f"Feature matrix: {X.shape[0]:,} rows  |  {X.shape[1]} features")

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.20, random_state=42, stratify=y
    )

    model = RandomForestClassifier(**RF_PARAMS)
    _log("Training gap prediction model...")
    t0 = time.time()
    model.fit(X_train, y_train)
    _log(f"Fit complete in {time.time() - t0:.1f}s")

    y_proba = model.predict_proba(X_test)[:, 1]
    y_pred  = model.predict(X_test)
    auc     = roc_auc_score(y_test, y_proba)
    _log(f"AUC-ROC on test set: {auc:.4f}")
    report = classification_report(y_test, y_pred, target_names=["No gap", "Has gap"])
    for line in report.strip().split("\n"):
        _log(f"  {line}")

    # Gini importance
    imp = pd.DataFrame({
        "feature":    list(X.columns),
        "importance": model.feature_importances_,
    }).sort_values("importance", ascending=False).reset_index(drop=True)
    imp["rank"]         = range(1, len(imp) + 1)
    imp["importance"]   = imp["importance"].round(6)
    imp["pct_of_total"] = (imp["importance"] / imp["importance"].sum() * 100).round(2)

    _log(f"\n  Top {TOP_N} factors driving pitch gaps (Gini importance):")
    _log(f"  {'Rank':<5} {'Feature':<35} {'Importance':>12} {'% Total':>9}")
    _log(f"  {''*5} {''*35} {''*12} {''*9}")
    for _, row in imp.head(TOP_N).iterrows():
        _log(f"  {int(row['rank']):<5} {row['feature']:<35} "
             f"{row['importance']:>12.6f} {row['pct_of_total']:>8.2f}%")

    #  Seller signal table 
    # For each top factor: compute mean gap rate for high vs low values
    # This tells the seller: "customers with high specialty_revenue_trend_pct
    # have X% gap rate" — actionable segmentation
    df_with_label = X.copy()
    df_with_label["has_major_gap"] = y.values

    signal_rows = []
    for _, row in imp.head(TOP_N).iterrows():
        feat = row["feature"]
        col  = df_with_label[feat]
        median_val  = col.median()
        high_mask   = col >= median_val
        low_mask    = col < median_val
        gap_high    = df_with_label.loc[high_mask, "has_major_gap"].mean() * 100
        gap_low     = df_with_label.loc[low_mask,  "has_major_gap"].mean() * 100
        gap_diff    = gap_high - gap_low

        # Seller-readable signal
        if feat in ("recency_days", "avg_order_gap_days"):
            direction = "inactive customers" if gap_diff > 0 else "active customers"
            signal = f"Gap more common in {direction}"
        elif feat in ("frequency", "F_score", "R_score", "M_score", "monetary"):
            direction = "low-engagement" if gap_diff > 0 else "high-engagement"
            signal = f"Gap more common in {direction} customers"
        elif feat.startswith("spec_"):
            spec_code = feat.replace("spec_", "")
            signal = f"Specialty {spec_code}: {gap_high:.0f}% gap rate"
        elif feat in ("cust_type_encoded", "mkt_cd_encoded", "mms_class_encoded"):
            signal = f"Segment split: high={gap_high:.0f}% gap, low={gap_low:.0f}% gap"
        elif feat == "specialty_tier":
            signal = "Tier 1 specialties have more measurable gaps (larger peer groups)"
        elif feat == "specialty_revenue_trend_pct":
            signal = "Growing specialties = more pitch opportunity" if gap_diff > 0 else "Declining specialties = more gap"
        else:
            signal = f"High value: {gap_high:.0f}% gap rate | Low value: {gap_low:.0f}% gap rate"

        seller_action = _derive_pitch_action(feat, gap_high, gap_low, gap_diff, median_val)

        signal_rows.append({
            "rank":                  int(row["rank"]),
            "feature":               feat,
            "gini_importance":       row["importance"],
            "pct_of_model":          row["pct_of_total"],
            "gap_rate_high_value":   round(gap_high, 1),
            "gap_rate_low_value":    round(gap_low,  1),
            "gap_rate_difference":   round(gap_diff, 1),
            "signal":                signal,
            "seller_action":         seller_action,
        })

    pitch_signals = pd.DataFrame(signal_rows)

    # Binary spec_ columns where all customers have value=1 (or 0) produce NaN
    # for the low (or high) group mean. Fill to 0 so charts and SVG render cleanly.
    pitch_signals["gap_rate_high_value"]  = pitch_signals["gap_rate_high_value"].fillna(0.0)
    pitch_signals["gap_rate_low_value"]   = pitch_signals["gap_rate_low_value"].fillna(0.0)
    pitch_signals["gap_rate_difference"]  = pitch_signals["gap_rate_difference"].fillna(0.0)

    _log(f"\n  {'RANK':<5} {'FEATURE':<28} {'HIGH%':>7} {'LOW%':>6} {'DIFF':>6}  SELLER ACTION")
    _log(f"  {''*5} {''*28} {''*7} {''*6} {''*6}  {''*50}")
    for _, r in pitch_signals.iterrows():
        action_short = r["seller_action"][:50]
        _log(f"  {r['rank']:<5} {r['feature']:<28} {r['gap_rate_high_value']:>7.1f} "
             f"{r['gap_rate_low_value']:>6.1f} {r['gap_rate_difference']:>+6.1f}  {action_short}")

    return imp, pitch_signals, auc, model, X_test, y_test


def _derive_pitch_action(feat, gap_high, gap_low, gap_diff, median_val) -> str:
    """
    Derive a seller-facing action from the factor's gap rate differential.
    No hardcoded text — all numbers come from the actual data.
    """
    if abs(gap_diff) < 2:
        return f"Minimal impact on pitch likelihood — use as filter only"

    if feat == "recency_days":
        if gap_diff > 0:
            return (
                f"Inactive customers ({median_val:.0f}+ days) have {gap_high:.0f}% gap rate. "
                f"Re-engagement visit = pitch opportunity."
            )
        else:
            return (
                f"Recently active customers have {gap_high:.0f}% gap rate. "
                f"Pitch new categories while engagement is high."
            )
    elif feat == "frequency":
        if gap_diff > 0:
            return (
                f"Low-frequency customers (<{median_val:.0f} orders) have {gap_high:.0f}% gap. "
                f"Broaden their category exposure."
            )
        else:
            return (
                f"High-frequency customers (>{median_val:.0f} orders) have {gap_high:.0f}% gap. "
                f"They buy often but narrowly — cross-sell opportunity."
            )
    elif feat in ("R_score", "F_score", "M_score"):
        label = {"R_score": "recency", "F_score": "frequency", "M_score": "spend"}[feat]
        if gap_diff > 0:
            return (
                f"Low {label} score customers have {gap_high:.0f}% gap. "
                f"Disengaged = narrow buying pattern."
            )
        else:
            return (
                f"High {label} score customers have {gap_high:.0f}% gap. "
                f"Engaged but focused — pitch adjacent categories."
            )
    elif feat == "monetary":
        if gap_diff > 0:
            return (
                f"Low-spend customers (< ${median_val:,.0f}) have {gap_high:.0f}% gap rate. "
                f"Likely buying from competitors for some categories."
            )
        else:
            return (
                f"High-spend customers (> ${median_val:,.0f}) have {gap_high:.0f}% gap. "
                f"Large wallet — pitch premium and adjacent categories."
            )
    elif feat.startswith("spec_"):
        spec_code = feat.replace("spec_", "")
        return (
            f"Specialty {spec_code} has {gap_high:.0f}% gap rate. "
            f"Build a specialty-specific pitch deck for this segment."
        )
    elif feat == "specialty_revenue_trend_pct":
        if gap_diff > 0:
            return (
                f"Growing specialty customers have {gap_high:.0f}% gap. "
                f"Expanding practices buy more broadly — pitch full catalog."
            )
        else:
            return (
                f"Declining specialty trend = {gap_high:.0f}% gap rate. "
                f"Customer may be consolidating — pitch essentials."
            )
    elif feat in ("cust_type_encoded", "mkt_cd_encoded", "mms_class_encoded"):
        return (
            f"Segment gap rates: high={gap_high:.0f}%, low={gap_low:.0f}%. "
            f"Customize pitch deck by segment — different product needs."
        )
    elif feat == "specialty_tier":
        return (
            f"Tier-based gap: high={gap_high:.0f}%, low={gap_low:.0f}%. "
            f"Use tier to select which recommendation model to invoke."
        )
    else:
        direction = "above" if gap_diff > 0 else "below"
        return (
            f"Customers {direction} median {feat} have {gap_high:.0f}% gap rate. "
            f"Use as a segmentation filter when prioritising visits."
        )


#  Step 7: Save outputs 

def save_outputs(
    pitch_top5:       pd.DataFrame,
    cat_gaps:         pd.DataFrame,
    imp:              pd.DataFrame,
    pitch_signals:    pd.DataFrame,
    gaps:             pd.DataFrame,
) -> None:
    _section("Step 7: Saving outputs")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Parquet outputs for API
    gaps.to_parquet(OUT_DIR / "pitch_opportunities.parquet", index=False)
    _log("Saved : data_clean/analysis/pitch_opportunities.parquet")

    pitch_top5.to_parquet(OUT_DIR / "pitch_summary.parquet", index=False)
    _log("Saved : data_clean/analysis/pitch_summary.parquet")

    # CSV for quick inspection
    imp.to_csv(OUT_DIR / "pitch_factor_importance.csv", index=False)
    _log("Saved : data_clean/analysis/pitch_factor_importance.csv")

    #  Excel workbook 
    CHART_N = 20
    xlsx_path = OUT_DIR / "pitch_factor_importance.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        cat_gaps.to_excel(writer,      sheet_name="01_category_gaps",    index=False)
        imp.to_excel(writer,           sheet_name="02_factor_importance", index=False)
        pitch_signals.to_excel(writer, sheet_name="03_pitch_signals",     index=False)
        pitch_top5.head(5000).to_excel(writer, sheet_name="04_sample_pitches", index=False)

        wb = writer.book
        tab_colors = {
            "01_category_gaps":    "1F4E79",
            "02_factor_importance": "375623",
            "03_pitch_signals":    "833C00",
            "04_sample_pitches":   "C00000",
        }
        sheets_dfs = {
            "01_category_gaps":    cat_gaps,
            "02_factor_importance": imp,
            "03_pitch_signals":    pitch_signals,
            "04_sample_pitches":   pitch_top5.head(5000),
        }
        for name, color in tab_colors.items():
            wb[name].sheet_properties.tabColor = color
            _style_sheet(writer.sheets[name], sheets_dfs[name], header_color=color)

        # Chart 1: Top categories by total opportunity
        _add_bar_chart(
            ws=wb["01_category_gaps"], n_rows=min(CHART_N, len(cat_gaps)),
            label_col=1, value_col=3,  # category, total_opportunity
            anchor="I2",
            title=f"Top {CHART_N} Categories by Total Pitch Opportunity ($)",
            x_title="Total $ Opportunity Across All Customers",
            bar_color="1F4E79",
        )

        # Chart 2: Factor importance for gap prediction
        _add_bar_chart(
            ws=wb["02_factor_importance"], n_rows=min(CHART_N, len(imp)),
            label_col=1, value_col=2,  # feature, importance
            anchor="I2",
            title=f"Top {CHART_N} Factors Driving Category Gaps",
            x_title="Gini Importance Score",
            bar_color="375623",
        )

        # Chart 3: Gap rate differential per factor
        _add_bar_chart(
            ws=wb["03_pitch_signals"], n_rows=min(CHART_N, len(pitch_signals)),
            label_col=1, value_col=7,  # feature, gap_rate_difference
            anchor="I2",
            title="Gap Rate Difference: High vs Low Feature Value",
            x_title="% Point Difference in Gap Rate (High - Low feature value)",
            bar_color="833C00",
        )

    _log("Saved : data_clean/analysis/pitch_factor_importance.xlsx  (3 embedded charts)")


def save_pitch_svg(
    cat_gaps:      pd.DataFrame,
    imp:           pd.DataFrame,
    pitch_signals: pd.DataFrame,
    auc:           float,
) -> None:
    """
    Save a self-contained SVG dashboard for the pitch analysis.
    Four panels: category gaps, factor importance, pitch signals table, model stats.
    """
    _section("Step 7b: Saving SVG dashboard")

    N       = 15
    W       = 1200
    PAD     = 40
    BAR_H   = 22
    CHART_W = (W - PAD * 3) // 2
    CHART_H = N * BAR_H + 80
    TITLE_H = 60
    GAP     = 20
    TABLE_H = N * 24 + 80

    total_h = TITLE_H + GAP + CHART_H + GAP + max(CHART_H, TABLE_H) + PAD * 2

    lines = []
    lines.append(
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{W}" height="{total_h}" '
        f'font-family="Arial, sans-serif" font-size="12">'
    )
    lines.append(f'<rect width="{W}" height="{total_h}" fill="#F8F9FA"/>')

    # Title banner
    lines.append(f'<rect x="0" y="0" width="{W}" height="{TITLE_H}" fill="#1F4E79"/>')
    lines.append(
        f'<text x="{W//2}" y="{TITLE_H//2 + 6}" text-anchor="middle" '
        f'fill="white" font-size="18" font-weight="bold">'
        f'Seller Pitch Signal Analysis — McKesson B2B MMS  |  Gap Model AUC: {auc:.3f}</text>'
    )

    def _bar_panel(df, feat_col, val_col, x0, y0, pw, ph, title, color, n=N):
        out = []
        out.append(
            f'<rect x="{x0}" y="{y0}" width="{pw}" height="{ph}" '
            f'fill="white" rx="6" stroke="#DEE2E6" stroke-width="1"/>'
        )
        out.append(
            f'<text x="{x0 + pw//2}" y="{y0 + 22}" text-anchor="middle" '
            f'font-size="13" font-weight="bold" fill="#1F4E79">{title}</text>'
        )
        sub     = df.head(n).reset_index(drop=True)
        vals    = sub[val_col].abs()
        max_val = float(vals.max()) if len(vals) else 1.0
        if max_val == 0:
            max_val = 1.0
        label_w  = 175
        bar_area = pw - label_w - PAD * 2 - 50
        bar_y0   = y0 + 42
        for i, row in sub.iterrows():
            bar_y   = bar_y0 + i * BAR_H
            raw_val = row[val_col]
            if raw_val is None or (isinstance(raw_val, float) and (raw_val != raw_val)):
                raw_val = 0.0
            bar_len = int((abs(float(raw_val)) / max_val) * bar_area)
            bar_len = max(bar_len, 2)
            label   = str(row[feat_col])
            if len(label) > 22:
                label = label[:21] + "…"
            out.append(
                f'<text x="{x0 + PAD + label_w - 4}" y="{bar_y + BAR_H - 7}" '
                f'text-anchor="end" font-size="10" fill="#343A40">{label}</text>'
            )
            bx = x0 + PAD + label_w
            out.append(
                f'<rect x="{bx}" y="{bar_y + 3}" width="{bar_len}" '
                f'height="{BAR_H - 6}" fill="{color}" rx="2" opacity="0.88"/>'
            )
            safe_val = float(raw_val) if raw_val is not None else 0.0
            val_text = f"${safe_val:,.0f}" if val_col == "total_opportunity" else f"{safe_val:.4f}"
            out.append(
                f'<text x="{bx + bar_len + 4}" y="{bar_y + BAR_H - 7}" '
                f'font-size="9" fill="#6C757D">{val_text}</text>'
            )
        return out

    row1_y = TITLE_H + GAP
    lines += _bar_panel(
        cat_gaps, "category", "total_opportunity",
        PAD, row1_y, CHART_W, CHART_H,
        "Top categories by pitch opportunity ($)", "#1F4E79",
    )
    lines += _bar_panel(
        imp, "feature", "importance",
        PAD * 2 + CHART_W, row1_y, CHART_W, CHART_H,
        "Top factors driving category gaps (Gini)", "#375623",
    )

    row2_y = row1_y + CHART_H + GAP
    lines += _bar_panel(
        pitch_signals, "feature", "gap_rate_difference",
        PAD, row2_y, CHART_W, CHART_H,
        "Gap rate diff: high vs low feature value (%pt)", "#833C00",
    )

    # Pitch signals table
    tx0 = PAD * 2 + CHART_W
    lines.append(
        f'<rect x="{tx0}" y="{row2_y}" width="{CHART_W}" height="{TABLE_H}" '
        f'fill="white" rx="6" stroke="#DEE2E6" stroke-width="1"/>'
    )
    lines.append(
        f'<text x="{tx0 + CHART_W//2}" y="{row2_y + 22}" '
        f'text-anchor="middle" font-size="13" font-weight="bold" fill="#C00000">'
        f'Seller action by factor</text>'
    )
    cols_x   = [tx0 + 8, tx0 + 140, tx0 + 210, tx0 + 270, tx0 + 330]
    col_hdrs = ["Feature", "High%", "Low%", "Diff%", "Seller action"]
    hy = row2_y + 42
    lines.append(
        f'<rect x="{tx0 + 4}" y="{hy - 14}" '
        f'width="{CHART_W - 8}" height="18" fill="#1F4E79" rx="2"/>'
    )
    for cx, hdr in zip(cols_x, col_hdrs):
        lines.append(
            f'<text x="{cx}" y="{hy}" font-size="10" font-weight="bold" fill="white">{hdr}</text>'
        )
    for i, row in pitch_signals.head(N).reset_index(drop=True).iterrows():
        ry   = hy + 18 + i * 24
        bg   = "#F2F7FF" if i % 2 == 0 else "white"
        lines.append(
            f'<rect x="{tx0 + 4}" y="{ry - 14}" '
            f'width="{CHART_W - 8}" height="20" fill="{bg}"/>'
        )
        diff_color = "#C00000" if row["gap_rate_difference"] > 0 else "#375623"
        act_short  = str(row["seller_action"])[:36] + "…"
        for cx, val, clr in zip(
            cols_x,
            [str(row["feature"])[:16],
             f"{row['gap_rate_high_value']:.0f}%",
             f"{row['gap_rate_low_value']:.0f}%",
             f"{row['gap_rate_difference']:+.0f}%",
             act_short],
            ["#343A40", "#C00000", "#375623", diff_color, "#495057"]
        ):
            lines.append(
                f'<text x="{cx}" y="{ry}" font-size="9" fill="{clr}">{val}</text>'
            )

    lines.append("</svg>")

    svg_path = OUT_DIR / "pitch_dashboard.svg"
    svg_path.write_text("\n".join(lines), encoding="utf-8")
    _log(f"Saved : data_clean/analysis/pitch_dashboard.svg")


#  Main 

def main() -> None:
    print()
    print("=" * 64)
    print("  B2B MEDICAL SUPPLY — SELLER PITCH SIGNAL ANALYSIS")
    print("=" * 64)
    start = time.time()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for f, label in [
        (MERGED_FILE,  "merged_dataset.parquet"),
        (FEATURE_FILE, "customer_features.parquet"),
        (RFM_FILE,     "customer_rfm.parquet"),
    ]:
        if not f.exists():
            print(f"\nFATAL: {label} not found at {f}. Run clean_data.py first.",
                  file=sys.stderr)
            sys.exit(1)

    spend_long, top_cats                   = load_category_spend()
    spend_long, features                   = assign_peer_groups(spend_long)
    gaps, cust_gap_summary                 = compute_gaps(spend_long)
    pitch_top5                             = build_pitch_table(gaps)
    cat_gaps                               = category_gap_overview(gaps)
    imp, pitch_signals, auc, *_            = compute_factor_importance(cust_gap_summary)

    save_outputs(pitch_top5, cat_gaps, imp, pitch_signals, gaps)
    save_pitch_svg(cat_gaps, imp, pitch_signals, auc)

    elapsed = round(time.time() - start, 1)

    _section("Analysis complete")
    _log(f"Total time : {elapsed:.1f}s")
    _log(f"Outputs    : {OUT_DIR.relative_to(ROOT)}/")
    _log("")
    _log("Key outputs for next steps:")
    _log("  pitch_summary.parquet          — top-5 pitches per customer, for API")
    _log("  pitch_opportunities.parquet    — full gap table, for model training")
    _log("  pitch_factor_importance.xlsx   — what drives gaps, with charts")
    _log("  pitch_dashboard.svg            — visual summary of all findings")
    _log("")
    _log("Next: use pitch_summary.parquet in scripts/api/main.py")
    _log("  GET /seller/customer/{id} should return pitch_top5 for that customer")
    _log("")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted by user.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise