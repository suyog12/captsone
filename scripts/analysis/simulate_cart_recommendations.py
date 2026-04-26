from __future__ import annotations

import sys
import time
import warnings
from pathlib import Path

import duckdb
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# Paths

ROOT       = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN = ROOT / "data_clean"
PRECOMP    = DATA_CLEAN / "serving" / "precomputed"
ANALYSIS   = DATA_CLEAN / "analysis"

MERGED_FILE       = DATA_CLEAN / "serving" / "merged_dataset.parquet"
RECS_FILE         = PRECOMP / "recommendations.parquet"
PRODUCTS_FILE     = PRECOMP / "product_segments.parquet"
PATTERNS_FILE     = PRECOMP / "customer_patterns.parquet"
SEGMENTS_FILE     = PRECOMP / "customer_segments.parquet"
COOC_FILE         = PRECOMP / "product_cooccurrence.parquet"
PB_EQUIV_FILE     = PRECOMP / "private_brand_equivalents.parquet"

OUT_TXT    = ANALYSIS / "cart_simulation.txt"
OUT_XLSX   = ANALYSIS / "cart_simulation.xlsx"
OUT_LOG    = ANALYSIS / "cart_simulation_terminal.log"


# Configuration

# Number of customers to simulate.  100 gives a representative mix across
# segments and size tiers without taking too long to run.
N_CUSTOMERS = 100

# Min unique items in a customer's order to use it as a cart.  Single-item
# orders are not interesting because the cart is the simplest possible case.
MIN_CART_ITEMS = 1

# Max cart size cap.  Real customer orders can have 50+ items; we don't need
# all of them to test cart logic.  Take up to 4.
MAX_CART_ITEMS = 4

# Min cooccurrence lift to surface as a complement.  Lower lift = weaker
# association.  1.5 means "items bought together 1.5x more often than chance".
MIN_LIFT_THRESHOLD = 1.5

# Min support (number of co-occurring orders) for cooccurrence to count.
# Below this, the lift is statistically noisy.
MIN_COOC_SUPPORT = 0.0001  # support_ab is a fraction (share of orders), not a count

# Random seed so results are reproducible
RANDOM_SEED = 42


# Logging

class _Tee:
    def __init__(self, *streams):
        self.streams = streams
    def write(self, data):
        for s in self.streams:
            try: s.write(data)
            except Exception: pass
    def flush(self):
        for s in self.streams:
            try: s.flush()
            except Exception: pass


def _setup_logging():
    OUT_LOG.parent.mkdir(parents=True, exist_ok=True)
    log_file = open(OUT_LOG, "w", encoding="utf-8", buffering=1)
    sys.stdout = _Tee(sys.__stdout__, log_file)
    sys.stderr = _Tee(sys.__stderr__, log_file)
    return log_file


def _s(title: str) -> None:
    print(f"\n{'-' * 70}\n  {title}\n{'-' * 70}", flush=True)


def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


# Step 1: Load data
def load_data() -> dict:
    _s("Step 1: Loading data")
    t0 = time.time()

    for f, label in [
        (RECS_FILE, "recommendations"),
        (MERGED_FILE, "merged_dataset"),
        (COOC_FILE, "product_cooccurrence"),
        (PB_EQUIV_FILE, "private_brand_equivalents"),
        (PRODUCTS_FILE, "product_segments"),
        (PATTERNS_FILE, "customer_patterns"),
        (SEGMENTS_FILE, "customer_segments"),
    ]:
        if not f.exists():
            print(f"\nFATAL: {label} file not found: {f}", file=sys.stderr)
            print("Run the full pipeline first.", file=sys.stderr)
            sys.exit(1)

    products = pd.read_parquet(PRODUCTS_FILE, columns=[
        "DIM_ITEM_E1_CURR_ID", "ITEM_DSC",
        "PROD_FMLY_LVL1_DSC", "PROD_CTGRY_LVL2_DSC",
        "SUPLR_ROLLUP_DSC", "is_private_brand",
    ])
    products["DIM_ITEM_E1_CURR_ID"] = products["DIM_ITEM_E1_CURR_ID"].astype("int64")
    products["is_mck_brand"] = products["is_private_brand"].fillna(0).astype(int)
    # Use SUPLR_ROLLUP_DSC as the manufacturer label
    products["MFR_NAME"] = products["SUPLR_ROLLUP_DSC"].astype(str)
    _log(f"product_segments         : {len(products):,} rows")

    patterns = pd.read_parquet(PATTERNS_FILE, columns=[
        "DIM_CUST_CURR_ID", "is_cold_start", "is_churned", "is_declining",
        "n_unique_products_total", "order_cadence_tier",
    ])
    patterns["DIM_CUST_CURR_ID"] = patterns["DIM_CUST_CURR_ID"].astype("int64")
    _log(f"customer_patterns        : {len(patterns):,} rows")

    segments = pd.read_parquet(SEGMENTS_FILE, columns=[
        "DIM_CUST_CURR_ID", "segment", "size_tier", "mkt_cd_clean",
    ])
    segments["DIM_CUST_CURR_ID"] = segments["DIM_CUST_CURR_ID"].astype("int64")
    _log(f"customer_segments        : {len(segments):,} rows")

    cooc = pd.read_parquet(COOC_FILE)
    # The actual column names in the parquet are product_a / product_b / support_ab,
    # not the older DIM_ITEM_E1_CURR_ID_A/_B / support naming. Alias them so the
    # rest of the script (which was written against the older names) keeps working.
    cooc = cooc.rename(columns={
        "product_a": "DIM_ITEM_E1_CURR_ID_A",
        "product_b": "DIM_ITEM_E1_CURR_ID_B",
        "support_ab": "support",
    })
    cooc["DIM_ITEM_E1_CURR_ID_A"] = cooc["DIM_ITEM_E1_CURR_ID_A"].astype("int64")
    cooc["DIM_ITEM_E1_CURR_ID_B"] = cooc["DIM_ITEM_E1_CURR_ID_B"].astype("int64")
    _log(f"product_cooccurrence     : {len(cooc):,} pairs")

    pb_equiv = pd.read_parquet(PB_EQUIV_FILE)
    # Actual columns: original_item_id / equivalent_item_id (with match_type to
    # split private_brand_upgrade vs medline_conversion). Alias to the older
    # _BRAND/_PB naming the rest of the script expects, treating "original" as
    # the branded item and "equivalent" as the McKesson alternative.
    pb_equiv = pb_equiv.rename(columns={
        "original_item_id":   "DIM_ITEM_E1_CURR_ID_BRAND",
        "equivalent_item_id": "DIM_ITEM_E1_CURR_ID_PB",
    })
    pb_equiv["DIM_ITEM_E1_CURR_ID_BRAND"] = pb_equiv["DIM_ITEM_E1_CURR_ID_BRAND"].astype("int64")
    pb_equiv["DIM_ITEM_E1_CURR_ID_PB"]    = pb_equiv["DIM_ITEM_E1_CURR_ID_PB"].astype("int64")
    _log(f"private_brand_equivalents: {len(pb_equiv):,} pairs")

    recs = pd.read_parquet(RECS_FILE)
    recs["DIM_CUST_CURR_ID"]    = recs["DIM_CUST_CURR_ID"].astype("int64")
    recs["DIM_ITEM_E1_CURR_ID"] = recs["DIM_ITEM_E1_CURR_ID"].astype("int64")
    _log(f"recommendations          : {len(recs):,} rows")

    _log(f"Step 1 done in {time.time()-t0:.1f}s")

    # Build customer pool with status + segment + size_tier joined together
    cust = patterns.merge(segments, on="DIM_CUST_CURR_ID", how="inner")
    cust["status"] = "stable_warm"
    cust.loc[cust["is_cold_start"] == 1, "status"] = "cold_start"
    cust.loc[(cust["is_cold_start"] == 0) & (cust["is_declining"] == 1), "status"] = "declining_warm"
    cust.loc[(cust["is_cold_start"] == 0) & (cust["is_churned"] == 1), "status"] = "churned_warm"

    return {
        "products": products,
        "customers": cust,
        "cooc": cooc,
        "pb_equiv": pb_equiv,
        "recs": recs,
    }


# Step 2: Pick stratified customers
def pick_customers(customers: pd.DataFrame, recs: pd.DataFrame) -> pd.DataFrame:
    """Stratified sample across (size_tier, status). Only customers with
    recommendations available."""
    _s("Step 2: Picking stratified customer sample")
    t0 = time.time()

    rng = np.random.default_rng(RANDOM_SEED)

    # Only consider customers who actually have recommendations
    cust_with_recs = set(recs["DIM_CUST_CURR_ID"].unique())
    pool = customers[customers["DIM_CUST_CURR_ID"].isin(cust_with_recs)].copy()

    _log(f"Customers with recommendations available: {len(pool):,}")

    # Stratify across (size_tier, status)
    strata_counts = pool.groupby(["size_tier", "status"]).size().reset_index(name="n")
    _log(f"Strata available:")
    for _, r in strata_counts.iterrows():
        _log(f"  size={r['size_tier']:<10} status={r['status']:<15}  n={r['n']:>8,}")

    # Allocate N_CUSTOMERS proportionally with min 2 per stratum if possible
    selected = []
    total_strata = len(strata_counts)
    base_per_stratum = max(2, N_CUSTOMERS // total_strata)
    remaining_budget = N_CUSTOMERS

    for _, stratum in strata_counts.iterrows():
        size = stratum["size_tier"]
        status = stratum["status"]
        avail = stratum["n"]
        take = min(base_per_stratum, avail, remaining_budget)
        if take <= 0:
            continue
        pool_strat = pool[(pool["size_tier"] == size) & (pool["status"] == status)]
        chosen = pool_strat.sample(n=take, random_state=int(rng.integers(0, 1_000_000)))
        selected.append(chosen)
        remaining_budget -= take

    if remaining_budget > 0:
        # fill remaining with random sample from largest strata
        already = pd.concat(selected) if selected else pd.DataFrame()
        leftover_pool = pool[~pool["DIM_CUST_CURR_ID"].isin(already["DIM_CUST_CURR_ID"])]
        if len(leftover_pool) > 0 and remaining_budget > 0:
            extra = leftover_pool.sample(
                n=min(remaining_budget, len(leftover_pool)),
                random_state=int(rng.integers(0, 1_000_000))
            )
            selected.append(extra)

    chosen = pd.concat(selected, ignore_index=True)
    _log(f"Selected {len(chosen)} customers across {chosen.groupby(['size_tier','status']).ngroups} strata")
    _log(f"Step 2 done in {time.time()-t0:.1f}s")

    return chosen


# Step 3: Build a real recent order per customer (the simulated cart)
def build_carts(chosen: pd.DataFrame, products: pd.DataFrame) -> dict:
    """For each chosen customer, find a real recent order and use up to
    MAX_CART_ITEMS items from it as the simulated cart."""
    _s("Step 3: Building simulated carts from real recent orders")
    t0 = time.time()

    cust_ids = chosen["DIM_CUST_CURR_ID"].tolist()
    cust_id_str = ",".join(str(c) for c in cust_ids)

    con = duckdb.connect()
    # For each customer, find their most recent order (by date), then take
    # up to MAX_CART_ITEMS distinct items from it (sorted by line spend desc
    # so we take the most "primary" items in the order).
    orders = con.execute(f"""
        WITH customer_lines AS (
            SELECT
                CAST(DIM_CUST_CURR_ID AS BIGINT)    AS DIM_CUST_CURR_ID,
                CAST(DIM_ITEM_E1_CURR_ID AS BIGINT) AS DIM_ITEM_E1_CURR_ID,
                MAKE_DATE(order_year, order_month, order_day) AS order_dt,
                UNIT_SLS_AMT
            FROM read_parquet('{MERGED_FILE.as_posix()}')
            WHERE DIM_CUST_CURR_ID IN ({cust_id_str})
              AND UNIT_SLS_AMT > 0
              AND DIM_ITEM_E1_CURR_ID IS NOT NULL
        ),
        last_order_per_cust AS (
            SELECT
                DIM_CUST_CURR_ID,
                MAX(order_dt) AS max_dt
            FROM customer_lines
            GROUP BY DIM_CUST_CURR_ID
        ),
        last_order_lines AS (
            SELECT
                cl.DIM_CUST_CURR_ID,
                cl.DIM_ITEM_E1_CURR_ID,
                cl.order_dt,
                SUM(cl.UNIT_SLS_AMT) AS line_amt
            FROM customer_lines cl
            JOIN last_order_per_cust lo
              ON cl.DIM_CUST_CURR_ID = lo.DIM_CUST_CURR_ID
             AND cl.order_dt          = lo.max_dt
            GROUP BY cl.DIM_CUST_CURR_ID, cl.DIM_ITEM_E1_CURR_ID, cl.order_dt
        )
        SELECT * FROM last_order_lines
    """).df()
    con.close()

    orders["DIM_CUST_CURR_ID"]    = orders["DIM_CUST_CURR_ID"].astype("int64")
    orders["DIM_ITEM_E1_CURR_ID"] = orders["DIM_ITEM_E1_CURR_ID"].astype("int64")

    # Build product lookup for descriptions/families/categories/MCK brand flag
    prod_lookup = products.set_index("DIM_ITEM_E1_CURR_ID").to_dict("index")

    carts = {}
    for cid, grp in orders.groupby("DIM_CUST_CURR_ID"):
        # Take up to MAX_CART_ITEMS items, sorted by line amount desc
        items = grp.nlargest(MAX_CART_ITEMS, "line_amt")["DIM_ITEM_E1_CURR_ID"].tolist()
        if len(items) < MIN_CART_ITEMS:
            continue
        order_dt = grp["order_dt"].iloc[0]
        items_meta = []
        for item_id in items:
            info = prod_lookup.get(int(item_id), {})
            items_meta.append({
                "item_id": int(item_id),
                "desc": str(info.get("ITEM_DSC", "?"))[:60],
                "family": info.get("PROD_FMLY_LVL1_DSC"),
                "category": info.get("PROD_CTGRY_LVL2_DSC"),
                "is_mck_brand": int(info.get("is_mck_brand", 0)),
                "mfr": str(info.get("MFR_NAME", ""))[:30],
            })
        carts[int(cid)] = {
            "order_dt": str(pd.to_datetime(order_dt).date()),
            "items": items_meta,
            "n_items_in_order": len(grp),  # actual order may have more
        }

    _log(f"Built carts for {len(carts)} customers")
    _log(f"  (some customers may have been excluded if no purchase history found)")

    # Cart size distribution
    sizes = [len(c["items"]) for c in carts.values()]
    if sizes:
        _log(f"  Cart size distribution: min={min(sizes)}, "
             f"avg={np.mean(sizes):.1f}, max={max(sizes)}")

    _log(f"Step 3 done in {time.time()-t0:.1f}s")
    return carts


# Step 4: Run cart-aware recommendation logic
def run_cart_logic(carts: dict,
                    cooc: pd.DataFrame,
                    pb_equiv: pd.DataFrame,
                    products: pd.DataFrame) -> dict:
    """For each cart, compute:
      - top 5 cart complements (from cooccurrence)
      - top 3 PB upgrades (for each cart item, what's the equivalent PB?)
      - top 3 Medline conversions (for each Medline cart item, what's the McK alternative?)
    """
    _s("Step 4: Running cart-aware recommendation logic")
    t0 = time.time()

    # Index cooccurrence by item_A for fast lookup
    cooc_filt = cooc[
        (cooc["lift"] >= MIN_LIFT_THRESHOLD) &
        (cooc["support"] >= MIN_COOC_SUPPORT)
    ].copy()
    _log(f"Cooccurrence pairs after filter (lift >= {MIN_LIFT_THRESHOLD}, support >= {MIN_COOC_SUPPORT}): {len(cooc_filt):,}")
    cooc_indexed = cooc_filt.set_index("DIM_ITEM_E1_CURR_ID_A")

    # Index PB equivalents by brand item
    pb_indexed = pb_equiv.set_index("DIM_ITEM_E1_CURR_ID_BRAND")
    # Filter to only branded->PB direction (where the brand item has McK PB equivalent)
    _log(f"PB equivalent pairs available: {len(pb_equiv):,}")

    prod_lookup = products.set_index("DIM_ITEM_E1_CURR_ID").to_dict("index")

    def _meta(item_id):
        info = prod_lookup.get(int(item_id), {})
        return {
            "desc": str(info.get("ITEM_DSC", "?"))[:60],
            "family": info.get("PROD_FMLY_LVL1_DSC"),
            "category": info.get("PROD_CTGRY_LVL2_DSC"),
            "is_mck_brand": int(info.get("is_mck_brand", 0)),
            "mfr": str(info.get("MFR_NAME", ""))[:30],
        }

    results = {}
    for cid, cart in carts.items():
        cart_item_ids = [it["item_id"] for it in cart["items"]]
        cart_id_set = set(cart_item_ids)
        cart_categories = set(it["category"] for it in cart["items"] if it["category"])

        # 1. CART COMPLEMENTS via cooccurrence
        complements_rows = []
        for cart_item in cart_item_ids:
            if cart_item not in cooc_indexed.index:
                continue
            partners = cooc_indexed.loc[[cart_item]]  # double-bracket guarantees DataFrame
            for _, row in partners.iterrows():
                partner_id = int(row["DIM_ITEM_E1_CURR_ID_B"])
                if partner_id in cart_id_set:
                    continue  # already in cart
                m = _meta(partner_id)
                complements_rows.append({
                    "trigger_item_id": cart_item,
                    "trigger_desc": _meta(cart_item)["desc"],
                    "partner_item_id": partner_id,
                    "partner_desc": m["desc"],
                    "partner_family": m["family"],
                    "partner_category": m["category"],
                    "partner_is_mck_brand": m["is_mck_brand"],
                    "partner_mfr": m["mfr"],
                    "lift": float(row["lift"]),
                    "support": float(row.get("support", 0) or 0),
                    "confidence": float(row.get("confidence_a_to_b", row.get("confidence", 0)) or 0),
                })
        complements_df = pd.DataFrame(complements_rows)
        if not complements_df.empty:
            # de-dup by partner_item, keeping the highest-lift trigger pair
            complements_df = complements_df.sort_values("lift", ascending=False)
            complements_df = complements_df.drop_duplicates(subset=["partner_item_id"], keep="first")
            complements_df = complements_df.head(5)

        # 2. PB UPGRADES for branded items in cart
        pb_rows = []
        for cart_item in cart_item_ids:
            if cart_item not in pb_indexed.index:
                continue
            cart_item_meta = _meta(cart_item)
            if cart_item_meta["is_mck_brand"] == 1:
                continue  # already MCK brand, no upgrade needed
            equivs = pb_indexed.loc[[cart_item]]
            for _, row in equivs.iterrows():
                pb_id = int(row["DIM_ITEM_E1_CURR_ID_PB"])
                m = _meta(pb_id)
                pb_rows.append({
                    "branded_item_id": cart_item,
                    "branded_desc": cart_item_meta["desc"],
                    "branded_mfr": cart_item_meta["mfr"],
                    "pb_item_id": pb_id,
                    "pb_desc": m["desc"],
                    "pb_mfr": m["mfr"],
                    "category_match": m["category"] == cart_item_meta["category"],
                })
        pb_df = pd.DataFrame(pb_rows)
        if not pb_df.empty:
            pb_df = pb_df.drop_duplicates(subset=["pb_item_id"]).head(3)

        # 3. MEDLINE CONVERSIONS - any Medline item in cart, suggest MCK equivalent
        medline_rows = []
        for cart_item in cart_item_ids:
            cart_item_meta = _meta(cart_item)
            if "MEDLINE" not in cart_item_meta["mfr"].upper():
                continue
            # Look in PB equivalents for an MCK alternative in same category
            if cart_item in pb_indexed.index:
                equivs = pb_indexed.loc[[cart_item]]
                for _, row in equivs.iterrows():
                    pb_id = int(row["DIM_ITEM_E1_CURR_ID_PB"])
                    m = _meta(pb_id)
                    medline_rows.append({
                        "medline_item_id": cart_item,
                        "medline_desc": cart_item_meta["desc"],
                        "mck_item_id": pb_id,
                        "mck_desc": m["desc"],
                        "mck_mfr": m["mfr"],
                    })
        medline_df = pd.DataFrame(medline_rows)
        if not medline_df.empty:
            medline_df = medline_df.drop_duplicates(subset=["mck_item_id"]).head(3)

        # Compute summary stats
        n_complements = len(complements_df)
        n_pb_upgrades = len(pb_df)
        n_medline_convs = len(medline_df)

        # Category alignment: of complements, what fraction are in same category as ANY cart item
        if n_complements > 0:
            comp_in_cart_cat = sum(
                1 for cat in complements_df["partner_category"]
                if cat in cart_categories
            )
            cat_align_rate = comp_in_cart_cat / n_complements
        else:
            cat_align_rate = None

        # MCK Brand penetration of complements
        if n_complements > 0:
            mck_brand_rate = complements_df["partner_is_mck_brand"].mean()
            avg_lift = complements_df["lift"].mean()
        else:
            mck_brand_rate = None
            avg_lift = None

        results[cid] = {
            "cart": cart,
            "complements": complements_df,
            "pb_upgrades": pb_df,
            "medline_conversions": medline_df,
            "summary": {
                "n_complements_surfaced": n_complements,
                "n_pb_upgrades_surfaced": n_pb_upgrades,
                "n_medline_conversions_surfaced": n_medline_convs,
                "category_alignment_rate": cat_align_rate,
                "mck_brand_rate_of_complements": mck_brand_rate,
                "avg_complement_lift": avg_lift,
            },
        }

    _log(f"Processed {len(results)} customers")

    # Aggregate stats
    n_with_complements = sum(1 for r in results.values() if r["summary"]["n_complements_surfaced"] > 0)
    n_with_pb = sum(1 for r in results.values() if r["summary"]["n_pb_upgrades_surfaced"] > 0)
    n_with_medline = sum(1 for r in results.values() if r["summary"]["n_medline_conversions_surfaced"] > 0)

    _log(f"")
    _log(f"OPPORTUNITY RATES (across {len(results)} customers):")
    _log(f"  Cart had at least 1 complement     : {n_with_complements:>3} ({n_with_complements/len(results)*100:.1f}%)")
    _log(f"  Cart had at least 1 PB upgrade     : {n_with_pb:>3} ({n_with_pb/len(results)*100:.1f}%)")
    _log(f"  Cart had at least 1 Medline conv   : {n_with_medline:>3} ({n_with_medline/len(results)*100:.1f}%)")

    # Avg lift across all complements
    all_lifts = []
    all_align = []
    all_mck = []
    for r in results.values():
        if r["summary"]["avg_complement_lift"] is not None:
            all_lifts.append(r["summary"]["avg_complement_lift"])
        if r["summary"]["category_alignment_rate"] is not None:
            all_align.append(r["summary"]["category_alignment_rate"])
        if r["summary"]["mck_brand_rate_of_complements"] is not None:
            all_mck.append(r["summary"]["mck_brand_rate_of_complements"])

    if all_lifts:
        _log(f"  Avg complement lift (per cust avg) : {np.mean(all_lifts):.2f}")
        _log(f"  Median complement lift             : {np.median(all_lifts):.2f}")
    if all_align:
        _log(f"  Avg category alignment             : {np.mean(all_align)*100:.1f}%")
    if all_mck:
        _log(f"  Avg MCK Brand penetration of comps : {np.mean(all_mck)*100:.1f}%")

    _log(f"Step 4 done in {time.time()-t0:.1f}s")

    return results


# Step 5: Build per-customer rows for output
def build_per_customer_rows(results: dict, chosen: pd.DataFrame) -> pd.DataFrame:
    rows = []
    chosen_lookup = chosen.set_index("DIM_CUST_CURR_ID").to_dict("index")
    for cid, r in results.items():
        ctx = chosen_lookup.get(cid, {})
        cart = r["cart"]
        s = r["summary"]
        cart_item_descs = " | ".join(it["desc"][:30] for it in cart["items"])
        cart_categories = ", ".join(set(
            it["category"] for it in cart["items"] if it["category"]
        ))[:80]
        # Top complement summary
        if not r["complements"].empty:
            top_comp = r["complements"].iloc[0]
            top_comp_str = f"{top_comp['partner_desc'][:40]} (lift={top_comp['lift']:.1f})"
        else:
            top_comp_str = ""
        # Top PB
        if not r["pb_upgrades"].empty:
            top_pb = r["pb_upgrades"].iloc[0]
            top_pb_str = f"{top_pb['branded_desc'][:25]} -> {top_pb['pb_desc'][:25]}"
        else:
            top_pb_str = ""
        # Top medline
        if not r["medline_conversions"].empty:
            top_med = r["medline_conversions"].iloc[0]
            top_med_str = f"{top_med['medline_desc'][:25]} -> {top_med['mck_desc'][:25]}"
        else:
            top_med_str = ""

        rows.append({
            "DIM_CUST_CURR_ID": cid,
            "size_tier":  ctx.get("size_tier", ""),
            "status":     ctx.get("status", ""),
            "segment":    ctx.get("segment", ""),
            "mkt":        ctx.get("mkt_cd_clean", ""),
            "cadence":    ctx.get("order_cadence_tier", ""),
            "cart_date":  cart["order_dt"],
            "cart_size":  len(cart["items"]),
            "cart_items": cart_item_descs,
            "cart_categories": cart_categories,
            "n_complements_surfaced": s["n_complements_surfaced"],
            "n_pb_upgrades_surfaced": s["n_pb_upgrades_surfaced"],
            "n_medline_convs_surfaced": s["n_medline_conversions_surfaced"],
            "avg_complement_lift": round(s["avg_complement_lift"], 2) if s["avg_complement_lift"] else None,
            "category_alignment_pct": round(s["category_alignment_rate"]*100, 1) if s["category_alignment_rate"] is not None else None,
            "complement_mck_brand_pct": round(s["mck_brand_rate_of_complements"]*100, 1) if s["mck_brand_rate_of_complements"] is not None else None,
            "top_complement": top_comp_str,
            "top_pb_upgrade": top_pb_str,
            "top_medline_conv": top_med_str,
        })
    return pd.DataFrame(rows)


# Step 6: Excel styling helper
def _style(ws, df, hc="1F4E79"):
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, 12), 60)


# Step 7: Save outputs
def save_outputs(results: dict, per_customer: pd.DataFrame, chosen: pd.DataFrame) -> None:
    _s("Step 5: Saving outputs")

    ANALYSIS.mkdir(parents=True, exist_ok=True)

    # Build summary aggregations
    n = len(results)
    n_with_complements = sum(1 for r in results.values() if r["summary"]["n_complements_surfaced"] > 0)
    n_with_pb = sum(1 for r in results.values() if r["summary"]["n_pb_upgrades_surfaced"] > 0)
    n_with_medline = sum(1 for r in results.values() if r["summary"]["n_medline_conversions_surfaced"] > 0)

    all_lifts = [r["summary"]["avg_complement_lift"] for r in results.values()
                  if r["summary"]["avg_complement_lift"] is not None]
    all_align = [r["summary"]["category_alignment_rate"] for r in results.values()
                  if r["summary"]["category_alignment_rate"] is not None]
    all_mck = [r["summary"]["mck_brand_rate_of_complements"] for r in results.values()
                if r["summary"]["mck_brand_rate_of_complements"] is not None]

    summary_rows = [
        {"metric": "SIMULATION CONFIG", "value": ""},
        {"metric": "  Customers simulated",     "value": str(n)},
        {"metric": "  Cart contents method",     "value": "Their actual most recent order, top items by spend"},
        {"metric": "  Max items per cart",       "value": str(MAX_CART_ITEMS)},
        {"metric": "  Min lift threshold",       "value": f"{MIN_LIFT_THRESHOLD}"},
        {"metric": "  Min support threshold",    "value": str(MIN_COOC_SUPPORT)},
        {"metric": "", "value": ""},
        {"metric": "OPPORTUNITY RATES", "value": ""},
        {"metric": "  Carts with cart complement(s)",
         "value": f"{n_with_complements} of {n}  ({n_with_complements/n*100:.1f}%)"},
        {"metric": "  Carts with PB upgrade opportunity",
         "value": f"{n_with_pb} of {n}  ({n_with_pb/n*100:.1f}%)"},
        {"metric": "  Carts with Medline conversion opportunity",
         "value": f"{n_with_medline} of {n}  ({n_with_medline/n*100:.1f}%)"},
        {"metric": "", "value": ""},
        {"metric": "QUALITY METRICS", "value": ""},
    ]
    if all_lifts:
        summary_rows.append({"metric": "  Avg cart complement lift",
                              "value": f"{np.mean(all_lifts):.2f}x"})
        summary_rows.append({"metric": "  Median cart complement lift",
                              "value": f"{np.median(all_lifts):.2f}x"})
    if all_align:
        summary_rows.append({"metric": "  Category alignment of complements",
                              "value": f"{np.mean(all_align)*100:.1f}%"})
    if all_mck:
        summary_rows.append({"metric": "  McKesson Brand share of complements",
                              "value": f"{np.mean(all_mck)*100:.1f}%"})

    summary_rows.append({"metric": "", "value": ""})
    summary_rows.append({"metric": "WHAT THIS TESTS", "value": ""})
    summary_rows.append({
        "metric": "  Method",
        "value": "Take real recent order from each customer, treat as cart, run cart-aware engine logic"
    })
    summary_rows.append({
        "metric": "  Why this is the right test",
        "value": "Engine has 2 parts: nightly batch top-10 + LIVE cart feature. This tests live cart feature."
    })
    summary_rows.append({
        "metric": "  What lift means",
        "value": "Lift > 1 = items co-occur more than chance. > 2 = strong. > 5 = very strong."
    })

    summary_df = pd.DataFrame(summary_rows)

    # Build by_size_tier breakdown
    if "size_tier" in per_customer.columns:
        by_size = per_customer.groupby("size_tier").agg(
            n_customers=("DIM_CUST_CURR_ID", "count"),
            avg_complements=("n_complements_surfaced", "mean"),
            avg_lift=("avg_complement_lift", "mean"),
            avg_cat_align_pct=("category_alignment_pct", "mean"),
            avg_mck_brand_pct=("complement_mck_brand_pct", "mean"),
        ).round(2).reset_index()
    else:
        by_size = pd.DataFrame()

    if "status" in per_customer.columns:
        by_status = per_customer.groupby("status").agg(
            n_customers=("DIM_CUST_CURR_ID", "count"),
            avg_complements=("n_complements_surfaced", "mean"),
            avg_lift=("avg_complement_lift", "mean"),
            avg_cat_align_pct=("category_alignment_pct", "mean"),
            avg_mck_brand_pct=("complement_mck_brand_pct", "mean"),
        ).round(2).reset_index()
    else:
        by_status = pd.DataFrame()

    # Build all-complements detail sheet (top 5 per customer)
    all_complements = []
    for cid, r in results.items():
        if r["complements"].empty:
            continue
        df = r["complements"].copy()
        df.insert(0, "DIM_CUST_CURR_ID", cid)
        all_complements.append(df)
    all_complements_df = pd.concat(all_complements, ignore_index=True) if all_complements else pd.DataFrame()

    # PB upgrades detail
    all_pb = []
    for cid, r in results.items():
        if r["pb_upgrades"].empty:
            continue
        df = r["pb_upgrades"].copy()
        df.insert(0, "DIM_CUST_CURR_ID", cid)
        all_pb.append(df)
    all_pb_df = pd.concat(all_pb, ignore_index=True) if all_pb else pd.DataFrame()

    # Medline conversion detail
    all_med = []
    for cid, r in results.items():
        if r["medline_conversions"].empty:
            continue
        df = r["medline_conversions"].copy()
        df.insert(0, "DIM_CUST_CURR_ID", cid)
        all_med.append(df)
    all_med_df = pd.concat(all_med, ignore_index=True) if all_med else pd.DataFrame()

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="00_summary", index=False)
        per_customer.to_excel(writer, sheet_name="01_per_customer", index=False)
        if not by_size.empty:
            by_size.to_excel(writer, sheet_name="02_by_size_tier", index=False)
        if not by_status.empty:
            by_status.to_excel(writer, sheet_name="03_by_status", index=False)
        if not all_complements_df.empty:
            all_complements_df.to_excel(writer, sheet_name="04_complements_detail", index=False)
        if not all_pb_df.empty:
            all_pb_df.to_excel(writer, sheet_name="05_pb_upgrades_detail", index=False)
        if not all_med_df.empty:
            all_med_df.to_excel(writer, sheet_name="06_medline_conv_detail", index=False)

        wb = writer.book
        _style(writer.sheets["00_summary"], summary_df, hc="002060")
        _style(writer.sheets["01_per_customer"], per_customer, hc="1F4E79")
        if not by_size.empty:
            _style(writer.sheets["02_by_size_tier"], by_size, hc="6F2DA8")
        if not by_status.empty:
            _style(writer.sheets["03_by_status"], by_status, hc="375623")
        if not all_complements_df.empty:
            _style(writer.sheets["04_complements_detail"], all_complements_df, hc="833C00")
        if not all_pb_df.empty:
            _style(writer.sheets["05_pb_upgrades_detail"], all_pb_df, hc="C00000")
        if not all_med_df.empty:
            _style(writer.sheets["06_medline_conv_detail"], all_med_df, hc="2F4F4F")

    size_kb = OUT_XLSX.stat().st_size / 1024
    _log(f"Saved: {OUT_XLSX.relative_to(ROOT)}  ({size_kb:.0f} KB)")

    # Text report
    lines = []
    lines.append("=" * 80)
    lines.append("  CART SIMULATION VALIDATION REPORT")
    lines.append("=" * 80)
    lines.append("")
    lines.append("WHAT WE DID")
    lines.append("-" * 80)
    lines.append(f"Picked {n} customers stratified across (size_tier, status). For each, took")
    lines.append(f"their actual most recent order and treated up to {MAX_CART_ITEMS} of those items as a")
    lines.append(f"simulated cart. Then ran the engine's cart-aware logic to see what it would")
    lines.append("surface as cart complements, PB upgrades, and Medline conversions.")
    lines.append("")
    lines.append("This is the right test for this engine because:")
    lines.append("  - The engine has TWO parts: nightly batch (precomputed top 10) and a")
    lines.append("    LIVE cart feature that reacts to cart contents in real time.")
    lines.append("  - This script tests the LIVE cart feature, the way it actually gets used.")
    lines.append("  - It does NOT try to predict purchases 60 days from now.")
    lines.append("")
    lines.append(f"Min cooccurrence lift to surface: {MIN_LIFT_THRESHOLD}x")
    lines.append(f"Min support (co-occurring orders): {MIN_COOC_SUPPORT}")
    lines.append("")
    lines.append("HEADLINE RESULTS")
    lines.append("-" * 80)
    lines.append(f"Customers with at least 1 cart complement surfaced : {n_with_complements} of {n}  ({n_with_complements/n*100:.1f}%)")
    lines.append(f"Customers with at least 1 PB upgrade opportunity   : {n_with_pb} of {n}  ({n_with_pb/n*100:.1f}%)")
    lines.append(f"Customers with at least 1 Medline conversion opp   : {n_with_medline} of {n}  ({n_with_medline/n*100:.1f}%)")
    lines.append("")
    if all_lifts:
        lines.append(f"Avg cart complement lift (across customers)        : {np.mean(all_lifts):.2f}x")
        lines.append(f"Median cart complement lift                        : {np.median(all_lifts):.2f}x")
    if all_align:
        lines.append(f"Avg category alignment of complements              : {np.mean(all_align)*100:.1f}%")
    if all_mck:
        lines.append(f"Avg McKesson Brand share of complements            : {np.mean(all_mck)*100:.1f}%")
    lines.append("")
    if not by_size.empty:
        lines.append("BY SIZE TIER")
        lines.append("-" * 80)
        lines.append(f"{'Size tier':<14} {'N':>4} {'Avg comp':>10} {'Avg lift':>10} {'Cat align':>10} {'MCK %':>8}")
        for _, r in by_size.iterrows():
            lines.append(f"{r['size_tier']:<14} {int(r['n_customers']):>4} {r['avg_complements']:>10.2f} {r['avg_lift'] if pd.notna(r['avg_lift']) else 0:>10.2f} {r['avg_cat_align_pct'] if pd.notna(r['avg_cat_align_pct']) else 0:>9.1f}% {r['avg_mck_brand_pct'] if pd.notna(r['avg_mck_brand_pct']) else 0:>7.1f}%")
        lines.append("")
    if not by_status.empty:
        lines.append("BY STATUS")
        lines.append("-" * 80)
        lines.append(f"{'Status':<18} {'N':>4} {'Avg comp':>10} {'Avg lift':>10} {'Cat align':>10} {'MCK %':>8}")
        for _, r in by_status.iterrows():
            lines.append(f"{r['status']:<18} {int(r['n_customers']):>4} {r['avg_complements']:>10.2f} {r['avg_lift'] if pd.notna(r['avg_lift']) else 0:>10.2f} {r['avg_cat_align_pct'] if pd.notna(r['avg_cat_align_pct']) else 0:>9.1f}% {r['avg_mck_brand_pct'] if pd.notna(r['avg_mck_brand_pct']) else 0:>7.1f}%")
        lines.append("")

    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")
    _log(f"Saved: {OUT_TXT.relative_to(ROOT)}")


# Main
def main() -> None:
    log_file = _setup_logging()

    try:
        print()
        print("=" * 80)
        print("  CART SIMULATION VALIDATION")
        print("=" * 80)
        print()
        print("  Tests the LIVE cart feature of the recommendation engine.")
        print("  Takes real recent orders from sample customers, treats them as carts,")
        print("  and shows what the engine would surface as cart complements,")
        print("  brand upgrades, and Medline conversions.")
        start = time.time()

        data = load_data()
        chosen = pick_customers(data["customers"], data["recs"])
        carts = build_carts(chosen, data["products"])
        results = run_cart_logic(
            carts, data["cooc"], data["pb_equiv"], data["products"]
        )
        per_customer = build_per_customer_rows(results, chosen)
        save_outputs(results, per_customer, chosen)

        _s("Complete")
        _log(f"Total time: {time.time() - start:.1f}s")
        _log(f"")
        _log(f"Outputs:")
        _log(f"  Terminal log : {OUT_LOG.relative_to(ROOT)}")
        _log(f"  XLSX         : {OUT_XLSX.relative_to(ROOT)}")
        _log(f"  TXT          : {OUT_TXT.relative_to(ROOT)}")
    finally:
        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__
        log_file.close()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise