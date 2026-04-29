from __future__ import annotations

import sys
import time
import warnings
from pathlib import Path

import duckdb
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# Paths

ROOT         = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN   = ROOT / "data_clean"
PRECOMP      = DATA_CLEAN / "serving" / "precomputed"
FEATURES     = DATA_CLEAN / "features"
ANALYSIS     = DATA_CLEAN / "analysis"

RECS_FILE         = PRECOMP  / "recommendations.parquet"
CUST_SEG_FILE     = PRECOMP  / "customer_segments.parquet"
CUST_PATT_FILE    = PRECOMP  / "customer_patterns.parquet"
FEATURE_FILE      = FEATURES / "customer_features.parquet"
PROD_SEG_FILE     = PRECOMP  / "product_segments.parquet"
COOCCUR_FILE      = PRECOMP  / "product_cooccurrence.parquet"
ITEM_SIM_FILE     = PRECOMP  / "item_similarity.parquet"
PB_EQUIV_FILE     = PRECOMP  / "private_brand_equivalents.parquet"
# Phase 6 inputs (used for additional context in the report)
REPLENISHMENT_FILE = PRECOMP / "customer_replenishment_candidates.parquet"
SEG_CADENCE_FILE   = PRECOMP / "product_segment_cadence.parquet"
MERGED_FILE       = DATA_CLEAN / "serving" / "merged_dataset.parquet"

OUT_XLSX          = ANALYSIS / "recommendation_sanity_check.xlsx"
OUT_TXT           = ANALYSIS / "recommendation_sanity_check.txt"
OUT_TERMINAL_LOG  = ANALYSIS / "recommendation_sanity_check_terminal.log"
OUT_CHARTS_DIR    = ANALYSIS / "charts" / "sanity_check"


# Configuration

CUSTOMERS_PER_GROUP = 1
MAX_CUSTOMERS_TOTAL = 60
HISTORY_TOP_N       = 15
SHOW_RECS_TOP_N     = 10
RECENT_ORDERS_N     = 3
CART_RECS_PER_TYPE  = 5
RANDOM_SEED         = 42

CART_MIN_LIFT       = 2.0
SIM_MIN_SCORE       = 0.10

# Phase 5 validation: customers with this many history products are subject
# to the specialty mismatch filter in recommendation_factors.py.
SPECIALTY_FILTER_MIN_HISTORY = 10

# Phase 6: expected signals/purposes (used to keep summary stats consistent
# even when a signal happens to have zero recs in the sample)
EXPECTED_SIGNALS = [
    "peer_gap", "popularity", "cart_complement", "item_similarity",
    "lapsed_recovery", "replenishment", "medline_conversion",
    "private_brand_upgrade",
]
EXPECTED_PURPOSES = [
    "new_product", "cross_sell", "win_back", "replenishment", "mckesson_substitute",
]


# Tee stdout to file so the long terminal output is captured for debugging

class _Tee:
    # Writes to multiple streams (real stdout + log file).
    # Used so the long sanity-check output can be reviewed after the run
    # without losing anything to terminal scrollback limits.
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for s in self.streams:
            try:
                s.write(data)
            except Exception:
                pass

    def flush(self):
        for s in self.streams:
            try:
                s.flush()
            except Exception:
                pass


def _setup_terminal_logging():
    OUT_TERMINAL_LOG.parent.mkdir(parents=True, exist_ok=True)
    log_file = open(OUT_TERMINAL_LOG, "w", encoding="utf-8", buffering=1)
    sys.stdout = _Tee(sys.__stdout__, log_file)
    sys.stderr = _Tee(sys.__stderr__, log_file)
    return log_file


# Logging

def _s(title: str) -> None:
    print(f"\n{'-' * 70}\n  {title}\n{'-' * 70}", flush=True)


def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


# Step 1: Load all data

def load_data() -> dict:
    _s("Step 1: Loading data")
    t0 = time.time()

    if not RECS_FILE.exists():
        print(f"\nFATAL: {RECS_FILE} not found.", file=sys.stderr)
        print("Run recommendation_factors.py first.", file=sys.stderr)
        sys.exit(1)

    recs = pd.read_parquet(RECS_FILE)
    _log(f"Recommendations           : {len(recs):,} rows")

    # Verify Phase 4 columns
    if "rec_purpose" not in recs.columns:
        _log(f"  WARNING: rec_purpose column missing. Run Phase 4+ recommendation_factors.py")
    if "is_mckesson_brand" not in recs.columns:
        _log(f"  WARNING: is_mckesson_brand column missing. Run Phase 4+ recommendation_factors.py")
    # Phase 5 column (optional - older runs won't have it)
    if "normalized_score" in recs.columns:
        _log(f"  Phase 5 normalized_score column present")
    else:
        _log(f"  NOTE: normalized_score column missing - this looks like a pre-Phase-5 file")
    # Phase 6 verification: replenishment signal present?
    if "primary_signal" in recs.columns:
        n_repl = int((recs["primary_signal"] == "replenishment").sum())
        if n_repl > 0:
            _log(f"  Phase 6 replenishment signal: {n_repl:,} recs ({n_repl/len(recs)*100:.1f}%)")
        else:
            _log(f"  NOTE: no replenishment recs found - either Phase 6 not run or no candidates")

    customers = pd.read_parquet(CUST_SEG_FILE)
    customers["DIM_CUST_CURR_ID"] = customers["DIM_CUST_CURR_ID"].astype("int64")

    patterns = pd.read_parquet(CUST_PATT_FILE)
    patterns["DIM_CUST_CURR_ID"] = patterns["DIM_CUST_CURR_ID"].astype("int64")

    feat_full = pd.read_parquet(FEATURE_FILE)
    keep_cols = ["DIM_CUST_CURR_ID", "median_monthly_spend",
                 "affordability_ceiling", "SPCLTY_CD"]
    if "SPCLTY_DSC" in feat_full.columns:
        keep_cols.append("SPCLTY_DSC")
    features = feat_full[[c for c in keep_cols if c in feat_full.columns]].copy()
    features["DIM_CUST_CURR_ID"] = features["DIM_CUST_CURR_ID"].astype("int64")

    products = pd.read_parquet(PROD_SEG_FILE, columns=[
        "DIM_ITEM_E1_CURR_ID", "ITEM_DSC",
        "PROD_FMLY_LVL1_DSC", "PROD_CTGRY_LVL2_DSC",
        "median_unit_price", "n_buyers", "is_private_brand",
    ])
    products["DIM_ITEM_E1_CURR_ID"] = products["DIM_ITEM_E1_CURR_ID"].astype("int64")
    products = products[
        products["ITEM_DSC"].notna() &
        (products["ITEM_DSC"].astype(str).str.strip() != "") &
        (products["ITEM_DSC"].astype(str).str.lower() != "nan")
    ].copy()

    cooccur = pd.read_parquet(COOCCUR_FILE)
    cooccur["product_a"] = cooccur["product_a"].astype("int64")
    cooccur["product_b"] = cooccur["product_b"].astype("int64")

    item_sim = pd.read_parquet(ITEM_SIM_FILE)
    item_sim["item_a"] = item_sim["item_a"].astype("int64")
    item_sim["item_b"] = item_sim["item_b"].astype("int64")

    pb_equiv = pd.read_parquet(PB_EQUIV_FILE)
    pb_equiv["original_item_id"]   = pb_equiv["original_item_id"].astype("int64")
    pb_equiv["equivalent_item_id"] = pb_equiv["equivalent_item_id"].astype("int64")

    if "price_anomaly" in pb_equiv.columns:
        n_before = len(pb_equiv)
        pb_equiv = pb_equiv[pb_equiv["price_anomaly"].fillna(0) == 0]
        _log(f"  Dropped {n_before - len(pb_equiv):,} pb_equiv price anomalies")

    # Phase 6: load replenishment candidates for context (overdue ratios etc.)
    # Used in per-customer report to show *why* a replenishment rec fired.
    if REPLENISHMENT_FILE.exists():
        replenishment = pd.read_parquet(REPLENISHMENT_FILE)
        rename_map = {}
        if "cust_id" in replenishment.columns:
            rename_map["cust_id"] = "DIM_CUST_CURR_ID"
        if "item_id" in replenishment.columns:
            rename_map["item_id"] = "DIM_ITEM_E1_CURR_ID"
        if rename_map:
            replenishment = replenishment.rename(columns=rename_map)
        replenishment["DIM_CUST_CURR_ID"]    = replenishment["DIM_CUST_CURR_ID"].astype("int64")
        replenishment["DIM_ITEM_E1_CURR_ID"] = replenishment["DIM_ITEM_E1_CURR_ID"].astype("int64")
        _log(f"  Replenishment context     : {len(replenishment):,} candidates")
    else:
        replenishment = pd.DataFrame()
        _log(f"  Replenishment context     : MISSING (Phase 6 file not found)")

    _log(f"All data loaded in {time.time()-t0:.1f}s")

    return {
        "recs": recs, "customers": customers, "patterns": patterns,
        "features": features, "products": products,
        "cooccur": cooccur, "item_sim": item_sim, "pb_equiv": pb_equiv,
        "replenishment": replenishment,
    }


# Step 2: Select sample customers

def select_sample_customers(data: dict) -> pd.DataFrame:
    _s("Step 2: Selecting sample customers (stratified by segment + status)")

    recs    = data["recs"]
    custs   = data["customers"]
    patts   = data["patterns"]
    feats   = data["features"]

    pool = custs.merge(patts, on="DIM_CUST_CURR_ID", how="inner", suffixes=("", "_p"))
    custs_with_recs = set(recs["DIM_CUST_CURR_ID"].unique())
    pool = pool[pool["DIM_CUST_CURR_ID"].isin(custs_with_recs)]
    pool = pool.merge(feats, on="DIM_CUST_CURR_ID", how="left")

    pool["status"] = "stable_warm"
    pool.loc[pool["is_cold_start"] == 1, "status"] = "cold_start"
    pool.loc[(pool["is_cold_start"] == 0) & (pool["is_declining"] == 1), "status"] = "declining_warm"
    pool.loc[(pool["is_cold_start"] == 0) & (pool["is_churned"] == 1), "status"] = "churned_warm"

    _log(f"Customer pool by status:")
    for s, n in pool["status"].value_counts().items():
        _log(f"  {s:<20}  {n:,}")

    rng = np.random.RandomState(RANDOM_SEED)
    samples = []
    segments = sorted(pool["segment"].dropna().unique().tolist())

    for seg in segments:
        seg_pool = pool[pool["segment"] == seg]
        for status in ["stable_warm", "declining_warm", "churned_warm", "cold_start"]:
            bucket = seg_pool[seg_pool["status"] == status]
            if len(bucket) == 0:
                continue
            picked = bucket.sample(
                n=min(CUSTOMERS_PER_GROUP, len(bucket)), random_state=rng
            )
            samples.append(picked)

    sample_df = pd.concat(samples, ignore_index=True) if samples else pd.DataFrame()

    if len(sample_df) > MAX_CUSTOMERS_TOTAL:
        sample_df = sample_df.sample(n=MAX_CUSTOMERS_TOTAL, random_state=rng)

    _log(f"")
    _log(f"Selected {len(sample_df):,} customers for review")
    _log(f"  Across {sample_df['segment'].nunique()} segments")
    _log(f"  Status distribution:")
    for s, n in sample_df["status"].value_counts().items():
        _log(f"    {s:<20}  {n}")

    return sample_df


# Step 3: Aggregate purchase history

def get_purchase_history(cust_ids: list) -> pd.DataFrame:
    _s("Step 3: Loading aggregate purchase history")
    t0 = time.time()

    con = duckdb.connect()
    cust_ids_sql = ", ".join(str(int(c)) for c in cust_ids)
    _log(f"Querying {len(cust_ids)} customers from merged_dataset...")

    history = con.execute(f"""
        SELECT
            CAST(DIM_CUST_CURR_ID AS BIGINT) AS DIM_CUST_CURR_ID,
            CAST(DIM_ITEM_E1_CURR_ID AS BIGINT) AS DIM_ITEM_E1_CURR_ID,
            COUNT(*) AS n_lines,
            SUM(ORDR_QTY) AS total_qty,
            SUM(UNIT_SLS_AMT) AS total_spend,
            MAX(MAKE_DATE(order_year, order_month, order_day)) AS last_order_date,
            MIN(MAKE_DATE(order_year, order_month, order_day)) AS first_order_date
        FROM read_parquet('{MERGED_FILE.as_posix()}')
        WHERE UNIT_SLS_AMT > 0
          AND DIM_CUST_CURR_ID IN ({cust_ids_sql})
        GROUP BY
            CAST(DIM_CUST_CURR_ID AS BIGINT),
            CAST(DIM_ITEM_E1_CURR_ID AS BIGINT)
    """).df()
    con.close()

    history["DIM_CUST_CURR_ID"]  = history["DIM_CUST_CURR_ID"].astype("int64")
    history["DIM_ITEM_E1_CURR_ID"] = history["DIM_ITEM_E1_CURR_ID"].astype("int64")
    history["last_order_date"]   = pd.to_datetime(history["last_order_date"])

    _log(f"Loaded {len(history):,} customer-product pairs in {time.time()-t0:.1f}s")
    return history


# Step 4: Get 3 most recent orders

def get_recent_orders(cust_ids: list) -> pd.DataFrame:
    _s("Step 4: Loading 3 most recent orders per customer")
    t0 = time.time()

    con = duckdb.connect()
    cust_ids_sql = ", ".join(str(int(c)) for c in cust_ids)
    _log(f"Querying recent line items for {len(cust_ids)} customers...")

    line_items = con.execute(f"""
        WITH line_items AS (
            SELECT
                CAST(DIM_CUST_CURR_ID AS BIGINT)   AS DIM_CUST_CURR_ID,
                CAST(DIM_ITEM_E1_CURR_ID AS BIGINT) AS DIM_ITEM_E1_CURR_ID,
                ORDR_QTY,
                UNIT_SLS_AMT,
                MAKE_DATE(order_year, order_month, order_day) AS order_date,
                CAST(strftime(MAKE_DATE(order_year, order_month, order_day), '%Y%W') AS BIGINT) AS order_week
            FROM read_parquet('{MERGED_FILE.as_posix()}')
            WHERE UNIT_SLS_AMT > 0
              AND DIM_CUST_CURR_ID IN ({cust_ids_sql})
              AND order_year IS NOT NULL
        ),
        order_summary AS (
            SELECT
                DIM_CUST_CURR_ID,
                order_week,
                MAX(order_date) AS order_date_summary,
                COUNT(*) AS line_count,
                SUM(UNIT_SLS_AMT) AS order_total
            FROM line_items
            GROUP BY DIM_CUST_CURR_ID, order_week
        ),
        ranked_orders AS (
            SELECT
                DIM_CUST_CURR_ID,
                order_week,
                order_date_summary,
                line_count,
                order_total,
                ROW_NUMBER() OVER (
                    PARTITION BY DIM_CUST_CURR_ID
                    ORDER BY order_date_summary DESC
                ) AS recency_rank
            FROM order_summary
        )
        SELECT
            li.DIM_CUST_CURR_ID,
            li.DIM_ITEM_E1_CURR_ID,
            li.ORDR_QTY,
            li.UNIT_SLS_AMT,
            li.order_date,
            li.order_week,
            ro.recency_rank,
            ro.order_total,
            ro.line_count
        FROM line_items li
        INNER JOIN ranked_orders ro
          ON li.DIM_CUST_CURR_ID = ro.DIM_CUST_CURR_ID
         AND li.order_week = ro.order_week
        WHERE ro.recency_rank <= {RECENT_ORDERS_N}
    """).df()
    con.close()

    line_items["DIM_CUST_CURR_ID"]   = line_items["DIM_CUST_CURR_ID"].astype("int64")
    line_items["DIM_ITEM_E1_CURR_ID"] = line_items["DIM_ITEM_E1_CURR_ID"].astype("int64")
    line_items["order_date"]   = pd.to_datetime(line_items["order_date"])
    line_items["recency_rank"] = line_items["recency_rank"].astype("int8")

    n_unique_orders = line_items.groupby(
        ["DIM_CUST_CURR_ID", "recency_rank"]
    ).ngroups
    _log(f"Loaded {len(line_items):,} line items across {n_unique_orders:,} orders "
         f"in {time.time()-t0:.1f}s")

    return line_items


# Step 5: Cart-aware recs simulator
# NOTE: replenishment is intentionally NOT a cart-aware signal - it operates
# on the customer's overall reorder cadence, not on items in a specific cart.
# So we do NOT add a replenishment branch here. The pre-computed top-10
# captures replenishment via the recommendations.parquet file.

def generate_cart_recs(
    cart_items: list,
    customer_id: int,
    products: pd.DataFrame,
    cooccur: pd.DataFrame,
    item_sim: pd.DataFrame,
    pb_equiv: pd.DataFrame,
    history_full: pd.DataFrame,
) -> dict:
    cart_set = set(int(i) for i in cart_items)
    customer_history_items = set(
        history_full[history_full["DIM_CUST_CURR_ID"] == customer_id][
            "DIM_ITEM_E1_CURR_ID"
        ].tolist()
    )

    products_lookup = products.set_index("DIM_ITEM_E1_CURR_ID")[
        ["ITEM_DSC", "PROD_FMLY_LVL1_DSC", "PROD_CTGRY_LVL2_DSC",
         "median_unit_price", "is_private_brand"]
    ].to_dict("index")

    def safe_lookup_name(pid: int) -> str:
        info = products_lookup.get(pid, {})
        name = info.get("ITEM_DSC", None)
        if name is None or pd.isna(name) or str(name).strip() == "" or str(name).lower() == "nan":
            return f"Item #{pid}"
        return str(name)[:40]

    # Cart complement
    cart_complement = []
    cooccur_filtered = cooccur[
        cooccur["product_a"].isin(cart_set) &
        ~cooccur["product_b"].isin(cart_set) &
        ~cooccur["product_b"].isin(customer_history_items) &
        (cooccur["lift"] >= CART_MIN_LIFT)
    ].copy()
    if len(cooccur_filtered) > 0:
        cooccur_filtered = cooccur_filtered.sort_values("lift", ascending=False)
        seen = set()
        for _, row in cooccur_filtered.iterrows():
            pid = int(row["product_b"])
            if pid in seen:
                continue
            seen.add(pid)
            info = products_lookup.get(pid, {})
            if info.get("ITEM_DSC") is None or pd.isna(info.get("ITEM_DSC")):
                continue
            cart_complement.append({
                "rec_item_id":  pid,
                "ITEM_DSC":     str(info.get("ITEM_DSC", "?"))[:50],
                "family":       str(info.get("PROD_FMLY_LVL1_DSC", "?"))[:25],
                "lift":         round(float(row["lift"]), 1),
                "support":      round(float(row.get("support_ab", 0)), 5),
                "median_price": float(info.get("median_unit_price", 0) or 0),
                "is_pb":        int(info.get("is_private_brand", 0) or 0),
                "triggered_by": safe_lookup_name(int(row["product_a"])),
            })
            if len(cart_complement) >= CART_RECS_PER_TYPE:
                break

    # Item similarity
    item_sim_recs = []
    sim_filtered = item_sim[
        item_sim["item_a"].isin(cart_set) &
        ~item_sim["item_b"].isin(cart_set) &
        ~item_sim["item_b"].isin(customer_history_items) &
        (item_sim["similarity"] >= SIM_MIN_SCORE)
    ].copy()
    if len(sim_filtered) > 0:
        sim_filtered = sim_filtered.sort_values("similarity", ascending=False)
        seen = set()
        for _, row in sim_filtered.iterrows():
            pid = int(row["item_b"])
            if pid in seen:
                continue
            seen.add(pid)
            info = products_lookup.get(pid, {})
            if info.get("ITEM_DSC") is None or pd.isna(info.get("ITEM_DSC")):
                continue
            item_sim_recs.append({
                "rec_item_id":  pid,
                "ITEM_DSC":     str(info.get("ITEM_DSC", "?"))[:50],
                "family":       str(info.get("PROD_FMLY_LVL1_DSC", "?"))[:25],
                "similarity":   round(float(row["similarity"]), 3),
                "median_price": float(info.get("median_unit_price", 0) or 0),
                "is_pb":        int(info.get("is_private_brand", 0) or 0),
                "triggered_by": safe_lookup_name(int(row["item_a"])),
            })
            if len(item_sim_recs) >= CART_RECS_PER_TYPE:
                break

    # Medline conversion
    medline_recs = []
    medline_in_cart = pb_equiv[
        (pb_equiv["match_type"] == "medline_conversion") &
        (pb_equiv["original_item_id"].isin(cart_set))
    ].copy()
    if len(medline_in_cart) > 0:
        medline_in_cart = medline_in_cart.sort_values("price_delta_pct", ascending=False)
        seen = set()
        for _, row in medline_in_cart.iterrows():
            pid = int(row["equivalent_item_id"])
            if pid in seen:
                continue
            seen.add(pid)
            info = products_lookup.get(pid, {})
            if info.get("ITEM_DSC") is None or pd.isna(info.get("ITEM_DSC")):
                continue
            medline_recs.append({
                "rec_item_id":     pid,
                "ITEM_DSC":        str(info.get("ITEM_DSC", "?"))[:50],
                "family":          str(info.get("PROD_FMLY_LVL1_DSC", "?"))[:25],
                "median_price":    float(info.get("median_unit_price", 0) or 0),
                "original_price":  round(float(row.get("original_unit_price", 0)), 2),
                "alt_price":       round(float(row.get("equivalent_unit_price", 0)), 2),
                "triggered_by_medline": safe_lookup_name(int(row["original_item_id"])),
            })
            if len(medline_recs) >= CART_RECS_PER_TYPE:
                break

    # PB upgrade
    pb_recs = []
    pb_in_cart = pb_equiv[
        (pb_equiv["match_type"] == "private_brand_upgrade") &
        (pb_equiv["original_item_id"].isin(cart_set))
    ].copy()
    if len(pb_in_cart) > 0:
        pb_in_cart = pb_in_cart.sort_values("price_delta_pct", ascending=True)
        seen = set()
        for _, row in pb_in_cart.iterrows():
            pid = int(row["equivalent_item_id"])
            if pid in seen:
                continue
            seen.add(pid)
            info = products_lookup.get(pid, {})
            if info.get("ITEM_DSC") is None or pd.isna(info.get("ITEM_DSC")):
                continue
            pb_recs.append({
                "rec_item_id":      pid,
                "ITEM_DSC":         str(info.get("ITEM_DSC", "?"))[:50],
                "family":           str(info.get("PROD_FMLY_LVL1_DSC", "?"))[:25],
                "savings_pct":      round(float(row.get("price_delta_pct", 0)) * -100, 1),
                "original_price":   round(float(row.get("original_unit_price", 0)), 2),
                "mck_brand_price":  round(float(row.get("equivalent_unit_price", 0)), 2),
                "triggered_by_natl": safe_lookup_name(int(row["original_item_id"])),
            })
            if len(pb_recs) >= CART_RECS_PER_TYPE:
                break

    return {
        "cart_complement": cart_complement,
        "item_similarity": item_sim_recs,
        "medline":         medline_recs,
        "pb_upgrade":      pb_recs,
    }


# Step 6: Build per-customer report

def build_customer_report(
    customer_row: pd.Series,
    history: pd.DataFrame,
    recent_orders: pd.DataFrame,
    recs: pd.DataFrame,
    products: pd.DataFrame,
    cooccur: pd.DataFrame,
    item_sim: pd.DataFrame,
    pb_equiv: pd.DataFrame,
    replenishment: pd.DataFrame,
) -> dict:
    cust_id = int(customer_row["DIM_CUST_CURR_ID"])

    profile_lines = [
        f"Customer ID:    {cust_id}",
        f"Segment:        {customer_row.get('segment', 'N/A')}",
        f"Size tier:      {customer_row.get('size_tier', 'N/A')}",
        f"Market:         {customer_row.get('mkt_cd_clean', 'N/A')}",
        f"Specialty:      {customer_row.get('SPCLTY_CD', 'N/A')}",
        f"Status:         {customer_row.get('status', 'N/A')}",
        f"Cadence:        {customer_row.get('order_cadence_tier', 'N/A')}",
        f"Total products: {customer_row.get('n_unique_products_total', 0):.0f}",
        f"Affordability ceiling: ${customer_row.get('affordability_ceiling', 0):,.0f}",
        f"Lapsed products: {customer_row.get('n_lapsed_products', 0):.0f}",
    ]

    # Historical purchases
    cust_history = history[history["DIM_CUST_CURR_ID"] == cust_id].copy()
    cust_history = cust_history.merge(
        products[["DIM_ITEM_E1_CURR_ID", "ITEM_DSC", "PROD_FMLY_LVL1_DSC",
                  "PROD_CTGRY_LVL2_DSC", "is_private_brand"]],
        on="DIM_ITEM_E1_CURR_ID", how="left"
    )
    cust_history = cust_history[cust_history["ITEM_DSC"].notna()]
    cust_history = cust_history.sort_values("total_spend", ascending=False).head(HISTORY_TOP_N)
    history_rows = []
    for _, h in cust_history.iterrows():
        history_rows.append({
            "rank":        len(history_rows) + 1,
            "ITEM_DSC":    str(h.get("ITEM_DSC", ""))[:50],
            "family":      str(h.get("PROD_FMLY_LVL1_DSC", ""))[:25],
            "category":    str(h.get("PROD_CTGRY_LVL2_DSC", ""))[:25],
            "total_qty":   float(h["total_qty"]),
            "total_spend": float(h["total_spend"]),
            "last_order":  h["last_order_date"].strftime("%Y-%m-%d") if pd.notna(h["last_order_date"]) else "",
            "is_pb":       int(h.get("is_private_brand", 0)) if pd.notna(h.get("is_private_brand")) else 0,
        })

    # Recent orders + cart simulations
    cust_orders = recent_orders[recent_orders["DIM_CUST_CURR_ID"] == cust_id].copy()
    cust_orders = cust_orders.merge(
        products[["DIM_ITEM_E1_CURR_ID", "ITEM_DSC", "PROD_FMLY_LVL1_DSC",
                  "is_private_brand"]],
        on="DIM_ITEM_E1_CURR_ID", how="left"
    )

    order_simulations = []
    for rank in range(1, RECENT_ORDERS_N + 1):
        order_lines = cust_orders[cust_orders["recency_rank"] == rank]
        if len(order_lines) == 0:
            continue

        first_line = order_lines.iloc[0]
        order_info = {
            "order_rank":   int(rank),
            "order_date":   first_line["order_date"].strftime("%Y-%m-%d") if pd.notna(first_line["order_date"]) else "?",
            "line_count":   int(first_line["line_count"]),
            "order_total":  float(first_line["order_total"]),
            "items":        [],
        }
        for _, line in order_lines.iterrows():
            item_name = line.get("ITEM_DSC", None)
            if item_name is None or pd.isna(item_name) or str(item_name).lower() == "nan":
                continue
            order_info["items"].append({
                "DIM_ITEM_E1_CURR_ID": int(line["DIM_ITEM_E1_CURR_ID"]),
                "ITEM_DSC":  str(item_name)[:50],
                "family":    str(line.get("PROD_FMLY_LVL1_DSC", "?"))[:25],
                "qty":       float(line["ORDR_QTY"]),
                "spend":     float(line["UNIT_SLS_AMT"]),
                "is_pb":     int(line.get("is_private_brand", 0)) if pd.notna(line.get("is_private_brand")) else 0,
            })

        cart_item_ids = order_lines["DIM_ITEM_E1_CURR_ID"].tolist()
        cart_recs = generate_cart_recs(
            cart_item_ids, cust_id, products, cooccur, item_sim, pb_equiv, history
        )
        order_info["cart_recs"] = cart_recs

        order_simulations.append(order_info)

    # Phase 6: pull replenishment context for this customer
    # (overdue ratios + segment cadence for items showing up as replenishment recs)
    repl_context = {}
    if len(replenishment) > 0:
        cust_repl = replenishment[replenishment["DIM_CUST_CURR_ID"] == cust_id]
        for _, r in cust_repl.iterrows():
            iid = int(r["DIM_ITEM_E1_CURR_ID"])
            repl_context[iid] = {
                "days_since_last": int(r.get("days_since_last", 0)) if pd.notna(r.get("days_since_last")) else 0,
                "segment_cadence": float(r.get("median_days_between_segment", 0)) if pd.notna(r.get("median_days_between_segment")) else 0,
                "overdue_ratio":   float(r.get("overdue_ratio", 0)) if pd.notna(r.get("overdue_ratio")) else 0,
                "peer_activity":   float(r.get("peer_activity_rate", 0)) if pd.notna(r.get("peer_activity_rate")) else 0,
            }

    # Pre-computed top 10 (Phase 5: also surface normalized_score; Phase 6: replenishment context)
    cust_recs = recs[recs["DIM_CUST_CURR_ID"] == cust_id].sort_values("rank").head(SHOW_RECS_TOP_N)
    rec_rows = []
    for _, r in cust_recs.iterrows():
        item_id = int(r["DIM_ITEM_E1_CURR_ID"])
        rc = repl_context.get(item_id, {})
        rec_rows.append({
            "rank":            int(r["rank"]),
            "ITEM_DSC":        str(r.get("ITEM_DSC", ""))[:50],
            "family":          str(r.get("PROD_FMLY_LVL1_DSC", ""))[:25],
            "category":        str(r.get("PROD_CTGRY_LVL2_DSC", ""))[:25],
            "primary_signal":  str(r.get("primary_signal", "")),
            "rec_purpose":     str(r.get("rec_purpose", "")),
            "is_mck_brand":    int(r.get("is_mckesson_brand", 0)) if pd.notna(r.get("is_mckesson_brand")) else 0,
            "confidence":      str(r.get("confidence_tier", "")),
            "score":           float(r.get("numeric_score", 0)),
            "norm_score":      float(r.get("normalized_score", 0)) if pd.notna(r.get("normalized_score")) else 0.0,
            "specialty_match": str(r.get("specialty_match", "")),
            "median_price":    float(r.get("median_unit_price", 0)) if pd.notna(r.get("median_unit_price")) else 0,
            "is_pb":           int(r.get("is_private_brand", 0)) if pd.notna(r.get("is_private_brand")) else 0,
            "pitch_reason":    str(r.get("pitch_reason", ""))[:120],
            # Phase 6 fields - only meaningful for replenishment recs, blank otherwise
            "repl_days_since_last": rc.get("days_since_last", 0),
            "repl_segment_cadence": rc.get("segment_cadence", 0),
            "repl_overdue_ratio":   rc.get("overdue_ratio", 0),
            "repl_peer_activity":   rc.get("peer_activity", 0),
        })

    # Quality assessment
    history_families = set(h["family"] for h in history_rows)
    history_categories = set(h["category"] for h in history_rows)
    rec_families   = set(r["family"] for r in rec_rows)
    rec_categories = set(r["category"] for r in rec_rows)

    family_overlap_pct = (
        100 * len(rec_families & history_families) / max(len(rec_families), 1)
    )
    category_overlap_pct = (
        100 * len(rec_categories & history_categories) / max(len(rec_categories), 1)
    )

    n_unique_signals = len(set(r["primary_signal"] for r in rec_rows))
    n_unique_purposes = len(set(r["rec_purpose"] for r in rec_rows))
    n_unique_families_in_recs = len(rec_families)
    n_mck_brand_recs = sum(1 for r in rec_rows if r["is_mck_brand"] == 1)

    # Phase 5 validation: count mismatches; classify history as rich/thin
    mismatch_count = sum(1 for r in rec_rows if r["specialty_match"] == "mismatch")
    n_history = len(history_rows)
    history_size_bucket = "rich" if n_history >= SPECIALTY_FILTER_MIN_HISTORY else "thin"
    # The Phase 5 specialty filter should keep mismatch_count == 0 for "rich" warm customers
    spec_filter_breach = (
        history_size_bucket == "rich"
        and mismatch_count > 0
        and customer_row.get("status") != "cold_start"
    )

    # Phase 6: replenishment-specific quality counts
    n_replenishment_recs = sum(1 for r in rec_rows if r["primary_signal"] == "replenishment")

    cart_summary = {
        "n_orders_with_cart_complement": sum(
            1 for o in order_simulations if len(o["cart_recs"]["cart_complement"]) > 0
        ),
        "n_orders_with_item_similarity": sum(
            1 for o in order_simulations if len(o["cart_recs"]["item_similarity"]) > 0
        ),
        "n_orders_with_medline_opp":     sum(
            1 for o in order_simulations if len(o["cart_recs"]["medline"]) > 0
        ),
        "n_orders_with_pb_opp":          sum(
            1 for o in order_simulations if len(o["cart_recs"]["pb_upgrade"]) > 0
        ),
    }

    quality = {
        "family_overlap_pct":     round(family_overlap_pct, 1),
        "category_overlap_pct":   round(category_overlap_pct, 1),
        "n_unique_families":      n_unique_families_in_recs,
        "n_unique_signals":       n_unique_signals,
        "n_unique_purposes":      n_unique_purposes,
        "n_mck_brand_recs":       n_mck_brand_recs,
        "n_replenishment_recs":   n_replenishment_recs,   # Phase 6
        "spclty_match_count":     sum(1 for r in rec_rows if r["specialty_match"] == "match"),
        "mismatch_count":         mismatch_count,
        "history_size_bucket":    history_size_bucket,
        "spec_filter_breach":     int(spec_filter_breach),
        "n_orders_simulated":     len(order_simulations),
        "n_orders_with_cart_recs": cart_summary["n_orders_with_cart_complement"],
        "n_orders_with_medline":  cart_summary["n_orders_with_medline_opp"],
        "n_orders_with_pb_upg":   cart_summary["n_orders_with_pb_opp"],
    }

    return {
        "customer_id":      cust_id,
        "profile":          profile_lines,
        "history":          history_rows,
        "recommendations":  rec_rows,
        "order_simulations": order_simulations,
        "quality":          quality,
        "history_families":   sorted(history_families),
        "rec_families":       sorted(rec_families),
    }


# Step 7: Terminal report

def print_terminal_report(reports: list) -> None:
    _s(f"Step 7: Terminal report ({len(reports)} customers)")

    for idx, rep in enumerate(reports, 1):
        print()
        print("=" * 80)
        print(f"  CUSTOMER {idx}/{len(reports)}: {rep['customer_id']}")
        print("=" * 80)

        print("\n  CUSTOMER PROFILE:")
        for line in rep["profile"]:
            print(f"    {line}")

        print(f"\n  HISTORICAL PURCHASES (top {min(8, len(rep['history']))}):")
        for h in rep["history"][:8]:
            pb_tag = " [PB]" if h["is_pb"] else ""
            print(f"    {h['rank']:>2}. {h['ITEM_DSC']:<52}{pb_tag}")
            print(f"        Family: {h['family']:<25}  Spent: ${h['total_spend']:>8,.0f}  "
                  f"Last: {h['last_order']}")

        print(f"\n  RECENT ORDERS + CART SIMULATIONS:")
        for order in rep["order_simulations"]:
            print(f"\n ORDER #{order['order_rank']} ({order['order_date']}, "
                  f"${order['order_total']:,.0f}, {order['line_count']} lines)")
            print(f"    Items in this order:")
            for item in order["items"][:5]:
                pb_tag = " [PB]" if item["is_pb"] else ""
                print(f"      - {item['ITEM_DSC']:<52}{pb_tag}")
            if len(order["items"]) > 5:
                print(f"      ... and {len(order['items']) - 5} more items")

            cr = order["cart_recs"]
            if cr["cart_complement"]:
                print(f"\n    >>> CART COMPLEMENT recs (high-lift partners):")
                for r in cr["cart_complement"]:
                    pb_tag = " [PB]" if r["is_pb"] else ""
                    print(f"      + {r['ITEM_DSC']:<52}{pb_tag}")
                    print(f"        Lift: {r['lift']}  Triggered by: {r['triggered_by']}")
            if cr["item_similarity"]:
                print(f"\n    >>> ITEM SIMILARITY recs:")
                for r in cr["item_similarity"]:
                    pb_tag = " [PB]" if r["is_pb"] else ""
                    print(f"      + {r['ITEM_DSC']:<52}{pb_tag}")
                    print(f"        Sim: {r['similarity']}  Triggered by: {r['triggered_by']}")
            if cr["medline"]:
                print(f"\n    >>> MEDLINE CONVERSION (cart contains Medline):")
                for r in cr["medline"]:
                    print(f"      + {r['ITEM_DSC']:<52} (replaces Medline)")
                    print(f"        ${r['original_price']:.2f} -> ${r['alt_price']:.2f}  "
                          f"Replaces: {r['triggered_by_medline']}")
            if cr["pb_upgrade"]:
                print(f"\n    >>> PRIVATE BRAND UPGRADE (national brand in cart):")
                for r in cr["pb_upgrade"]:
                    print(f"      + {r['ITEM_DSC']:<52}")
                    print(f"        Save {r['savings_pct']:.0f}%: ${r['original_price']:.2f} -> "
                          f"${r['mck_brand_price']:.2f}  Replaces: {r['triggered_by_natl']}")

        # Phase 5: shows raw + normalized score
        # Phase 6: replenishment recs get extra context line
        print(f"\n  PRE-COMPUTED TOP 10 RECOMMENDATIONS:")
        for r in rep["recommendations"]:
            mck_tag = " [MCK]" if r["is_mck_brand"] else ""
            print(f"    Rank {r['rank']:>2}: {r['ITEM_DSC']:<50}{mck_tag}")
            print(f"        Signal: {r['primary_signal']:<22}  "
                  f"Purpose: {r['rec_purpose']:<22}  "
                  f"Score: {r['score']:>7.2f}  Norm: {r['norm_score']:.3f}")
            # Phase 6: extra context for replenishment recs
            if r["primary_signal"] == "replenishment" and r["repl_overdue_ratio"] > 0:
                print(f"        REPLENISHMENT: {r['repl_days_since_last']}d since last  "
                      f"(segment cadence {r['repl_segment_cadence']:.0f}d, "
                      f"overdue {r['repl_overdue_ratio']:.2f}x, "
                      f"peer activity {r['repl_peer_activity']:.0%})")

        q = rep["quality"]
        print(f"\n  QUALITY ASSESSMENT:")
        print(f"    Family overlap with history    : {q['family_overlap_pct']:.0f}%")
        print(f"    Category overlap with history  : {q['category_overlap_pct']:.0f}%")
        print(f"    Unique signal types in top 10  : {q['n_unique_signals']}")
        print(f"    Unique rec purposes in top 10  : {q['n_unique_purposes']}")
        print(f"    McKesson Brand recs in top 10  : {q['n_mck_brand_recs']}/10")
        print(f"    Replenishment recs in top 10   : {q['n_replenishment_recs']}/10")
        print(f"    Specialty matches in top 10    : {q['spclty_match_count']}/10")
        print(f"    Specialty mismatches in top 10 : {q['mismatch_count']}/10  "
              f"(history: {q['history_size_bucket']})")
        if q["spec_filter_breach"]:
            print(f"    *** PHASE 5 BREACH: warm customer with rich history "
                  f"({len(rep['history'])} products) has {q['mismatch_count']} mismatches ***")
        print(f"    Orders simulated               : {q['n_orders_simulated']}")
        print(f"    Orders with cart-complement opp: {q['n_orders_with_cart_recs']}")
        print(f"    Orders with Medline conv opp   : {q['n_orders_with_medline']}")
        print(f"    Orders with PB upgrade opp     : {q['n_orders_with_pb_upg']}")


# Step 8: Charts (Phase 5: adds normalization-validation chart;
#                 Phase 6: adds replenishment-coverage chart, updates color maps)

def generate_charts(reports: list, recs: pd.DataFrame) -> None:
    _s("Step 8: Generating charts")
    OUT_CHARTS_DIR.mkdir(parents=True, exist_ok=True)

    plt.rcParams.update({"font.family": "Arial", "font.size": 9})

    # Chart 1: Signal distribution
    fig, ax = plt.subplots(figsize=(8, 5))
    sig_counts = recs["primary_signal"].value_counts()
    bars = ax.bar(range(len(sig_counts)), sig_counts.values, color="#4472C4")
    ax.set_xticks(range(len(sig_counts)))
    ax.set_xticklabels(sig_counts.index, rotation=30, ha="right")
    ax.set_ylabel("Count")
    ax.set_title("Signal Distribution Across All Recommendations")
    for bar, n in zip(bars, sig_counts.values):
        pct = n / len(recs) * 100
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                f"{n:,}\n({pct:.1f}%)", ha="center", va="bottom", fontsize=8)
    plt.tight_layout()
    plt.savefig(OUT_CHARTS_DIR / "01_signal_distribution.png", dpi=120, bbox_inches="tight")
    plt.close()
    _log("Saved: 01_signal_distribution.png")

    # Chart 2: Rec purpose distribution (Phase 6: added replenishment color)
    if "rec_purpose" in recs.columns:
        fig, ax = plt.subplots(figsize=(9, 5))
        purp_counts = recs["rec_purpose"].value_counts()
        colors = {
            "new_product":         "#4472C4",
            "win_back":            "#ED7D31",
            "cross_sell":          "#FFC000",
            "mckesson_substitute": "#70AD47",
            "replenishment":       "#7030A0",   # Phase 6 - distinct purple
        }
        bar_colors = [colors.get(c, "#888888") for c in purp_counts.index]
        bars = ax.bar(range(len(purp_counts)), purp_counts.values, color=bar_colors)
        ax.set_xticks(range(len(purp_counts)))
        ax.set_xticklabels(purp_counts.index, rotation=15, ha="right")
        ax.set_ylabel("Count")
        ax.set_title("Recommendation Purpose Distribution (Business View)")
        for bar, n in zip(bars, purp_counts.values):
            pct = n / len(recs) * 100
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                    f"{n:,}\n({pct:.1f}%)", ha="center", va="bottom", fontsize=8)
        plt.tight_layout()
        plt.savefig(OUT_CHARTS_DIR / "02_rec_purpose_distribution.png", dpi=120, bbox_inches="tight")
        plt.close()
        _log("Saved: 02_rec_purpose_distribution.png")

    # Chart 3: McKesson Brand penetration by signal type
    if "is_mckesson_brand" in recs.columns:
        fig, ax = plt.subplots(figsize=(10, 6))
        by_sig = recs.groupby("primary_signal").agg(
            total=("DIM_ITEM_E1_CURR_ID", "count"),
            mck_count=("is_mckesson_brand", "sum"),
        )
        by_sig["mck_pct"] = (by_sig["mck_count"] / by_sig["total"] * 100)
        by_sig = by_sig.sort_values("mck_pct", ascending=False)
        bars = ax.bar(range(len(by_sig)), by_sig["mck_pct"].values, color="#70AD47")
        ax.set_xticks(range(len(by_sig)))
        ax.set_xticklabels(by_sig.index, rotation=30, ha="right")
        ax.set_ylabel("McKesson Brand %")
        ax.set_title("McKesson Brand Penetration by Signal Type")
        ax.set_ylim(0, 105)
        overall_pct = recs["is_mckesson_brand"].mean() * 100
        ax.axhline(overall_pct, color="red", linestyle="--",
                   label=f"Overall: {overall_pct:.1f}%")
        ax.legend()
        for bar, pct, total, mck in zip(bars, by_sig["mck_pct"], by_sig["total"], by_sig["mck_count"]):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                    f"{pct:.0f}%\n({mck:,}/{total:,})",
                    ha="center", va="bottom", fontsize=7)
        plt.tight_layout()
        plt.savefig(OUT_CHARTS_DIR / "03_mckesson_brand_by_signal.png", dpi=120, bbox_inches="tight")
        plt.close()
        _log("Saved: 03_mckesson_brand_by_signal.png")

    # Chart 4: Confidence tier distribution
    fig, ax = plt.subplots(figsize=(7, 4))
    conf_counts = recs["confidence_tier"].value_counts()
    colors = {"high": "#70AD47", "medium": "#FFC000", "low": "#C00000"}
    bar_colors = [colors.get(c, "#888888") for c in conf_counts.index]
    bars = ax.bar(range(len(conf_counts)), conf_counts.values, color=bar_colors)
    ax.set_xticks(range(len(conf_counts)))
    ax.set_xticklabels(conf_counts.index)
    ax.set_ylabel("Count")
    ax.set_title("Confidence Tier Distribution")
    for bar, n in zip(bars, conf_counts.values):
        pct = n / len(recs) * 100
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                f"{n:,}\n({pct:.1f}%)", ha="center", va="bottom", fontsize=8)
    plt.tight_layout()
    plt.savefig(OUT_CHARTS_DIR / "04_confidence_distribution.png", dpi=120, bbox_inches="tight")
    plt.close()
    _log("Saved: 04_confidence_distribution.png")

    # Chart 5: Family diversity histogram
    fig, ax = plt.subplots(figsize=(8, 5))
    fam_diversity = recs.groupby("DIM_CUST_CURR_ID")["PROD_FMLY_LVL1_DSC"].nunique()
    ax.hist(fam_diversity, bins=range(0, 12), color="#4472C4", edgecolor="black", alpha=0.8)
    ax.set_xlabel("Number of unique product families in customer's top 10")
    ax.set_ylabel("Number of customers")
    ax.set_title("Family Diversity per Customer")
    ax.axvline(fam_diversity.median(), color="red", linestyle="--",
               label=f"Median: {fam_diversity.median():.0f}")
    ax.legend()
    plt.tight_layout()
    plt.savefig(OUT_CHARTS_DIR / "05_family_diversity.png", dpi=120, bbox_inches="tight")
    plt.close()
    _log("Saved: 05_family_diversity.png")

    # Chart 6: McKesson Brand count per customer
    if "is_mckesson_brand" in recs.columns:
        fig, ax = plt.subplots(figsize=(9, 5))
        mck_per_cust = recs.groupby("DIM_CUST_CURR_ID")["is_mckesson_brand"].sum()
        ax.hist(mck_per_cust, bins=range(0, 12), color="#70AD47", edgecolor="black", alpha=0.8)
        ax.set_xlabel("Number of McKesson Brand recs in customer's top 10")
        ax.set_ylabel("Number of customers")
        ax.set_title("McKesson Brand Penetration per Customer")
        ax.axvline(mck_per_cust.median(), color="red", linestyle="--",
                   label=f"Median: {mck_per_cust.median():.0f}")
        ax.legend()
        plt.tight_layout()
        plt.savefig(OUT_CHARTS_DIR / "06_mck_brand_per_customer.png", dpi=120, bbox_inches="tight")
        plt.close()
        _log("Saved: 06_mck_brand_per_customer.png")

    # Chart 7: Family overlap with history
    fig, ax = plt.subplots(figsize=(10, 6))
    overlaps = [r["quality"]["family_overlap_pct"] for r in reports]
    cust_labels = [f"C{i+1}" for i in range(len(reports))]
    ax.bar(cust_labels, overlaps, color="#5B9BD5", alpha=0.8)
    ax.axhline(50, color="red", linestyle="--", alpha=0.6, label="50% threshold")
    ax.set_xlabel("Sample customers")
    ax.set_ylabel("Family overlap with history (%)")
    ax.set_title("Recommendation Family Overlap with Customer Purchase History")
    ax.set_ylim(0, 105)
    plt.xticks(rotation=90, fontsize=7)
    ax.legend()
    plt.tight_layout()
    plt.savefig(OUT_CHARTS_DIR / "07_family_overlap.png", dpi=120, bbox_inches="tight")
    plt.close()
    _log("Saved: 07_family_overlap.png")

    # Chart 8: Cart simulation opportunities by status
    fig, ax = plt.subplots(figsize=(9, 5))
    status_groups = {}
    for rep in reports:
        status = "unknown"
        for line in rep["profile"]:
            if line.startswith("Status:"):
                status = line.split(":", 1)[1].strip()
                break
        if status not in status_groups:
            status_groups[status] = {
                "cart_recs": 0, "medline": 0, "pb_upg": 0, "n": 0
            }
        q = rep["quality"]
        status_groups[status]["cart_recs"] += q["n_orders_with_cart_recs"]
        status_groups[status]["medline"]   += q["n_orders_with_medline"]
        status_groups[status]["pb_upg"]    += q["n_orders_with_pb_upg"]
        status_groups[status]["n"] += 1

    x = np.arange(len(status_groups))
    width = 0.27
    statuses = list(status_groups.keys())
    cart_vals    = [status_groups[s]["cart_recs"] for s in statuses]
    medline_vals = [status_groups[s]["medline"]   for s in statuses]
    pb_vals      = [status_groups[s]["pb_upg"]    for s in statuses]

    ax.bar(x - width, cart_vals,    width, label="Cart Complement", color="#4472C4")
    ax.bar(x,         medline_vals, width, label="Medline Conv",    color="#ED7D31")
    ax.bar(x + width, pb_vals,      width, label="PB Upgrade",      color="#70AD47")
    ax.set_xticks(x)
    ax.set_xticklabels(statuses)
    ax.set_ylabel("Number of orders with opportunity")
    ax.set_title("Cart Simulation Opportunities by Customer Status")
    ax.legend()
    plt.tight_layout()
    plt.savefig(OUT_CHARTS_DIR / "08_cart_opportunities_by_status.png",
                dpi=120, bbox_inches="tight")
    plt.close()
    _log("Saved: 08_cart_opportunities_by_status.png")

    # Chart 9 (Phase 5 validation): raw vs normalized score per signal.
    if "normalized_score" in recs.columns:
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        sig_order = sorted(recs["primary_signal"].unique())

        raw_data = [recs[recs["primary_signal"] == s]["numeric_score"].values for s in sig_order]
        norm_data = [recs[recs["primary_signal"] == s]["normalized_score"].values for s in sig_order]

        bp1 = ax1.boxplot(raw_data, labels=sig_order, patch_artist=True, showfliers=False)
        for patch in bp1["boxes"]:
            patch.set_facecolor("#ED7D31")
        ax1.set_title("RAW numeric_score by signal\n(varying scales across signals)")
        ax1.set_ylabel("numeric_score")
        ax1.tick_params(axis="x", rotation=30)
        for label in ax1.get_xticklabels():
            label.set_ha("right")

        bp2 = ax2.boxplot(norm_data, labels=sig_order, patch_artist=True, showfliers=False)
        for patch in bp2["boxes"]:
            patch.set_facecolor("#70AD47")
        ax2.set_title("NORMALIZED score by signal\n(0-1 percentile - fair cross-signal comparison)")
        ax2.set_ylabel("normalized_score (0-1 percentile within signal)")
        ax2.set_ylim(0, 1.05)
        ax2.tick_params(axis="x", rotation=30)
        for label in ax2.get_xticklabels():
            label.set_ha("right")
        ax2.axhline(0.5, color="red", linestyle="--", alpha=0.5, label="Median (expected)")
        ax2.legend()

        plt.tight_layout()
        plt.savefig(OUT_CHARTS_DIR / "09_score_normalization.png",
                    dpi=120, bbox_inches="tight")
        plt.close()
        _log("Saved: 09_score_normalization.png  (Phase 5 validation)")
    else:
        _log("Skipped: 09_score_normalization.png (normalized_score column missing)")

    # Chart 10 (NEW Phase 6): replenishment overdue distribution and peer activity
    # Sanity check that replenishment is firing on real overdue cases backed
    # by active peer cohorts, not just on edge cases.
    repl_recs = recs[recs["primary_signal"] == "replenishment"]
    if len(repl_recs) > 0:
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

        # Left: rec count by signal incl. replenishment
        sig_counts = recs["primary_signal"].value_counts()
        # highlight replenishment
        bar_colors = ["#7030A0" if s == "replenishment" else "#5B9BD5" for s in sig_counts.index]
        bars = ax1.bar(range(len(sig_counts)), sig_counts.values, color=bar_colors)
        ax1.set_xticks(range(len(sig_counts)))
        ax1.set_xticklabels(sig_counts.index, rotation=30, ha="right")
        ax1.set_ylabel("Total recs")
        ax1.set_title("Phase 6: Replenishment slot share\n(highlighted in purple)")
        for bar, n in zip(bars, sig_counts.values):
            pct = n / len(recs) * 100
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                     f"{pct:.1f}%", ha="center", va="bottom", fontsize=8)

        # Right: replenishment confidence tier breakdown
        repl_conf = repl_recs["confidence_tier"].value_counts()
        conf_colors = {"high": "#70AD47", "medium": "#FFC000", "low": "#C00000"}
        bar_colors = [conf_colors.get(c, "#888888") for c in repl_conf.index]
        bars = ax2.bar(range(len(repl_conf)), repl_conf.values, color=bar_colors)
        ax2.set_xticks(range(len(repl_conf)))
        ax2.set_xticklabels(repl_conf.index)
        ax2.set_ylabel("Replenishment recs")
        ax2.set_title("Replenishment confidence tier\n(high = clean overdue + active peers)")
        for bar, n in zip(bars, repl_conf.values):
            pct = n / len(repl_recs) * 100
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                     f"{n:,}\n({pct:.1f}%)", ha="center", va="bottom", fontsize=8)

        plt.tight_layout()
        plt.savefig(OUT_CHARTS_DIR / "10_replenishment_validation.png",
                    dpi=120, bbox_inches="tight")
        plt.close()
        _log("Saved: 10_replenishment_validation.png  (Phase 6 validation)")
    else:
        _log("Skipped: 10_replenishment_validation.png (no replenishment recs found)")


# Excel styling helper

def _style_sheet(ws, df: pd.DataFrame, hc: str = "1F4E79") -> None:
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = bdr

    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = bdr

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(
            col_cells[0].column)].width = min(max(w + 2, 12), 55)


# Step 9: Save xlsx
# Phase 5: summary now includes mismatch_count + history bucket
# Phase 6: summary includes n_replenishment_recs; new sheet 07_replenishment

def save_xlsx(reports: list, recs_full: pd.DataFrame) -> None:
    _s("Step 9: Building xlsx report")
    t0 = time.time()

    # Sheet 1: Summary (Phase 5 + Phase 6 columns)
    summary_rows = []
    for rep in reports:
        status = "?"
        seg = "?"
        spc = "?"
        for line in rep["profile"]:
            if line.startswith("Status:"):
                status = line.split(":", 1)[1].strip()
            elif line.startswith("Segment:"):
                seg = line.split(":", 1)[1].strip()
            elif line.startswith("Specialty:"):
                spc = line.split(":", 1)[1].strip()

        q = rep["quality"]
        summary_rows.append({
            "customer_id":          rep["customer_id"],
            "segment":              seg,
            "status":               status,
            "specialty":            spc,
            "n_history_products":   len(rep["history"]),
            "history_size_bucket":  q["history_size_bucket"],
            "n_orders_simulated":   q["n_orders_simulated"],
            "n_recommendations":    len(rep["recommendations"]),
            "family_overlap_pct":   q["family_overlap_pct"],
            "unique_rec_families":  q["n_unique_families"],
            "unique_signals":       q["n_unique_signals"],
            "unique_purposes":      q["n_unique_purposes"],
            "mckesson_brand_count": q["n_mck_brand_recs"],
            "replenishment_count":  q["n_replenishment_recs"],   # Phase 6
            "spclty_matches":       q["spclty_match_count"],
            "spclty_mismatches":    q["mismatch_count"],
            "spec_filter_breach":   q["spec_filter_breach"],
            "orders_with_cart_recs":  q["n_orders_with_cart_recs"],
            "orders_with_medline":    q["n_orders_with_medline"],
            "orders_with_pb_upgrade": q["n_orders_with_pb_upg"],
        })
    summary_df = pd.DataFrame(summary_rows)

    # Sheet 2: Profiles
    profiles_data = []
    for rep in reports:
        d = {"customer_id": rep["customer_id"]}
        for line in rep["profile"]:
            if ":" in line:
                k, v = line.split(":", 1)
                d[k.strip()] = v.strip()
        profiles_data.append(d)
    profiles_df = pd.DataFrame(profiles_data)

    # Sheet 3: History
    history_rows = []
    for rep in reports:
        for h in rep["history"]:
            history_rows.append({"customer_id": rep["customer_id"], **h})
    history_df = pd.DataFrame(history_rows)

    # Sheet 4: Top 10 Recommendations (Phase 5: norm_score; Phase 6: repl context)
    rec_rows = []
    for rep in reports:
        for r in rep["recommendations"]:
            rec_rows.append({"customer_id": rep["customer_id"], **r})
    recs_df = pd.DataFrame(rec_rows)

    # Sheet 5: Order line items
    order_lines = []
    for rep in reports:
        for order in rep["order_simulations"]:
            for item in order["items"]:
                order_lines.append({
                    "customer_id":  rep["customer_id"],
                    "order_rank":   order["order_rank"],
                    "order_date":   order["order_date"],
                    "order_total":  round(order["order_total"], 2),
                    "DIM_ITEM_E1_CURR_ID": item["DIM_ITEM_E1_CURR_ID"],
                    "ITEM_DSC":     item["ITEM_DSC"],
                    "family":       item["family"],
                    "qty":          item["qty"],
                    "spend":        round(item["spend"], 2),
                    "is_pb":        item["is_pb"],
                })
    orders_df = pd.DataFrame(order_lines)

    # Sheet 6: Cart simulations
    cart_rows = []
    for rep in reports:
        cust_id = rep["customer_id"]
        for order in rep["order_simulations"]:
            order_rank = order["order_rank"]
            order_date = order["order_date"]
            for rec_type, rec_list in order["cart_recs"].items():
                for r in rec_list:
                    base = {
                        "customer_id":   cust_id,
                        "order_rank":    order_rank,
                        "order_date":    order_date,
                        "rec_type":      rec_type,
                        "rec_item_id":   r.get("rec_item_id", ""),
                        "ITEM_DSC":      r.get("ITEM_DSC", ""),
                        "family":        r.get("family", ""),
                    }
                    if rec_type == "cart_complement":
                        base["score"] = r.get("lift", 0)
                        base["note"]  = f"Triggered by: {r.get('triggered_by', '')[:30]}"
                    elif rec_type == "item_similarity":
                        base["score"] = r.get("similarity", 0)
                        base["note"]  = f"Triggered by: {r.get('triggered_by', '')[:30]}"
                    elif rec_type == "medline":
                        base["score"] = 0
                        base["note"]  = (f"Replaces Medline: {r.get('triggered_by_medline', '')[:30]}"
                                         f" | ${r.get('original_price', 0):.2f} -> "
                                         f"${r.get('alt_price', 0):.2f}")
                    elif rec_type == "pb_upgrade":
                        base["score"] = r.get("savings_pct", 0)
                        base["note"]  = (f"Save {r.get('savings_pct', 0):.0f}% on: "
                                         f"{r.get('triggered_by_natl', '')[:30]}")
                    cart_rows.append(base)
    cart_df = pd.DataFrame(cart_rows)

    # Sheet 7 (NEW Phase 6): Replenishment recs only - shows all the detail
    # for inspecting the new signal
    repl_only_rows = []
    for rep in reports:
        for r in rep["recommendations"]:
            if r["primary_signal"] == "replenishment":
                repl_only_rows.append({
                    "customer_id":          rep["customer_id"],
                    "rank":                 r["rank"],
                    "ITEM_DSC":             r["ITEM_DSC"],
                    "family":               r["family"],
                    "category":             r["category"],
                    "is_mck_brand":         r["is_mck_brand"],
                    "score":                round(r["score"], 2),
                    "norm_score":           round(r["norm_score"], 3),
                    "confidence":           r["confidence"],
                    "specialty_match":      r["specialty_match"],
                    "days_since_last":      r["repl_days_since_last"],
                    "segment_cadence_days": round(r["repl_segment_cadence"], 0),
                    "overdue_ratio":        round(r["repl_overdue_ratio"], 2),
                    "peer_activity_pct":    round(r["repl_peer_activity"] * 100, 1),
                    "pitch_reason":         r["pitch_reason"],
                })
    repl_df = pd.DataFrame(repl_only_rows) if repl_only_rows else pd.DataFrame(
        [{"note": "No replenishment recs in sample"}]
    )

    # Sheet 0: Aggregate business metrics
    business_metrics = []
    business_metrics.append({"metric": "Total recommendations", "value": f"{len(recs_full):,}"})
    business_metrics.append({"metric": "Unique customers", "value": f"{recs_full['DIM_CUST_CURR_ID'].nunique():,}"})
    business_metrics.append({"metric": "", "value": ""})
    business_metrics.append({"metric": "REC PURPOSE DISTRIBUTION", "value": ""})
    if "rec_purpose" in recs_full.columns:
        for purp, n in recs_full["rec_purpose"].value_counts().items():
            pct = n / len(recs_full) * 100
            business_metrics.append({"metric": f"  {purp}", "value": f"{n:,} ({pct:.1f}%)"})
    business_metrics.append({"metric": "", "value": ""})
    business_metrics.append({"metric": "PRIMARY SIGNAL DISTRIBUTION", "value": ""})
    for sig, n in recs_full["primary_signal"].value_counts().items():
        pct = n / len(recs_full) * 100
        business_metrics.append({"metric": f"  {sig}", "value": f"{n:,} ({pct:.1f}%)"})
    business_metrics.append({"metric": "", "value": ""})
    business_metrics.append({"metric": "MCKESSON BRAND PENETRATION", "value": ""})
    if "is_mckesson_brand" in recs_full.columns:
        n_mck = int(recs_full["is_mckesson_brand"].sum())
        pct_mck = n_mck / len(recs_full) * 100
        business_metrics.append({"metric": "  Total McKesson Brand recs", "value": f"{n_mck:,} ({pct_mck:.1f}%)"})
        business_metrics.append({"metric": "", "value": ""})
        business_metrics.append({"metric": "  By signal type:", "value": ""})
        by_sig = recs_full.groupby("primary_signal").agg(
            total=("DIM_ITEM_E1_CURR_ID", "count"),
            mck_count=("is_mckesson_brand", "sum"),
        )
        by_sig["mck_pct"] = (by_sig["mck_count"] / by_sig["total"] * 100).round(1)
        by_sig = by_sig.sort_values("mck_pct", ascending=False)
        for sig, row in by_sig.iterrows():
            business_metrics.append({
                "metric": f"    {sig}",
                "value": f"{int(row['mck_count']):,}/{int(row['total']):,} ({row['mck_pct']:.1f}%)"
            })

    # Phase 5 validation block
    business_metrics.append({"metric": "", "value": ""})
    business_metrics.append({"metric": "PHASE 5 VALIDATION", "value": ""})
    if "normalized_score" in recs_full.columns:
        norm_by_sig = recs_full.groupby("primary_signal").agg(
            raw_med=("numeric_score", "median"),
            norm_med=("normalized_score", "median"),
            n=("numeric_score", "size"),
        ).round(3)
        business_metrics.append({"metric": "  Score normalization (per-signal medians):", "value": ""})
        for sig, row in norm_by_sig.iterrows():
            business_metrics.append({
                "metric": f"    {sig}",
                "value": f"raw={row['raw_med']:.2f}  norm={row['norm_med']:.3f}  n={int(row['n']):,}"
            })
    else:
        business_metrics.append({"metric": "  normalized_score column", "value": "MISSING (pre-Phase-5 file)"})

    # Specialty filter validation
    n_breach = sum(r["quality"]["spec_filter_breach"] for r in reports)
    business_metrics.append({"metric": "", "value": ""})
    business_metrics.append({"metric": "  Specialty filter breaches", "value": f"{n_breach}/{len(reports)} customers"})
    business_metrics.append({
        "metric": "  (warm + rich-history customers)",
        "value": "0 breaches expected" if n_breach == 0 else "INVESTIGATE - filter has a hole"
    })

    # Phase 6 validation block
    business_metrics.append({"metric": "", "value": ""})
    business_metrics.append({"metric": "PHASE 6 VALIDATION (replenishment)", "value": ""})
    repl_recs = recs_full[recs_full["primary_signal"] == "replenishment"]
    if len(repl_recs) > 0:
        n_repl_total = len(repl_recs)
        n_repl_custs = repl_recs["DIM_CUST_CURR_ID"].nunique()
        pct_total = n_repl_total / len(recs_full) * 100
        business_metrics.append({
            "metric": "  Total replenishment recs",
            "value": f"{n_repl_total:,} ({pct_total:.1f}% of all recs)"
        })
        business_metrics.append({
            "metric": "  Customers with replenishment recs",
            "value": f"{n_repl_custs:,}"
        })
        if "confidence_tier" in repl_recs.columns:
            for tier in ["high", "medium", "low"]:
                n_t = int((repl_recs["confidence_tier"] == tier).sum())
                p_t = n_t / max(n_repl_total, 1) * 100
                business_metrics.append({
                    "metric": f"    Confidence tier: {tier}",
                    "value": f"{n_t:,} ({p_t:.1f}%)"
                })
        # Sample-level: customers in our sample with replenishment recs
        n_sample_with_repl = sum(1 for r in reports if r["quality"]["n_replenishment_recs"] > 0)
        avg_repl = np.mean([r["quality"]["n_replenishment_recs"] for r in reports])
        business_metrics.append({
            "metric": "  Sample customers w/ replenishment recs",
            "value": f"{n_sample_with_repl}/{len(reports)} (avg {avg_repl:.1f}/cust)"
        })
    else:
        business_metrics.append({"metric": "  Replenishment recs", "value": "NONE - check Phase 6 pipeline"})

    business_metrics_df = pd.DataFrame(business_metrics)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        business_metrics_df.to_excel(writer, sheet_name="00_business_metrics", index=False)
        summary_df.to_excel(writer,   sheet_name="01_summary",        index=False)
        profiles_df.to_excel(writer,  sheet_name="02_profiles",       index=False)
        history_df.to_excel(writer,   sheet_name="03_history",        index=False)
        recs_df.to_excel(writer,      sheet_name="04_top10_recs",     index=False)
        orders_df.to_excel(writer,    sheet_name="05_order_lines",    index=False)
        cart_df.to_excel(writer,      sheet_name="06_cart_simulations", index=False)
        repl_df.to_excel(writer,      sheet_name="07_replenishment",  index=False)   # Phase 6

        wb = writer.book
        _style_sheet(writer.sheets["00_business_metrics"], business_metrics_df, hc="002060")
        _style_sheet(writer.sheets["01_summary"],         summary_df,  hc="1F4E79")
        _style_sheet(writer.sheets["02_profiles"],        profiles_df, hc="2E75B6")
        _style_sheet(writer.sheets["03_history"],         history_df,  hc="375623")
        _style_sheet(writer.sheets["04_top10_recs"],      recs_df,     hc="833C00")
        _style_sheet(writer.sheets["05_order_lines"],     orders_df,   hc="6F2DA8")
        _style_sheet(writer.sheets["06_cart_simulations"], cart_df,    hc="C00000")
        _style_sheet(writer.sheets["07_replenishment"],   repl_df,     hc="7030A0")   # Phase 6

        wb["00_business_metrics"].sheet_properties.tabColor = "002060"
        wb["01_summary"].sheet_properties.tabColor          = "1F4E79"
        wb["02_profiles"].sheet_properties.tabColor         = "2E75B6"
        wb["03_history"].sheet_properties.tabColor          = "375623"
        wb["04_top10_recs"].sheet_properties.tabColor       = "833C00"
        wb["05_order_lines"].sheet_properties.tabColor      = "6F2DA8"
        wb["06_cart_simulations"].sheet_properties.tabColor = "C00000"
        wb["07_replenishment"].sheet_properties.tabColor    = "7030A0"

    size_kb = OUT_XLSX.stat().st_size / 1024
    _log(f"Saved: {OUT_XLSX.relative_to(ROOT)}  ({size_kb:.0f} KB, 8 sheets)  in {time.time()-t0:.1f}s")


# Step 10: Save text report

def save_text_report(reports: list, recs_full: pd.DataFrame) -> None:
    _s("Step 10: Saving text report")

    OUT_TXT.parent.mkdir(parents=True, exist_ok=True)
    lines = []
    lines.append("=" * 80)
    lines.append("  RECOMMENDATION SANITY CHECK REPORT (Phase 6)")
    lines.append("=" * 80)
    lines.append(f"  Total customers reviewed: {len(reports)}")
    lines.append("")

    # Aggregate business metrics at the top
    lines.append("=" * 80)
    lines.append("  AGGREGATE BUSINESS METRICS")
    lines.append("=" * 80)
    lines.append(f"  Total recommendations: {len(recs_full):,}")
    lines.append(f"  Unique customers: {recs_full['DIM_CUST_CURR_ID'].nunique():,}")
    lines.append("")
    if "rec_purpose" in recs_full.columns:
        lines.append("  Rec purpose distribution:")
        for purp, n in recs_full["rec_purpose"].value_counts().items():
            pct = n / len(recs_full) * 100
            lines.append(f"    {purp:<22}  {n:>10,} ({pct:.1f}%)")
        lines.append("")
    lines.append("  Primary signal distribution:")
    for sig, n in recs_full["primary_signal"].value_counts().items():
        pct = n / len(recs_full) * 100
        lines.append(f"    {sig:<22}  {n:>10,} ({pct:.1f}%)")
    lines.append("")
    if "is_mckesson_brand" in recs_full.columns:
        n_mck = int(recs_full["is_mckesson_brand"].sum())
        pct_mck = n_mck / len(recs_full) * 100
        lines.append(f"  McKesson Brand penetration: {n_mck:,} / {len(recs_full):,} ({pct_mck:.1f}%)")
        lines.append("")
        lines.append("  McKesson Brand by signal type:")
        by_sig = recs_full.groupby("primary_signal").agg(
            total=("DIM_ITEM_E1_CURR_ID", "count"),
            mck_count=("is_mckesson_brand", "sum"),
        )
        by_sig["mck_pct"] = (by_sig["mck_count"] / by_sig["total"] * 100).round(1)
        by_sig = by_sig.sort_values("mck_pct", ascending=False)
        for sig, row in by_sig.iterrows():
            lines.append(f"    {sig:<22}  "
                         f"{int(row['mck_count']):>8,}/{int(row['total']):<8,}  "
                         f"({row['mck_pct']:>5.1f}%)")
    lines.append("")

    # Phase 5 validation block
    lines.append("=" * 80)
    lines.append("  PHASE 5 VALIDATION")
    lines.append("=" * 80)
    if "normalized_score" in recs_full.columns:
        lines.append("  Score normalization (per-signal medians):")
        norm_by_sig = recs_full.groupby("primary_signal").agg(
            raw_med=("numeric_score", "median"),
            norm_med=("normalized_score", "median"),
            n=("numeric_score", "size"),
        ).round(3)
        for sig, row in norm_by_sig.iterrows():
            lines.append(f"    {sig:<22}  raw={row['raw_med']:>7.2f}  "
                         f"norm={row['norm_med']:.3f}  n={int(row['n']):,}")
        lines.append("")
        lines.append("  Expected: norm_med ~0.5 for all signals (means ranker treats them fairly)")
    else:
        lines.append("  WARNING: normalized_score column missing - this is a pre-Phase-5 file")
    lines.append("")

    n_breach = sum(r["quality"]["spec_filter_breach"] for r in reports)
    lines.append(f"  Specialty filter breaches: {n_breach}/{len(reports)} customers")
    if n_breach > 0:
        lines.append("  *** Warm customers with rich history that still got mismatch recs: ***")
        for rep in reports:
            if rep["quality"]["spec_filter_breach"]:
                lines.append(f"      Customer {rep['customer_id']}  "
                             f"history={len(rep['history'])} products  "
                             f"mismatches={rep['quality']['mismatch_count']}/10")
    else:
        lines.append("  GOOD: zero breaches - the specialty filter is working as designed.")
    lines.append("")

    # Phase 6 validation block
    lines.append("=" * 80)
    lines.append("  PHASE 6 VALIDATION (replenishment signal)")
    lines.append("=" * 80)
    repl_recs = recs_full[recs_full["primary_signal"] == "replenishment"]
    if len(repl_recs) > 0:
        n_repl = len(repl_recs)
        pct_repl = n_repl / len(recs_full) * 100
        n_custs = repl_recs["DIM_CUST_CURR_ID"].nunique()
        lines.append(f"  Total replenishment recs        : {n_repl:,} ({pct_repl:.1f}% of all recs)")
        lines.append(f"  Customers with replenishment    : {n_custs:,}")
        lines.append(f"  Avg replenishment recs/customer : {n_repl/max(n_custs,1):.2f}")
        if "confidence_tier" in repl_recs.columns:
            lines.append(f"  Confidence tier breakdown:")
            for tier in ["high", "medium", "low"]:
                n_t = int((repl_recs["confidence_tier"] == tier).sum())
                p_t = n_t / max(n_repl, 1) * 100
                lines.append(f"    {tier:<8}  {n_t:>8,} ({p_t:.1f}%)")
        # Sample-level
        n_sample_with_repl = sum(1 for r in reports if r["quality"]["n_replenishment_recs"] > 0)
        avg_repl = np.mean([r["quality"]["n_replenishment_recs"] for r in reports])
        lines.append("")
        lines.append(f"  Sample customers w/ replenishment recs: {n_sample_with_repl}/{len(reports)}")
        lines.append(f"  Average replenishment recs in sample  : {avg_repl:.1f} per customer")
    else:
        lines.append("  WARNING: no replenishment recs found - Phase 6 may not have run")
    lines.append("")

    for idx, rep in enumerate(reports, 1):
        lines.append("=" * 80)
        lines.append(f"  CUSTOMER {idx}/{len(reports)}: {rep['customer_id']}")
        lines.append("=" * 80)

        lines.append("")
        lines.append("  CUSTOMER PROFILE:")
        for line in rep["profile"]:
            lines.append(f"    {line}")

        lines.append("")
        lines.append(f"  HISTORICAL PURCHASES (top {len(rep['history'])}):")
        for h in rep["history"]:
            pb_tag = " [PB]" if h["is_pb"] else ""
            lines.append(f"    {h['rank']:>2}. {h['ITEM_DSC']:<52}{pb_tag}")
            lines.append(f"        Family: {h['family']:<25}  Spent: ${h['total_spend']:>8,.0f}  "
                         f"Last: {h['last_order']}")

        lines.append("")
        lines.append("  RECENT ORDERS WITH CART-AWARE RECOMMENDATIONS:")
        for order in rep["order_simulations"]:
            lines.append("")
            lines.append(f" ORDER #{order['order_rank']} ({order['order_date']}, "
                         f"${order['order_total']:,.0f}, {order['line_count']} lines)")
            lines.append(f"    Items in this order:")
            for item in order["items"]:
                pb_tag = " [PB]" if item["is_pb"] else ""
                lines.append(f"      - {item['ITEM_DSC']:<52}{pb_tag}  "
                             f"(qty {item['qty']:.0f}, ${item['spend']:.2f})")

            cr = order["cart_recs"]
            if cr["cart_complement"]:
                lines.append("")
                lines.append("    CART COMPLEMENT recs:")
                for r in cr["cart_complement"]:
                    pb_tag = " [PB]" if r["is_pb"] else ""
                    lines.append(f"      + {r['ITEM_DSC']:<52}{pb_tag}")
                    lines.append(f"        Lift: {r['lift']}  ${r['median_price']:.2f}  "
                                 f"Triggered by: {r['triggered_by']}")
            if cr["item_similarity"]:
                lines.append("")
                lines.append("    ITEM SIMILARITY recs:")
                for r in cr["item_similarity"]:
                    pb_tag = " [PB]" if r["is_pb"] else ""
                    lines.append(f"      + {r['ITEM_DSC']:<52}{pb_tag}")
                    lines.append(f"        Sim: {r['similarity']}  ${r['median_price']:.2f}  "
                                 f"Triggered by: {r['triggered_by']}")
            if cr["medline"]:
                lines.append("")
                lines.append("    MEDLINE CONVERSION (cart contains Medline):")
                for r in cr["medline"]:
                    lines.append(f"      + {r['ITEM_DSC']:<52} (replaces Medline)")
                    lines.append(f"        ${r['original_price']:.2f} -> ${r['alt_price']:.2f}  "
                                 f"Replaces: {r['triggered_by_medline']}")
            if cr["pb_upgrade"]:
                lines.append("")
                lines.append("    PRIVATE BRAND UPGRADE (national brand in cart):")
                for r in cr["pb_upgrade"]:
                    lines.append(f"      + {r['ITEM_DSC']:<52}")
                    lines.append(f"        Save {r['savings_pct']:.0f}%: ${r['original_price']:.2f} -> "
                                 f"${r['mck_brand_price']:.2f}  Replaces: {r['triggered_by_natl']}")

        lines.append("")
        lines.append(f"  PRE-COMPUTED TOP 10 RECOMMENDATIONS:")
        for r in rep["recommendations"]:
            mck_tag = " [MCK]" if r["is_mck_brand"] else ""
            lines.append(f"    Rank {r['rank']:>2}: {r['ITEM_DSC']:<50}{mck_tag}")
            lines.append(f"        Signal: {r['primary_signal']:<22}  "
                         f"Purpose: {r['rec_purpose']:<22}")
            lines.append(f"        Conf: {r['confidence']:<7}  "
                         f"Score: {r['score']:>7.2f}  Norm: {r['norm_score']:.3f}  "
                         f"Specialty: {r['specialty_match']}")
            # Phase 6: replenishment context line
            if r["primary_signal"] == "replenishment" and r["repl_overdue_ratio"] > 0:
                lines.append(f"        REPLENISHMENT: {r['repl_days_since_last']}d since last  "
                             f"(segment cadence {r['repl_segment_cadence']:.0f}d, "
                             f"overdue {r['repl_overdue_ratio']:.2f}x, "
                             f"peer activity {r['repl_peer_activity']:.0%})")
            lines.append(f"        Reason: {r['pitch_reason']}")

        q = rep["quality"]
        lines.append("")
        lines.append(f"  QUALITY ASSESSMENT:")
        lines.append(f"    Family overlap with history     : {q['family_overlap_pct']:.0f}%")
        lines.append(f"    Category overlap with history   : {q['category_overlap_pct']:.0f}%")
        lines.append(f"    Unique signal types in top 10   : {q['n_unique_signals']}")
        lines.append(f"    Unique rec purposes in top 10   : {q['n_unique_purposes']}")
        lines.append(f"    McKesson Brand recs in top 10   : {q['n_mck_brand_recs']}/10")
        lines.append(f"    Replenishment recs in top 10    : {q['n_replenishment_recs']}/10")
        lines.append(f"    Specialty matches in top 10     : {q['spclty_match_count']}/10")
        lines.append(f"    Specialty mismatches in top 10  : {q['mismatch_count']}/10  "
                     f"(history: {q['history_size_bucket']})")
        if q["spec_filter_breach"]:
            lines.append(f"    *** PHASE 5 BREACH: warm + rich history but {q['mismatch_count']} mismatches ***")
        lines.append(f"    Orders simulated                : {q['n_orders_simulated']}")
        lines.append(f"    Orders with cart-complement opp : {q['n_orders_with_cart_recs']}")
        lines.append(f"    Orders with Medline conv opp    : {q['n_orders_with_medline']}")
        lines.append(f"    Orders with PB upgrade opp      : {q['n_orders_with_pb_upg']}")
        lines.append("")

    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")
    size_kb = OUT_TXT.stat().st_size / 1024
    _log(f"Saved: {OUT_TXT.relative_to(ROOT)}  ({size_kb:.0f} KB)")


# Main

def main() -> None:
    log_file = _setup_terminal_logging()

    try:
        print()
        print("=" * 80)
        print("  RECOMMENDATION SANITY CHECK (Phase 6)")
        print("=" * 80)
        start = time.time()

        data = load_data()
        sample_df = select_sample_customers(data)

        if len(sample_df) == 0:
            print("\nFATAL: No sample customers selected", file=sys.stderr)
            sys.exit(1)

        cust_ids = sample_df["DIM_CUST_CURR_ID"].tolist()
        history = get_purchase_history(cust_ids)
        recent_orders = get_recent_orders(cust_ids)

        _s(f"Step 5-6: Building reports for {len(sample_df):,} customers")
        t0 = time.time()
        reports = []
        for _, row in sample_df.iterrows():
            rep = build_customer_report(
                row, history, recent_orders, data["recs"], data["products"],
                data["cooccur"], data["item_sim"], data["pb_equiv"],
                data["replenishment"],
            )
            reports.append(rep)
        _log(f"Built {len(reports)} customer reports in {time.time()-t0:.1f}s")

        print_terminal_report(reports)
        generate_charts(reports, data["recs"])
        save_xlsx(reports, data["recs"])
        save_text_report(reports, data["recs"])

        _s("Complete")
        _log(f"Total time: {time.time() - start:.1f}s")
        _log("")
        _log("Outputs:")
        _log(f"  Terminal log : {OUT_TERMINAL_LOG.relative_to(ROOT)}")
        _log(f"  XLSX         : {OUT_XLSX.relative_to(ROOT)}  (8 sheets, includes 07_replenishment)")
        _log(f"  TXT          : {OUT_TXT.relative_to(ROOT)}")
        _log(f"  Charts (10)  : {OUT_CHARTS_DIR.relative_to(ROOT)}/")
    finally:
        # Restore stdout/stderr and close log file no matter what
        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__
        log_file.close()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise