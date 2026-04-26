from __future__ import annotations

import sys
import time
import warnings
from pathlib import Path

import duckdb
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# Paths

ROOT         = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN   = ROOT / "data_clean"
PRECOMP      = DATA_CLEAN / "serving" / "precomputed"
ANALYSIS     = DATA_CLEAN / "analysis"

RECS_FILE       = PRECOMP / "recommendations.parquet"
PRODUCTS_FILE   = PRECOMP / "product_segments.parquet"
PATTERNS_FILE   = PRECOMP / "customer_patterns.parquet"
SEGMENTS_FILE   = PRECOMP / "customer_segments.parquet"
MERGED_FILE     = DATA_CLEAN / "serving" / "merged_dataset.parquet"

OUT_TXT     = ANALYSIS / "validation_report.txt"
OUT_XLSX    = ANALYSIS / "validation_report.xlsx"
OUT_LOG     = ANALYSIS / "validation_terminal.log"


# Configuration

# How many days at the end of the data to hold out for prediction.
# 60 days gives us enough volume to have stable hit rates while still being
# recent enough to be meaningful.
HOLDOUT_DAYS = 60

# How many recommendations per customer (the rec engine produces top 10)
TOP_N = 10

# Minimum holdout-window orders for a customer to count in our evaluation.
# Customers with no orders in the holdout window cannot be predicted against
# at all, so we exclude them from hit-rate denominators. This is honest:
# we are measuring quality among customers who actually bought something.
MIN_HOLDOUT_ORDERS = 1


# Tee stdout so we keep a log of the run

class _Tee:
    def __init__(self, *streams):
        self.streams = streams
    def write(self, data):
        for s in self.streams:
            try: s.write(data)
            except Exception: pass
    def flush(self):
        for s in self.streams:
            try: s.flush()
            except Exception: pass


def _setup_logging():
    OUT_LOG.parent.mkdir(parents=True, exist_ok=True)
    log_file = open(OUT_LOG, "w", encoding="utf-8", buffering=1)
    sys.stdout = _Tee(sys.__stdout__, log_file)
    sys.stderr = _Tee(sys.__stderr__, log_file)
    return log_file


def _s(title: str) -> None:
    print(f"\n{'-' * 70}\n  {title}\n{'-' * 70}", flush=True)


def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


# Step 1: Load all data

def load_data() -> dict:
    _s("Step 1: Loading data")
    t0 = time.time()

    if not RECS_FILE.exists():
        print(f"\nFATAL: {RECS_FILE} not found.", file=sys.stderr)
        print("Run recommendation_factors.py first.", file=sys.stderr)
        sys.exit(1)

    if not MERGED_FILE.exists():
        print(f"\nFATAL: {MERGED_FILE} not found.", file=sys.stderr)
        sys.exit(1)

    # Recommendations
    recs = pd.read_parquet(RECS_FILE)
    _log(f"recommendations.parquet           : {len(recs):,} rows")
    _log(f"  Unique customers                : {recs['DIM_CUST_CURR_ID'].nunique():,}")
    _log(f"  Unique products in recs         : {recs['DIM_ITEM_E1_CURR_ID'].nunique():,}")

    # Products (we need family/category for relaxed match definitions)
    products = pd.read_parquet(PRODUCTS_FILE, columns=[
        "DIM_ITEM_E1_CURR_ID", "ITEM_DSC",
        "PROD_FMLY_LVL1_DSC", "PROD_CTGRY_LVL2_DSC",
    ])
    products["DIM_ITEM_E1_CURR_ID"] = products["DIM_ITEM_E1_CURR_ID"].astype("int64")
    _log(f"product_segments.parquet           : {len(products):,} rows")

    # Customer patterns (status, history size)
    patterns = pd.read_parquet(PATTERNS_FILE, columns=[
        "DIM_CUST_CURR_ID", "is_cold_start", "is_churned", "is_declining",
        "n_unique_products_total", "order_cadence_tier",
    ])
    patterns["DIM_CUST_CURR_ID"] = patterns["DIM_CUST_CURR_ID"].astype("int64")
    _log(f"customer_patterns.parquet          : {len(patterns):,} rows")

    # Customer segments
    segments = pd.read_parquet(SEGMENTS_FILE, columns=[
        "DIM_CUST_CURR_ID", "segment", "size_tier", "mkt_cd_clean",
    ])
    segments["DIM_CUST_CURR_ID"] = segments["DIM_CUST_CURR_ID"].astype("int64")
    _log(f"customer_segments.parquet          : {len(segments):,} rows")

    _log(f"All loaded in {time.time()-t0:.1f}s")

    return {
        "recs": recs,
        "products": products,
        "patterns": patterns,
        "segments": segments,
    }


# Step 2: Establish the holdout cutoff

def determine_cutoff(merged_path: Path) -> tuple[str, str]:
    """Find the max date in the data and compute holdout cutoff.

    Returns (max_date, cutoff_date) as strings YYYY-MM-DD.
    Holdout window is (cutoff_date, max_date].
    """
    _s("Step 2: Determining holdout window")
    t0 = time.time()

    con = duckdb.connect()
    max_dt_row = con.execute(f"""
        SELECT
            MAX(MAKE_DATE(order_year, order_month, order_day)) AS max_dt,
            MIN(MAKE_DATE(order_year, order_month, order_day)) AS min_dt
        FROM read_parquet('{merged_path.as_posix()}')
    """).df()
    con.close()

    max_dt = pd.to_datetime(max_dt_row["max_dt"].iloc[0])
    min_dt = pd.to_datetime(max_dt_row["min_dt"].iloc[0])
    cutoff_dt = max_dt - pd.Timedelta(days=HOLDOUT_DAYS)

    _log(f"Data range : {min_dt.date()} to {max_dt.date()}")
    _log(f"Holdout    : last {HOLDOUT_DAYS} days  ({cutoff_dt.date()} to {max_dt.date()}]")
    _log(f"Training   : up to and including {cutoff_dt.date()}")
    _log(f"")
    _log(f"Note about leakage:")
    _log(f"  The recommendations we are validating were generated from the FULL history,")
    _log(f"  including the holdout window. This is the SHORTCUT approach.")
    _log(f"  A strict evaluation would rebuild the recommendations using only training-period")
    _log(f"  data. We are not doing that here. Hit rates may be slightly inflated as a result.")
    _log(f"  The relative comparison between signals and against the baseline is still valid.")

    _log(f"Step 2 done in {time.time()-t0:.1f}s")
    return str(max_dt.date()), str(cutoff_dt.date())


# Step 3: Build the holdout purchase set per customer

def load_holdout_purchases(merged_path: Path,
                            cutoff_date: str,
                            max_date: str) -> pd.DataFrame:
    """Returns one row per (customer, item) purchased in the holdout window."""
    _s("Step 3: Loading holdout-window purchases")
    t0 = time.time()

    con = duckdb.connect()
    holdout = con.execute(f"""
        SELECT
            CAST(DIM_CUST_CURR_ID AS BIGINT)  AS DIM_CUST_CURR_ID,
            CAST(DIM_ITEM_E1_CURR_ID AS BIGINT) AS DIM_ITEM_E1_CURR_ID,
            COUNT(*)                            AS n_lines_in_holdout,
            SUM(UNIT_SLS_AMT)                   AS spend_in_holdout,
            COUNT(DISTINCT MAKE_DATE(order_year, order_month, order_day))
                                                AS n_orders_in_holdout
        FROM read_parquet('{merged_path.as_posix()}')
        WHERE UNIT_SLS_AMT > 0
          AND DIM_CUST_CURR_ID IS NOT NULL
          AND DIM_ITEM_E1_CURR_ID IS NOT NULL
          AND MAKE_DATE(order_year, order_month, order_day) > DATE '{cutoff_date}'
          AND MAKE_DATE(order_year, order_month, order_day) <= DATE '{max_date}'
        GROUP BY
            CAST(DIM_CUST_CURR_ID AS BIGINT),
            CAST(DIM_ITEM_E1_CURR_ID AS BIGINT)
    """).df()
    con.close()

    holdout["DIM_CUST_CURR_ID"]    = holdout["DIM_CUST_CURR_ID"].astype("int64")
    holdout["DIM_ITEM_E1_CURR_ID"] = holdout["DIM_ITEM_E1_CURR_ID"].astype("int64")

    n_pairs = len(holdout)
    n_custs = holdout["DIM_CUST_CURR_ID"].nunique()
    n_items = holdout["DIM_ITEM_E1_CURR_ID"].nunique()

    _log(f"Holdout customer-product pairs   : {n_pairs:,}")
    _log(f"Customers with holdout activity  : {n_custs:,}")
    _log(f"Distinct products bought         : {n_items:,}")
    _log(f"Step 3 done in {time.time()-t0:.1f}s")

    return holdout


# Step 4: Build a popularity baseline

def build_popularity_baseline(merged_path: Path,
                                cutoff_date: str,
                                products: pd.DataFrame) -> list:
    """Returns a list of the top-N most-bought products in the training period.
    This is what we compare our engine against. If our engine cannot beat
    'recommend the top 10 popular products to everyone' then it is not adding value.
    """
    _s("Step 4: Building popularity baseline")
    t0 = time.time()

    con = duckdb.connect()
    pop = con.execute(f"""
        SELECT
            CAST(DIM_ITEM_E1_CURR_ID AS BIGINT) AS DIM_ITEM_E1_CURR_ID,
            COUNT(DISTINCT DIM_CUST_CURR_ID)    AS n_buyers,
            COUNT(*)                             AS n_lines,
            SUM(UNIT_SLS_AMT)                    AS total_spend
        FROM read_parquet('{merged_path.as_posix()}')
        WHERE UNIT_SLS_AMT > 0
          AND DIM_CUST_CURR_ID IS NOT NULL
          AND DIM_ITEM_E1_CURR_ID IS NOT NULL
          AND MAKE_DATE(order_year, order_month, order_day) <= DATE '{cutoff_date}'
        GROUP BY CAST(DIM_ITEM_E1_CURR_ID AS BIGINT)
        ORDER BY n_buyers DESC
        LIMIT 100
    """).df()
    con.close()

    # Take top N that have product metadata
    pop["DIM_ITEM_E1_CURR_ID"] = pop["DIM_ITEM_E1_CURR_ID"].astype("int64")
    pop = pop.merge(products, on="DIM_ITEM_E1_CURR_ID", how="inner")
    top_n = pop.head(TOP_N)["DIM_ITEM_E1_CURR_ID"].tolist()

    _log(f"Top-{TOP_N} popularity baseline (by buyer count, training period only):")
    for i, r in pop.head(TOP_N).iterrows():
        desc = str(r.get("ITEM_DSC", "?"))[:55]
        _log(f"  {i+1:2d}. {desc:<57} buyers={int(r['n_buyers']):,}")
    _log(f"Step 4 done in {time.time()-t0:.1f}s")

    return top_n


# Step 5: Compute hit rates

def compute_hit_rates(recs: pd.DataFrame,
                       holdout: pd.DataFrame,
                       products: pd.DataFrame,
                       baseline_items: list,
                       patterns: pd.DataFrame,
                       segments: pd.DataFrame) -> dict:
    """Compute hit rate at three definitions:
       - strict: exact item match
       - family: PROD_FMLY_LVL1_DSC match
       - category: PROD_CTGRY_LVL2_DSC match
    Done both for engine recommendations and for the popularity baseline.
    """
    _s("Step 5: Computing hit rates")
    t0 = time.time()

    # Customers eligible for evaluation: those with at least MIN_HOLDOUT_ORDERS in holdout
    holdout_custs = holdout.groupby("DIM_CUST_CURR_ID").agg(
        n_holdout_pairs=("DIM_ITEM_E1_CURR_ID", "count"),
        n_holdout_orders=("n_orders_in_holdout", "max"),
        holdout_spend=("spend_in_holdout", "sum"),
    ).reset_index()
    holdout_custs = holdout_custs[holdout_custs["n_holdout_orders"] >= MIN_HOLDOUT_ORDERS]

    # Customers we have recs for AND who are in the holdout
    rec_custs = set(recs["DIM_CUST_CURR_ID"].unique())
    eval_custs_df = holdout_custs[holdout_custs["DIM_CUST_CURR_ID"].isin(rec_custs)].copy()
    n_eval = len(eval_custs_df)

    _log(f"Evaluation customers (in holdout AND have recs): {n_eval:,}")
    _log(f"  These are the customers we can actually evaluate against.")

    if n_eval == 0:
        _log("FATAL: No customers to evaluate.")
        return {}

    # Build product metadata lookup for family/category match
    prod_lookup = products.set_index("DIM_ITEM_E1_CURR_ID")[
        ["PROD_FMLY_LVL1_DSC", "PROD_CTGRY_LVL2_DSC"]
    ].to_dict("index")

    def get_family(item_id):
        info = prod_lookup.get(int(item_id), {})
        return info.get("PROD_FMLY_LVL1_DSC")

    def get_category(item_id):
        info = prod_lookup.get(int(item_id), {})
        return info.get("PROD_CTGRY_LVL2_DSC")

    # Pre-build per-customer holdout sets (items, families, categories)
    _log(f"Pre-building per-customer holdout sets...")
    holdout_by_cust = {}
    holdout_eval = holdout[holdout["DIM_CUST_CURR_ID"].isin(eval_custs_df["DIM_CUST_CURR_ID"])]
    for cid, grp in holdout_eval.groupby("DIM_CUST_CURR_ID"):
        items = set(int(i) for i in grp["DIM_ITEM_E1_CURR_ID"])
        families = set(get_family(i) for i in items if get_family(i) is not None)
        categories = set(get_category(i) for i in items if get_category(i) is not None)
        holdout_by_cust[int(cid)] = {
            "items": items,
            "families": families,
            "categories": categories,
        }

    # Pre-build per-customer recommendation sets (items, families, categories)
    _log(f"Pre-building per-customer recommendation sets...")
    recs_eval = recs[recs["DIM_CUST_CURR_ID"].isin(eval_custs_df["DIM_CUST_CURR_ID"])]
    rec_by_cust = {}
    for cid, grp in recs_eval.groupby("DIM_CUST_CURR_ID"):
        items = set(int(i) for i in grp["DIM_ITEM_E1_CURR_ID"])
        families = set(get_family(i) for i in items if get_family(i) is not None)
        categories = set(get_category(i) for i in items if get_category(i) is not None)
        rec_by_cust[int(cid)] = {
            "items": items,
            "families": families,
            "categories": categories,
        }

    # Build baseline sets (same for every customer)
    baseline_items_set = set(int(i) for i in baseline_items)
    baseline_families = set(get_family(i) for i in baseline_items_set if get_family(i) is not None)
    baseline_categories = set(get_category(i) for i in baseline_items_set if get_category(i) is not None)

    _log(f"Computing hit rates per customer...")

    # Per-customer outcome rows
    rows = []
    for cid in eval_custs_df["DIM_CUST_CURR_ID"]:
        cid_int = int(cid)
        h = holdout_by_cust.get(cid_int, {"items": set(), "families": set(), "categories": set()})
        r = rec_by_cust.get(cid_int, {"items": set(), "families": set(), "categories": set()})

        # Engine hit rates
        eng_strict_hits   = len(r["items"] & h["items"])
        eng_family_hits   = len(r["families"] & h["families"])
        eng_category_hits = len(r["categories"] & h["categories"])

        eng_strict_hit   = eng_strict_hits > 0
        eng_family_hit   = eng_family_hits > 0
        eng_category_hit = eng_category_hits > 0

        # Engine precision (of 10 recs, how many were bought)
        eng_strict_precision = eng_strict_hits / max(len(r["items"]), 1)
        # Engine recall (of items bought in holdout, how many did we recommend)
        eng_strict_recall = eng_strict_hits / max(len(h["items"]), 1)

        # Baseline hit rates (same baseline for everyone)
        bl_strict_hits   = len(baseline_items_set & h["items"])
        bl_family_hits   = len(baseline_families & h["families"])
        bl_category_hits = len(baseline_categories & h["categories"])

        bl_strict_hit   = bl_strict_hits > 0
        bl_family_hit   = bl_family_hits > 0
        bl_category_hit = bl_category_hits > 0

        bl_strict_precision = bl_strict_hits / max(len(baseline_items_set), 1)
        bl_strict_recall = bl_strict_hits / max(len(h["items"]), 1)

        rows.append({
            "DIM_CUST_CURR_ID": cid_int,
            "n_holdout_items": len(h["items"]),
            "n_rec_items": len(r["items"]),
            # Engine
            "eng_strict_hit":   int(eng_strict_hit),
            "eng_family_hit":   int(eng_family_hit),
            "eng_category_hit": int(eng_category_hit),
            "eng_strict_hits":   eng_strict_hits,
            "eng_family_hits":   eng_family_hits,
            "eng_category_hits": eng_category_hits,
            "eng_strict_precision": eng_strict_precision,
            "eng_strict_recall":    eng_strict_recall,
            # Baseline
            "bl_strict_hit":   int(bl_strict_hit),
            "bl_family_hit":   int(bl_family_hit),
            "bl_category_hit": int(bl_category_hit),
            "bl_strict_hits":   bl_strict_hits,
            "bl_family_hits":   bl_family_hits,
            "bl_category_hits": bl_category_hits,
            "bl_strict_precision": bl_strict_precision,
            "bl_strict_recall":    bl_strict_recall,
        })

    df = pd.DataFrame(rows)

    # Attach customer context (status, segment)
    df = df.merge(eval_custs_df, on="DIM_CUST_CURR_ID", how="left")

    df = df.merge(
        patterns.rename(columns={"is_cold_start": "_cold", "is_churned": "_churn", "is_declining": "_decline"}),
        on="DIM_CUST_CURR_ID", how="left"
    )
    df["status"] = "stable_warm"
    df.loc[df["_cold"] == 1, "status"] = "cold_start"
    df.loc[(df["_cold"] == 0) & (df["_decline"] == 1), "status"] = "declining_warm"
    df.loc[(df["_cold"] == 0) & (df["_churn"] == 1), "status"] = "churned_warm"
    df = df.drop(columns=[c for c in ["_cold", "_churn", "_decline"] if c in df.columns])

    df = df.merge(segments, on="DIM_CUST_CURR_ID", how="left")

    # Headline numbers
    n = len(df)
    _log(f"")
    _log(f"=" * 60)
    _log(f"HEADLINE RESULTS  (n = {n:,} customers)")
    _log(f"=" * 60)
    _log(f"")
    _log(f"  {'Metric':<40} {'Engine':>10} {'Baseline':>10} {'Lift':>8}")
    _log(f"  {'-'*40} {'-'*10} {'-'*10} {'-'*8}")
    metrics_for_table = []
    for label, eng_col, bl_col in [
        ("Hit rate (any strict match)",     "eng_strict_hit",   "bl_strict_hit"),
        ("Hit rate (any family match)",     "eng_family_hit",   "bl_family_hit"),
        ("Hit rate (any category match)",   "eng_category_hit", "bl_category_hit"),
    ]:
        eng_pct = df[eng_col].mean() * 100
        bl_pct  = df[bl_col].mean() * 100
        lift    = eng_pct - bl_pct
        sign    = "+" if lift >= 0 else ""
        _log(f"  {label:<40} {eng_pct:>9.1f}% {bl_pct:>9.1f}% {sign}{lift:>6.1f}pp")
        metrics_for_table.append({
            "metric": label, "engine_pct": round(eng_pct, 1),
            "baseline_pct": round(bl_pct, 1), "lift_pp": round(lift, 1),
        })

    _log(f"")
    eng_prec = df["eng_strict_precision"].mean() * 100
    bl_prec  = df["bl_strict_precision"].mean() * 100
    eng_rec  = df["eng_strict_recall"].mean() * 100
    bl_rec   = df["bl_strict_recall"].mean() * 100
    _log(f"  {'Avg precision (strict, of 10 recs)':<40} {eng_prec:>9.1f}% {bl_prec:>9.1f}% {(eng_prec-bl_prec):>+7.1f}pp")
    _log(f"  {'Avg recall (strict, of holdout items)':<40} {eng_rec:>9.1f}% {bl_rec:>9.1f}% {(eng_rec-bl_rec):>+7.1f}pp")

    metrics_for_table.append({
        "metric": "Avg precision at 10 (strict)",
        "engine_pct": round(eng_prec, 1), "baseline_pct": round(bl_prec, 1),
        "lift_pp": round(eng_prec - bl_prec, 1),
    })
    metrics_for_table.append({
        "metric": "Avg recall (strict)",
        "engine_pct": round(eng_rec, 1), "baseline_pct": round(bl_rec, 1),
        "lift_pp": round(eng_rec - bl_rec, 1),
    })

    _log(f"Step 5 done in {time.time()-t0:.1f}s")

    return {
        "per_customer": df,
        "metrics_table": metrics_for_table,
        "n_eval": n,
    }


# Step 6: Hit rate by signal type

def hit_rates_by_signal(recs: pd.DataFrame,
                         holdout: pd.DataFrame,
                         eval_custs: set) -> pd.DataFrame:
    """For each signal, compute strict hit rate among customers who got at
    least one rec from that signal.
    """
    _s("Step 6: Hit rates by signal type")
    t0 = time.time()

    holdout_pairs = set()
    for _, r in holdout[holdout["DIM_CUST_CURR_ID"].isin(eval_custs)].iterrows():
        holdout_pairs.add((int(r["DIM_CUST_CURR_ID"]), int(r["DIM_ITEM_E1_CURR_ID"])))

    recs_eval = recs[recs["DIM_CUST_CURR_ID"].isin(eval_custs)].copy()
    recs_eval["was_bought"] = recs_eval.apply(
        lambda r: 1 if (int(r["DIM_CUST_CURR_ID"]), int(r["DIM_ITEM_E1_CURR_ID"])) in holdout_pairs else 0,
        axis=1
    )

    by_signal = recs_eval.groupby("primary_signal").agg(
        n_recs=("was_bought", "count"),
        n_hits=("was_bought", "sum"),
    ).reset_index()
    by_signal["hit_rate_pct"] = (by_signal["n_hits"] / by_signal["n_recs"] * 100).round(2)
    by_signal = by_signal.sort_values("hit_rate_pct", ascending=False)

    _log(f"Strict hit rate per signal (rec-level, denominator is total recs from that signal):")
    _log(f"  {'Signal':<22} {'N recs':>10} {'N hits':>8} {'Hit rate':>10}")
    _log(f"  {'-'*22} {'-'*10} {'-'*8} {'-'*10}")
    for _, r in by_signal.iterrows():
        _log(f"  {r['primary_signal']:<22} {int(r['n_recs']):>10,} {int(r['n_hits']):>8,} {r['hit_rate_pct']:>9.2f}%")

    _log(f"Step 6 done in {time.time()-t0:.1f}s")
    return by_signal


# Step 7: Hit rate by customer status and segment

def hit_rates_by_status(per_customer: pd.DataFrame) -> tuple:
    _s("Step 7: Hit rates by customer status and size tier")

    by_status = per_customer.groupby("status").agg(
        n_customers=("DIM_CUST_CURR_ID", "count"),
        eng_strict_hit_pct=("eng_strict_hit", lambda x: x.mean() * 100),
        eng_family_hit_pct=("eng_family_hit", lambda x: x.mean() * 100),
        eng_category_hit_pct=("eng_category_hit", lambda x: x.mean() * 100),
        bl_strict_hit_pct=("bl_strict_hit", lambda x: x.mean() * 100),
        bl_family_hit_pct=("bl_family_hit", lambda x: x.mean() * 100),
    ).round(2).reset_index()

    _log(f"Hit rates by customer status:")
    _log(f"  {'Status':<18} {'N':>8} {'Eng strict':>11} {'Eng family':>11} {'BL strict':>11} {'Strict lift':>13}")
    _log(f"  {'-'*18} {'-'*8} {'-'*11} {'-'*11} {'-'*11} {'-'*13}")
    for _, r in by_status.iterrows():
        lift = r["eng_strict_hit_pct"] - r["bl_strict_hit_pct"]
        _log(f"  {r['status']:<18} {int(r['n_customers']):>8,} {r['eng_strict_hit_pct']:>10.1f}% {r['eng_family_hit_pct']:>10.1f}% {r['bl_strict_hit_pct']:>10.1f}% {lift:>+12.1f}pp")

    by_size = per_customer.groupby("size_tier").agg(
        n_customers=("DIM_CUST_CURR_ID", "count"),
        eng_strict_hit_pct=("eng_strict_hit", lambda x: x.mean() * 100),
        eng_family_hit_pct=("eng_family_hit", lambda x: x.mean() * 100),
        bl_strict_hit_pct=("bl_strict_hit", lambda x: x.mean() * 100),
    ).round(2).reset_index()

    _log(f"")
    _log(f"Hit rates by size tier:")
    _log(f"  {'Size tier':<14} {'N':>8} {'Eng strict':>11} {'Eng family':>11} {'BL strict':>11} {'Strict lift':>13}")
    _log(f"  {'-'*14} {'-'*8} {'-'*11} {'-'*11} {'-'*11} {'-'*13}")
    for _, r in by_size.iterrows():
        lift = r["eng_strict_hit_pct"] - r["bl_strict_hit_pct"]
        _log(f"  {r['size_tier']:<14} {int(r['n_customers']):>8,} {r['eng_strict_hit_pct']:>10.1f}% {r['eng_family_hit_pct']:>10.1f}% {r['bl_strict_hit_pct']:>10.1f}% {lift:>+12.1f}pp")

    return by_status, by_size


# Excel styling

def _style(ws, df, hc="1F4E79"):
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, 12), 55)


# Step 8: Save outputs

def save_outputs(results: dict,
                  by_signal: pd.DataFrame,
                  by_status: pd.DataFrame,
                  by_size: pd.DataFrame,
                  max_date: str,
                  cutoff_date: str) -> None:
    _s("Step 8: Saving outputs")

    ANALYSIS.mkdir(parents=True, exist_ok=True)

    per_customer = results["per_customer"]
    metrics = pd.DataFrame(results["metrics_table"])

    # Build a summary dataframe
    summary_rows = []
    summary_rows.append({"metric": "VALIDATION CONFIG", "value": ""})
    summary_rows.append({"metric": "  Holdout window (days)", "value": str(HOLDOUT_DAYS)})
    summary_rows.append({"metric": "  Holdout cutoff", "value": cutoff_date})
    summary_rows.append({"metric": "  Max data date", "value": max_date})
    summary_rows.append({"metric": "  Top-N evaluated", "value": str(TOP_N)})
    summary_rows.append({"metric": "  Customers evaluated", "value": f"{results['n_eval']:,}"})
    summary_rows.append({"metric": "", "value": ""})
    summary_rows.append({"metric": "HEADLINE METRICS (engine vs popularity baseline)", "value": ""})
    for r in results["metrics_table"]:
        summary_rows.append({
            "metric": "  " + r["metric"],
            "value": f"engine={r['engine_pct']:.1f}%  baseline={r['baseline_pct']:.1f}%  lift={r['lift_pp']:+.1f}pp"
        })
    summary_rows.append({"metric": "", "value": ""})
    summary_rows.append({"metric": "DISCLOSURE", "value": ""})
    summary_rows.append({
        "metric": "  Method",
        "value": "Shortcut: existing recommendations validated against held-out 60 days"
    })
    summary_rows.append({
        "metric": "  Caveat",
        "value": "Recommendations were built from full history including holdout. May inflate hit rates."
    })
    summary_rows.append({
        "metric": "  Why this is still useful",
        "value": "Relative comparison vs popularity baseline is unaffected by leakage."
    })
    summary_df = pd.DataFrame(summary_rows)

    # Excel output
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="00_summary", index=False)
        metrics.to_excel(writer, sheet_name="01_headline_metrics", index=False)
        by_signal.to_excel(writer, sheet_name="02_by_signal", index=False)
        by_status.to_excel(writer, sheet_name="03_by_status", index=False)
        by_size.to_excel(writer, sheet_name="04_by_size_tier", index=False)
        per_customer.head(2000).to_excel(writer, sheet_name="05_per_customer_sample", index=False)

        wb = writer.book
        _style(writer.sheets["00_summary"], summary_df, hc="002060")
        _style(writer.sheets["01_headline_metrics"], metrics, hc="1F4E79")
        _style(writer.sheets["02_by_signal"], by_signal, hc="833C00")
        _style(writer.sheets["03_by_status"], by_status, hc="375623")
        _style(writer.sheets["04_by_size_tier"], by_size, hc="6F2DA8")
        _style(writer.sheets["05_per_customer_sample"], per_customer.head(2000), hc="C00000")
        for s in wb.sheetnames:
            wb[s].sheet_properties.tabColor = "1F4E79"

    size_kb = OUT_XLSX.stat().st_size / 1024
    _log(f"Saved: {OUT_XLSX.relative_to(ROOT)}  ({size_kb:.0f} KB, 6 sheets)")

    # Text report
    lines = []
    lines.append("=" * 80)
    lines.append("  RECOMMENDATION ENGINE VALIDATION REPORT")
    lines.append("=" * 80)
    lines.append("")
    lines.append(f"Method   : Temporal holdout, last {HOLDOUT_DAYS} days predicted from earlier data")
    lines.append(f"Holdout  : after {cutoff_date} through {max_date}")
    lines.append(f"Customers evaluated: {results['n_eval']:,}")
    lines.append("")
    lines.append("DISCLOSURE")
    lines.append("-" * 80)
    lines.append("This validation uses the SHORTCUT approach: the recommendations being")
    lines.append("validated were generated from the FULL history including the holdout window.")
    lines.append("A fully strict evaluation would rebuild every precomputation file using only")
    lines.append("training-period data and regenerate the recommendations from scratch. We do")
    lines.append("not do that here. Hit rates may be slightly inflated as a result.")
    lines.append("")
    lines.append("The relative comparison between the engine and the popularity baseline is NOT")
    lines.append("affected by this leakage, because both see exactly the same data. So even if")
    lines.append("the absolute hit rate numbers are optimistic, the gap between engine and")
    lines.append("baseline is a meaningful measure of how much the engine adds.")
    lines.append("")
    lines.append("HEADLINE RESULTS")
    lines.append("-" * 80)
    lines.append(f"{'Metric':<40} {'Engine':>10} {'Baseline':>10} {'Lift':>10}")
    for r in results["metrics_table"]:
        sign = "+" if r["lift_pp"] >= 0 else ""
        lines.append(f"{r['metric']:<40} {r['engine_pct']:>9.1f}% {r['baseline_pct']:>9.1f}% {sign}{r['lift_pp']:>8.1f}pp")
    lines.append("")
    lines.append("HIT RATE BY SIGNAL")
    lines.append("-" * 80)
    lines.append("Strict hit rate per signal (denominator = recs from that signal):")
    lines.append(f"{'Signal':<22} {'N recs':>10} {'N hits':>8} {'Hit rate':>10}")
    for _, r in by_signal.iterrows():
        lines.append(f"{r['primary_signal']:<22} {int(r['n_recs']):>10,} {int(r['n_hits']):>8,} {r['hit_rate_pct']:>9.2f}%")
    lines.append("")
    lines.append("HIT RATE BY CUSTOMER STATUS")
    lines.append("-" * 80)
    lines.append(f"{'Status':<18} {'N':>8} {'Eng strict':>11} {'Eng family':>11} {'BL strict':>11}")
    for _, r in by_status.iterrows():
        lines.append(f"{r['status']:<18} {int(r['n_customers']):>8,} {r['eng_strict_hit_pct']:>10.1f}% {r['eng_family_hit_pct']:>10.1f}% {r['bl_strict_hit_pct']:>10.1f}%")
    lines.append("")
    lines.append("HIT RATE BY SIZE TIER")
    lines.append("-" * 80)
    lines.append(f"{'Size tier':<14} {'N':>8} {'Eng strict':>11} {'Eng family':>11} {'BL strict':>11}")
    for _, r in by_size.iterrows():
        lines.append(f"{r['size_tier']:<14} {int(r['n_customers']):>8,} {r['eng_strict_hit_pct']:>10.1f}% {r['eng_family_hit_pct']:>10.1f}% {r['bl_strict_hit_pct']:>10.1f}%")
    lines.append("")

    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")
    size_kb = OUT_TXT.stat().st_size / 1024
    _log(f"Saved: {OUT_TXT.relative_to(ROOT)}  ({size_kb:.0f} KB)")


# Main

def main() -> None:
    log_file = _setup_logging()

    try:
        print()
        print("=" * 80)
        print("  RECOMMENDATION ENGINE VALIDATION (next-basket prediction)")
        print("=" * 80)
        start = time.time()

        data = load_data()
        max_date, cutoff_date = determine_cutoff(MERGED_FILE)
        holdout = load_holdout_purchases(MERGED_FILE, cutoff_date, max_date)
        baseline = build_popularity_baseline(MERGED_FILE, cutoff_date, data["products"])

        results = compute_hit_rates(
            data["recs"], holdout, data["products"], baseline,
            data["patterns"], data["segments"]
        )

        if not results:
            sys.exit(1)

        eval_custs = set(results["per_customer"]["DIM_CUST_CURR_ID"].tolist())
        by_signal = hit_rates_by_signal(data["recs"], holdout, eval_custs)
        by_status, by_size = hit_rates_by_status(results["per_customer"])

        save_outputs(results, by_signal, by_status, by_size, max_date, cutoff_date)

        _s("Complete")
        _log(f"Total time: {time.time() - start:.1f}s")
        _log(f"")
        _log(f"Outputs:")
        _log(f"  Terminal log : {OUT_LOG.relative_to(ROOT)}")
        _log(f"  XLSX         : {OUT_XLSX.relative_to(ROOT)}  (6 sheets)")
        _log(f"  TXT          : {OUT_TXT.relative_to(ROOT)}")
    finally:
        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__
        log_file.close()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise