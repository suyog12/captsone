from __future__ import annotations

import sys
import time
import warnings
from collections import defaultdict
from itertools import combinations
from pathlib import Path
from typing import Optional

import duckdb
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore")


# Paths

ROOT        = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN  = ROOT / "data_clean"
MERGED_FILE = DATA_CLEAN / "serving"  / "merged_dataset.parquet"
SEG_FILE    = DATA_CLEAN / "serving"  / "precomputed" / "customer_segments.parquet"
FEAT_FILE   = DATA_CLEAN / "features" / "customer_features.parquet"
OUT_PRECOMP = DATA_CLEAN / "serving"  / "precomputed"
OUT_ANALYSIS = DATA_CLEAN / "analysis"


# Configuration

# Fiscal years to analyse. Both years read for robust pattern statistics.
# Must match FISCAL_YEARS in recommendation_factors.py so the sequence data
# they produce matches the transaction data the recommender sees.
FISCAL_YEARS = ("FY2425", "FY2526")

# Minimum support: a category pair must co-occur in this fraction of segment
# customers to be recorded as an association rule. Set low because healthcare
# purchasing is inherently niche — a 5% co-occurrence in wound care is strong.
MIN_SUPPORT       = 0.05

# Minimum confidence: given that a customer bought category A, this is the
# floor probability that they also bought category B within the window.
MIN_CONFIDENCE    = 0.20

# Co-occurrence window in days. Two categories are "co-purchased" if orders
# from each appear within this window. 90 days captures a quarterly cycle.
CO_OCCUR_WINDOW   = 90

# Minimum customers in a segment to run pattern analysis. Segments smaller
# than this do not have enough data for reliable association rules.
MIN_SEGMENT_SIZE  = 50

# Top-N association rules to keep per segment.
TOP_RULES_PER_SEG = 30

# Top-N category sequences to keep per segment.
TOP_SEQS_PER_SEG  = 20

# Excluded suppliers — products from these suppliers are never surfaced in
# pattern-driven recommendations. Matches EXCLUDED_SUPPLIERS in
# recommendation_factors.py.
EXCLUDED_SUPPLIERS = {"MEDLINE", "MEDLINE INDUSTRIES"}

# Excluded product families — administrative or non-pitchable categories.
# Matches EXCLUDED_FAMILIES in recommendation_factors.py.
EXCLUDED_FAMILIES = {"Fee", "Unknown", "NaN", "nan", ""}


# Logging

def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)

def _section(title: str) -> None:
    print(f"\n{'-' * 64}\n  {title}\n{'-' * 64}", flush=True)


# DuckDB helpers

def _pq(p: Path) -> str:
    return "'" + p.as_posix() + "'"


# Excel helpers

def _style(ws, df: pd.DataFrame, hc: str = "1F4E79") -> None:
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(
            col_cells[0].column)].width = min(max(w + 2, 12), 55)


def _chart(ws, n, lc, vc, anchor, title, xtitle, color="1F4E79") -> None:
    ch = BarChart()
    ch.type = "bar"
    ch.grouping = "clustered"
    ch.title = title
    ch.x_axis.title = xtitle
    ch.legend = None
    ch.width = 26
    ch.height = 16
    ch.style = 2
    ch.add_data(Reference(ws, min_col=vc, min_row=1, max_row=n + 1),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=lc, min_row=2, max_row=n + 1))
    ch.series[0].graphicalProperties.solidFill = color
    ch.series[0].graphicalProperties.line.solidFill = color
    ws.add_chart(ch, anchor)


# Step 1: Load inputs

def load_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    _section("Step 1: Loading inputs")

    for path, label in [
        (MERGED_FILE, "merged_dataset.parquet"),
        (SEG_FILE,    "customer_segments.parquet"),
        (FEAT_FILE,   "customer_features.parquet"),
    ]:
        if not path.exists():
            print(f"\nFATAL: {label} not found at {path}", file=sys.stderr)
            print("Run clean_data.py then segment_customers.py first.", file=sys.stderr)
            sys.exit(1)

    segments = pd.read_parquet(SEG_FILE, columns=["DIM_CUST_CURR_ID", "segment"])
    _log(f"Segments loaded  : {len(segments):,} customers")

    seg_counts = segments["segment"].value_counts()
    for seg, cnt in seg_counts.items():
        _log(f"  {seg:<30} {cnt:>8,} customers")

    features = pd.read_parquet(FEAT_FILE, columns=[
        "DIM_CUST_CURR_ID", "SPCLTY_CD", "R_score", "F_score",
        "monetary", "recency_days",
    ])
    _log(f"Features loaded  : {len(features):,} customers")

    # Load transactions spanning both fiscal years. Four columns needed:
    #   cust_id, product_family, first_order_date, order_count.
    # DuckDB processes the 7.4 GB file in-process so pandas never holds it.
    _log("Loading transactions via DuckDB (both FY2425 and FY2526, takes ~90s)...")
    t0  = time.time()
    con = duckdb.connect()

    # Detect which columns are available for product family and supplier
    desc = con.execute(
        f"DESCRIBE SELECT * FROM read_parquet({_pq(MERGED_FILE)}) LIMIT 0"
    ).df()
    available = set(desc["column_name"].tolist())

    family_col = next(
        (c for c in ["PROD_FMLY_LVL1_DSC", "PROD_FMLY_LVL1_CD", "PROD_CTGRY_LVL2_DSC"]
         if c in available), None
    )
    supplier_col = next(
        (c for c in ["SUPLR_ROLLUP_DSC", "SUPLR_DSC"] if c in available), None
    )

    if family_col is None:
        print("\nFATAL: No product family column found in merged_dataset.", file=sys.stderr)
        print(f"Available columns: {sorted(available)}", file=sys.stderr)
        sys.exit(1)

    _log(f"Product family column : {family_col}")
    _log(f"Supplier column       : {supplier_col or 'absent — Medline filter skipped'}")

    # Fiscal year filter — both years
    fy_sql = ", ".join(f"'{fy}'" for fy in FISCAL_YEARS)
    _log(f"Fiscal years          : {list(FISCAL_YEARS)}")

    # Family exclusion — Fee, Unknown, etc.
    excl_families_sql = ", ".join(f"'{f}'" for f in sorted(EXCLUDED_FAMILIES))
    _log(f"Excluded families     : {sorted(EXCLUDED_FAMILIES)}")

    # Supplier exclusion — Medline
    supplier_filter = ""
    if supplier_col:
        excl_list = ", ".join(f"'{s}'" for s in EXCLUDED_SUPPLIERS)
        supplier_filter = (
            f"AND UPPER(COALESCE({supplier_col}, '')) NOT IN ({excl_list})"
        )

    # Aggregate in DuckDB before pulling to pandas to avoid OOM.
    # DIM_ORDR_DT_ID is a YYYYMMDD integer. We compute a proper date_days value
    # (days since 1970-01-01) via MAKE_DATE so the 90-day co-occurrence window
    # uses accurate arithmetic, not the old y*365+m*30 approximation.
    txn = con.execute(f"""
        SELECT
            CAST(DIM_CUST_CURR_ID AS BIGINT)              AS cust_id,
            COALESCE({family_col}, 'Unknown')             AS product_family,
            MIN(DIM_ORDR_DT_ID)                           AS order_date_id,
            DATE_DIFF(
                'day',
                MAKE_DATE(1970, 1, 1),
                MAKE_DATE(
                    CAST(MIN(DIM_ORDR_DT_ID) / 10000 AS INTEGER),
                    CAST((MIN(DIM_ORDR_DT_ID) % 10000) / 100 AS INTEGER),
                    CAST(MIN(DIM_ORDR_DT_ID) % 100 AS INTEGER)
                )
            )                                             AS date_days,
            COUNT(DISTINCT ORDR_NUM)                      AS order_count
        FROM read_parquet({_pq(MERGED_FILE)})
        WHERE UNIT_SLS_AMT > 0
          AND fiscal_year IN ({fy_sql})
          AND DIM_CUST_CURR_ID IS NOT NULL
          AND {family_col} IS NOT NULL
          AND COALESCE({family_col}, 'Unknown') NOT IN ({excl_families_sql})
          {supplier_filter}
        GROUP BY
            CAST(DIM_CUST_CURR_ID AS BIGINT),
            COALESCE({family_col}, 'Unknown')
    """).df()
    con.close()

    _log(f"Customer-family pairs loaded: {len(txn):,} rows in {time.time()-t0:.1f}s")
    _log(f"Unique customers       : {txn['cust_id'].nunique():,}")
    _log(f"Unique product families: {txn['product_family'].nunique():,}")

    # Attach segment labels to transactions
    txn = txn.merge(
        segments.rename(columns={"DIM_CUST_CURR_ID": "cust_id"}),
        on="cust_id", how="left"
    )
    unmatched = txn["segment"].isna().sum()
    if unmatched > 0:
        _log(f"Warning: {unmatched:,} transaction rows have no segment match — dropped")
        txn = txn.dropna(subset=["segment"])

    return txn, segments, features


# Step 2: Category co-occurrence

def compute_category_cooccurrence(
    txn: pd.DataFrame,
    segments: pd.DataFrame,
) -> pd.DataFrame:
    # For each segment, find category pairs that co-occur within CO_OCCUR_WINDOW
    # days in the same customer. Compute support (co-occurrence frequency),
    # confidence (P(B|A)), and lift (how much more common than random).

    _section("Step 2: Category co-occurrence within segments")

    all_rules = []

    for seg in txn["segment"].unique():
        seg_txn = txn[txn["segment"] == seg]
        n_custs = seg_txn["cust_id"].nunique()

        if n_custs < MIN_SEGMENT_SIZE:
            _log(f"  {seg:<30} skipped ({n_custs} customers < {MIN_SEGMENT_SIZE} minimum)")
            continue

        _log(f"  {seg:<30} {n_custs:>8,} customers")

        # Per customer: dict of category -> earliest date (in days since epoch)
        cust_cat_day: dict[int, dict[str, int]] = defaultdict(dict)
        for _, row in seg_txn.iterrows():
            cid = row["cust_id"]
            fam = row["product_family"]
            day = int(row["date_days"])
            if fam not in cust_cat_day[cid] or day < cust_cat_day[cid][fam]:
                cust_cat_day[cid][fam] = day

        # Count how many customers bought each category (support denominator)
        cat_count: dict[str, int] = defaultdict(int)
        for cid, cat_days in cust_cat_day.items():
            for cat in cat_days:
                cat_count[cat] += 1

        # Count co-occurrences within the window
        pair_count: dict[tuple[str, str], int] = defaultdict(int)
        for cid, cat_days in cust_cat_day.items():
            cats = sorted(cat_days.keys())
            for a, b in combinations(cats, 2):
                day_a = cat_days[a]
                day_b = cat_days[b]
                if abs(day_a - day_b) <= CO_OCCUR_WINDOW:
                    key = (a, b) if a < b else (b, a)
                    pair_count[key] += 1

        for (cat_a, cat_b), co_count in pair_count.items():
            support    = co_count / n_custs
            conf_a_b   = co_count / max(cat_count[cat_a], 1)
            conf_b_a   = co_count / max(cat_count[cat_b], 1)
            lift       = support / max(
                (cat_count[cat_a] / n_custs) * (cat_count[cat_b] / n_custs),
                1e-9
            )

            if support < MIN_SUPPORT:
                continue
            if max(conf_a_b, conf_b_a) < MIN_CONFIDENCE:
                continue

            all_rules.append({
                "segment":        seg,
                "n_segment_custs": n_custs,
                "category_a":     cat_a,
                "category_b":     cat_b,
                "co_occurrence":  co_count,
                "support":        round(support, 4),
                "confidence_a_b": round(conf_a_b, 4),
                "confidence_b_a": round(conf_b_a, 4),
                "lift":           round(lift, 4),
                "seller_action":  _cooccur_action(cat_a, cat_b, conf_a_b, conf_b_a),
            })

    if not all_rules:
        _log("No co-occurrence rules met the minimum thresholds.")
        return pd.DataFrame()

    rules_df = pd.DataFrame(all_rules)

    # Keep top rules per segment by lift (lift > 1 = not random co-occurrence)
    rules_df = (
        rules_df[rules_df["lift"] > 1.0]
        .sort_values(["segment", "lift"], ascending=[True, False])
        .groupby("segment")
        .head(TOP_RULES_PER_SEG)
        .reset_index(drop=True)
    )

    _log(f"\n  Total association rules (support>={MIN_SUPPORT:.0%}, "
         f"confidence>={MIN_CONFIDENCE:.0%}, lift>1): {len(rules_df):,}")

    for seg in rules_df["segment"].unique():
        seg_rules = rules_df[rules_df["segment"] == seg]
        _log(f"\n  {seg} — top 5 rules by lift:")
        for _, r in seg_rules.head(5).iterrows():
            _log(f"    {r['category_a'][:25]:<26} + {r['category_b'][:25]:<26} "
                 f"lift={r['lift']:.2f}  conf={r['confidence_a_b']:.0%}")

    return rules_df


def _cooccur_action(cat_a: str, cat_b: str,
                    conf_ab: float, conf_ba: float) -> str:
    # Describe the co-occurrence rule in seller language.
    if conf_ab >= conf_ba:
        primary_from, primary_to, conf = cat_a, cat_b, conf_ab
    else:
        primary_from, primary_to, conf = cat_b, cat_a, conf_ba

    return (
        f"{conf:.0%} of customers who buy {primary_from[:30]} "
        f"also buy {primary_to[:30]} within {CO_OCCUR_WINDOW} days. "
        f"Cross-sell pitch: suggest {primary_to[:30]} at same visit."
    )


# Step 3: Category sequences

def compute_category_sequences(txn: pd.DataFrame) -> pd.DataFrame:
    # For each segment, find first-order category transitions.
    # "After category A, customer typically buys B next."
    # Used by recommendation_factors.py for pitch message enrichment.

    _section("Step 3: Category purchase sequences within segments")

    txn = txn.copy().sort_values(["cust_id", "order_date_id"])

    all_sequences = []

    for seg in txn["segment"].unique():
        seg_txn = txn[txn["segment"] == seg]
        n_custs = seg_txn["cust_id"].nunique()

        if n_custs < MIN_SEGMENT_SIZE:
            continue

        _log(f"  {seg:<30} {n_custs:>8,} customers")

        # Per customer: ordered list of categories by first appearance date
        cust_first_cat: dict[int, dict[str, int]] = defaultdict(dict)
        for _, row in seg_txn.iterrows():
            cid  = row["cust_id"]
            fam  = row["product_family"]
            date = int(row["order_date_id"])
            if fam not in cust_first_cat[cid]:
                cust_first_cat[cid][fam] = date
            else:
                cust_first_cat[cid][fam] = min(cust_first_cat[cid][fam], date)

        # Build category sequence per customer sorted by first purchase date
        transition_count: dict[tuple[str, str], int] = defaultdict(int)
        from_count: dict[str, int] = defaultdict(int)

        for cid, cat_dates in cust_first_cat.items():
            ordered = sorted(cat_dates.items(), key=lambda x: x[1])
            for i in range(len(ordered) - 1):
                from_cat = ordered[i][0]
                to_cat   = ordered[i + 1][0]
                if from_cat != to_cat:
                    transition_count[(from_cat, to_cat)] += 1
                    from_count[from_cat] += 1

        for (from_cat, to_cat), count in transition_count.items():
            prob = count / max(from_count[from_cat], 1)
            support = count / n_custs

            if support < MIN_SUPPORT or prob < MIN_CONFIDENCE:
                continue

            all_sequences.append({
                "segment":          seg,
                "n_segment_custs":  n_custs,
                "from_category":    from_cat,
                "to_category":      to_cat,
                "transition_count": count,
                "transition_prob":  round(prob, 4),
                "support":          round(support, 4),
                "seller_action": (
                    f"{prob:.0%} of {seg.replace('_',' ')} customers who start buying "
                    f"{from_cat[:30]} next buy {to_cat[:30]}. "
                    f"Pitch {to_cat[:30]} at the same visit as or shortly after {from_cat[:30]}."
                ),
            })

    if not all_sequences:
        _log("No category sequences met the minimum thresholds.")
        return pd.DataFrame()

    seq_df = (
        pd.DataFrame(all_sequences)
        .sort_values(["segment", "transition_prob"], ascending=[True, False])
        .groupby("segment")
        .head(TOP_SEQS_PER_SEG)
        .reset_index(drop=True)
    )

    _log(f"\n  Total sequences (support>={MIN_SUPPORT:.0%}, "
         f"confidence>={MIN_CONFIDENCE:.0%}): {len(seq_df):,}")

    for seg in seq_df["segment"].unique():
        s = seq_df[seq_df["segment"] == seg]
        _log(f"\n  {seg} — top 5 next-category sequences:")
        for _, r in s.head(5).iterrows():
            _log(f"    {r['from_category'][:28]:<30} -> {r['to_category'][:28]:<30} "
                 f"prob={r['transition_prob']:.0%}")

    return seq_df


# Step 4: Segment category profiles

def compute_segment_category_profiles(txn: pd.DataFrame) -> pd.DataFrame:
    # Category adoption rate per segment — shows which families are core
    # and which represent expansion opportunity. Read-only aggregation.

    _section("Step 4: Category revenue profiles per segment")

    profile = (
        txn.groupby(["segment", "product_family"])
        .agg(
            n_customers = ("cust_id", "nunique"),
            order_count = ("order_date_id", "nunique"),
        )
        .reset_index()
    )

    seg_sizes = txn.groupby("segment")["cust_id"].nunique().rename("seg_total_custs")
    profile = profile.merge(seg_sizes, on="segment", how="left")
    profile["adoption_rate"] = (
        profile["n_customers"] / profile["seg_total_custs"]
    ).round(4)

    profile = profile.sort_values(
        ["segment", "adoption_rate"], ascending=[True, False]
    ).reset_index(drop=True)

    _log(f"Category-segment profiles: {len(profile):,} rows")
    _log(f"Segments covered: {profile['segment'].nunique()}")

    for seg in profile["segment"].unique():
        top = profile[profile["segment"] == seg].head(5)
        _log(f"\n  {seg} — top 5 categories by adoption:")
        for _, r in top.iterrows():
            _log(f"    {r['product_family'][:40]:<42} {r['adoption_rate']:.0%} "
                 f"of customers ({r['n_customers']:,})")

    return profile


# Step 5: Save outputs

def _normalise_rules_schema(df: pd.DataFrame) -> pd.DataFrame:
    # Enforce consistent types for the backend
    df = df.copy()
    for c in ["n_segment_custs", "co_occurrence"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype("int64")
    for c in ["support", "confidence_a_b", "confidence_b_a", "lift"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype("float64")
    for c in ["segment", "category_a", "category_b", "seller_action"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)
    return df


def _normalise_seq_schema(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in ["n_segment_custs", "transition_count"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype("int64")
    for c in ["transition_prob", "support"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype("float64")
    for c in ["segment", "from_category", "to_category", "seller_action"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)
    return df


def _normalise_profile_schema(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in ["n_customers", "order_count", "seg_total_custs"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype("int64")
    if "adoption_rate" in df.columns:
        df["adoption_rate"] = pd.to_numeric(
            df["adoption_rate"], errors="coerce"
        ).fillna(0.0).astype("float64")
    for c in ["segment", "product_family"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)
    return df


def save_outputs(
    rules_df:  pd.DataFrame,
    seq_df:    pd.DataFrame,
    profile_df: pd.DataFrame,
) -> None:
    _section("Step 5: Saving outputs")

    OUT_PRECOMP.mkdir(parents=True, exist_ok=True)
    OUT_ANALYSIS.mkdir(parents=True, exist_ok=True)

    # API-serving parquets with enforced types
    if len(rules_df) > 0:
        rules_df = _normalise_rules_schema(rules_df)
        rules_df = rules_df.sort_values(["segment", "lift"], ascending=[True, False])
        out = OUT_PRECOMP / "segment_patterns.parquet"
        rules_df.to_parquet(out, index=False)
        _log(f"Saved: {out.relative_to(ROOT)}")
    else:
        _log("Warning: no association rules to save")

    if len(seq_df) > 0:
        seq_df = _normalise_seq_schema(seq_df)
        seq_df = seq_df.sort_values(
            ["segment", "transition_prob"], ascending=[True, False]
        )
        out = OUT_PRECOMP / "segment_sequences.parquet"
        seq_df.to_parquet(out, index=False)
        _log(f"Saved: {out.relative_to(ROOT)}")
    else:
        _log("Warning: no category sequences to save")

    if len(profile_df) > 0:
        profile_df = _normalise_profile_schema(profile_df)
        profile_df = profile_df.sort_values(
            ["segment", "adoption_rate"], ascending=[True, False]
        )
        out = OUT_PRECOMP / "segment_category_profiles.parquet"
        profile_df.to_parquet(out, index=False)
        _log(f"Saved: {out.relative_to(ROOT)}")

    # Excel report
    xlsx_path = OUT_ANALYSIS / "segment_patterns.xlsx"
    sheets = {}
    if len(rules_df) > 0:
        sheets["01_association_rules"] = ("1F4E79", rules_df)
    if len(seq_df) > 0:
        sheets["02_category_sequences"] = ("375623", seq_df)
    if len(profile_df) > 0:
        sheets["03_segment_profiles"] = ("7030A0", profile_df)

    if not sheets:
        _log("No data to write to Excel report.")
        return

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet_name, (color, df) in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            wb = writer.book
            wb[sheet_name].sheet_properties.tabColor = color
            _style(writer.sheets[sheet_name], df, hc=color)

        wb = writer.book

        # Chart: top rules by lift per segment
        if "01_association_rules" in wb.sheetnames and len(rules_df) > 0:
            chart_data = (
                rules_df.groupby("segment")["lift"]
                .mean()
                .reset_index()
                .rename(columns={"lift": "avg_lift"})
                .sort_values("avg_lift", ascending=False)
            )
            chart_data.to_excel(writer, sheet_name="_chart_rules", index=False)
            wb["_chart_rules"].sheet_properties.tabColor = "CCCCCC"
            _style(writer.sheets["_chart_rules"], chart_data, hc="CCCCCC")
            _chart(
                wb["_chart_rules"],
                n=len(chart_data),
                lc=1, vc=2,
                anchor="D2",
                title="Average association rule lift per segment",
                xtitle="Average lift (>1 = non-random co-occurrence)",
                color="1F4E79",
            )

        # Chart: top sequences by probability per segment
        if "02_category_sequences" in wb.sheetnames and len(seq_df) > 0:
            seq_chart = (
                seq_df.groupby("segment")["transition_prob"]
                .mean()
                .reset_index()
                .rename(columns={"transition_prob": "avg_transition_prob"})
                .sort_values("avg_transition_prob", ascending=False)
            )
            seq_chart.to_excel(writer, sheet_name="_chart_seqs", index=False)
            wb["_chart_seqs"].sheet_properties.tabColor = "CCCCCC"
            _style(writer.sheets["_chart_seqs"], seq_chart, hc="CCCCCC")
            _chart(
                wb["_chart_seqs"],
                n=len(seq_chart),
                lc=1, vc=2,
                anchor="D2",
                title="Average category transition probability per segment",
                xtitle="Average next-category probability",
                color="375623",
            )

    _log(f"Saved: {xlsx_path.relative_to(ROOT)}  (2 charts)")


# Main

def main() -> None:
    print()
    print("=" * 64)
    print("  SEGMENT BUYING PATTERN ANALYSIS")
    print("=" * 64)
    start = time.time()

    OUT_PRECOMP.mkdir(parents=True, exist_ok=True)
    OUT_ANALYSIS.mkdir(parents=True, exist_ok=True)

    txn, segments, features = load_inputs()

    rules_df   = compute_category_cooccurrence(txn, segments)
    seq_df     = compute_category_sequences(txn)
    profile_df = compute_segment_category_profiles(txn)

    save_outputs(rules_df, seq_df, profile_df)

    _section("Complete")
    _log(f"Total time: {time.time() - start:.1f}s")
    _log("")
    _log("Outputs written to data_clean/serving/precomputed/:")
    _log("  segment_patterns.parquet          association rules per segment")
    _log("  segment_sequences.parquet         next-category sequences per segment")
    _log("  segment_category_profiles.parquet category adoption rates per segment")
    _log("")
    _log("segment_sequences.parquet is read by recommendation_factors.py to")
    _log("enrich pitch messages with 'customers in this segment next buy X'.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise