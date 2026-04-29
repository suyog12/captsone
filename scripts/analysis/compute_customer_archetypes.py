from __future__ import annotations

import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Path configuration

ROOT = Path(__file__).resolve().parents[2]
DATA_CLEAN = ROOT / "data_clean"

FEATURES_FILE   = DATA_CLEAN / "features"  / "customer_features.parquet"
CUSTOMERS_FILE  = DATA_CLEAN / "customer"  / "customers_clean.parquet"
OUTPUT_REPORT   = DATA_CLEAN / "analysis"  / "customer_archetypes_report.xlsx"


# Logging

def _section(title):
    print(f"\n{'-' * 70}")
    print(f"  {title}")
    print(f"{'-' * 70}", flush=True)


def _log(msg):
    print(f"  {msg}", flush=True)


# Archetype taxonomy and classification rules

# Each archetype has:
#   - code: short identifier used in archetype column
#   - label: human-readable name for display
#   - description: defense narrative
#   - keywords: substring matches in SPCLTY_DSC (case-insensitive)

ARCHETYPES = [
    # Order matters - first match wins, so most specific first
    {
        "code": "government",
        "label": "Government / Military",
        "description": "Federal/state/military procurement; bulk contract pricing",
        "keywords": ["GOVT", "GOVERNMENT", "HOMELAND", "MILITARY",
                     "VA HOSPITAL", " VA ", "DEPT OF DEFENSE", "DOD",
                     "FEDERAL", "STATE OF ", "COUNTY OF ", "SHERIFF"],
    },
    {
        "code": "marketplace_reseller",
        "label": "Marketplace / Reseller",
        "description": "E-commerce / B2B resellers; high volume, broad catalog",
        "keywords": ["AMAZON", "MARKETPLACE", "RESELLER", "WHOLESALE",
                     "DISTRIBUT"],
    },
    {
        "code": "home_infusion",
        "label": "Home Infusion Pharmacy",
        "description": "503A pharmacies serving home infusion patients; specialty Rx, IV equipment",
        "keywords": ["HOME INFUSION", "INFUSION PHARMACY", "INFUSION CARE"],
    },
    {
        "code": "home_care_provider",
        "label": "Home Care / Hospice",
        "description": "DME providers, home health agencies, hospice; mobile patient care",
        "keywords": ["HOME MEDICAL", "HOME HEALTH", "HOME HOSPICE",
                     "HOSPICE", "HOME CARE"],
    },
    {
        "code": "hospital_acute",
        "label": "Hospital / Acute Care",
        "description": "Inpatient and emergency facilities; surgical and emergency supplies",
        "keywords": ["HOSPITAL", "MEDICAL CENTER", "ACUTE CARE",
                     "EMERGENCY MEDICINE", "EMERGENCY DEPT"],
    },
    {
        "code": "skilled_nursing",
        "label": "Skilled Nursing / Long-Term Care",
        "description": "SNFs, assisted living, intermediate care; high-volume daily care supplies",
        "keywords": ["SKILLED", "NURSING HOME", "ASSISTED LIVING",
                     "LONG TERM", "NURSING FACILITY", "INTERMEDIATE CARE",
                     "ALF "],
    },
    {
        "code": "surgery_center",
        "label": "Surgery Center",
        "description": "Ambulatory surgery centers; procedural supplies and OR consumables",
        "keywords": ["SURGERY CENTER", "AMBULATORY SURG", "SURGICAL CENTER",
                     "GENERAL SURGERY", "PLASTIC SURGERY"],
    },
    {
        "code": "lab_pathology",
        "label": "Lab / Pathology",
        "description": "Reference labs, hospital labs, pathology; specialty consumables",
        "keywords": ["REFERENCE LAB", "HOSPITAL LAB", "PATHOLOGY",
                     "DIAGNOSTIC LAB", "CLINICAL LAB"],
    },
    {
        "code": "veterinary",
        "label": "Veterinary",
        "description": "Animal health; different catalog needs from human medicine",
        "keywords": ["VETERIN", " VET ", "ANIMAL HOSP"],
    },
    {
        "code": "educational",
        "label": "Educational / Research",
        "description": "Universities, schools, research institutions; sporadic purchasing",
        "keywords": ["EDUCATIONAL", "UNIVERSITY", "COLLEGE", "SCHOOL",
                     "RESEARCH"],
    },
    {
        "code": "pharmacy",
        "label": "Pharmacy",
        "description": "Retail and specialty pharmacies; Rx-focused catalog",
        "keywords": ["PHARMACY", "PHARMACEUTICAL", "RETAIL PHARM"],
    },
    {
        "code": "multispecialty_group",
        "label": "Multispecialty Group Practice",
        "description": "Large group practices spanning specialties; broad catalog needs",
        "keywords": ["MULTIPLE SPECIALTY", "MULTI-SPECIAL", "GROUP PRACTICE",
                     "GROUP PRACT"],
    },
    {
        "code": "community_health",
        "label": "Community Health Center",
        "description": "FQHCs, safety-net clinics; primary care + chronic disease management",
        "keywords": ["COMMUNITY HEALTH", "FQHC", "FEDERALLY QUAL"],
    },
    {
        "code": "pediatric",
        "label": "Pediatric Practice",
        "description": "Children's specialty clinics; vaccines, pediatric Rx",
        "keywords": ["PEDIATRIC", "CHILDREN", "NEONATAL"],
    },
    {
        "code": "primary_care",
        "label": "Primary Care",
        "description": "Family practice, internal medicine, general practitioners",
        "keywords": ["FAMILY PRACTICE", "FAMILY MEDICINE", "INTERNAL MED",
                     "GENERAL PRAC", "FAMILY MED"],
    },
    {
        "code": "specialty_clinic",
        "label": "Specialty Clinic",
        "description": "Single-specialty clinics; deep but narrow catalog needs",
        "keywords": [],   # catch-all for any specialty not matched above
    },
    {
        "code": "unknown",
        "label": "Unknown / Unclassified",
        "description": "Specialty not provided in data",
        "keywords": [],   # for null/blank specialty values
    },
]


# Classification function

def classify_specialty(spec):
    if pd.isna(spec) or spec == "" or spec is None:
        return "unknown"

    s = str(spec).upper().strip()

    # Walk archetypes in order; first match wins
    for arch in ARCHETYPES:
        if arch["code"] == "specialty_clinic":
            continue   # catch-all, applied last
        if arch["code"] == "unknown":
            continue   # null handled above
        for kw in arch["keywords"]:
            if kw in s:
                return arch["code"]

    # No match - it's a specialty clinic of some kind
    return "specialty_clinic"


def classify_dataframe(df):
    return df["SPCLTY_DSC"].apply(classify_specialty)


# Excel styling

def _style(ws, df, header_color="1F4E79"):
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=header_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, 10), 50)


# Step 1: Load features and customer dimension

def step1_load_data():
    _section("Step 1: Loading data")

    if not FEATURES_FILE.exists():
        print(f"FATAL: {FEATURES_FILE} not found", file=sys.stderr)
        print("Run scripts/cleaning/clean_data.py first.", file=sys.stderr)
        sys.exit(1)

    if not CUSTOMERS_FILE.exists():
        print(f"FATAL: {CUSTOMERS_FILE} not found", file=sys.stderr)
        print("Run scripts/cleaning/clean_data.py first.", file=sys.stderr)
        sys.exit(1)

    feats = pd.read_parquet(FEATURES_FILE)
    _log(f"Loaded {len(feats):,} customers from features parquet ({feats.shape[1]} cols)")

    custs = pd.read_parquet(
        CUSTOMERS_FILE,
        columns=["DIM_CUST_CURR_ID", "SPCLTY_DSC", "CUST_NAME"]
    )
    _log(f"Loaded {len(custs):,} customers from dimension parquet")

    # Join specialty descriptions onto features
    df = feats.merge(custs, on="DIM_CUST_CURR_ID", how="left")
    _log(f"After join: {len(df):,} rows")

    # Sanity check - how many customers have specialty descriptions?
    n_with_spec = df["SPCLTY_DSC"].notna().sum()
    _log(f"Customers with specialty descriptions: {n_with_spec:,} ({100*n_with_spec/len(df):.1f}%)")

    return df


# Step 2: Classify customers into archetypes

def step2_classify(df):
    _section("Step 2: Classifying customers into archetypes")

    t0 = time.time()
    df["archetype"] = classify_dataframe(df)

    # Map to human-readable label
    label_map = {a["code"]: a["label"] for a in ARCHETYPES}
    df["archetype_label"] = df["archetype"].map(label_map)

    elapsed = time.time() - t0
    _log(f"Classification done in {elapsed:.1f}s")

    # Distribution
    counts = df["archetype"].value_counts()
    _log(f"")
    _log(f"  Archetype distribution:")
    for arch in ARCHETYPES:
        n = counts.get(arch["code"], 0)
        pct = 100 * n / len(df)
        _log(f"    {arch['code']:<25} {n:>9,}  ({pct:>5.2f}%)  - {arch['label']}")

    return df


# Step 3: Build archetype profile statistics

def step3_archetype_profiles(df):
    _section("Step 3: Building archetype profile statistics")

    profiles = df.groupby("archetype").agg(
        n_customers     = ("DIM_CUST_CURR_ID",     "count"),
        total_spend     = ("monetary",              "sum"),
        median_spend    = ("monetary",              "median"),
        mean_spend      = ("monetary",              "mean"),
        median_volume   = ("median_monthly_volume", "median"),
        median_freq     = ("frequency",             "median"),
        median_recency  = ("recency_days",          "median"),
        median_n_cats   = ("n_categories_bought",   "median"),
        churn_rate      = ("churn_label",           lambda s: (s == 1).mean() * 100),
    ).round(2).reset_index()

    label_map = {a["code"]: a["label"] for a in ARCHETYPES}
    desc_map  = {a["code"]: a["description"] for a in ARCHETYPES}

    profiles["archetype_label"] = profiles["archetype"].map(label_map)
    profiles["description"]     = profiles["archetype"].map(desc_map)

    profiles = profiles.sort_values("total_spend", ascending=False)

    profiles = profiles[[
        "archetype", "archetype_label", "description",
        "n_customers", "total_spend",
        "median_spend", "mean_spend",
        "median_volume", "median_freq", "median_recency",
        "median_n_cats", "churn_rate",
    ]]

    _log(f"Built profiles for {len(profiles)} archetypes")
    return profiles


# Step 4: Build archetype x size_tier crosstab

def step4_crosstabs(df):
    _section("Step 4: Building crosstabs")

    # archetype x size_tier
    ct_size = pd.crosstab(df["archetype"], df["size_tier"], margins=True, margins_name="TOTAL")
    ct_size = ct_size.reset_index()
    _log(f"  Built archetype x size_tier crosstab: {ct_size.shape}")

    # archetype x MKT_CD
    ct_mkt = pd.crosstab(df["archetype"], df["MKT_CD"], margins=True, margins_name="TOTAL")
    ct_mkt = ct_mkt.reset_index()
    _log(f"  Built archetype x MKT_CD crosstab: {ct_mkt.shape}")

    # archetype x supplier_profile (loyalty stance)
    if "supplier_profile" in df.columns:
        ct_sup = pd.crosstab(df["archetype"], df["supplier_profile"], margins=True, margins_name="TOTAL")
        ct_sup = ct_sup.reset_index()
        _log(f"  Built archetype x supplier_profile crosstab: {ct_sup.shape}")
    else:
        ct_sup = pd.DataFrame()

    return ct_size, ct_mkt, ct_sup


# Step 5: Build top specialties per archetype (validation aid)

def step5_top_specialties(df):
    _section("Step 5: Top specialties per archetype")

    rows = []
    for arch_code in df["archetype"].unique():
        sub = df[df["archetype"] == arch_code]
        top = sub["SPCLTY_DSC"].value_counts().head(10)
        for rank, (spec, count) in enumerate(top.items(), 1):
            rows.append({
                "archetype":   arch_code,
                "rank":        rank,
                "specialty":   spec if pd.notna(spec) else "(blank)",
                "n_customers": count,
            })

    df_top = pd.DataFrame(rows)
    _log(f"  Built top-10 specialties for each archetype: {len(df_top)} rows")
    return df_top


# Step 6: Save outputs

def step6_save(df, profiles, ct_size, ct_mkt, ct_sup, top_specs):
    _section("Step 6: Saving outputs")

    # Write back to features parquet with the new columns
    out_cols = list(df.columns)
    # Drop SPCLTY_DSC and CUST_NAME we joined in - they belong to the customer dim,
    # not the features parquet
    drop_cols = ["SPCLTY_DSC", "CUST_NAME"]
    out_df = df.drop(columns=[c for c in drop_cols if c in df.columns])

    out_df.to_parquet(FEATURES_FILE, index=False)
    _log(f"  Updated: {FEATURES_FILE.name}  ({FEATURES_FILE.stat().st_size / 1024 / 1024:.1f} MB)")
    _log(f"    Now has columns: archetype, archetype_label (plus existing 56)")

    # Write Excel report
    OUTPUT_REPORT.parent.mkdir(parents=True, exist_ok=True)

    # Summary sheet
    total_revenue = df["monetary"].sum()
    summary_rows = [
        ("Total customers",                          f"{len(df):,}"),
        ("Total revenue",                            f"${total_revenue:,.0f}"),
        ("Distinct archetypes",                       f"{df['archetype'].nunique()}"),
        ("Customers with specialty data",             f"{df['SPCLTY_DSC'].notna().sum():,}"),
        ("Customers classified as 'unknown'",        f"{(df['archetype'] == 'unknown').sum():,}"),
        ("Customers classified as 'specialty_clinic' (catch-all)",
                                                      f"{(df['archetype'] == 'specialty_clinic').sum():,}"),
        ("",                                          ""),
        ("Top archetype by revenue",                  ""),
    ]
    top_arch = profiles.iloc[0]
    summary_rows.append((f"  - {top_arch['archetype_label']}",
                         f"${top_arch['total_spend']:,.0f}"))
    summary_rows.append((f"  - n_customers",
                         f"{int(top_arch['n_customers']):,}"))
    summary_df = pd.DataFrame(summary_rows, columns=["metric", "value"])

    with pd.ExcelWriter(OUTPUT_REPORT, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="01_summary", index=False)
        _style(writer.sheets["01_summary"], summary_df, "1F4E79")

        profiles.to_excel(writer, sheet_name="02_archetype_profiles", index=False)
        _style(writer.sheets["02_archetype_profiles"], profiles, "375623")

        ct_size.to_excel(writer, sheet_name="03_archetype_x_size_tier", index=False)
        _style(writer.sheets["03_archetype_x_size_tier"], ct_size, "C00000")

        ct_mkt.to_excel(writer, sheet_name="04_archetype_x_market", index=False)
        _style(writer.sheets["04_archetype_x_market"], ct_mkt, "7030A0")

        if not ct_sup.empty:
            ct_sup.to_excel(writer, sheet_name="05_archetype_x_supplier", index=False)
            _style(writer.sheets["05_archetype_x_supplier"], ct_sup, "1F6B75")

        top_specs.to_excel(writer, sheet_name="06_top_specialties_per_arch", index=False)
        _style(writer.sheets["06_top_specialties_per_arch"], top_specs, "833C00")

    _log(f"  Saved report: {OUTPUT_REPORT.name}  ({OUTPUT_REPORT.stat().st_size / 1024:.0f} KB)")

    return OUTPUT_REPORT


# Main

def main():
    print()
    print("=" * 70)
    print("  CUSTOMER ARCHETYPE CLASSIFICATION")
    print("=" * 70)
    start = time.time()

    df              = step1_load_data()
    df              = step2_classify(df)
    profiles        = step3_archetype_profiles(df)
    ct_size, ct_mkt, ct_sup = step4_crosstabs(df)
    top_specs       = step5_top_specialties(df)
    out_path        = step6_save(df, profiles, ct_size, ct_mkt, ct_sup, top_specs)

    elapsed = round(time.time() - start, 1)
    _section(f"Complete in {elapsed}s")
    print()
    print(f"  Two new columns added to customer_features.parquet:")
    print(f"    - archetype       (e.g., 'multispecialty_group')")
    print(f"    - archetype_label (e.g., 'Multispecialty Group Practice')")
    print()
    print(f"  Report: {out_path.name}")
    print(f"    01_summary                   - high-level stats")
    print(f"    02_archetype_profiles        - profile per archetype (revenue, churn, etc.)")
    print(f"    03_archetype_x_size_tier     - crosstab of archetype vs size_tier")
    print(f"    04_archetype_x_market        - crosstab of archetype vs MKT_CD")
    print(f"    05_archetype_x_supplier      - crosstab of archetype vs supplier_profile")
    print(f"    06_top_specialties_per_arch  - top SPCLTY_DSC values within each archetype")
    print()
    print(f"  IMPORTANT: No other scripts need to be re-run. The archetype columns")
    print(f"  are added to features parquet but not used by any other script.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL: {exc}", file=sys.stderr)
        raise