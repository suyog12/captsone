from __future__ import annotations

import sys
import time
import warnings
from pathlib import Path

import duckdb
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# Paths

ROOT         = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN   = ROOT / "data_clean"
MERGED_FILE  = DATA_CLEAN / "serving"  / "merged_dataset.parquet"
PRODUCT_FILE = DATA_CLEAN / "product"  / "products_clean.parquet"

OUT_PRECOMP  = DATA_CLEAN / "serving"  / "precomputed"
OUT_ANALYSIS = DATA_CLEAN / "analysis"
CHART_DIR    = OUT_ANALYSIS / "charts" / "product_cooccurrence"

SLIM_OUT      = OUT_PRECOMP  / "product_cooccurrence.parquet"
FULL_OUT      = OUT_ANALYSIS / "product_cooccurrence_with_metadata.parquet"
XLSX_OUT      = OUT_ANALYSIS / "product_cooccurrence_analysis.xlsx"


# Configuration

FISCAL_YEARS            = ("FY2425", "FY2526")
MIN_BUYERS_PER_PRODUCT  = 50      # Match other files — products with < 50 buyers excluded
MIN_ORDERS_TOGETHER     = 20      # Minimum co-occurrences to include a pair
MIN_LIFT                = 1.5     # Pair must be 50% more likely than random
MIN_SUPPORT             = 0.0001  # 0.01% of orders
TOP_N_PAIRS_PER_PRODUCT = 20      # Keep top 20 co-purchased per product

EXCLUDED_SUPPLIERS = {"MEDLINE", "MEDLINE INDUSTRIES"}
EXCLUDED_FAMILIES  = {"Fee", "Unknown", "NaN", "nan", ""}


# Chart configuration

CHART_DPI = 120
CHART_COLORS = {
    "primary":   "#1F4E79",
    "secondary": "#2E75B6",
    "accent":    "#375623",
    "highlight": "#833C00",
    "warning":   "#C00000",
    "neutral":   "#7F7F7F",
}


# Logging

def _s(title: str) -> None:
    print(f"\n{'-' * 64}\n  {title}\n{'-' * 64}", flush=True)


def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


def _pq(p: Path) -> str:
    return "'" + p.as_posix() + "'"


# Step 1-4: Pipeline using a single DuckDB connection

def run_pipeline() -> pd.DataFrame:
    _s("Step 1: Building order-level data (same customer + same week = one order)")
    t0 = time.time()

    con = duckdb.connect()

    # Detect columns
    desc = con.execute(
        f"DESCRIBE SELECT * FROM read_parquet({_pq(MERGED_FILE)}) LIMIT 0"
    ).df()
    available = set(desc["column_name"].tolist())

    family_col = next(
        (c for c in ["PROD_FMLY_LVL1_DSC", "PROD_FMLY_LVL1_CD"]
         if c in available), None
    )
    supplier_col = next(
        (c for c in ["SUPLR_ROLLUP_DSC", "SUPLR_DSC"] if c in available), None
    )
    has_date_parts = all(
        c in available for c in ["order_year", "order_month", "order_day"]
    )

    _log(f"Product family column : {family_col}")
    _log(f"Supplier column       : {supplier_col}")
    _log(f"Date parts available  : {has_date_parts}")

    if not has_date_parts:
        print("\nFATAL: need order_year, order_month, order_day columns",
              file=sys.stderr)
        sys.exit(1)

    # Filters
    fy_sql = ", ".join(f"'{fy}'" for fy in FISCAL_YEARS)
    family_filter = ""
    if family_col:
        excl_fams = ", ".join(f"'{f}'" for f in sorted(EXCLUDED_FAMILIES))
        family_filter = f"AND COALESCE({family_col}, 'Unknown') NOT IN ({excl_fams})"

    supplier_filter = ""
    if supplier_col:
        excl_sups = ", ".join(f"'{s}'" for s in EXCLUDED_SUPPLIERS)
        supplier_filter = f"AND UPPER(COALESCE({supplier_col}, '')) NOT IN ({excl_sups})"

    # Build filtered transactions with order_id (cust + year_week)
    _log("Building filtered transaction table with order_id...")
    con.execute(f"""
        CREATE TEMPORARY TABLE filtered_txn AS
        SELECT
            CAST(DIM_CUST_CURR_ID AS BIGINT) AS cust_id,
            CAST(DIM_ITEM_E1_CURR_ID AS BIGINT) AS item_id,
            CAST(order_year AS BIGINT) * 100 +
                CAST(
                    EXTRACT(WEEK FROM MAKE_DATE(order_year, order_month, order_day))
                    AS BIGINT
                ) AS year_week
        FROM read_parquet({_pq(MERGED_FILE)})
        WHERE UNIT_SLS_AMT > 0
          AND fiscal_year IN ({fy_sql})
          AND DIM_CUST_CURR_ID IS NOT NULL
          AND DIM_ITEM_E1_CURR_ID IS NOT NULL
          {family_filter}
          {supplier_filter}
    """)

    n_txn = con.execute("SELECT COUNT(*) FROM filtered_txn").fetchone()[0]
    _log(f"  Filtered transactions: {n_txn:,}")

    # Filter to products with enough buyers
    _log(f"Filtering to products with >= {MIN_BUYERS_PER_PRODUCT} buyers...")
    con.execute(f"""
        CREATE TEMPORARY TABLE eligible_items AS
        SELECT item_id
        FROM filtered_txn
        GROUP BY item_id
        HAVING COUNT(DISTINCT cust_id) >= {MIN_BUYERS_PER_PRODUCT}
    """)

    n_eligible = con.execute("SELECT COUNT(*) FROM eligible_items").fetchone()[0]
    _log(f"  Eligible products: {n_eligible:,}")

    # Build orders table: (order_id, item_id) - one row per unique pair
    _log("Building orders table (distinct order-item pairs)...")
    con.execute("""
        CREATE TEMPORARY TABLE orders AS
        SELECT DISTINCT
            (cust_id * 1000000 + year_week) AS order_id,
            item_id
        FROM filtered_txn
        WHERE item_id IN (SELECT item_id FROM eligible_items)
    """)

    total_orders = con.execute(
        "SELECT COUNT(DISTINCT order_id) FROM orders"
    ).fetchone()[0]
    total_products = con.execute(
        "SELECT COUNT(DISTINCT item_id) FROM orders"
    ).fetchone()[0]
    total_pairs_in_orders = con.execute("SELECT COUNT(*) FROM orders").fetchone()[0]

    _log(f"  Total unique orders          : {total_orders:,}")
    _log(f"  Total unique products        : {total_products:,}")
    _log(f"  Total order-item pairs       : {total_pairs_in_orders:,}")
    _log(f"  Avg items per order          : {total_pairs_in_orders/total_orders:.1f}")

    _log(f"Step 1 done in {time.time()-t0:.1f}s")

    # Step 2: Compute order count per product (for support_a calculation)
    _s("Step 2: Computing order count per product")
    t0 = time.time()

    _log("Computing n_orders_per_product...")
    con.execute("""
        CREATE TEMPORARY TABLE product_order_counts AS
        SELECT
            item_id,
            COUNT(DISTINCT order_id) AS n_orders
        FROM orders
        GROUP BY item_id
    """)

    n_products_counted = con.execute(
        "SELECT COUNT(*) FROM product_order_counts"
    ).fetchone()[0]
    _log(f"  Products with order counts: {n_products_counted:,}")
    _log(f"Step 2 done in {time.time()-t0:.1f}s")

    # Step 3: Generate product pairs via self-join on orders
    _s("Step 3: Computing product co-occurrence pairs (self-join)")
    t0 = time.time()

    _log(f"Running self-join: this is the heaviest step (could take 5-15 min)...")
    _log(f"  Filter thresholds: MIN_ORDERS_TOGETHER={MIN_ORDERS_TOGETHER}, "
         f"MIN_SUPPORT={MIN_SUPPORT}, MIN_LIFT={MIN_LIFT}")

    # The self-join: find pairs of items that appear in the same order
    # We enforce item_a < item_b to avoid duplicate pairs (A,B) and (B,A)
    # We do this with symmetric rules later for recommendation
    con.execute(f"""
        CREATE TEMPORARY TABLE cooccurrence_raw AS
        SELECT
            a.item_id AS item_a,
            b.item_id AS item_b,
            COUNT(*) AS n_orders_together
        FROM orders a
        INNER JOIN orders b
          ON a.order_id = b.order_id
          AND a.item_id < b.item_id
        GROUP BY a.item_id, b.item_id
        HAVING COUNT(*) >= {MIN_ORDERS_TOGETHER}
    """)

    n_raw_pairs = con.execute(
        "SELECT COUNT(*) FROM cooccurrence_raw"
    ).fetchone()[0]
    _log(f"  Raw co-occurrence pairs (>= {MIN_ORDERS_TOGETHER} orders): {n_raw_pairs:,}")

    # Step 4: Compute support, confidence, lift
    _log("Computing support, confidence, lift...")
    con.execute(f"""
        CREATE TEMPORARY TABLE cooccurrence_stats AS
        SELECT
            co.item_a,
            co.item_b,
            co.n_orders_together,
            pa.n_orders AS n_orders_a,
            pb.n_orders AS n_orders_b,
            CAST(co.n_orders_together AS DOUBLE) / {total_orders} AS support_ab,
            CAST(pa.n_orders AS DOUBLE) / {total_orders} AS support_a,
            CAST(pb.n_orders AS DOUBLE) / {total_orders} AS support_b,
            CAST(co.n_orders_together AS DOUBLE) / pa.n_orders AS confidence_a_to_b,
            CAST(co.n_orders_together AS DOUBLE) / pb.n_orders AS confidence_b_to_a,
            (CAST(co.n_orders_together AS DOUBLE) / {total_orders}) /
                ((CAST(pa.n_orders AS DOUBLE) / {total_orders}) *
                 (CAST(pb.n_orders AS DOUBLE) / {total_orders})) AS lift
        FROM cooccurrence_raw co
        JOIN product_order_counts pa ON co.item_a = pa.item_id
        JOIN product_order_counts pb ON co.item_b = pb.item_id
        WHERE CAST(co.n_orders_together AS DOUBLE) / {total_orders} >= {MIN_SUPPORT}
          AND (CAST(co.n_orders_together AS DOUBLE) / {total_orders}) /
              ((CAST(pa.n_orders AS DOUBLE) / {total_orders}) *
               (CAST(pb.n_orders AS DOUBLE) / {total_orders})) > {MIN_LIFT}
    """)

    n_filtered = con.execute(
        "SELECT COUNT(*) FROM cooccurrence_stats"
    ).fetchone()[0]
    _log(f"  After support/lift filters: {n_filtered:,} pairs")
    _log(f"Step 3 done in {time.time()-t0:.1f}s")

    # Step 4: Make pairs directional (A->B and B->A) so each product has its own rankings
    _s("Step 4: Creating directional pairs and ranking top-N per product")
    t0 = time.time()

    _log("Creating directional pairs (A->B and B->A)...")
    con.execute("""
        CREATE TEMPORARY TABLE cooccurrence_directional AS
        -- A -> B direction
        SELECT
            item_a AS product_a,
            item_b AS product_b,
            n_orders_together,
            n_orders_a AS n_orders_product_a,
            n_orders_b AS n_orders_product_b,
            support_ab,
            confidence_a_to_b AS confidence,
            lift
        FROM cooccurrence_stats
        UNION ALL
        -- B -> A direction (swap)
        SELECT
            item_b AS product_a,
            item_a AS product_b,
            n_orders_together,
            n_orders_b AS n_orders_product_a,
            n_orders_a AS n_orders_product_b,
            support_ab,
            confidence_b_to_a AS confidence,
            lift
        FROM cooccurrence_stats
    """)

    n_directional = con.execute(
        "SELECT COUNT(*) FROM cooccurrence_directional"
    ).fetchone()[0]
    _log(f"  Directional pairs: {n_directional:,}")

    # Rank top-N pairs per product_a (ranked by lift descending, then confidence)
    _log(f"Ranking top-{TOP_N_PAIRS_PER_PRODUCT} per product...")
    pairs_df = con.execute(f"""
        SELECT *
        FROM (
            SELECT
                product_a,
                product_b,
                n_orders_together,
                n_orders_product_a,
                n_orders_product_b,
                support_ab,
                confidence,
                lift,
                ROW_NUMBER() OVER (
                    PARTITION BY product_a
                    ORDER BY lift DESC, confidence DESC
                ) AS rank
            FROM cooccurrence_directional
        )
        WHERE rank <= {TOP_N_PAIRS_PER_PRODUCT}
        ORDER BY product_a, rank
    """).df()

    _log(f"  Final top-{TOP_N_PAIRS_PER_PRODUCT} pairs: {len(pairs_df):,}")
    _log(f"  Products with at least 1 pair: {pairs_df['product_a'].nunique():,}")
    _log(f"Step 4 done in {time.time()-t0:.1f}s")

    con.close()

    # Enforce types
    pairs_df["product_a"]          = pairs_df["product_a"].astype("int64")
    pairs_df["product_b"]          = pairs_df["product_b"].astype("int64")
    pairs_df["rank"]                = pairs_df["rank"].astype("int32")
    pairs_df["n_orders_together"]  = pairs_df["n_orders_together"].astype("int32")
    pairs_df["n_orders_product_a"] = pairs_df["n_orders_product_a"].astype("int32")
    pairs_df["n_orders_product_b"] = pairs_df["n_orders_product_b"].astype("int32")
    pairs_df["support_ab"]          = pairs_df["support_ab"].astype("float32")
    pairs_df["confidence"]          = pairs_df["confidence"].astype("float32")
    pairs_df["lift"]                = pairs_df["lift"].astype("float32")

    return pairs_df


# Step 5: Enrich with product metadata

def enrich_with_metadata(pairs_df: pd.DataFrame) -> pd.DataFrame:
    _s("Step 5: Enriching with product metadata")

    if not PRODUCT_FILE.exists():
        _log("products_clean.parquet not found - skipping metadata enrichment")
        return pairs_df

    products = pd.read_parquet(PRODUCT_FILE, columns=[
        "DIM_ITEM_E1_CURR_ID", "ITEM_DSC",
        "PROD_FMLY_LVL1_DSC", "PROD_CTGRY_LVL2_DSC",
        "SUPLR_ROLLUP_DSC", "is_private_brand",
    ])
    products["DIM_ITEM_E1_CURR_ID"] = products["DIM_ITEM_E1_CURR_ID"].astype("int64")

    # Enrich product_a
    a_meta = products.rename(columns={
        "DIM_ITEM_E1_CURR_ID": "product_a",
        "ITEM_DSC":            "product_a_desc",
        "PROD_FMLY_LVL1_DSC":  "product_a_family",
        "PROD_CTGRY_LVL2_DSC": "product_a_category",
        "SUPLR_ROLLUP_DSC":    "product_a_supplier",
        "is_private_brand":    "product_a_private_brand",
    })
    pairs_df = pairs_df.merge(a_meta, on="product_a", how="left")

    # Enrich product_b
    b_meta = products.rename(columns={
        "DIM_ITEM_E1_CURR_ID": "product_b",
        "ITEM_DSC":            "product_b_desc",
        "PROD_FMLY_LVL1_DSC":  "product_b_family",
        "PROD_CTGRY_LVL2_DSC": "product_b_category",
        "SUPLR_ROLLUP_DSC":    "product_b_supplier",
        "is_private_brand":    "product_b_private_brand",
    })
    pairs_df = pairs_df.merge(b_meta, on="product_b", how="left")

    # Flag same-family pairs
    pairs_df["same_family"] = (
        pairs_df["product_a_family"] == pairs_df["product_b_family"]
    ).astype(int)

    _log(f"Metadata merged: {len(pairs_df):,} rows")
    _log(f"  Same-family pairs: {pairs_df['same_family'].sum():,} "
         f"({pairs_df['same_family'].mean()*100:.1f}%)")
    _log(f"  Cross-family pairs: {(1-pairs_df['same_family']).sum():,} "
         f"({(1-pairs_df['same_family'].mean())*100:.1f}%)")

    return pairs_df


# Step 6: Save outputs

def save_outputs(pairs_df: pd.DataFrame) -> None:
    _s("Step 6: Saving outputs")

    OUT_PRECOMP.mkdir(parents=True, exist_ok=True)
    OUT_ANALYSIS.mkdir(parents=True, exist_ok=True)

    # Sort for consistent output
    pairs_df = pairs_df.sort_values(
        ["product_a", "rank"], ascending=[True, True]
    ).reset_index(drop=True)

    # Slim version for recommendation engine
    slim_cols = [
        "product_a", "product_b", "rank",
        "n_orders_together",
        "support_ab", "confidence", "lift",
        "same_family",
    ]
    slim_cols = [c for c in slim_cols if c in pairs_df.columns]
    slim = pairs_df[slim_cols].copy()

    slim.to_parquet(SLIM_OUT, index=False)
    size_mb = SLIM_OUT.stat().st_size / (1024 * 1024)
    _log(f"Saved slim version: {SLIM_OUT.relative_to(ROOT)}  ({size_mb:.1f} MB)")

    # Full version with metadata
    pairs_df.to_parquet(FULL_OUT, index=False)
    size_mb = FULL_OUT.stat().st_size / (1024 * 1024)
    _log(f"Saved full version: {FULL_OUT.relative_to(ROOT)}  ({size_mb:.1f} MB)")


# Step 7: Print samples and stats

def print_samples(pairs_df: pd.DataFrame) -> None:
    _s("Step 7: Sample co-occurrence results")

    if "product_a_desc" not in pairs_df.columns:
        _log("Metadata not available — skipping samples")
        return

    _log(f"Total co-purchase pairs: {len(pairs_df):,}")
    _log(f"Unique product_a count : {pairs_df['product_a'].nunique():,}")
    _log(f"Avg pairs per product  : {len(pairs_df) / pairs_df['product_a'].nunique():.1f}")
    _log(f"Lift range  : [{pairs_df['lift'].min():.2f}, {pairs_df['lift'].max():.2f}]")
    _log(f"Median lift : {pairs_df['lift'].median():.2f}")
    _log(f"Same-family rate: {pairs_df['same_family'].mean()*100:.1f}%")
    _log("")

    # Pick 5 popular products to show their top co-purchases
    popular_products = pairs_df["product_a"].value_counts().head(5).index.tolist()

    for item_a in popular_products:
        top5 = pairs_df[pairs_df["product_a"] == item_a].head(5)
        if len(top5) == 0:
            continue

        desc_a = str(top5.iloc[0].get("product_a_desc", "?"))[:55]
        fam_a  = str(top5.iloc[0].get("product_a_family", "?"))[:35]

        _log(f"")
        _log(f"  Product A: {desc_a}")
        _log(f"  Family   : {fam_a}")
        _log(f"  Top 5 co-purchased products:")
        for _, r in top5.iterrows():
            desc_b = str(r.get("product_b_desc", "?"))[:55]
            fam_b  = str(r.get("product_b_family", "?"))[:25]
            same = " [same family]" if r["same_family"] == 1 else ""
            _log(f"    conf={r['confidence']:.2f}  lift={r['lift']:.2f}  "
                 f"{desc_b:<57} ({fam_b}){same}")


# Chart builders

def build_charts(pairs_df: pd.DataFrame) -> list[Path]:
    _s("Step 8: Building analysis charts")
    CHART_DIR.mkdir(parents=True, exist_ok=True)

    paths = []

    # Chart 1: Lift distribution
    _log("Chart 1: lift distribution")
    fig, ax = plt.subplots(figsize=(11, 6), dpi=CHART_DPI)
    ax.hist(pairs_df["lift"].clip(upper=20), bins=80,
            color=CHART_COLORS["primary"], alpha=0.85, edgecolor="white")
    ax.axvline(2.0, color=CHART_COLORS["warning"], linestyle="--",
               linewidth=1.5, label="Strong threshold (lift > 2.0)")
    ax.axvline(pairs_df["lift"].median(), color=CHART_COLORS["accent"],
               linestyle="--", linewidth=1.5,
               label=f"Median: {pairs_df['lift'].median():.2f}")
    ax.set_xlabel("Lift (capped at 20 for display)", fontsize=11)
    ax.set_ylabel("Number of pairs", fontsize=11)
    ax.set_title("Co-occurrence Lift Distribution",
                 fontsize=12, fontweight="bold")
    ax.set_yscale("log")
    ax.grid(True, alpha=0.3, axis="y")
    ax.legend(fontsize=10)
    plt.tight_layout()
    p = CHART_DIR / "01_lift_distribution.png"
    plt.savefig(p, bbox_inches="tight")
    plt.close()
    paths.append(p)

    # Chart 2: Confidence distribution
    _log("Chart 2: confidence distribution")
    fig, ax = plt.subplots(figsize=(11, 6), dpi=CHART_DPI)
    ax.hist(pairs_df["confidence"], bins=50,
            color=CHART_COLORS["secondary"], alpha=0.85, edgecolor="white")
    ax.axvline(pairs_df["confidence"].median(), color=CHART_COLORS["warning"],
               linestyle="--", linewidth=1.5,
               label=f"Median: {pairs_df['confidence'].median():.2f}")
    ax.set_xlabel("Confidence (P(B in order | A in order))", fontsize=11)
    ax.set_ylabel("Number of pairs", fontsize=11)
    ax.set_title("Co-occurrence Confidence Distribution",
                 fontsize=12, fontweight="bold")
    ax.grid(True, alpha=0.3, axis="y")
    ax.legend(fontsize=10)
    plt.tight_layout()
    p = CHART_DIR / "02_confidence_distribution.png"
    plt.savefig(p, bbox_inches="tight")
    plt.close()
    paths.append(p)

    # Chart 3: Same-family rate by lift tier
    _log("Chart 3: same-family rate by lift tier")
    fig, ax = plt.subplots(figsize=(11, 6), dpi=CHART_DPI)
    tiers = [
        (1.5, 2.0,   "Medium (1.5-2.0)"),
        (2.0, 3.0,   "Strong (2.0-3.0)"),
        (3.0, 5.0,   "Very Strong (3.0-5.0)"),
        (5.0, 10.0,  "Exceptional (5.0-10.0)"),
        (10.0, 1000, "Extreme (>10.0)"),
    ]
    tier_names  = []
    same_family_rates = []
    tier_counts = []
    for lo, hi, name in tiers:
        tier = pairs_df[(pairs_df["lift"] >= lo) & (pairs_df["lift"] < hi)]
        if len(tier) == 0:
            continue
        tier_names.append(f"{name}\n(n={len(tier):,})")
        same_family_rates.append(tier["same_family"].mean() * 100)
        tier_counts.append(len(tier))

    bars = ax.bar(tier_names, same_family_rates,
                  color=CHART_COLORS["accent"], alpha=0.85,
                  edgecolor="white", linewidth=1)
    for bar, rate in zip(bars, same_family_rates):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() * 1.02,
                f"{rate:.1f}%", ha="center", fontsize=10)

    ax.axhline(50, color=CHART_COLORS["warning"], linestyle="--",
               linewidth=1.5, alpha=0.7, label="50% reference line")
    ax.set_ylabel("% same-family pairs", fontsize=11)
    ax.set_title("Same-Family Rate by Lift Tier\n"
                 "(higher lift should correlate with more same-family pairs)",
                 fontsize=12, fontweight="bold")
    ax.grid(True, alpha=0.3, axis="y")
    ax.legend(fontsize=10)
    ax.set_ylim(0, 110)
    plt.tight_layout()
    p = CHART_DIR / "03_same_family_by_lift.png"
    plt.savefig(p, bbox_inches="tight")
    plt.close()
    paths.append(p)

    # Chart 4: Pairs per product distribution
    _log("Chart 4: pairs per product distribution")
    fig, ax = plt.subplots(figsize=(11, 6), dpi=CHART_DPI)
    pairs_per = pairs_df["product_a"].value_counts()
    ax.hist(pairs_per, bins=range(0, TOP_N_PAIRS_PER_PRODUCT + 2),
            color=CHART_COLORS["highlight"], alpha=0.85, edgecolor="white")
    ax.axvline(pairs_per.median(), color=CHART_COLORS["warning"],
               linestyle="--", linewidth=1.5,
               label=f"Median: {pairs_per.median():.0f}")
    ax.set_xlabel("Number of co-purchased products per product", fontsize=11)
    ax.set_ylabel("Number of products", fontsize=11)
    ax.set_title("Co-purchased Pairs per Product",
                 fontsize=12, fontweight="bold")
    ax.grid(True, alpha=0.3, axis="y")
    ax.legend(fontsize=10)
    plt.tight_layout()
    p = CHART_DIR / "04_pairs_per_product.png"
    plt.savefig(p, bbox_inches="tight")
    plt.close()
    paths.append(p)

    # Chart 5: Support vs lift scatter
    _log("Chart 5: support vs lift scatter")
    fig, ax = plt.subplots(figsize=(11, 7), dpi=CHART_DPI)
    sample = pairs_df.sample(n=min(20000, len(pairs_df)), random_state=42)
    ax.scatter(sample["support_ab"], sample["lift"].clip(upper=20),
               c=sample["same_family"], cmap="RdYlBu_r",
               alpha=0.4, s=10, edgecolors="none")
    ax.set_xscale("log")
    ax.set_yscale("log")
    ax.set_xlabel("Support (% of all orders containing both, log scale)", fontsize=11)
    ax.set_ylabel("Lift (capped at 20, log scale)", fontsize=11)
    ax.set_title("Pair Support vs Lift\n"
                 "(color: blue=cross-family, red=same-family)",
                 fontsize=12, fontweight="bold")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    p = CHART_DIR / "05_support_vs_lift.png"
    plt.savefig(p, bbox_inches="tight")
    plt.close()
    paths.append(p)

    _log(f"")
    _log(f"Saved {len(paths)} charts to {CHART_DIR.relative_to(ROOT)}")
    for p in paths:
        size_kb = p.stat().st_size / 1024
        _log(f"  {p.name}  ({size_kb:.0f} KB)")

    return paths


# Excel styling helper

def _style_sheet(ws, df: pd.DataFrame, hc: str = "1F4E79") -> None:
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = bdr

    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = bdr

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(
            col_cells[0].column)].width = min(max(w + 2, 12), 55)


# Step 9: Build xlsx

def build_xlsx(pairs_df: pd.DataFrame) -> None:
    _s("Step 9: Building multi-sheet xlsx analysis")
    t0 = time.time()

    # Summary sheet
    total_pairs = len(pairs_df)
    unique_a = pairs_df["product_a"].nunique()
    median_lift = pairs_df["lift"].median()
    max_lift = pairs_df["lift"].max()
    same_family_rate = pairs_df["same_family"].mean() * 100 if "same_family" in pairs_df.columns else 0

    rows = [
        {"metric": "=== TOTALS ===", "value": ""},
        {"metric": "Total directional pairs", "value": f"{total_pairs:,}"},
        {"metric": "Unique products with pairs", "value": f"{unique_a:,}"},
        {"metric": "Avg pairs per product",
         "value": f"{total_pairs / max(unique_a, 1):.1f}"},
        {"metric": "", "value": ""},
        {"metric": "=== LIFT ===", "value": ""},
        {"metric": "Min lift",    "value": f"{pairs_df['lift'].min():.2f}"},
        {"metric": "Median lift", "value": f"{median_lift:.2f}"},
        {"metric": "P90 lift",
         "value": f"{pairs_df['lift'].quantile(0.90):.2f}"},
        {"metric": "Max lift",    "value": f"{max_lift:.2f}"},
        {"metric": "", "value": ""},
        {"metric": "=== LIFT TIERS ===", "value": ""},
    ]

    for lo, hi, name in [
        (1.5, 2.0, "Medium (1.5-2.0)"),
        (2.0, 3.0, "Strong (2.0-3.0)"),
        (3.0, 5.0, "Very strong (3.0-5.0)"),
        (5.0, 10.0, "Exceptional (5.0-10.0)"),
        (10.0, 1e9, "Extreme (>10.0)"),
    ]:
        tier = pairs_df[(pairs_df["lift"] >= lo) & (pairs_df["lift"] < hi)]
        rows.append({
            "metric": f"  {name}",
            "value": f"{len(tier):,} pairs ({len(tier)/total_pairs*100:.1f}%)"
        })

    rows.append({"metric": "", "value": ""})
    rows.append({"metric": "=== CONFIDENCE ===", "value": ""})
    rows.append({"metric": "Median confidence",
                 "value": f"{pairs_df['confidence'].median():.3f}"})
    rows.append({"metric": "P90 confidence",
                 "value": f"{pairs_df['confidence'].quantile(0.90):.3f}"})

    rows.append({"metric": "", "value": ""})
    rows.append({"metric": "=== FAMILY RELATIONSHIP ===", "value": ""})
    if "same_family" in pairs_df.columns:
        n_same = int(pairs_df["same_family"].sum())
        rows.append({
            "metric": "Same-family pairs",
            "value": f"{n_same:,} ({n_same/total_pairs*100:.1f}%)"
        })
        rows.append({
            "metric": "Cross-family pairs",
            "value": f"{total_pairs-n_same:,} ({(1-n_same/total_pairs)*100:.1f}%)"
        })

    summary_df = pd.DataFrame(rows)

    # Sheet builders
    display_cols = [
        "product_a", "product_a_desc", "product_a_family",
        "product_b", "product_b_desc", "product_b_family",
        "rank", "n_orders_together",
        "n_orders_product_a", "n_orders_product_b",
        "support_ab", "confidence", "lift",
        "same_family",
    ]
    display_cols = [c for c in display_cols if c in pairs_df.columns]

    top_strong = pairs_df.nlargest(500, "lift")[display_cols].copy()
    top_confident = pairs_df.nlargest(500, "confidence")[display_cols].copy()
    top_volume = pairs_df.nlargest(500, "n_orders_together")[display_cols].copy()
    same_family_top = pairs_df[pairs_df.get("same_family", 0) == 1].nlargest(
        500, "lift")[display_cols].copy() if "same_family" in pairs_df.columns else pd.DataFrame()
    cross_family_top = pairs_df[pairs_df.get("same_family", 0) == 0].nlargest(
        500, "lift")[display_cols].copy() if "same_family" in pairs_df.columns else pd.DataFrame()

    _log(f"  Summary sheet  : {len(summary_df)} rows")
    _log(f"  Top strong     : {len(top_strong):,} rows (highest lift)")
    _log(f"  Top confident  : {len(top_confident):,} rows (highest confidence)")
    _log(f"  Top volume     : {len(top_volume):,} rows (most co-occurrences)")
    _log(f"  Same family    : {len(same_family_top):,} rows")
    _log(f"  Cross family   : {len(cross_family_top):,} rows")

    with pd.ExcelWriter(XLSX_OUT, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="01_summary", index=False)
        top_strong.to_excel(writer, sheet_name="02_top_strongest_lift", index=False)
        top_confident.to_excel(writer, sheet_name="03_top_confidence", index=False)
        top_volume.to_excel(writer, sheet_name="04_top_volume", index=False)
        if len(same_family_top) > 0:
            same_family_top.to_excel(writer, sheet_name="05_same_family", index=False)
        if len(cross_family_top) > 0:
            cross_family_top.to_excel(writer, sheet_name="06_cross_family", index=False)

        wb = writer.book
        _style_sheet(writer.sheets["01_summary"], summary_df, hc="1F4E79")
        wb["01_summary"].sheet_properties.tabColor = "1F4E79"

        _style_sheet(writer.sheets["02_top_strongest_lift"], top_strong, hc="C00000")
        wb["02_top_strongest_lift"].sheet_properties.tabColor = "C00000"

        _style_sheet(writer.sheets["03_top_confidence"], top_confident, hc="833C00")
        wb["03_top_confidence"].sheet_properties.tabColor = "833C00"

        _style_sheet(writer.sheets["04_top_volume"], top_volume, hc="2E75B6")
        wb["04_top_volume"].sheet_properties.tabColor = "2E75B6"

        if "05_same_family" in wb.sheetnames:
            _style_sheet(writer.sheets["05_same_family"], same_family_top, hc="375623")
            wb["05_same_family"].sheet_properties.tabColor = "375623"

        if "06_cross_family" in wb.sheetnames:
            _style_sheet(writer.sheets["06_cross_family"], cross_family_top, hc="7030A0")
            wb["06_cross_family"].sheet_properties.tabColor = "7030A0"

    size_mb = XLSX_OUT.stat().st_size / (1024 * 1024)
    _log(f"")
    _log(f"Saved: {XLSX_OUT.relative_to(ROOT)}  ({size_mb:.2f} MB, 6 sheets)  "
         f"in {time.time()-t0:.1f}s")


# Main

def main() -> None:
    print()
    print("=" * 64)
    print("  PRODUCT CO-OCCURRENCE (Same-Order Basket Analysis)")
    print("=" * 64)
    start = time.time()

    OUT_PRECOMP.mkdir(parents=True, exist_ok=True)
    OUT_ANALYSIS.mkdir(parents=True, exist_ok=True)
    CHART_DIR.mkdir(parents=True, exist_ok=True)

    pairs_df = run_pipeline()
    pairs_df = enrich_with_metadata(pairs_df)
    save_outputs(pairs_df)
    print_samples(pairs_df)
    build_charts(pairs_df)
    build_xlsx(pairs_df)

    _s("Complete")
    _log(f"Total time: {time.time() - start:.1f}s")
    _log("")
    _log(f"Key outputs:")
    _log(f"  product_cooccurrence.parquet              (slim, for recommendation engine)")
    _log(f"  product_cooccurrence_with_metadata.parquet (full, for inspection)")
    _log(f"  product_cooccurrence_analysis.xlsx        (6 sheets)")
    _log(f"  charts/product_cooccurrence/              (5 PNG files)")
    _log("")
    _log("Definition of 'same order': same customer + same ISO week")
    _log("Methodology: Agrawal 1993 association rules at product-order level")
    _log("")
    _log("Used by recommendation_factors.py for cart-based recommendations:")
    _log("  'If customer has Product X in cart, recommend Product Y'")
    _log("  Based on: customers who buy X in an order also buy Y in same order")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise