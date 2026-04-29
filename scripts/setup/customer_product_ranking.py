from __future__ import annotations

import sys
import time
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Path configuration

ROOT = Path(__file__).resolve().parent
DATA_CLEAN = ROOT / "data_clean"

FEATURES_FILE   = DATA_CLEAN / "features"  / "customer_features.parquet"
CUSTOMERS_FILE  = DATA_CLEAN / "customer"  / "customers_clean.parquet"
PRODUCTS_FILE   = DATA_CLEAN / "product"   / "products_clean.parquet"
MERGED_FILE     = DATA_CLEAN / "serving"   / "merged_dataset.parquet"

OUT_DIR         = ROOT / "rankings_report"


# Configuration

# A product is considered "in active inventory" if at least this many distinct
# customers have bought it. Below this it's effectively dead/long-tail.
MIN_ACTIVE_BUYERS = 1   # 1 = include everything sold even once
                        # 50 = matches the engine's filter (cleaner report)

# Quantity column - PRMRY_QTY is UoM-normalized
VOLUME_QTY_COL      = "PRMRY_QTY"
VOLUME_QTY_FALLBACK = "SHIP_QTY"


# Logging

def _section(title):
    print(f"\n{'-' * 70}")
    print(f"  {title}")
    print(f"{'-' * 70}", flush=True)


def _log(msg):
    print(f"  {msg}", flush=True)


# DuckDB

def _con(memory_gb=4, threads=2):
    con = duckdb.connect(":memory:")
    con.execute(f"SET memory_limit = '{memory_gb}GB'")
    con.execute(f"SET threads = {threads}")
    con.execute("SET preserve_insertion_order = false")
    return con


# Excel styling

def _style(ws, df, header_color="1F4E79"):
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=header_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, 10), 50)


# Step 1: Load and rank customers

def step1_rank_customers():
    # Load customer_features.parquet, join with customers_clean.parquet for
    # readable names, then rank on three dimensions:
    #   - rank_by_spend   - based on monetary (lifetime spend) descending
    #   - rank_by_volume  - based on lifetime volume descending
    #   - combined_rank   - based on (rank_by_spend + rank_by_volume) ascending
    # The combined rank is a simple Borda-count blend - lower = bigger customer.

    _section("Step 1: Loading and ranking customers")

    if not FEATURES_FILE.exists():
        print(f"FATAL: {FEATURES_FILE} not found. Run clean_data.py first.",
              file=sys.stderr)
        sys.exit(1)

    feats = pd.read_parquet(FEATURES_FILE)
    _log(f"Loaded {len(feats):,} customers from features parquet")

    # We need lifetime volume (sum of PRMRY_QTY across all months).
    # The features parquet has median_monthly_volume, not lifetime. Compute it.
    if MERGED_FILE.exists():
        _log("Computing lifetime volume per customer from merged_dataset...")
        qty_expr = (
            f"GREATEST(0, COALESCE(NULLIF({VOLUME_QTY_COL}, 0), "
            f"{VOLUME_QTY_FALLBACK}, 0))"
        )
        con = _con(memory_gb=5, threads=2)
        lifetime = con.execute(f"""
            SELECT
                CAST(DIM_CUST_CURR_ID AS BIGINT)  AS DIM_CUST_CURR_ID,
                ROUND(SUM(UNIT_SLS_AMT), 2)        AS lifetime_spend,
                ROUND(SUM({qty_expr}), 2)          AS lifetime_volume,
                COUNT(DISTINCT ORDR_NUM)           AS lifetime_orders,
                COUNT(*)                           AS lifetime_lines
            FROM read_parquet('{MERGED_FILE.as_posix()}')
            WHERE UNIT_SLS_AMT > 0
              AND DIM_CUST_CURR_ID IS NOT NULL
            GROUP BY CAST(DIM_CUST_CURR_ID AS BIGINT)
        """).df()
        con.close()
        _log(f"  Computed lifetime stats for {len(lifetime):,} customers")
    else:
        _log("merged_dataset.parquet not found; using monetary as fallback")
        lifetime = feats[["DIM_CUST_CURR_ID", "monetary", "frequency"]].rename(
            columns={"monetary": "lifetime_spend", "frequency": "lifetime_orders"}
        )
        lifetime["lifetime_volume"] = 0
        lifetime["lifetime_lines"] = 0

    # Join with customer name/profile fields for readable output
    if CUSTOMERS_FILE.exists():
        custs = pd.read_parquet(
            CUSTOMERS_FILE,
            columns=["DIM_CUST_CURR_ID", "CUST_NAME", "CUST_TYPE_DSC",
                     "SPCLTY_DSC", "MKT_CD", "STATE", "CITY"]
        )
        _log(f"  Joining customer names from customers_clean.parquet")
    else:
        custs = pd.DataFrame()
        _log(f"  customers_clean.parquet not found - names will be blank")

    # Build the master customer dataframe
    df = feats[[
        "DIM_CUST_CURR_ID",
        "MKT_CD", "SPCLTY_CD", "STATE",
        "size_tier", "spend_tier", "volume_tier",
        "median_monthly_spend", "median_monthly_volume",
        "active_months_last_12",
        "monetary", "frequency", "recency_days",
        "n_categories_bought", "supplier_profile",
        "affordability_ceiling", "churn_label",
    ]].copy()

    df = df.merge(lifetime, on="DIM_CUST_CURR_ID", how="left")
    df["lifetime_spend"]  = df["lifetime_spend"].fillna(0).astype(float)
    df["lifetime_volume"] = df["lifetime_volume"].fillna(0).astype(float)
    df["lifetime_orders"] = df["lifetime_orders"].fillna(0).astype(int)
    df["lifetime_lines"]  = df["lifetime_lines"].fillna(0).astype(int)

    if not custs.empty:
        df = df.merge(custs, on="DIM_CUST_CURR_ID", how="left",
                      suffixes=("", "_dim"))
        # Reconcile MKT_CD / STATE if both exist (features value wins)
        for c in ["MKT_CD_dim", "STATE_dim"]:
            if c in df.columns:
                df = df.drop(columns=[c])

    # Compute the three rank columns
    df["rank_by_spend"]  = df["lifetime_spend"].rank(method="min",  ascending=False).astype(int)
    df["rank_by_volume"] = df["lifetime_volume"].rank(method="min", ascending=False).astype(int)
    df["combined_rank_score"] = df["rank_by_spend"] + df["rank_by_volume"]
    df["combined_rank"] = df["combined_rank_score"].rank(method="min", ascending=True).astype(int)

    # Reorder columns for human-readable output
    col_order = [
        "combined_rank", "rank_by_spend", "rank_by_volume",
        "DIM_CUST_CURR_ID",
        "CUST_NAME", "CUST_TYPE_DSC", "SPCLTY_DSC",
        "MKT_CD", "STATE", "CITY",
        "size_tier", "spend_tier", "volume_tier",
        "lifetime_spend", "lifetime_volume", "lifetime_orders", "lifetime_lines",
        "median_monthly_spend", "median_monthly_volume",
        "monetary", "frequency", "recency_days",
        "active_months_last_12", "n_categories_bought",
        "supplier_profile", "affordability_ceiling", "churn_label",
    ]
    df = df[[c for c in col_order if c in df.columns]]

    df = df.sort_values("combined_rank").reset_index(drop=True)

    _log(f"  Customers ranked: {len(df):,}")
    _log(f"  Top customer: rank={df.iloc[0]['combined_rank']}  "
         f"spend=${df.iloc[0]['lifetime_spend']:,.0f}  "
         f"volume={df.iloc[0]['lifetime_volume']:,.0f}")

    return df


# Step 2: Load and rank products / inventory

def step2_rank_products():
    # Build a per-product ranking sorted by sales count descending.
    # Sales count = number of distinct orders containing this product.
    # Also include total revenue, total units, distinct buyers, avg price,
    # primary market, and the brand/supplier flags.

    _section("Step 2: Loading and ranking products")

    if not MERGED_FILE.exists():
        print(f"FATAL: {MERGED_FILE} not found. Run clean_data.py first.",
              file=sys.stderr)
        sys.exit(1)

    qty_expr = (
        f"GREATEST(0, COALESCE(NULLIF({VOLUME_QTY_COL}, 0), "
        f"{VOLUME_QTY_FALLBACK}, 0))"
    )

    _log("Aggregating per-product sales metrics from merged_dataset...")
    con = _con(memory_gb=6, threads=2)

    # Aggregate per product
    prod = con.execute(f"""
        SELECT
            DIM_ITEM_E1_CURR_ID                    AS item_id,
            COUNT(DISTINCT ORDR_NUM)                AS sales_count,
            COUNT(*)                                AS line_count,
            COUNT(DISTINCT DIM_CUST_CURR_ID)        AS distinct_buyers,
            ROUND(SUM(UNIT_SLS_AMT), 2)             AS total_revenue,
            ROUND(SUM({qty_expr}), 2)               AS total_volume,
            ROUND(AVG(UNIT_SLS_AMT), 2)             AS avg_line_revenue,
            ROUND(SUM(UNIT_SLS_AMT) / NULLIF(SUM({qty_expr}), 0), 4)
                                                    AS avg_unit_price,
            MAX(DIM_ORDR_DT_ID)                     AS last_sold_date_id,
            MIN(DIM_ORDR_DT_ID)                     AS first_sold_date_id
        FROM read_parquet('{MERGED_FILE.as_posix()}')
        WHERE UNIT_SLS_AMT > 0
        GROUP BY DIM_ITEM_E1_CURR_ID
        ORDER BY sales_count DESC
    """).df()
    _log(f"  Aggregated {len(prod):,} products")

    # Compute primary market per product (the market with highest buyer count)
    _log("Computing primary market per product...")
    prim_mkt = con.execute(f"""
        WITH market_counts AS (
            SELECT
                DIM_ITEM_E1_CURR_ID                AS item_id,
                MKT_CD                              AS mkt_cd,
                COUNT(DISTINCT DIM_CUST_CURR_ID)    AS n_buyers
            FROM read_parquet('{MERGED_FILE.as_posix()}')
            WHERE UNIT_SLS_AMT > 0
              AND MKT_CD IS NOT NULL
            GROUP BY DIM_ITEM_E1_CURR_ID, MKT_CD
        ),
        ranked AS (
            SELECT
                item_id, mkt_cd, n_buyers,
                ROW_NUMBER() OVER (
                    PARTITION BY item_id ORDER BY n_buyers DESC
                ) AS rk
            FROM market_counts
        )
        SELECT
            item_id,
            mkt_cd                                  AS primary_market,
            n_buyers                                AS primary_market_buyers
        FROM ranked WHERE rk = 1
    """).df()
    con.close()

    prod = prod.merge(prim_mkt, on="item_id", how="left")
    _log(f"  Joined primary market for {len(prod):,} products")

    # Filter to active inventory
    if MIN_ACTIVE_BUYERS > 1:
        before = len(prod)
        prod = prod[prod["distinct_buyers"] >= MIN_ACTIVE_BUYERS].copy()
        _log(f"  Filtered to >={MIN_ACTIVE_BUYERS} buyers: "
             f"{len(prod):,} (dropped {before - len(prod):,})")

    # Join product metadata for descriptions and brand info
    if PRODUCTS_FILE.exists():
        _log("Joining product metadata (descriptions, brand flags)...")
        prods_meta = pd.read_parquet(
            PRODUCTS_FILE,
            columns=["DIM_ITEM_E1_CURR_ID", "ITEM_E1_NUM", "ITEM_DSC",
                     "PROD_FMLY_LVL1_DSC", "PROD_CTGRY_LVL2_DSC",
                     "SUPLR_DSC", "SUPLR_ROLLUP_DSC",
                     "is_private_brand", "is_discontinued", "is_generic"]
        ).rename(columns={"DIM_ITEM_E1_CURR_ID": "item_id"})
        prod = prod.merge(prods_meta, on="item_id", how="left")
        _log(f"  After metadata join: {len(prod):,} products")

    # Convert date_id (YYYYMMDD) to readable date
    def _id_to_date(d):
        try:
            s = str(int(d))
            return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
        except Exception:
            return ""
    prod["last_sold_date"]  = prod["last_sold_date_id"].apply(_id_to_date)
    prod["first_sold_date"] = prod["first_sold_date_id"].apply(_id_to_date)

    # Flag Medline products explicitly
    if "SUPLR_ROLLUP_DSC" in prod.columns:
        prod["is_medline"] = (
            prod["SUPLR_ROLLUP_DSC"] == "MEDLINE INDUSTRIES"
        ).astype(int)
    else:
        prod["is_medline"] = 0

    # Compute rank by sales_count
    prod["rank_by_sales_count"] = prod["sales_count"].rank(
        method="min", ascending=False
    ).astype(int)
    prod["rank_by_revenue"] = prod["total_revenue"].rank(
        method="min", ascending=False
    ).astype(int)
    prod["rank_by_volume"] = prod["total_volume"].rank(
        method="min", ascending=False
    ).astype(int)
    prod["rank_by_buyers"] = prod["distinct_buyers"].rank(
        method="min", ascending=False
    ).astype(int)

    # Reorder columns for human-readable output
    col_order = [
        "rank_by_sales_count", "rank_by_revenue", "rank_by_volume", "rank_by_buyers",
        "item_id", "ITEM_E1_NUM", "ITEM_DSC",
        "PROD_FMLY_LVL1_DSC", "PROD_CTGRY_LVL2_DSC",
        "sales_count", "line_count",
        "distinct_buyers",
        "total_revenue", "total_volume",
        "avg_unit_price", "avg_line_revenue",
        "primary_market", "primary_market_buyers",
        "first_sold_date", "last_sold_date",
        "SUPLR_DSC", "SUPLR_ROLLUP_DSC",
        "is_private_brand", "is_medline", "is_discontinued", "is_generic",
    ]
    prod = prod[[c for c in col_order if c in prod.columns]]
    prod = prod.sort_values("rank_by_sales_count").reset_index(drop=True)

    _log(f"  Top product: {prod.iloc[0].get('ITEM_DSC', 'N/A')[:50]}")
    _log(f"  Top sales count: {prod.iloc[0]['sales_count']:,}")
    _log(f"  Top revenue:     ${prod.iloc[0]['total_revenue']:,.0f}")

    return prod


# Step 3: Save outputs

def step3_save_outputs(customers_df, products_df):
    _section("Step 3: Saving outputs")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Save CSVs first - these are fast to load anywhere and don't have row limits
    cust_csv = OUT_DIR / "customer_ranking_full.csv"
    prod_csv = OUT_DIR / "product_ranking_full.csv"
    customers_df.to_csv(cust_csv, index=False)
    products_df.to_csv(prod_csv, index=False)
    _log(f"  Saved: {cust_csv.name}  ({cust_csv.stat().st_size / 1024 / 1024:.1f} MB)")
    _log(f"  Saved: {prod_csv.name}  ({prod_csv.stat().st_size / 1024 / 1024:.1f} MB)")

    # Build top 100 summaries for the Excel workbook
    top_100_customers = customers_df.head(100).copy()
    top_100_products  = products_df.head(100).copy()

    # Build size_tier summary
    if "size_tier" in customers_df.columns:
        tier_summary = customers_df.groupby("size_tier").agg(
            n_customers       = ("DIM_CUST_CURR_ID", "count"),
            total_spend       = ("lifetime_spend",   "sum"),
            total_volume      = ("lifetime_volume",  "sum"),
            avg_spend         = ("lifetime_spend",   "mean"),
            avg_volume        = ("lifetime_volume",  "mean"),
            median_spend      = ("lifetime_spend",   "median"),
            median_volume     = ("lifetime_volume",  "median"),
        ).round(2).reset_index()
        # Sort by tier order: enterprise > large > mid > small > new
        tier_order = ["enterprise", "large", "mid", "small", "new"]
        tier_summary["_ord"] = tier_summary["size_tier"].map(
            {t: i for i, t in enumerate(tier_order)}
        ).fillna(99)
        tier_summary = tier_summary.sort_values("_ord").drop(columns="_ord")
    else:
        tier_summary = pd.DataFrame()

    # Build product family summary
    if "PROD_FMLY_LVL1_DSC" in products_df.columns:
        family_summary = products_df.groupby("PROD_FMLY_LVL1_DSC").agg(
            n_products        = ("item_id",        "count"),
            total_sales_count = ("sales_count",    "sum"),
            total_revenue     = ("total_revenue",  "sum"),
            total_volume      = ("total_volume",   "sum"),
            total_buyers      = ("distinct_buyers","sum"),
            avg_unit_price    = ("avg_unit_price", "mean"),
        ).round(2).reset_index()
        family_summary = family_summary.sort_values("total_revenue", ascending=False)
    else:
        family_summary = pd.DataFrame()

    # Top 1000 in Excel for browsability (full data lives in CSV)
    top_1000_customers = customers_df.head(1000).copy()
    top_1000_products  = products_df.head(1000).copy()

    # Build summary statistics sheet
    summary_rows = [
        ("Total customers",               f"{len(customers_df):,}"),
        ("Total products in inventory",   f"{len(products_df):,}"),
        ("Customers with $0 lifetime",
            f"{int((customers_df['lifetime_spend'] == 0).sum()):,}"),
        ("Customers with 0 volume",
            f"{int((customers_df['lifetime_volume'] == 0).sum()):,}"),
        ("Total lifetime revenue",
            f"${customers_df['lifetime_spend'].sum():,.0f}"),
        ("Total lifetime volume",
            f"{customers_df['lifetime_volume'].sum():,.0f}"),
        ("Top customer lifetime spend",
            f"${customers_df['lifetime_spend'].max():,.0f}"),
        ("Top product sales count",
            f"{products_df['sales_count'].max():,}"),
        ("Top product revenue",
            f"${products_df['total_revenue'].max():,.0f}"),
        ("Active product threshold",
            f">= {MIN_ACTIVE_BUYERS} buyers"),
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["metric", "value"])

    # Excel workbook with multiple sheets for browsing
    xlsx_path = OUT_DIR / "rankings_report.xlsx"
    _log(f"  Building Excel workbook (this can take 30-60s)...")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="01_summary", index=False)
        _style(writer.sheets["01_summary"], summary_df, "1F4E79")

        if not tier_summary.empty:
            tier_summary.to_excel(writer, sheet_name="02_by_size_tier", index=False)
            _style(writer.sheets["02_by_size_tier"], tier_summary, "375623")

        top_100_customers.to_excel(writer, sheet_name="03_top100_customers", index=False)
        _style(writer.sheets["03_top100_customers"], top_100_customers, "C00000")

        top_1000_customers.to_excel(writer, sheet_name="04_top1000_customers", index=False)
        _style(writer.sheets["04_top1000_customers"], top_1000_customers, "C00000")

        top_100_products.to_excel(writer, sheet_name="05_top100_products", index=False)
        _style(writer.sheets["05_top100_products"], top_100_products, "1F6B75")

        top_1000_products.to_excel(writer, sheet_name="06_top1000_products", index=False)
        _style(writer.sheets["06_top1000_products"], top_1000_products, "1F6B75")

        if not family_summary.empty:
            family_summary.to_excel(writer, sheet_name="07_by_product_family", index=False)
            _style(writer.sheets["07_by_product_family"], family_summary, "7030A0")

    _log(f"  Saved: {xlsx_path.name}  ({xlsx_path.stat().st_size / 1024 / 1024:.1f} MB)")

    return [xlsx_path, cust_csv, prod_csv]


# Main

def main():
    print()
    print("=" * 70)
    print("  CUSTOMER + PRODUCT RANKING REPORT")
    print("=" * 70)
    start = time.time()

    customers_df = step1_rank_customers()
    products_df  = step2_rank_products()
    out_files    = step3_save_outputs(customers_df, products_df)

    elapsed = round(time.time() - start, 1)
    _section(f"Complete in {elapsed}s")
    print()
    print(f"  Output directory: {OUT_DIR}")
    print()
    print(f"  Files produced:")
    for f in out_files:
        size_mb = f.stat().st_size / 1024 / 1024
        print(f"    {f.name}  ({size_mb:.1f} MB)")
    print()
    print(f"  The Excel workbook has 7 sheets:")
    print(f"    01_summary              - high-level stats")
    print(f"    02_by_size_tier         - aggregated by size_tier")
    print(f"    03_top100_customers     - top 100 customers (browsable)")
    print(f"    04_top1000_customers    - top 1000 customers")
    print(f"    05_top100_products      - top 100 products by sales count")
    print(f"    06_top1000_products     - top 1000 products")
    print(f"    07_by_product_family    - aggregated by product family")
    print()
    print(f"  Full rankings (all {len(customers_df):,} customers and "
          f"{len(products_df):,} products) are in the CSV files.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL: {exc}", file=sys.stderr)
        raise