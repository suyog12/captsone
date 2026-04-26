from __future__ import annotations

import datetime
import shutil
import sys
import time
from pathlib import Path
from typing import Optional

import duckdb
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Path configuration

ROOT = Path(__file__).resolve().parent.parent.parent

DATA_RAW   = ROOT / "data_raw"
DATA_CLEAN = ROOT / "data_clean"

OUT_CUSTOMER = DATA_CLEAN / "customer"
OUT_PRODUCT  = DATA_CLEAN / "product"
OUT_SALES    = DATA_CLEAN / "sales"
OUT_FEATURES = DATA_CLEAN / "features"
OUT_SERVING  = DATA_CLEAN / "serving"
OUT_AUDIT    = DATA_CLEAN / "audit"


# Pipeline configuration

EXCLUDED_ORDER_NUMS: frozenset[int] = frozenset({61408700, 61737401, 35996955})

MAX_ORDER_LINE_VALUE: float = 10_000_000.0
TIER_B_THRESHOLD: float = 50_000.0
TIER_C_THRESHOLD: float = 500_000.0

TIER1_MIN_CUSTOMERS: int = 1_000
TIER2_MIN_CUSTOMERS: int = 10

STATE_MIN_CUSTOMERS: int = 500

# Supplier profile thresholds
MEDLINE_SUPPLIER_NAME: str = "MEDLINE INDUSTRIES"
MEDLINE_THRESHOLD: float = 0.50
MCKESSON_PRIVATE_THRESHOLD: float = 0.50

# Size tier thresholds (median monthly spend in dollars)
SIZE_TIER_SMALL_MAX  = 500
SIZE_TIER_MID_MAX    = 2_500
SIZE_TIER_LARGE_MAX  = 15_000

# New customer definition
MIN_ACTIVE_MONTHS_FOR_TIER = 2

# Affordability ceiling multipliers by size tier
AFFORDABILITY_MULTIPLIER = {
    "new":        3.0,
    "small":      1.5,
    "mid":        1.8,
    "large":      2.0,
    "enterprise": 2.5,
}

_DROP_NULL_CUSTOMER: list[str] = [
    "BID_NUM", "CCS_CD", "CCS_DSC", "CCS_DT",
    "EC_ACCT_SPECIF_BID_CD", "EC_ACCT_TYPE_CD", "EC_BUYING_GRP_NAME",
    "EC_BUYING_GRP_NUM", "EC_BUY_PLAN_CD", "EC_HNDLNG_CD",
    "EC_PROMO_BID_CD", "EC_SCNDRY_BID_CD", "EC_SPR_GRP_NAME",
    "EC_SPR_GRP_NUM", "MDM_GUID_ID", "MDM_PARTY_ID",
    "MSTR_GRP_ADMIN_NAME", "MSTR_GRP_CD", "MSTR_GRP_NAME",
    "MSTR_GRP_NUM", "MSTR_GRP_TYPE_CD", "PYMNT_TERMS_EC_CD",
    "PYMNT_TERMS_EC_DSC", "RPT_NAMING_CD", "SITE_ADDRSS_STATE",
]

_DROP_NULL_PRODUCT: list[str] = [
    "BY_TRANS_MODE", "CYP_BRND_CD", "GHX_MAJOR_CD", "GHX_MAJOR_DSC",
    "GHX_MATCH_LEVEL_CD", "GHX_MINOR_CD", "GHX_MINOR_DSC",
    "HPIS_SKU", "TARIFF_CD", "TRACEABLE_TYPE",
]

_DROP_SPARSE_CUSTOMER: list[str] = [
    "GOVT_GOTIT_CNTRCT_NUM", "SISTER_340B_TYPE_CD", "CUST_SUPPLD_ACCT_NUM",
    "CUST_SPCLTY_ORG_CD",    "CUST_SPCLTY_ORG_DSC",
    "DEA_LIC_EXPR_DT",       "DEA_LIC_NUM",
    "ADDRSS_LINE3",          "MED_LIC_STATE_CD",    "CCS_E1_CD",
    "CNVRSN_DT",             "FRMLRY_CTLG_CD",      "FRMLRY_CTLG_DSC",
    "TIER_CD",               "TIER_DSC",
    "GRP_ADMIN_NAME",        "SUPER_GRP_ADMIN_NAME",
    "EDI_TRDNG_PRTNR_ID",    "GOVT_CLASS_CD",
]

_DROP_SPARSE_PRODUCT: list[str] = [
    "FED_HZRD_CD",       "EC_VIP_COST_AMT",   "MAC_COST_AMT",
    "GCN_DOSE_FRM",      "GCN_NUM",            "PHARMA_ITEM_NUM",
    "INBND_PEDIGREE_CD", "ITEM_STATUS_CD",     "ITEM_STATUS_DSC",
    "NDC_NUM",           "ALT_CTLG_NUM",
    "MFG_STATUS_CD",     "MFG_STATUS_DSC",     "MFG_BKORDR_DSC",
    "MFG_BKORDR_DUE_DT", "MFG_BKORDR_REL_DT",
    "PROD_TYPE_CD",      "PROD_TYPE_DSC",
]

KEEP_CUSTOMER_COLS: list[str] = [
    "DIM_CUST_CURR_ID", "CUST_NUM",      "CUST_NAME",
    "CUST_TYPE_CD",     "CUST_TYPE_DSC",
    "SPCLTY_CD",        "SPCLTY_DSC",
    "MKT_CD",
    "MMS_CLASS_CD",     "MMS_CLASS_DSC",
    "MMS_SGMNT_CD",     "MMS_SUB_CLASS_CD",
    "ACTV_FLG",
    "ZIP",              "STATE",         "CITY",  "CNTRY_CD",
]

KEEP_PRODUCT_COLS: list[str] = [
    "DIM_ITEM_E1_CURR_ID", "ITEM_E1_NUM",           "ITEM_DSC",
    "PRVT_BRND_FLG",
    "PROD_FMLY_LVL1_CD",   "PROD_FMLY_LVL1_DSC",
    "PROD_CTGRY_LVL2_CD",  "PROD_CTGRY_LVL2_DSC",
    "PROD_GRP_LVL3_CD",    "PROD_GRP_LVL3_DSC",
    "PROD_SUB_CTGRY_LVL4_CD", "PROD_SUB_CTGRY_LVL4_DSC",
    "SUPLR_DSC",           "SUPLR_ROLLUP_DSC",
    "SLS_GRP_CD",          "SLS_GRP_DSC",
    "GNRC_FLG",            "DISCTND_FLG",
]

_SALES_KEEP_COLS: list[str] = [
    "DIM_CUST_CURR_ID", "DIM_ITEM_E1_CURR_ID",
    "ORDR_NUM", "ORDR_LINE_NUM",
    "ORDR_QTY", "UNIT_SLS_AMT",
    "DIM_ORDR_DT_ID",
    "ORDR_MTHD_DSC", "ORDR_SRC_DSC",
    "MCK_FLG", "INV_NUM", "SHIP_QTY", "PRMRY_QTY",
]


# Logging utilities

def _section(title: str) -> None:
    print(f"\n{'─' * 62}", flush=True)
    print(f"  {title}", flush=True)
    print(f"{'─' * 62}", flush=True)


def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


# DuckDB helpers

def _con(memory_gb: int = 4, threads: int = 2, spill_dir: Optional[Path] = None) -> duckdb.DuckDBPyConnection:
    con = duckdb.connect(":memory:")
    con.execute(f"SET memory_limit = '{memory_gb}GB'")
    con.execute(f"SET threads = {threads}")
    con.execute("SET preserve_insertion_order = false")
    if spill_dir:
        con.execute(f"SET temp_directory = '{spill_dir.as_posix()}'")
    return con


def _disk_con(db_path: Path, memory_gb: int = 4, threads: int = 2,
              spill_dir: Optional[Path] = None) -> duckdb.DuckDBPyConnection:
    con = duckdb.connect(str(db_path))
    con.execute(f"SET memory_limit = '{memory_gb}GB'")
    con.execute(f"SET threads = {threads}")
    con.execute("SET preserve_insertion_order = false")
    if spill_dir:
        con.execute(f"SET temp_directory = '{spill_dir.as_posix()}'")
    return con


def _pq(path: Path) -> str:
    return "'" + path.as_posix() + "'"


def _glob(folder: Path) -> str:
    return "'" + folder.as_posix() + "/*.parquet'"


# DataFrame utilities

def _drop_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    return df.drop(columns=[c for c in cols if c in df.columns])


def _keep_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    return df[[c for c in cols if c in df.columns]]


def _flag_to_int(series: pd.Series) -> pd.Series:
    return (
        series.astype(str).str.strip().str.upper()
        .map({"Y": 1, "1": 1, "TRUE": 1, "T": 1, "YES": 1})
        .fillna(0).astype(int)
    )


# Path discovery

def _discover_folders() -> dict:
    if not DATA_RAW.exists():
        raise FileNotFoundError(
            f"data_raw/ not found at {DATA_RAW}\n"
            "Rename the source Data folder to data_raw before running."
        )

    cust_folder: Optional[Path] = None
    item_folder: Optional[Path] = None
    fact_folders: dict[str, Path] = {}

    for folder in sorted(DATA_RAW.iterdir()):
        if not folder.is_dir():
            continue
        n = folder.name.lower()
        if "dim" in n and "cust" in n:
            cust_folder = folder
        elif "dim" in n and "item" in n:
            item_folder = folder
        elif ("fct" in n or "sales" in n) and "dim" not in n:
            fy = "FY2425" if "2425" in n else "FY2526"
            fact_folders[fy] = folder

    if not cust_folder:
        raise FileNotFoundError("Customer dimension folder not found in data_raw/")
    if not item_folder:
        raise FileNotFoundError("Product dimension folder not found in data_raw/")
    if not fact_folders:
        raise FileNotFoundError("Sales fact folders not found in data_raw/")

    return {"cust_folder": cust_folder, "item_folder": item_folder,
            "fact_folders": fact_folders}


# Step 1: Validate source paths

def step1_validate(folders: dict) -> None:
    _section("Step 1: Validating source paths")
    _log(f"Customer  : {folders['cust_folder'].name}  "
         f"({len(list(folders['cust_folder'].glob('*.parquet')))} files)")
    _log(f"Product   : {folders['item_folder'].name}  "
         f"({len(list(folders['item_folder'].glob('*.parquet')))} files)")
    for fy, folder in sorted(folders["fact_folders"].items()):
        _log(f"Sales {fy}: {folder.name}  "
             f"({len(list(folder.glob('*.parquet')))} files)")


# Step 2: Clean customer dimension

def step2_clean_customers(
    cust_folder: Path,
) -> tuple[pd.DataFrame, list[str], list[str], pd.DataFrame]:
    _section("Step 2: Cleaning customer dimension")

    con = _con(memory_gb=4)
    df  = con.execute(f"SELECT * FROM read_parquet({_glob(cust_folder)})").df()
    con.close()

    raw_cols = list(df.columns)
    _log(f"Loaded : {len(df):>12,} rows  |  {len(df.columns)} columns")

    n_before = len(df.columns)
    df = _drop_cols(df, _DROP_NULL_CUSTOMER + _DROP_SPARSE_CUSTOMER)
    _log(f"Dropped: {n_before - len(df.columns)} null/sparse columns")

    df = _keep_cols(df, KEEP_CUSTOMER_COLS)
    _log(f"Kept   : {len(df.columns)} ML key columns")

    n_before = len(df)
    df = df.drop_duplicates(subset=["DIM_CUST_CURR_ID"])
    _log(f"Dedup  : {n_before - len(df):,} duplicate rows removed")

    null_zip_df = pd.DataFrame()
    if "ZIP" in df.columns:
        df["_zip_raw"] = df["ZIP"].copy()

        def _parse_zip(v) -> Optional[str]:
            if v is None:
                return None
            s = str(v).strip()
            if s.endswith(".0"):
                s = s[:-2]
            if s.lower() in ("nan", "none", "nat", "null", ""):
                return None
            digits = "".join(c for c in s if c.isdigit())
            if len(digits) == 4:
                return digits.zfill(5)
            if len(digits) in (5, 9):
                return digits
            return None

        df["ZIP"]    = df["ZIP"].apply(_parse_zip)
        null_mask    = df["ZIP"].isna()
        null_zip_df  = df.loc[null_mask, [
            c for c in ["DIM_CUST_CURR_ID", "CUST_TYPE_CD",
                        "SPCLTY_CD", "STATE", "CITY", "_zip_raw"]
            if c in df.columns
        ]].rename(columns={"_zip_raw": "zip_original"}).copy()
        df = df.drop(columns=["_zip_raw"])

        _log(f"ZIP    : {(~null_mask).sum():,} valid  |  "
             f"{null_mask.sum():,} invalid set to None")

    if "STATE" in df.columns:
        df["STATE"] = (
            df["STATE"].astype(str).str.strip().str.upper()
            .where(lambda s: s.str.match(r"^[A-Z]{2}$"), other=None)
        )

    if "ACTV_FLG" in df.columns:
        raw = df["ACTV_FLG"].astype(str).str.strip().str.upper()
        df["ACTV_FLG"] = raw.map(
            lambda v: "Y" if v in ("Y", "1", "TRUE", "T", "YES") else
                      "N" if v in ("N", "0", "FALSE", "F", "NO") else None
        )

    if "CUST_TYPE_CD" in df.columns:
        df["CUST_TYPE_CD"] = df["CUST_TYPE_CD"].where(
            df["CUST_TYPE_CD"].isin(["S", "X", "B"]), other=None
        )
        n_bad = df["CUST_TYPE_CD"].isna().sum()
        _log(f"Type   : {(~df['CUST_TYPE_CD'].isna()).sum():,} valid S/X/B  |  "
             f"{n_bad:,} non-standard set to None")

    _log(f"Result : {len(df):>12,} rows  |  {len(df.columns)} columns")
    df.to_parquet(OUT_CUSTOMER / "customers_clean.parquet", index=False)
    _log(f"Saved  : {(OUT_CUSTOMER / 'customers_clean.parquet').relative_to(ROOT)}")

    return df, raw_cols, list(df.columns), null_zip_df


# Step 3: Clean product dimension

def step3_clean_products(
    item_folder: Path,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    _section("Step 3: Cleaning product dimension")

    con = _con(memory_gb=4)
    df  = con.execute(f"SELECT * FROM read_parquet({_glob(item_folder)})").df()
    con.close()

    raw_cols = list(df.columns)
    _log(f"Loaded : {len(df):>12,} rows  |  {len(df.columns)} columns")

    n_before = len(df.columns)
    df = _drop_cols(df, _DROP_NULL_PRODUCT + _DROP_SPARSE_PRODUCT)
    _log(f"Dropped: {n_before - len(df.columns)} null/sparse columns")

    df = _keep_cols(df, KEEP_PRODUCT_COLS)
    _log(f"Kept   : {len(df.columns)} ML key columns")

    n_before = len(df)
    df = df.drop_duplicates(subset=["DIM_ITEM_E1_CURR_ID"])
    _log(f"Dedup  : {n_before - len(df):,} duplicate rows removed")

    for src, dst in [
        ("PRVT_BRND_FLG", "is_private_brand"),
        ("DISCTND_FLG",   "is_discontinued"),
        ("GNRC_FLG",      "is_generic"),
    ]:
        if src in df.columns:
            df[dst] = _flag_to_int(df[src])

    _log(f"Result : {len(df):>12,} rows  |  {len(df.columns)} columns")
    df.to_parquet(OUT_PRODUCT / "products_clean.parquet", index=False)
    _log(f"Saved  : {(OUT_PRODUCT / 'products_clean.parquet').relative_to(ROOT)}")

    return df, raw_cols, list(df.columns)


# Step 4: Clean sales facts

def step4_clean_sales(
    fact_folders: dict[str, Path],
) -> tuple[pd.DataFrame, list[dict], pd.DataFrame]:
    _section("Step 4: Cleaning sales facts")
    _log(f"Blocked orders: {', '.join(str(o) for o in sorted(EXCLUDED_ORDER_NUMS))}")
    _log(f"Revenue cap   : ${MAX_ORDER_LINE_VALUE:,.0f} per order line")

    stage_dir   = OUT_SALES / "_stage"
    stage_dir.mkdir(parents=True, exist_ok=True)

    col_sql     = ", ".join(f'"{c}"' for c in _SALES_KEEP_COLS)
    blocked_sql = ", ".join(str(o) for o in EXCLUDED_ORDER_NUMS)

    duplicate_log:   list[dict]         = []
    excluded_rows:   list[pd.DataFrame] = []
    grand_raw = grand_dupes = grand_blocked = grand_capped = grand_returns = 0

    fiscal_years = sorted(fact_folders.keys())

    # 4A: per-file cleaning

    for fy, folder in ((fy, fact_folders[fy]) for fy in fiscal_years):
        _log(f"\n  Fiscal year: {fy}  ({folder.name})")
        files = sorted(folder.glob("*.parquet"))

        for idx, fp in enumerate(files, 1):
            print(f"    [{idx:3d}/{len(files)}] {fp.name:<52s}", end="  ", flush=True)

            db_path = stage_dir / f"_work_{fy}_{idx}.duckdb"
            con     = _disk_con(db_path, memory_gb=4, threads=1, spill_dir=stage_dir)
            fp_q    = _pq(fp)

            raw = con.execute(f"SELECT COUNT(*) FROM read_parquet({fp_q})").fetchone()[0]
            grand_raw += raw

            r = con.execute(f"""
                SELECT
                    COUNT(*) FILTER (WHERE ORDR_NUM IN ({blocked_sql}))    AS blocked,
                    COUNT(*) FILTER (WHERE ORDR_NUM NOT IN ({blocked_sql})
                                       AND UNIT_SLS_AMT IS NOT NULL
                                       AND UNIT_SLS_AMT > {MAX_ORDER_LINE_VALUE}) AS capped
                FROM read_parquet({fp_q})
            """).fetchone()
            n_blocked, n_capped = int(r[0]), int(r[1])
            grand_blocked += n_blocked
            grand_capped  += n_capped

            if n_blocked + n_capped > 0:
                excl = con.execute(f"""
                    SELECT '{fy}' AS fiscal_year, ORDR_NUM, UNIT_SLS_AMT,
                        CASE WHEN ORDR_NUM IN ({blocked_sql})
                             THEN 'Blocked order number'
                             ELSE 'Revenue cap exceeded'
                        END AS exclusion_reason
                    FROM read_parquet({fp_q})
                    WHERE ORDR_NUM IN ({blocked_sql})
                       OR (UNIT_SLS_AMT IS NOT NULL
                           AND UNIT_SLS_AMT > {MAX_ORDER_LINE_VALUE})
                """).df()
                excluded_rows.append(excl)

            clean_inner = f"""
                WITH projected AS (
                    SELECT {col_sql}
                    FROM   read_parquet({fp_q})
                    WHERE  ORDR_NUM NOT IN ({blocked_sql})
                      AND  (UNIT_SLS_AMT IS NULL
                            OR UNIT_SLS_AMT <= {MAX_ORDER_LINE_VALUE})
                ),
                deduped AS (
                    SELECT * EXCLUDE (_rn)
                    FROM (
                        SELECT *,
                            ROW_NUMBER() OVER (
                                PARTITION BY ORDR_NUM, ORDR_LINE_NUM
                                ORDER BY DIM_ORDR_DT_ID DESC
                            ) AS _rn
                        FROM projected
                    ) WHERE _rn = 1
                ),
                enriched AS (
                    SELECT *,
                        '{fy}'                                          AS fiscal_year,
                        CAST(DIM_ORDR_DT_ID / 10000       AS INTEGER)  AS order_year,
                        CAST((DIM_ORDR_DT_ID % 10000)/100 AS INTEGER)  AS order_month,
                        CAST(DIM_ORDR_DT_ID % 100         AS INTEGER)  AS order_day
                    FROM deduped
                )
                SELECT * FROM enriched
            """

            txn_path = stage_dir / f"txn_{fy}_{idx:04d}.parquet"
            con.execute(f"""
                COPY (
                    SELECT * FROM ({clean_inner}) t WHERE ORDR_QTY >= 0
                ) TO {_pq(txn_path)} (FORMAT PARQUET, COMPRESSION SNAPPY)
            """)

            ret_count = con.execute(
                f"SELECT COUNT(*) FROM ({clean_inner}) t WHERE ORDR_QTY < 0"
            ).fetchone()[0]

            if ret_count > 0:
                ret_path = stage_dir / f"ret_{fy}_{idx:04d}.parquet"
                con.execute(f"""
                    COPY (
                        SELECT * FROM ({clean_inner}) t WHERE ORDR_QTY < 0
                    ) TO {_pq(ret_path)} (FORMAT PARQUET, COMPRESSION SNAPPY)
                """)
                grand_returns += ret_count

            clean_count = con.execute(
                f"SELECT COUNT(*) FROM read_parquet({_pq(txn_path)})"
            ).fetchone()[0]

            dupes = max(0, raw - n_blocked - n_capped - clean_count - ret_count)
            grand_dupes += dupes

            duplicate_log.append({
                "fiscal_year":        fy,
                "file_name":          fp.name,
                "raw_rows":           raw,
                "blocked_removed":    n_blocked,
                "cap_removed":        n_capped,
                "duplicates_removed": dupes,
                "clean_rows":         clean_count,
            })

            con.close()
            db_path.unlink(missing_ok=True)

            print(f"raw={raw:,}  clean={clean_count:,}  "
                  f"dupes={dupes}  blocked={n_blocked}  capped={n_capped}",
                  flush=True)

    _log(f"\n  Raw rows loaded     : {grand_raw:>14,}")
    _log(f"  Duplicates removed  : {grand_dupes:>14,}")
    _log(f"  Blocked orders      : {grand_blocked:>14,}")
    _log(f"  Cap exclusions      : {grand_capped:>14,}")
    _log(f"  Returns separated   : {grand_returns:>14,}")

    # 4B: cross-file dedup

    fy_txn_paths: list[Path] = []
    fy_ret_paths: list[Path] = []
    total_cross  = 0
    total_final  = 0

    _dedup_key_cols  = ["ORDR_NUM", "ORDR_LINE_NUM"]
    _dedup_agg_cols  = [
        c for c in _SALES_KEEP_COLS if c not in _dedup_key_cols
    ] + ["fiscal_year", "order_year", "order_month", "order_day"]

    agg_sql = ",\n            ".join(
        f'MAX("{c}") AS "{c}"' for c in _dedup_agg_cols
    )

    for fy in fiscal_years:
        _log(f"\n  Cross-file dedup: {fy}  (GROUP BY strategy)...")

        fy_stage_files = sorted(stage_dir.glob(f"txn_{fy}_*.parquet"))
        if not fy_stage_files:
            continue

        file_list_sql = ", ".join(f"\'{p.as_posix()}\'" for p in fy_stage_files)
        fy_out        = OUT_SALES / f"transactions_clean_{fy}.parquet"
        db_path       = stage_dir / f"_dedup_{fy}.duckdb"

        con = _disk_con(db_path, memory_gb=4, threads=1, spill_dir=stage_dir)

        before = con.execute(
            f"SELECT COUNT(*) FROM read_parquet([{file_list_sql}])"
        ).fetchone()[0]

        con.execute(f"""
            COPY (
                SELECT
                    "ORDR_NUM",
                    "ORDR_LINE_NUM",
                    {agg_sql}
                FROM read_parquet([{file_list_sql}])
                GROUP BY "ORDR_NUM", "ORDR_LINE_NUM"
            ) TO {_pq(fy_out)} (FORMAT PARQUET, COMPRESSION SNAPPY)
        """)

        after       = con.execute(
            f"SELECT COUNT(*) FROM read_parquet({_pq(fy_out)})"
        ).fetchone()[0]
        cross_dupes = before - after
        total_cross += cross_dupes
        total_final += after

        if cross_dupes > 0:
            _log(f"    {fy} cross-file duplicates removed: {cross_dupes:,}")
        _log(f"    {fy} final transactions: {after:,}")

        fy_txn_paths.append(fy_out)
        con.close()
        db_path.unlink(missing_ok=True)

        ret_stage_files = sorted(stage_dir.glob(f"ret_{fy}_*.parquet"))
        if ret_stage_files:
            ret_list_sql = ", ".join(f"\'{p.as_posix()}\'" for p in ret_stage_files)
            ret_out      = OUT_SALES / f"returns_clean_{fy}.parquet"
            ret_con      = _con(memory_gb=2, threads=1)
            ret_con.execute(f"""
                COPY (SELECT * FROM read_parquet([{ret_list_sql}]))
                TO {_pq(ret_out)} (FORMAT PARQUET, COMPRESSION SNAPPY)
            """)
            ret_con.close()
            fy_ret_paths.append(ret_out)

    if total_cross > 0:
        _log(f"  Total cross-file duplicates: {total_cross:,}")
    _log(f"  Total final transactions  : {total_final:,}")

    # 4C: build returns/excluded audit DataFrames

    returns = pd.concat(
        [pd.read_parquet(p) for p in fy_ret_paths], ignore_index=True
    ) if fy_ret_paths else pd.DataFrame()

    excluded_df = (
        pd.concat(excluded_rows, ignore_index=True)
        if excluded_rows else pd.DataFrame()
    )

    for p in fy_txn_paths:
        _log(f"Saved  : {p.relative_to(ROOT)}")
    for p in fy_ret_paths:
        _log(f"Saved  : {p.relative_to(ROOT)}")

    try:
        shutil.rmtree(stage_dir)
    except Exception:
        pass

    return returns, duplicate_log, excluded_df


# Step 5: Specialty tier assignment

def step5_specialty_tiers(
    fact_paths: list[Path],
    customers:  pd.DataFrame,
) -> pd.DataFrame:
    _section("Step 5: Assigning specialty tiers")

    cust_spec_path = OUT_FEATURES / "_tmp_cust_spec.parquet"
    (customers[["DIM_CUST_CURR_ID", "SPCLTY_CD", "SPCLTY_DSC"]]
     .dropna(subset=["SPCLTY_CD"])
     .to_parquet(cust_spec_path, index=False))

    txn_paths_sql = ", ".join(f"'{p.as_posix()}'" for p in fact_paths)

    con = _con(memory_gb=6, threads=2)
    spec_stats = con.execute(f"""
        SELECT
            c.SPCLTY_CD,
            c.SPCLTY_DSC,
            COUNT(DISTINCT s.DIM_CUST_CURR_ID) AS unique_customers,
            COUNT(DISTINCT s.ORDR_NUM)          AS unique_orders,
            ROUND(SUM(s.UNIT_SLS_AMT), 2)       AS total_revenue
        FROM read_parquet([{txn_paths_sql}]) s
        JOIN read_parquet({_pq(cust_spec_path)}) c
          ON s.DIM_CUST_CURR_ID = c.DIM_CUST_CURR_ID
        GROUP BY c.SPCLTY_CD, c.SPCLTY_DSC
        ORDER BY unique_customers DESC
    """).df()
    con.close()
    cust_spec_path.unlink(missing_ok=True)

    total_rev = spec_stats["total_revenue"].sum()
    spec_stats["pct_of_total_revenue"]  = (spec_stats["total_revenue"] / total_rev * 100).round(4)
    spec_stats["avg_revenue_per_order"] = (
        spec_stats["total_revenue"] /
        spec_stats["unique_orders"].replace(0, pd.NA)
    ).round(2)

    def _tier(n: int) -> int:
        if n >= TIER1_MIN_CUSTOMERS: return 1
        if n >= TIER2_MIN_CUSTOMERS: return 2
        return 3

    spec_stats["specialty_tier"] = spec_stats["unique_customers"].apply(_tier)

    tier1 = spec_stats[spec_stats["specialty_tier"] == 1].copy()
    tier1["_first_word"] = tier1["SPCLTY_DSC"].str.split().str[0].str.upper()
    fw_map     = tier1.set_index("_first_word")["SPCLTY_CD"].to_dict()
    t1_words   = {
        row["SPCLTY_CD"]: set(str(row["SPCLTY_DSC"]).upper().split())
        for _, row in tier1.iterrows()
    }

    def _fallback(dsc) -> Optional[str]:
        if pd.isna(dsc) or not str(dsc).strip():
            return None
        dsc_upper = str(dsc).upper()
        words     = dsc_upper.split()

        hit = fw_map.get(words[0]) if words else None
        if hit:
            return hit

        dsc_word_set = set(words)
        for cd, t1_word_set in t1_words.items():
            if dsc_word_set & t1_word_set:
                return cd

        if words:
            for cd, t1_word_set in t1_words.items():
                if words[0] in t1_word_set:
                    return cd

        return None

    t3 = spec_stats["specialty_tier"] == 3
    spec_stats["tier3_fallback_spclty_cd"] = None
    spec_stats.loc[t3, "tier3_fallback_spclty_cd"] = (
        spec_stats.loc[t3, "SPCLTY_DSC"].apply(_fallback)
    )

    counts = spec_stats.groupby("specialty_tier").size()
    mapped = spec_stats.loc[t3, "tier3_fallback_spclty_cd"].notna().sum()
    _log(f"Tier 1 (>={TIER1_MIN_CUSTOMERS:,}): {counts.get(1,0):>4} specialties")
    _log(f"Tier 2 ({TIER2_MIN_CUSTOMERS}-{TIER1_MIN_CUSTOMERS-1})       : {counts.get(2,0):>4} specialties")
    _log(f"Tier 3 (<{TIER2_MIN_CUSTOMERS})            : {counts.get(3,0):>4} specialties  ({mapped} mapped to Tier 1)")

    out = OUT_FEATURES / "specialty_tiers.parquet"
    spec_stats.to_parquet(out, index=False)
    _log(f"Saved  : {out.relative_to(ROOT)}")

    return spec_stats


# Step 6: RFM scoring and churn labelling

def step6_rfm(
    fact_paths: list[Path],
    customers:  pd.DataFrame,
) -> pd.DataFrame:
    _section("Step 6: Calculating RFM scores, churn labels, and behavioural features")

    fy25_path = next((p for p in fact_paths if "FY2425" in p.name), None)
    fy26_path = next((p for p in fact_paths if "FY2526" in p.name), None)

    if not fy25_path or not fy26_path:
        raise FileNotFoundError("Could not locate per-FY transaction files for RFM.")

    con = _con(memory_gb=4, threads=2)

    ref_id = int(con.execute(f"""
        SELECT MAX(DIM_ORDR_DT_ID)
        FROM read_parquet(['{fy25_path.as_posix()}', '{fy26_path.as_posix()}'])
    """).fetchone()[0])
    ref_y  = ref_id // 10000
    ref_m  = (ref_id % 10000) // 100
    ref_d  = ref_id % 100
    _log(f"Reference date (dataset max): {ref_y}-{ref_m:02d}-{ref_d:02d}")

    fy25_ids = set(con.execute(
        f"SELECT DISTINCT DIM_CUST_CURR_ID FROM read_parquet('{fy25_path.as_posix()}')"
    ).df()["DIM_CUST_CURR_ID"])

    fy26_ids = set(con.execute(
        f"SELECT DISTINCT DIM_CUST_CURR_ID FROM read_parquet('{fy26_path.as_posix()}')"
    ).df()["DIM_CUST_CURR_ID"])

    churned  = fy25_ids - fy26_ids
    retained = fy25_ids & fy26_ids
    new_in26 = fy26_ids - fy25_ids

    _log(f"FY2425 customers     : {len(fy25_ids):,}")
    _log(f"FY2526 customers     : {len(fy26_ids):,}")
    _log(f"Churned  (label=1)   : {len(churned):,}")
    _log(f"Retained (label=0)   : {len(retained):,}")
    _log(f"New FY2526 (label=-1): {len(new_in26):,}")

    fact_cols_df = con.execute(
        f"DESCRIBE SELECT * FROM read_parquet('{fy25_path.as_posix()}') LIMIT 0"
    ).df()
    fact_cols = set(fact_cols_df["column_name"].tolist())

    if "PROD_FMLY_LVL1_DSC" in fact_cols:
        family_col = "PROD_FMLY_LVL1_DSC"
    elif "PROD_FMLY_LVL1_CD" in fact_cols:
        family_col = "PROD_FMLY_LVL1_CD"
    else:
        family_col = None

    _log(f"Product family column for breadth/HHI: {family_col or 'absent — using item ID as proxy'}")

    both_years_sql = f"['{fy25_path.as_posix()}', '{fy26_path.as_posix()}']"

    if family_col:
        rfm = con.execute(f"""
            SELECT
                DIM_CUST_CURR_ID,
                DATE_DIFF(
                    'day',
                    MAKE_DATE(
                        CAST(MAX(DIM_ORDR_DT_ID) / 10000 AS INTEGER),
                        CAST((MAX(DIM_ORDR_DT_ID) % 10000) / 100 AS INTEGER),
                        CAST(MAX(DIM_ORDR_DT_ID) % 100 AS INTEGER)
                    ),
                    MAKE_DATE({ref_y}, {ref_m}, {ref_d})
                )                                           AS recency_days,
                COUNT(DISTINCT ORDR_NUM)                    AS frequency,
                ROUND(SUM(UNIT_SLS_AMT), 2)                 AS monetary,
                MAX(DIM_ORDR_DT_ID)                         AS last_order_date_id,
                COUNT(DISTINCT COALESCE({family_col}, 'Unknown'))
                                                            AS n_categories_bought,
                ROUND(
                    SUM(UNIT_SLS_AMT) / NULLIF(COUNT(DISTINCT ORDR_NUM), 0),
                    2
                )                                           AS avg_revenue_per_order
            FROM read_parquet({both_years_sql})
            WHERE UNIT_SLS_AMT > 0
            GROUP BY DIM_CUST_CURR_ID
        """).df()
    else:
        rfm = con.execute(f"""
            SELECT
                DIM_CUST_CURR_ID,
                DATE_DIFF(
                    'day',
                    MAKE_DATE(
                        CAST(MAX(DIM_ORDR_DT_ID) / 10000 AS INTEGER),
                        CAST((MAX(DIM_ORDR_DT_ID) % 10000) / 100 AS INTEGER),
                        CAST(MAX(DIM_ORDR_DT_ID) % 100 AS INTEGER)
                    ),
                    MAKE_DATE({ref_y}, {ref_m}, {ref_d})
                )                                           AS recency_days,
                COUNT(DISTINCT ORDR_NUM)                    AS frequency,
                ROUND(SUM(UNIT_SLS_AMT), 2)                 AS monetary,
                MAX(DIM_ORDR_DT_ID)                         AS last_order_date_id,
                COUNT(DISTINCT DIM_ITEM_E1_CURR_ID)         AS n_categories_bought,
                ROUND(
                    SUM(UNIT_SLS_AMT) / NULLIF(COUNT(DISTINCT ORDR_NUM), 0),
                    2
                )                                           AS avg_revenue_per_order
            FROM read_parquet({both_years_sql})
            WHERE UNIT_SLS_AMT > 0
            GROUP BY DIM_CUST_CURR_ID
        """).df()
        _log("Warning: n_categories_bought is product count, not family count (fallback mode)")

    if family_col:
        _log("Computing category HHI (spend concentration) across both fiscal years...")
        hhi_con = _con(memory_gb=4, threads=2)
        family_spend = hhi_con.execute(f"""
            SELECT
                DIM_CUST_CURR_ID,
                COALESCE({family_col}, 'Unknown')   AS family,
                SUM(UNIT_SLS_AMT)                   AS family_spend
            FROM read_parquet({both_years_sql})
            WHERE UNIT_SLS_AMT > 0
            GROUP BY DIM_CUST_CURR_ID, COALESCE({family_col}, 'Unknown')
        """).df()
        hhi_con.close()

        cust_total = family_spend.groupby("DIM_CUST_CURR_ID")["family_spend"].sum()
        family_spend = family_spend.merge(
            cust_total.rename("total_spend"),
            on="DIM_CUST_CURR_ID", how="left"
        )
        family_spend["share_sq"] = (
            family_spend["family_spend"] / family_spend["total_spend"].clip(lower=1)
        ) ** 2
        hhi = (
            family_spend.groupby("DIM_CUST_CURR_ID")["share_sq"]
            .sum()
            .reset_index()
            .rename(columns={"share_sq": "category_hhi"})
        )
        hhi["category_hhi"] = hhi["category_hhi"].round(4)
        rfm = rfm.merge(hhi, on="DIM_CUST_CURR_ID", how="left")
        rfm["category_hhi"] = rfm["category_hhi"].fillna(1.0)
    else:
        rfm["category_hhi"] = 1.0

    con.close()

    _log("Computing order gap averages and cycle regularity across both fiscal years...")
    gap_con  = _con(memory_gb=3, threads=1)
    dates_df = gap_con.execute(f"""
        SELECT DIM_CUST_CURR_ID, DIM_ORDR_DT_ID
        FROM   read_parquet({both_years_sql})
    """).df()
    gap_con.close()

    def _to_date(d: int) -> Optional[datetime.date]:
        try:
            s = str(int(d))
            return datetime.date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except Exception:
            return None

    def _gap_stats(grp: pd.Series) -> tuple[float, float]:
        dates = sorted({_to_date(d) for d in grp if _to_date(d) is not None})
        if len(dates) < 2:
            return 0.0, 0.0
        gaps = [
            (dates[i + 1] - dates[i]).days
            for i in range(len(dates) - 1)
        ]
        return float(np.mean(gaps)), float(np.std(gaps))

    gap_stats = (
        dates_df.groupby("DIM_CUST_CURR_ID")["DIM_ORDR_DT_ID"]
        .apply(_gap_stats)
        .reset_index()
    )
    gap_stats[["avg_order_gap_days", "cycle_regularity"]] = pd.DataFrame(
        gap_stats["DIM_ORDR_DT_ID"].tolist(), index=gap_stats.index
    )
    gap_stats = gap_stats.drop(columns=["DIM_ORDR_DT_ID"])
    gap_stats["cycle_regularity"]   = gap_stats["cycle_regularity"].round(2)
    gap_stats["avg_order_gap_days"] = gap_stats["avg_order_gap_days"].round(2)
    del dates_df

    rfm = rfm.merge(gap_stats, on="DIM_CUST_CURR_ID", how="left")
    rfm["recency_days"]          = rfm["recency_days"].clip(lower=0)
    rfm["avg_order_gap_days"]    = rfm["avg_order_gap_days"].fillna(0)
    rfm["cycle_regularity"]      = rfm["cycle_regularity"].fillna(0)
    rfm["avg_revenue_per_order"] = rfm["avg_revenue_per_order"].fillna(0)
    rfm["n_categories_bought"]   = rfm["n_categories_bought"].fillna(1)
    rfm["category_hhi"]          = rfm["category_hhi"].fillna(1.0)

    rfm["churn_label"] = rfm["DIM_CUST_CURR_ID"].map(
        lambda cid: 1 if cid in churned else 0 if cid in retained else -1
    )

    rfm["R_score"] = pd.qcut(
        rfm["recency_days"], q=5, labels=[5, 4, 3, 2, 1], duplicates="drop"
    ).astype("Int64")
    rfm["F_score"] = pd.qcut(
        rfm["frequency"].rank(method="first"), q=5, labels=[1, 2, 3, 4, 5],
        duplicates="drop"
    ).astype("Int64")
    rfm["M_score"] = pd.qcut(
        rfm["monetary"].rank(method="first"), q=5, labels=[1, 2, 3, 4, 5],
        duplicates="drop"
    ).astype("Int64")
    rfm["RFM_score"] = (
        rfm["R_score"].astype(str) +
        rfm["F_score"].astype(str) +
        rfm["M_score"].astype(str)
    )

    _log(f"RFM + behavioural features computed for {len(rfm):,} customers")
    _log(f"  recency_days           median: {rfm['recency_days'].median():.0f} days")
    _log(f"  frequency              median: {rfm['frequency'].median():.0f} orders")
    _log(f"  monetary               median: ${rfm['monetary'].median():,.0f}")
    _log(f"  avg_revenue_per_order  median: ${rfm['avg_revenue_per_order'].median():,.0f}")
    _log(f"  n_categories_bought    median: {rfm['n_categories_bought'].median():.0f}")
    _log(f"  category_hhi           median: {rfm['category_hhi'].median():.3f}")
    _log(f"  avg_order_gap_days     median: {rfm['avg_order_gap_days'].median():.1f} days")
    _log(f"  cycle_regularity       median: {rfm['cycle_regularity'].median():.1f} days std dev")

    out = OUT_FEATURES / "customer_rfm.parquet"
    rfm.to_parquet(out, index=False)
    _log(f"Saved  : {out.relative_to(ROOT)}")

    return rfm


# Step 6b: Patch category features and compute supplier profile

def step6b_patch_category_features(
    features: pd.DataFrame,
    merged_path: Path,
) -> pd.DataFrame:
    _section("Step 6b: Patching category features and computing supplier profile")

    if not merged_path.exists():
        _log("merged_dataset.parquet not found — skipping patch")
        return features

    con = _con(memory_gb=5, threads=2)

    desc = con.execute(
        f"DESCRIBE SELECT * FROM read_parquet('{merged_path.as_posix()}') LIMIT 0"
    ).df()
    available = set(desc["column_name"].tolist())

    family_col = next(
        (c for c in ["PROD_FMLY_LVL1_DSC", "PROD_FMLY_LVL1_CD"] if c in available),
        None
    )

    if family_col is None:
        _log("No product family column in merged_dataset — category patch skipped")
    else:
        _log(f"Product family column in merged_dataset: {family_col}")

        n_cats = con.execute(f"""
            SELECT
                CAST(DIM_CUST_CURR_ID AS BIGINT)             AS cust_id,
                COUNT(DISTINCT COALESCE({family_col}, 'Unknown')) AS n_categories_bought
            FROM read_parquet('{merged_path.as_posix()}')
            WHERE UNIT_SLS_AMT > 0
              AND DIM_CUST_CURR_ID IS NOT NULL
            GROUP BY CAST(DIM_CUST_CURR_ID AS BIGINT)
        """).df()
        _log(f"n_categories_bought computed for {len(n_cats):,} customers")
        _log(f"  median: {n_cats['n_categories_bought'].median():.1f}  "
             f"max: {n_cats['n_categories_bought'].max()}")

        family_spend = con.execute(f"""
            SELECT
                CAST(DIM_CUST_CURR_ID AS BIGINT)            AS cust_id,
                COALESCE({family_col}, 'Unknown')            AS family,
                SUM(UNIT_SLS_AMT)                            AS family_spend
            FROM read_parquet('{merged_path.as_posix()}')
            WHERE UNIT_SLS_AMT > 0
              AND DIM_CUST_CURR_ID IS NOT NULL
            GROUP BY CAST(DIM_CUST_CURR_ID AS BIGINT), COALESCE({family_col}, 'Unknown')
        """).df()

        cust_total = family_spend.groupby("cust_id")["family_spend"].sum()
        family_spend = family_spend.merge(
            cust_total.rename("total_spend"), on="cust_id", how="left"
        )
        family_spend["share_sq"] = (
            family_spend["family_spend"] / family_spend["total_spend"].clip(lower=1)
        ) ** 2
        hhi = (
            family_spend.groupby("cust_id")["share_sq"]
            .sum()
            .reset_index()
            .rename(columns={"share_sq": "category_hhi"})
        )
        hhi["category_hhi"] = hhi["category_hhi"].round(4)
        _log(f"category_hhi computed for {len(hhi):,} customers")
        _log(f"  median: {hhi['category_hhi'].median():.3f}  "
             f"mean: {hhi['category_hhi'].mean():.3f}")

        features = features.copy()
        features["DIM_CUST_CURR_ID_int"] = features["DIM_CUST_CURR_ID"].astype("Int64")

        n_cats["cust_id"] = n_cats["cust_id"].astype("Int64")
        hhi["cust_id"]    = hhi["cust_id"].astype("Int64")

        before_n   = features["n_categories_bought"].median()
        before_hhi = features["category_hhi"].median()

        features = features.merge(
            n_cats.rename(columns={"cust_id": "DIM_CUST_CURR_ID_int",
                                    "n_categories_bought": "_n_cats_patched"}),
            on="DIM_CUST_CURR_ID_int", how="left"
        )
        patched_mask = features["_n_cats_patched"].notna()
        features.loc[patched_mask, "n_categories_bought"] = (
            features.loc[patched_mask, "_n_cats_patched"]
        )
        features = features.drop(columns=["_n_cats_patched"])

        features = features.merge(
            hhi.rename(columns={"cust_id": "DIM_CUST_CURR_ID_int",
                                 "category_hhi": "_hhi_patched"}),
            on="DIM_CUST_CURR_ID_int", how="left"
        )
        patched_mask = features["_hhi_patched"].notna()
        features.loc[patched_mask, "category_hhi"] = (
            features.loc[patched_mask, "_hhi_patched"]
        )
        features = features.drop(columns=["_hhi_patched"])

        _log(f"n_categories_bought  before patch median: {before_n:.1f}  "
             f"after: {features['n_categories_bought'].median():.1f}")
        _log(f"category_hhi         before patch median: {before_hhi:.3f}  "
             f"after: {features['category_hhi'].median():.3f}")

    _log("")
    _log("Computing supplier_profile tag from merged dataset...")

    has_supplier = "SUPLR_ROLLUP_DSC" in available
    has_private  = "is_private_brand" in available

    if not (has_supplier and has_private):
        _log(f"Required columns missing (SUPLR_ROLLUP_DSC: {has_supplier}, "
             f"is_private_brand: {has_private}) — defaulting all customers to 'mixed'")
        con.close()
        if "DIM_CUST_CURR_ID_int" in features.columns:
            features = features.drop(columns=["DIM_CUST_CURR_ID_int"])
        features["supplier_profile"] = "mixed"
        out = OUT_FEATURES / "customer_features.parquet"
        features.to_parquet(out, index=False)
        _log(f"Saved  : {out.relative_to(ROOT)}  (patched)")
        return features

    supplier_spend = con.execute(f"""
        SELECT
            CAST(DIM_CUST_CURR_ID AS BIGINT) AS cust_id,
            SUM(UNIT_SLS_AMT)                AS total_spend,
            SUM(CASE WHEN SUPLR_ROLLUP_DSC = '{MEDLINE_SUPPLIER_NAME}'
                     THEN UNIT_SLS_AMT ELSE 0 END) AS medline_spend,
            SUM(CASE WHEN is_private_brand = 1
                     THEN UNIT_SLS_AMT ELSE 0 END) AS private_brand_spend
        FROM read_parquet('{merged_path.as_posix()}')
        WHERE UNIT_SLS_AMT > 0
          AND DIM_CUST_CURR_ID IS NOT NULL
        GROUP BY CAST(DIM_CUST_CURR_ID AS BIGINT)
    """).df()
    con.close()

    supplier_spend["pct_medline"] = (
        supplier_spend["medline_spend"] /
        supplier_spend["total_spend"].clip(lower=1)
    )
    supplier_spend["pct_private_brand"] = (
        supplier_spend["private_brand_spend"] /
        supplier_spend["total_spend"].clip(lower=1)
    )

    def _classify(row) -> str:
        if row["pct_medline"] >= MEDLINE_THRESHOLD:
            return "medline_only"
        if row["pct_private_brand"] >= MCKESSON_PRIVATE_THRESHOLD:
            return "mckesson_primary"
        return "mixed"

    supplier_spend["supplier_profile"] = supplier_spend.apply(_classify, axis=1)

    profile_counts = supplier_spend["supplier_profile"].value_counts()
    _log(f"Supplier profile distribution:")
    for profile_name in ["medline_only", "mckesson_primary", "mixed"]:
        count = int(profile_counts.get(profile_name, 0))
        pct = round(count / len(supplier_spend) * 100, 2) if len(supplier_spend) else 0
        _log(f"  {profile_name:<20} {count:>8,} customers  ({pct}%)")

    if "DIM_CUST_CURR_ID_int" not in features.columns:
        features = features.copy()
        features["DIM_CUST_CURR_ID_int"] = features["DIM_CUST_CURR_ID"].astype("Int64")

    supplier_spend["cust_id"] = supplier_spend["cust_id"].astype("Int64")

    features = features.merge(
        supplier_spend[["cust_id", "supplier_profile"]].rename(
            columns={"cust_id": "DIM_CUST_CURR_ID_int"}
        ),
        on="DIM_CUST_CURR_ID_int", how="left"
    )

    features["supplier_profile"] = features["supplier_profile"].fillna("mixed")

    n_unmatched = int((features["supplier_profile"] == "mixed").sum() -
                       int(profile_counts.get("mixed", 0)))
    if n_unmatched > 0:
        _log(f"  {n_unmatched:,} feature-matrix customers not in merged dataset "
             "— defaulted to 'mixed'")

    features = features.drop(columns=["DIM_CUST_CURR_ID_int"])

    out = OUT_FEATURES / "customer_features.parquet"
    features.to_parquet(out, index=False)
    _log(f"Saved  : {out.relative_to(ROOT)}  (patched + supplier_profile)")

    return features


# Step 6c: Compute size tier and affordability ceiling

def step6c_size_tier_and_affordability(
    features:    pd.DataFrame,
    merged_path: Path,
) -> pd.DataFrame:
    _section("Step 6c: Computing size tier and affordability ceiling")

    if not merged_path.exists():
        _log("merged_dataset.parquet not found — skipping size tier")
        features["median_monthly_spend"]    = 0.0
        features["active_months_last_12"]   = 0
        features["size_tier"]               = "new"
        features["affordability_ceiling"]   = 0.0
        return features

    con = _con(memory_gb=5, threads=2)

    # Compute monthly spend per customer per year-month
    _log("Computing monthly spend per customer...")
    monthly_spend = con.execute(f"""
        SELECT
            CAST(DIM_CUST_CURR_ID AS BIGINT)                 AS cust_id,
            order_year,
            order_month,
            SUM(UNIT_SLS_AMT)                                 AS month_spend
        FROM read_parquet('{merged_path.as_posix()}')
        WHERE UNIT_SLS_AMT > 0
          AND DIM_CUST_CURR_ID IS NOT NULL
        GROUP BY CAST(DIM_CUST_CURR_ID AS BIGINT), order_year, order_month
    """).df()

    _log(f"Monthly spend rows: {len(monthly_spend):,}")

    # Compute median monthly spend and active months per customer
    size_stats = (
        monthly_spend.groupby("cust_id")["month_spend"]
        .agg(["median", "count"])
        .reset_index()
        .rename(columns={
            "median": "median_monthly_spend",
            "count":  "active_months_total",
        })
    )
    size_stats["median_monthly_spend"] = size_stats["median_monthly_spend"].round(2)

    _log(f"Customers with spending history: {len(size_stats):,}")
    _log(f"  median_monthly_spend p25:    ${size_stats['median_monthly_spend'].quantile(0.25):,.2f}")
    _log(f"  median_monthly_spend median: ${size_stats['median_monthly_spend'].median():,.2f}")
    _log(f"  median_monthly_spend p75:    ${size_stats['median_monthly_spend'].quantile(0.75):,.2f}")
    _log(f"  median_monthly_spend p95:    ${size_stats['median_monthly_spend'].quantile(0.95):,.2f}")
    _log(f"  median_monthly_spend max:    ${size_stats['median_monthly_spend'].max():,.2f}")

    # Count active months in the last 12 months only
    max_ym = con.execute(f"""
        SELECT MAX(order_year * 100 + order_month)
        FROM read_parquet('{merged_path.as_posix()}')
        WHERE UNIT_SLS_AMT > 0
    """).fetchone()[0]

    max_year  = int(max_ym // 100)
    max_month = int(max_ym % 100)
    cutoff_minus_12 = max_year * 100 + max_month - 100

    _log(f"Dataset max year-month: {max_year}-{max_month:02d}")
    _log(f"Active months counted from: year-month {cutoff_minus_12}")

    active_last_12 = con.execute(f"""
        SELECT
            CAST(DIM_CUST_CURR_ID AS BIGINT)                 AS cust_id,
            COUNT(DISTINCT order_year * 100 + order_month)   AS active_months_last_12
        FROM read_parquet('{merged_path.as_posix()}')
        WHERE UNIT_SLS_AMT > 0
          AND DIM_CUST_CURR_ID IS NOT NULL
          AND (order_year * 100 + order_month) > {cutoff_minus_12}
        GROUP BY CAST(DIM_CUST_CURR_ID AS BIGINT)
    """).df()

    con.close()

    _log(f"Customers active in last 12 months: {len(active_last_12):,}")

    size_stats = size_stats.merge(active_last_12, on="cust_id", how="left")
    size_stats["active_months_last_12"] = size_stats["active_months_last_12"].fillna(0).astype(int)

    # Assign size_tier
    def _classify_size(row) -> str:
        if row["active_months_last_12"] < MIN_ACTIVE_MONTHS_FOR_TIER:
            return "new"
        m = row["median_monthly_spend"]
        if m < SIZE_TIER_SMALL_MAX:
            return "small"
        if m < SIZE_TIER_MID_MAX:
            return "mid"
        if m < SIZE_TIER_LARGE_MAX:
            return "large"
        return "enterprise"

    size_stats["size_tier"] = size_stats.apply(_classify_size, axis=1)

    # Compute affordability_ceiling
    size_stats["affordability_multiplier"] = (
        size_stats["size_tier"].map(AFFORDABILITY_MULTIPLIER)
    )
    size_stats["affordability_ceiling"] = (
        size_stats["median_monthly_spend"] * size_stats["affordability_multiplier"]
    ).round(2)

    tier_counts = size_stats["size_tier"].value_counts()
    _log("")
    _log("Size tier distribution:")
    for tier in ["new", "small", "mid", "large", "enterprise"]:
        count = int(tier_counts.get(tier, 0))
        pct = round(count / len(size_stats) * 100, 2) if len(size_stats) else 0
        _log(f"  {tier:<12} {count:>8,} customers  ({pct}%)")

    # Merge into features DataFrame
    features = features.copy()
    features["DIM_CUST_CURR_ID_int"] = features["DIM_CUST_CURR_ID"].astype("Int64")
    size_stats["cust_id"] = size_stats["cust_id"].astype("Int64")

    features = features.merge(
        size_stats[[
            "cust_id", "median_monthly_spend", "active_months_last_12",
            "size_tier", "affordability_ceiling"
        ]].rename(columns={"cust_id": "DIM_CUST_CURR_ID_int"}),
        on="DIM_CUST_CURR_ID_int", how="left"
    )

    features["median_monthly_spend"]  = features["median_monthly_spend"].fillna(0.0)
    features["active_months_last_12"] = features["active_months_last_12"].fillna(0).astype(int)
    features["size_tier"]             = features["size_tier"].fillna("new")
    features["affordability_ceiling"] = features["affordability_ceiling"].fillna(0.0)

    features = features.drop(columns=["DIM_CUST_CURR_ID_int"])

    _log("")
    _log(f"Affordability ceiling distribution:")
    _log(f"  p25:    ${features['affordability_ceiling'].quantile(0.25):,.2f}")
    _log(f"  median: ${features['affordability_ceiling'].median():,.2f}")
    _log(f"  p75:    ${features['affordability_ceiling'].quantile(0.75):,.2f}")
    _log(f"  p95:    ${features['affordability_ceiling'].quantile(0.95):,.2f}")

    out = OUT_FEATURES / "customer_features.parquet"
    features.to_parquet(out, index=False)
    _log(f"Saved  : {out.relative_to(ROOT)}  (with size_tier and affordability)")

    return features


# Step 7: Encode categorical features

def step7_encode_features(
    fact_paths:  list[Path],
    customers:   pd.DataFrame,
    rfm:         pd.DataFrame,
    spec_tiers:  pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    _section("Step 7: Encoding categorical features")

    features = rfm[[
        "DIM_CUST_CURR_ID", "recency_days", "frequency", "monetary",
        "avg_order_gap_days", "R_score", "F_score", "M_score",
        "RFM_score", "churn_label",
        "avg_revenue_per_order",
        "n_categories_bought",
        "category_hhi",
        "cycle_regularity",
    ]].copy()

    profile_cols = ["DIM_CUST_CURR_ID", "CUST_TYPE_CD", "SPCLTY_CD",
                    "MKT_CD", "MMS_CLASS_CD", "STATE"]
    features = features.merge(
        customers[[c for c in profile_cols if c in customers.columns]],
        on="DIM_CUST_CURR_ID", how="left",
    )

    spec_tiers_for_merge = spec_tiers.copy()
    if "avg_revenue_per_order" in spec_tiers_for_merge.columns:
        spec_tiers_for_merge = spec_tiers_for_merge.rename(
            columns={"avg_revenue_per_order": "spec_avg_revenue_per_order"}
        )
    tier_cols = ["SPCLTY_CD", "specialty_tier",
                 "pct_of_total_revenue", "spec_avg_revenue_per_order"]
    features = features.merge(
        spec_tiers_for_merge[[c for c in tier_cols if c in spec_tiers_for_merge.columns]],
        on="SPCLTY_CD", how="left",
    )

    if "CUST_TYPE_CD" in features.columns:
        features["cust_type_encoded"] = (
            features["CUST_TYPE_CD"].map({"S": 0, "X": 1, "B": 2})
            .fillna(-1).astype(int)
        )
    for col, dst in [("MKT_CD", "mkt_cd_encoded"), ("MMS_CLASS_CD", "mms_class_encoded")]:
        if col in features.columns:
            m = {v: i for i, v in enumerate(sorted(features[col].dropna().unique()))}
            features[dst] = features[col].map(m).fillna(-1).astype(int)

    state_groups = pd.DataFrame()
    if "STATE" in features.columns:
        counts      = features["STATE"].value_counts()
        valid       = set(counts[counts >= STATE_MIN_CUSTOMERS].index)
        features["state_grouped"] = features["STATE"].apply(
            lambda s: s if s in valid else "Other"
        )
        sm = {v: i for i, v in enumerate(sorted(features["state_grouped"].dropna().unique()))}
        features["state_encoded"] = features["state_grouped"].map(sm).fillna(-1).astype(int)
        state_groups = pd.DataFrame({
            "state":            counts.index,
            "unique_customers": counts.values,
            "group_assignment": [s if s in valid else "Other" for s in counts.index],
        })
        _log(f"STATE  : {len(valid)} individual  |  {len(counts) - len(valid)} grouped as Other")

    top20 = (
        spec_tiers.sort_values("total_revenue", ascending=False)
        .head(20)["SPCLTY_CD"].tolist()
    )
    for s in top20:
        features[f"spec_{s}"] = (features["SPCLTY_CD"] == s).astype(int)

    cust_spec_path = OUT_FEATURES / "_tmp_cust_spec2.parquet"
    (customers[["DIM_CUST_CURR_ID", "SPCLTY_CD"]]
     .dropna(subset=["SPCLTY_CD"])
     .to_parquet(cust_spec_path, index=False))

    fy25_path = next(p for p in fact_paths if "FY2425" in p.name)
    fy26_path = next(p for p in fact_paths if "FY2526" in p.name)

    trend_con   = _con(memory_gb=4, threads=2)
    spec_trend  = trend_con.execute(f"""
        WITH fy25 AS (
            SELECT c.SPCLTY_CD, SUM(s.UNIT_SLS_AMT) AS rev_fy25
            FROM   read_parquet('{fy25_path.as_posix()}') s
            JOIN   read_parquet({_pq(cust_spec_path)}) c
              ON   s.DIM_CUST_CURR_ID = c.DIM_CUST_CURR_ID
            GROUP  BY c.SPCLTY_CD
        ),
        fy26 AS (
            SELECT c.SPCLTY_CD, SUM(s.UNIT_SLS_AMT) AS rev_fy26
            FROM   read_parquet('{fy26_path.as_posix()}') s
            JOIN   read_parquet({_pq(cust_spec_path)}) c
              ON   s.DIM_CUST_CURR_ID = c.DIM_CUST_CURR_ID
            GROUP  BY c.SPCLTY_CD
        )
        SELECT
            COALESCE(a.SPCLTY_CD, b.SPCLTY_CD) AS SPCLTY_CD,
            COALESCE(a.rev_fy25, 0)             AS rev_fy25,
            COALESCE(b.rev_fy26, 0)             AS rev_fy26,
            ROUND(
                (COALESCE(b.rev_fy26, 0) - COALESCE(a.rev_fy25, 0))
                / NULLIF(COALESCE(a.rev_fy25, 0), 0) * 100,
                2
            ) AS specialty_revenue_trend_pct
        FROM fy25 a FULL OUTER JOIN fy26 b ON a.SPCLTY_CD = b.SPCLTY_CD
    """).df()
    trend_con.close()
    cust_spec_path.unlink(missing_ok=True)

    features = features.merge(
        spec_trend[["SPCLTY_CD", "specialty_revenue_trend_pct"]],
        on="SPCLTY_CD", how="left",
    )

    _log(f"Feature matrix: {len(features):,} customers  |  {len(features.columns)} features")
    out = OUT_FEATURES / "customer_features.parquet"
    features.to_parquet(out, index=False)
    _log(f"Saved  : {out.relative_to(ROOT)}")

    return features, state_groups


# Step 8: Build merged serving dataset

def step8_serving_dataset(
    fact_paths:  list[Path],
    customers:   pd.DataFrame,
    products:    pd.DataFrame,
    spec_tiers:  pd.DataFrame,
) -> int:
    _section("Step 8: Building merged serving dataset")

    cust_path = OUT_SERVING / "_tmp_cust.parquet"
    prod_path = OUT_SERVING / "_tmp_prod.parquet"
    tier_path = OUT_SERVING / "_tmp_tier.parquet"

    cust_cols = ["DIM_CUST_CURR_ID", "CUST_TYPE_CD", "CUST_TYPE_DSC",
                 "SPCLTY_CD", "SPCLTY_DSC", "MKT_CD",
                 "MMS_CLASS_CD", "MMS_CLASS_DSC", "ZIP", "STATE", "CITY"]
    prod_cols = ["DIM_ITEM_E1_CURR_ID", "ITEM_DSC",
                 "is_private_brand", "PRVT_BRND_FLG",
                 "PROD_FMLY_LVL1_CD", "PROD_FMLY_LVL1_DSC",
                 "PROD_CTGRY_LVL2_CD", "PROD_CTGRY_LVL2_DSC",
                 "SUPLR_DSC", "SUPLR_ROLLUP_DSC", "is_discontinued"]

    customers[[c for c in cust_cols if c in customers.columns]].to_parquet(cust_path, index=False)
    products[[c  for c in prod_cols if c in products.columns]].to_parquet(prod_path,  index=False)
    spec_tiers[["SPCLTY_CD", "specialty_tier"]].dropna().to_parquet(tier_path, index=False)

    txn_list_sql = ", ".join(f"'{p.as_posix()}'" for p in fact_paths)
    out          = OUT_SERVING / "merged_dataset.parquet"

    con = _con(memory_gb=6, threads=2, spill_dir=OUT_SERVING)
    con.execute(f"""
        COPY (
            SELECT
                s.*,
                c.CUST_TYPE_CD, c.CUST_TYPE_DSC,
                c.SPCLTY_CD,    c.SPCLTY_DSC,
                c.MKT_CD,       c.MMS_CLASS_CD, c.MMS_CLASS_DSC,
                c.ZIP,          c.STATE,         c.CITY,
                p.ITEM_DSC,
                p.is_private_brand, p.PRVT_BRND_FLG,
                p.PROD_FMLY_LVL1_CD,  p.PROD_FMLY_LVL1_DSC,
                p.PROD_CTGRY_LVL2_CD, p.PROD_CTGRY_LVL2_DSC,
                p.SUPLR_DSC,   p.SUPLR_ROLLUP_DSC, p.is_discontinued,
                t.specialty_tier
            FROM read_parquet([{txn_list_sql}]) s
            LEFT JOIN read_parquet({_pq(cust_path)}) c
               ON s.DIM_CUST_CURR_ID    = c.DIM_CUST_CURR_ID
            LEFT JOIN read_parquet({_pq(prod_path)}) p
               ON s.DIM_ITEM_E1_CURR_ID = p.DIM_ITEM_E1_CURR_ID
            LEFT JOIN read_parquet({_pq(tier_path)}) t
               ON c.SPCLTY_CD           = t.SPCLTY_CD
        ) TO {_pq(out)} (FORMAT PARQUET, COMPRESSION SNAPPY)
    """)
    row_count = con.execute(f"SELECT COUNT(*) FROM read_parquet({_pq(out)})").fetchone()[0]
    con.close()

    for p in [cust_path, prod_path, tier_path]:
        p.unlink(missing_ok=True)

    _log(f"Result : {row_count:>14,} rows")
    _log(f"Saved  : {out.relative_to(ROOT)}")

    return row_count


# Step 9: Write audit Excel files

def _style_ws(ws, df: pd.DataFrame) -> None:
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, col)
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, 10), 55)


def _save_audit(df: pd.DataFrame, path: Path, sheet: str) -> None:
    if df.empty:
        df = pd.DataFrame([{"note": "No data"}])
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name=sheet[:31], index=False)
        _style_ws(w.sheets[sheet[:31]], df)
    _log(f"Audit  : {path.relative_to(ROOT)}")


def step9_write_audit(
    cust_raw_cols:  list[str],
    cust_kept_cols: list[str],
    prod_raw_cols:  list[str],
    prod_kept_cols: list[str],
    null_zip_df:    pd.DataFrame,
    excluded_df:    pd.DataFrame,
    duplicate_log:  list[dict],
    returns:        pd.DataFrame,
    spec_tiers:     pd.DataFrame,
    state_groups:   pd.DataFrame,
    features:       pd.DataFrame,
    run_log:        list[dict],
) -> None:
    _section("Step 9: Writing audit files")

    rows = []
    for col in cust_raw_cols:
        if col not in cust_kept_cols:
            rows.append({"table": "customer_dim", "column": col,
                         "reason": "100% null" if col in _DROP_NULL_CUSTOMER else
                                   "High null >80%" if col in _DROP_SPARSE_CUSTOMER else
                                   "Not an ML feature"})
    for col in prod_raw_cols:
        if col not in prod_kept_cols:
            rows.append({"table": "product_dim", "column": col,
                         "reason": "100% null" if col in _DROP_NULL_PRODUCT else
                                   "High null >80%" if col in _DROP_SPARSE_PRODUCT else
                                   "Not an ML feature"})
    _save_audit(pd.DataFrame(rows), OUT_AUDIT / "01_dropped_columns.xlsx", "Dropped Columns")

    nz = null_zip_df.copy() if not null_zip_df.empty else pd.DataFrame([{"note": "None found"}])
    if not null_zip_df.empty:
        nz["recommendation"] = "Falls back to state-level geographic recommendations"
    _save_audit(nz, OUT_AUDIT / "02_null_zip_customers.xlsx", "Null ZIP")

    ex = excluded_df.copy() if not excluded_df.empty else pd.DataFrame({
        "ORDR_NUM": list(EXCLUDED_ORDER_NUMS),
        "exclusion_reason": ["Confirmed data entry error"] * len(EXCLUDED_ORDER_NUMS),
    })
    _save_audit(ex, OUT_AUDIT / "03_excluded_outlier_orders.xlsx", "Excluded Orders")

    _save_audit(
        pd.DataFrame(duplicate_log) if duplicate_log else pd.DataFrame(),
        OUT_AUDIT / "04_duplicate_rows.xlsx", "Duplicates"
    )

    if not returns.empty:
        ret_sum = (
            returns.groupby("fiscal_year")
            .agg(return_rows=("ORDR_NUM","count"),
                 unique_orders=("ORDR_NUM","nunique"),
                 total_return_value=("UNIT_SLS_AMT", lambda x: round(abs(x.sum()),2)))
            .reset_index()
        )
        ret_sum["note"] = "Full data at data_clean/sales/returns_clean_*.parquet"
    else:
        ret_sum = pd.DataFrame([{"note": "No return rows found"}])
    _save_audit(ret_sum, OUT_AUDIT / "05_returns_summary.xlsx", "Returns Summary")

    t3 = spec_tiers.loc[spec_tiers["specialty_tier"] == 3,
         [c for c in ["SPCLTY_CD","SPCLTY_DSC","unique_customers",
                      "specialty_tier","tier3_fallback_spclty_cd"]
          if c in spec_tiers.columns]].copy()
    if not t3.empty:
        t3["fallback_found"] = t3["tier3_fallback_spclty_cd"].notna()
    _save_audit(t3, OUT_AUDIT / "06_tier3_specialty_mapping.xlsx", "Tier 3 Mapping")

    _save_audit(
        state_groups.copy() if not state_groups.empty else pd.DataFrame(),
        OUT_AUDIT / "07_state_grouping.xlsx", "State Grouping"
    )

    if "churn_label" in features.columns:
        churn = features["churn_label"].value_counts().reset_index()
        churn.columns = ["churn_label", "customer_count"]
        churn["label_meaning"] = churn["churn_label"].map({
            1:  "Churned — in FY24-25, absent from FY25-26",
            0:  "Retained — ordered in both fiscal years",
            -1: "New in FY25-26 — excluded from churn model training",
        })
        churn["pct_of_total"] = (churn["customer_count"] / churn["customer_count"].sum() * 100).round(2)
    else:
        churn = pd.DataFrame([{"note": "churn_label not found"}])
    _save_audit(churn, OUT_AUDIT / "08_churn_labels.xlsx", "Churn Labels")

    # Supplier profile audit
    if "supplier_profile" in features.columns:
        sp = features["supplier_profile"].value_counts().reset_index()
        sp.columns = ["supplier_profile", "customer_count"]
        sp["pct_of_total"] = (sp["customer_count"] / sp["customer_count"].sum() * 100).round(2)
        sp["rule_applied"] = sp["supplier_profile"].map({
            "medline_only":
                f"spend_share_on_MEDLINE_INDUSTRIES >= {int(MEDLINE_THRESHOLD*100)}%",
            "mckesson_primary":
                f"spend_share_on_private_brand >= {int(MCKESSON_PRIVATE_THRESHOLD*100)}%",
            "mixed":
                "default bucket — everything else",
        })
    else:
        sp = pd.DataFrame([{"note": "supplier_profile not found"}])
    _save_audit(sp, OUT_AUDIT / "08b_supplier_profile.xlsx", "Supplier Profile")

    # Size tier audit
    if "size_tier" in features.columns:
        st = features["size_tier"].value_counts().reset_index()
        st.columns = ["size_tier", "customer_count"]
        st["pct_of_total"] = (st["customer_count"] / st["customer_count"].sum() * 100).round(2)
        st["rule_applied"] = st["size_tier"].map({
            "new":        f"active_months_last_12 < {MIN_ACTIVE_MONTHS_FOR_TIER}",
            "small":      f"median_monthly_spend < ${SIZE_TIER_SMALL_MAX:,}",
            "mid":        f"${SIZE_TIER_SMALL_MAX:,} <= median_monthly_spend < ${SIZE_TIER_MID_MAX:,}",
            "large":      f"${SIZE_TIER_MID_MAX:,} <= median_monthly_spend < ${SIZE_TIER_LARGE_MAX:,}",
            "enterprise": f"median_monthly_spend >= ${SIZE_TIER_LARGE_MAX:,}",
        })
    else:
        st = pd.DataFrame([{"note": "size_tier not found"}])
    _save_audit(st, OUT_AUDIT / "08c_size_tier_distribution.xlsx", "Size Tier")

    _save_audit(
        pd.DataFrame(run_log) if run_log else pd.DataFrame(),
        OUT_AUDIT / "09_cleaning_run_log.xlsx", "Run Log"
    )


# Step 9b: Team-facing Excel summary report

def step9b_excel_report(
    customers:   pd.DataFrame,
    products:    pd.DataFrame,
    fact_paths:  list[Path],
    spec_tiers:  pd.DataFrame,
    features:    pd.DataFrame,
    elapsed:     float,
) -> None:
    _section("Step 9b: Writing team Excel summary report")

    out_path = DATA_CLEAN / "cleaning_summary_report.xlsx"
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _h(cell, bg="1F4E79"):
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    def _d(cell, bg="FFFFFF", center=False):
        cell.font      = Font(name="Arial", size=9, color="000000")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        cell.border    = border

    def _sheet(ws, df, hdr_bg="1F4E79", tab="1F4E79"):
        ws.sheet_properties.tabColor = tab
        if df.empty:
            ws.cell(1, 1, "No data")
            return
        for ci, col in enumerate(df.columns, 1):
            _h(ws.cell(1, ci, str(col)), bg=hdr_bg)
        for ri, row in enumerate(df.itertuples(index=False), 2):
            bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, val if pd.notna(val) else "")
                _d(c, bg=bg, center=isinstance(val, (int, float)))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_cells in ws.columns:
            w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w+2,12),55)

    churn1 = int((features["churn_label"] == 1).sum()) if "churn_label" in features.columns else 0
    churn0 = int((features["churn_label"] == 0).sum()) if "churn_label" in features.columns else 0
    rate   = round(churn1 / max(churn1 + churn0, 1) * 100, 2)

    con       = _con(memory_gb=2, threads=1)
    txn_list  = ", ".join(f"'{p.as_posix()}'" for p in fact_paths)
    txn_stats = con.execute(f"""
        SELECT fiscal_year,
               COUNT(*)                   AS transaction_rows,
               COUNT(DISTINCT ORDR_NUM)   AS unique_orders,
               COUNT(DISTINCT DIM_CUST_CURR_ID)  AS unique_customers,
               COUNT(DISTINCT DIM_ITEM_E1_CURR_ID) AS unique_products,
               ROUND(SUM(UNIT_SLS_AMT),2) AS total_revenue,
               ROUND(AVG(UNIT_SLS_AMT),2) AS avg_order_value
        FROM read_parquet([{txn_list}])
        GROUP BY fiscal_year ORDER BY fiscal_year
    """).df()
    con.close()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        wb = writer.book

        summary = pd.DataFrame([
            ("PIPELINE",    "Run duration (sec)",    round(elapsed, 1)),
            ("PIPELINE",    "Revenue cap applied",   f"${MAX_ORDER_LINE_VALUE:,.0f}"),
            ("PIPELINE",    "Blocked order numbers", ", ".join(str(o) for o in sorted(EXCLUDED_ORDER_NUMS))),
            ("CUSTOMERS",   "Rows after dedup",      f"{len(customers):,}"),
            ("CUSTOMERS",   "Columns kept",          len(customers.columns)),
            ("CUSTOMERS",   "Columns dropped",       "44 (25 fully null + 19 sparse)"),
            ("PRODUCTS",    "Rows after dedup",      f"{len(products):,}"),
            ("PRODUCTS",    "Columns kept",          len(products.columns)),
            ("PRODUCTS",    "Columns dropped",       "28 (10 fully null + 18 sparse)"),
            ("SPECIALTIES", "Total specialties",     f"{len(spec_tiers):,}"),
            ("SPECIALTIES", "Tier 1 (>=1000 custs)", f"{(spec_tiers['specialty_tier']==1).sum()}"),
            ("SPECIALTIES", "Tier 2 (10-999 custs)", f"{(spec_tiers['specialty_tier']==2).sum()}"),
            ("SPECIALTIES", "Tier 3 (<10 custs)",    f"{(spec_tiers['specialty_tier']==3).sum()}"),
            ("CHURN MODEL", "Churned label=1",       f"{churn1:,}"),
            ("CHURN MODEL", "Retained label=0",      f"{churn0:,}"),
            ("CHURN MODEL", "Churn rate",            f"{rate}%"),
            ("FEATURES",    "ML feature columns",    len(features.columns)),
            ("FEATURES",    "Customers with RFM",    f"{len(features):,}"),
        ], columns=["category", "metric", "value"])

        if "supplier_profile" in features.columns:
            sp_counts = features["supplier_profile"].value_counts()
            sp_rows = pd.DataFrame([
                ("SUPPLIER", "Medline-only customers",      f"{int(sp_counts.get('medline_only', 0)):,}"),
                ("SUPPLIER", "McKesson-primary customers",  f"{int(sp_counts.get('mckesson_primary', 0)):,}"),
                ("SUPPLIER", "Mixed customers",             f"{int(sp_counts.get('mixed', 0)):,}"),
            ], columns=["category", "metric", "value"])
            summary = pd.concat([summary, sp_rows], ignore_index=True)

        if "size_tier" in features.columns:
            st_counts = features["size_tier"].value_counts()
            st_rows = pd.DataFrame([
                ("SIZE", "New customers (<2 months active)",  f"{int(st_counts.get('new', 0)):,}"),
                ("SIZE", "Small tier customers",              f"{int(st_counts.get('small', 0)):,}"),
                ("SIZE", "Mid tier customers",                f"{int(st_counts.get('mid', 0)):,}"),
                ("SIZE", "Large tier customers",              f"{int(st_counts.get('large', 0)):,}"),
                ("SIZE", "Enterprise tier customers",         f"{int(st_counts.get('enterprise', 0)):,}"),
            ], columns=["category", "metric", "value"])
            summary = pd.concat([summary, st_rows], ignore_index=True)

        summary.to_excel(writer, sheet_name="01_pipeline_summary", index=False)
        _sheet(writer.sheets["01_pipeline_summary"], summary, "1F4E79", "1F4E79")

        cust_stats = pd.DataFrame([{
            "column": c, "dtype": str(customers[c].dtype),
            "null_count": int(customers[c].isna().sum()),
            "null_pct": round(customers[c].isna().sum() / len(customers) * 100, 2),
            "unique_values": int(customers[c].nunique()),
        } for c in customers.columns])
        cust_stats.to_excel(writer, sheet_name="02_customer_stats", index=False)
        _sheet(writer.sheets["02_customer_stats"], cust_stats, "375623", "375623")

        prod_stats = pd.DataFrame([{
            "column": c, "dtype": str(products[c].dtype),
            "null_count": int(products[c].isna().sum()),
            "null_pct": round(products[c].isna().sum() / len(products) * 100, 2),
            "unique_values": int(products[c].nunique()),
        } for c in products.columns])
        prod_stats.to_excel(writer, sheet_name="03_product_stats", index=False)
        _sheet(writer.sheets["03_product_stats"], prod_stats, "7030A0", "7030A0")

        txn_stats.to_excel(writer, sheet_name="04_sales_stats", index=False)
        _sheet(writer.sheets["04_sales_stats"], txn_stats, "C00000", "C00000")

        spec_out = spec_tiers[[c for c in [
            "SPCLTY_CD","SPCLTY_DSC","unique_customers","unique_orders",
            "total_revenue","pct_of_total_revenue","avg_revenue_per_order",
            "specialty_tier","tier3_fallback_spclty_cd"] if c in spec_tiers.columns
        ]].sort_values("total_revenue", ascending=False)
        spec_out.to_excel(writer, sheet_name="05_specialty_tiers", index=False)
        _sheet(writer.sheets["05_specialty_tiers"], spec_out, "1F6B75", "1F6B75")

        if "churn_label" in features.columns:
            ch = features["churn_label"].value_counts().reset_index()
            ch.columns = ["churn_label", "customer_count"]
            ch["label_meaning"] = ch["churn_label"].map({
                1: "Churned — in FY24-25, absent from FY25-26",
                0: "Retained — ordered in both fiscal years",
                -1: "New in FY25-26 — excluded from churn model training",
            })
            ch["pct_of_total"] = (ch["customer_count"] / ch["customer_count"].sum() * 100).round(2)
        else:
            ch = pd.DataFrame([{"note": "Churn labels not found"}])
        ch.to_excel(writer, sheet_name="06_churn_overview", index=False)
        _sheet(writer.sheets["06_churn_overview"], ch, "C00000", "C00000")

        feat_ref = pd.DataFrame({
            "column_name":   features.columns,
            "dtype":         [str(features[c].dtype) for c in features.columns],
            "null_count":    [int(features[c].isna().sum()) for c in features.columns],
            "null_pct":      [round(features[c].isna().sum()/len(features)*100,2) for c in features.columns],
            "unique_values": [int(features[c].nunique()) for c in features.columns],
        })
        feat_ref.to_excel(writer, sheet_name="07_feature_columns", index=False)
        _sheet(writer.sheets["07_feature_columns"], feat_ref, "833C00", "833C00")

    _log(f"Report : {out_path.relative_to(ROOT)}")


# Step 10: Final summary

def step10_summary(
    customers:  pd.DataFrame,
    products:   pd.DataFrame,
    fact_paths: list[Path],
    spec_tiers: pd.DataFrame,
    features:   pd.DataFrame,
    elapsed:    float,
) -> None:
    _section("Step 10: Pipeline complete")

    outputs = [
        OUT_CUSTOMER / "customers_clean.parquet",
        OUT_PRODUCT  / "products_clean.parquet",
    ] + fact_paths + [
        OUT_FEATURES / "specialty_tiers.parquet",
        OUT_FEATURES / "customer_rfm.parquet",
        OUT_FEATURES / "customer_features.parquet",
        OUT_SERVING  / "merged_dataset.parquet",
    ]

    print()
    print(f"  {'FILE':<58} {'SIZE':>8}")
    print(f"  {'─'*58} {'─'*8}")
    for f in outputs:
        if f.exists():
            mb = round(f.stat().st_size / (1_024 * 1_024), 1)
            print(f"  {str(f.relative_to(ROOT)):<58} {mb:>6.1f} MB")
        else:
            print(f"  {str(f.relative_to(ROOT)):<58}  MISSING")

    churn1 = (features["churn_label"] == 1).sum() if "churn_label" in features.columns else 0
    churn0 = (features["churn_label"] == 0).sum() if "churn_label" in features.columns else 0
    rate   = round(churn1 / max(churn1 + churn0, 1) * 100, 2)

    print()
    metrics = [
        ("Unique customers",        f"{customers['DIM_CUST_CURR_ID'].nunique():,}"),
        ("Unique products",         f"{products['DIM_ITEM_E1_CURR_ID'].nunique():,}"),
        ("Unique specialties",      f"{spec_tiers['SPCLTY_CD'].nunique():,}"),
        ("Tier 1 specialties",      f"{(spec_tiers['specialty_tier']==1).sum()}"),
        ("Tier 2 specialties",      f"{(spec_tiers['specialty_tier']==2).sum()}"),
        ("Tier 3 specialties",      f"{(spec_tiers['specialty_tier']==3).sum()}"),
        ("Customers with RFM",      f"{len(features):,}"),
        ("Churned  (label=1)",      f"{churn1:,}  ({rate}%)"),
        ("Retained (label=0)",      f"{churn0:,}"),
        ("Pipeline time",           f"{elapsed:.1f}s"),
    ]
    for label, value in metrics:
        print(f"  {label:<35} {value}")

    if "supplier_profile" in features.columns:
        sp = features["supplier_profile"].value_counts()
        print()
        print("  Supplier profile distribution:")
        for p in ["medline_only", "mckesson_primary", "mixed"]:
            print(f"    {p:<20} {int(sp.get(p, 0)):>8,}")

    if "size_tier" in features.columns:
        st = features["size_tier"].value_counts()
        print()
        print("  Size tier distribution:")
        for t in ["new", "small", "mid", "large", "enterprise"]:
            print(f"    {t:<12} {int(st.get(t, 0)):>8,}")

# Main

def main() -> None:
    print()
    print("=" * 62)
    print("  B2B MEDICAL SUPPLY — DATA CLEANING PIPELINE")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}")
    print("=" * 62)
    start = time.time()

    for d in [OUT_CUSTOMER, OUT_PRODUCT, OUT_SALES,
              OUT_FEATURES, OUT_SERVING, OUT_AUDIT]:
        d.mkdir(parents=True, exist_ok=True)

    run_log: list[dict] = []

    def _rlog(step: str, msg: str) -> None:
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        run_log.append({"timestamp": ts, "step": step, "message": msg})
        _log(f"[{ts}] {step}: {msg}")

    folders = _discover_folders()
    step1_validate(folders)
    _rlog("Step 1", "Source paths validated")

    customers, cust_raw, cust_kept, null_zip_df = step2_clean_customers(folders["cust_folder"])
    _rlog("Step 2", f"{len(customers):,} customers  |  {len(cust_kept)} columns")

    products, prod_raw, prod_kept = step3_clean_products(folders["item_folder"])
    _rlog("Step 3", f"{len(products):,} products  |  {len(prod_kept)} columns")

    returns, dup_log, excluded_df = step4_clean_sales(folders["fact_folders"])

    fact_paths = sorted(
        [p for p in OUT_SALES.glob("transactions_clean_FY*.parquet")],
        key=lambda p: p.name,
    )
    if not fact_paths:
        raise FileNotFoundError("Step 4 output parquets not found in data_clean/sales/")

    txn_count_con = _con(memory_gb=2, threads=1)
    txn_count_sql = ", ".join(f"'{p.as_posix()}'" for p in fact_paths)
    total_txn     = txn_count_con.execute(
        f"SELECT COUNT(*) FROM read_parquet([{txn_count_sql}])"
    ).fetchone()[0]
    txn_count_con.close()

    _rlog("Step 4", f"{total_txn:,} transactions on disk  |  {len(returns):,} returns")

    spec_tiers = step5_specialty_tiers(fact_paths, customers)
    _rlog("Step 5", f"{len(spec_tiers):,} specialties tiered")

    rfm = step6_rfm(fact_paths, customers)
    _rlog("Step 6", f"RFM scored for {len(rfm):,} customers")

    features, state_groups = step7_encode_features(fact_paths, customers, rfm, spec_tiers)
    _rlog("Step 7", f"{len(features):,} customers  |  {len(features.columns)} features")

    merged_rows = step8_serving_dataset(fact_paths, customers, products, spec_tiers)
    _rlog("Step 8", f"{merged_rows:,} rows in merged serving dataset")

    features = step6b_patch_category_features(
        features, OUT_SERVING / "merged_dataset.parquet"
    )
    _rlog("Step 6b", f"category features patched + supplier_profile added  |  "
                      f"{len(features.columns)} features")

    features = step6c_size_tier_and_affordability(
        features, OUT_SERVING / "merged_dataset.parquet"
    )
    _rlog("Step 6c", f"size_tier and affordability_ceiling computed  |  "
                      f"{len(features.columns)} features")

    step9_write_audit(
        cust_raw_cols  = cust_raw,
        cust_kept_cols = cust_kept,
        prod_raw_cols  = prod_raw,
        prod_kept_cols = prod_kept,
        null_zip_df    = null_zip_df,
        excluded_df    = excluded_df,
        duplicate_log  = dup_log,
        returns        = returns,
        spec_tiers     = spec_tiers,
        state_groups   = state_groups,
        features       = features,
        run_log        = run_log,
    )
    _rlog("Step 9", "Audit files written")

    elapsed = round(time.time() - start, 1)

    step9b_excel_report(
        customers  = customers,
        products   = products,
        fact_paths = fact_paths,
        spec_tiers = spec_tiers,
        features   = features,
        elapsed    = elapsed,
    )
    _rlog("Step 9b", "Team summary report written")

    step10_summary(customers, products, fact_paths, spec_tiers, features, elapsed)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted by user.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise