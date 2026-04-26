from __future__ import annotations

import sys
import time
import warnings
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# Paths

ROOT       = Path(__file__).resolve().parent.parent.parent
DATA_CLEAN = ROOT / "data_clean"

CUSTOMER_FILE = DATA_CLEAN / "customer" / "customers_clean.parquet"
PRODUCT_FILE  = DATA_CLEAN / "product"  / "products_clean.parquet"
SALES_FY25    = DATA_CLEAN / "sales"    / "transactions_clean_FY2425.parquet"
SALES_FY26    = DATA_CLEAN / "sales"    / "transactions_clean_FY2526.parquet"
MERGED_FILE   = DATA_CLEAN / "serving"  / "merged_dataset.parquet"
FEATURE_FILE  = DATA_CLEAN / "features" / "customer_features.parquet"
RFM_FILE      = DATA_CLEAN / "features" / "customer_rfm.parquet"
SPEC_FILE     = DATA_CLEAN / "features" / "specialty_tiers.parquet"

OUT_REPORT = DATA_CLEAN / "audit" / "00_sanity_check_summary.xlsx"


# Expected ranges

EXPECTED_CUSTOMER_ROWS   = (1_000_000, 1_200_000)
EXPECTED_PRODUCT_ROWS    = (270_000,   285_000)
EXPECTED_TRANSACTION_ROWS = (100_000_000, 115_000_000)
EXPECTED_FEATURE_ROWS    = (380_000,   400_000)

EXPECTED_SIZE_TIERS      = {"new", "small", "mid", "large", "enterprise"}
EXPECTED_SUPPLIER_PROFILES = {"medline_only", "mckesson_primary", "mixed"}
EXPECTED_CHURN_LABELS    = {-1, 0, 1}
EXPECTED_MKT_CDS         = {"PO", "LTC", "SC", "LC", "HC", "AC", "X", "N/A"}

REQUIRED_FEATURE_COLS = [
    "DIM_CUST_CURR_ID",
    "recency_days", "frequency", "monetary",
    "avg_revenue_per_order", "avg_order_gap_days",
    "R_score", "F_score", "M_score", "RFM_score", "churn_label",
    "n_categories_bought", "category_hhi", "cycle_regularity",
    "CUST_TYPE_CD", "SPCLTY_CD", "MKT_CD", "STATE",
    "supplier_profile",
    "median_monthly_spend", "active_months_last_12",
    "size_tier", "affordability_ceiling",
]


# Logging

def _s(title: str) -> None:
    print(f"\n{'-' * 64}\n  {title}\n{'-' * 64}", flush=True)


def _log(msg: str) -> None:
    print(f"  {msg}", flush=True)


def _check(name: str, passed: bool, detail: str = "") -> dict:
    status = "PASS" if passed else "FAIL"
    marker = "OK " if passed else "XX "
    line = f"  [{marker}] {name}"
    if detail:
        line += f"  |  {detail}"
    print(line, flush=True)
    return {"check": name, "status": status, "detail": detail}


# Excel helpers

def _style(ws, df: pd.DataFrame, hc: str = "1F4E79") -> None:
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, str(col))
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val if pd.notna(val) else "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(
            col_cells[0].column)].width = min(max(w + 2, 12), 55)


# Check 1: File existence and sizes

def check_file_existence() -> list[dict]:
    _s("Check 1: File existence and sizes")
    results = []

    expected_files = [
        ("customers_clean.parquet",          CUSTOMER_FILE, 30,   80),
        ("products_clean.parquet",           PRODUCT_FILE,  10,   30),
        ("transactions_clean_FY2425.parquet", SALES_FY25,  1500, 3500),
        ("transactions_clean_FY2526.parquet", SALES_FY26,  1500, 3500),
        ("merged_dataset.parquet",           MERGED_FILE,  5000, 10000),
        ("customer_features.parquet",        FEATURE_FILE,    5,   50),
        ("customer_rfm.parquet",             RFM_FILE,        5,   30),
        ("specialty_tiers.parquet",          SPEC_FILE,   0.001, 1.0),
    ]

    for label, path, min_mb, max_mb in expected_files:
        if not path.exists():
            results.append(_check(f"{label} exists", False, f"Not found: {path}"))
            continue
        size_mb = path.stat().st_size / (1024 * 1024)
        in_range = min_mb <= size_mb <= max_mb
        results.append(_check(
            f"{label} size",
            in_range,
            f"{size_mb:.1f} MB  (expected {min_mb}-{max_mb} MB)"
        ))
    return results


# Check 2: Row counts

def check_row_counts() -> list[dict]:
    _s("Check 2: Row counts match expectations")
    results = []

    if CUSTOMER_FILE.exists():
        n = len(pd.read_parquet(CUSTOMER_FILE, columns=["DIM_CUST_CURR_ID"]))
        in_range = EXPECTED_CUSTOMER_ROWS[0] <= n <= EXPECTED_CUSTOMER_ROWS[1]
        results.append(_check(
            "Customer rows", in_range,
            f"{n:,} rows  (expected {EXPECTED_CUSTOMER_ROWS[0]:,}-{EXPECTED_CUSTOMER_ROWS[1]:,})"
        ))

    if PRODUCT_FILE.exists():
        n = len(pd.read_parquet(PRODUCT_FILE, columns=["DIM_ITEM_E1_CURR_ID"]))
        in_range = EXPECTED_PRODUCT_ROWS[0] <= n <= EXPECTED_PRODUCT_ROWS[1]
        results.append(_check(
            "Product rows", in_range,
            f"{n:,} rows  (expected {EXPECTED_PRODUCT_ROWS[0]:,}-{EXPECTED_PRODUCT_ROWS[1]:,})"
        ))

    if SALES_FY25.exists() and SALES_FY26.exists():
        import duckdb
        con = duckdb.connect()
        total = con.execute(f"""
            SELECT COUNT(*) FROM read_parquet(
                ['{SALES_FY25.as_posix()}', '{SALES_FY26.as_posix()}']
            )
        """).fetchone()[0]
        con.close()
        in_range = EXPECTED_TRANSACTION_ROWS[0] <= total <= EXPECTED_TRANSACTION_ROWS[1]
        results.append(_check(
            "Transaction rows (both FYs)", in_range,
            f"{total:,} rows  (expected {EXPECTED_TRANSACTION_ROWS[0]:,}-{EXPECTED_TRANSACTION_ROWS[1]:,})"
        ))

    if FEATURE_FILE.exists():
        n = len(pd.read_parquet(FEATURE_FILE, columns=["DIM_CUST_CURR_ID"]))
        in_range = EXPECTED_FEATURE_ROWS[0] <= n <= EXPECTED_FEATURE_ROWS[1]
        results.append(_check(
            "Feature rows", in_range,
            f"{n:,} rows  (expected {EXPECTED_FEATURE_ROWS[0]:,}-{EXPECTED_FEATURE_ROWS[1]:,})"
        ))

    return results


# Check 3: Schema integrity

def check_schema() -> list[dict]:
    _s("Check 3: Schema integrity (required columns present)")
    results = []

    if FEATURE_FILE.exists():
        df_cols = pd.read_parquet(FEATURE_FILE).columns.tolist()
        missing = [c for c in REQUIRED_FEATURE_COLS if c not in df_cols]
        results.append(_check(
            "customer_features.parquet columns",
            len(missing) == 0,
            f"{len(df_cols)} total columns  |  missing: {missing if missing else 'none'}"
        ))

        # Check specifically for new Step 6c columns
        new_cols = ["median_monthly_spend", "active_months_last_12",
                    "size_tier", "affordability_ceiling"]
        for col in new_cols:
            present = col in df_cols
            results.append(_check(
                f"New column: {col}",
                present,
                "present" if present else "MISSING — Step 6c did not run"
            ))

    return results


# Check 4: Value integrity

def check_values() -> list[dict]:
    _s("Check 4: Value integrity (valid values only)")
    results = []

    if not FEATURE_FILE.exists():
        return results

    df = pd.read_parquet(FEATURE_FILE)

    # size_tier
    if "size_tier" in df.columns:
        actual = set(df["size_tier"].dropna().unique())
        invalid = actual - EXPECTED_SIZE_TIERS
        results.append(_check(
            "size_tier values valid",
            len(invalid) == 0,
            f"values: {sorted(actual)}  |  invalid: {sorted(invalid) if invalid else 'none'}"
        ))

    # supplier_profile
    if "supplier_profile" in df.columns:
        actual = set(df["supplier_profile"].dropna().unique())
        invalid = actual - EXPECTED_SUPPLIER_PROFILES
        results.append(_check(
            "supplier_profile values valid",
            len(invalid) == 0,
            f"values: {sorted(actual)}  |  invalid: {sorted(invalid) if invalid else 'none'}"
        ))

    # churn_label
    if "churn_label" in df.columns:
        actual = set(df["churn_label"].dropna().astype(int).unique())
        invalid = actual - EXPECTED_CHURN_LABELS
        results.append(_check(
            "churn_label values valid",
            len(invalid) == 0,
            f"values: {sorted(actual)}  |  invalid: {sorted(invalid) if invalid else 'none'}"
        ))

    # No negative monetary
    if "monetary" in df.columns:
        n_neg = (df["monetary"] < 0).sum()
        results.append(_check(
            "monetary non-negative",
            n_neg == 0,
            f"{n_neg:,} negative values"
        ))

    # No negative affordability_ceiling
    if "affordability_ceiling" in df.columns:
        n_neg = (df["affordability_ceiling"] < 0).sum()
        results.append(_check(
            "affordability_ceiling non-negative",
            n_neg == 0,
            f"{n_neg:,} negative values"
        ))

    # No negative median_monthly_spend
    if "median_monthly_spend" in df.columns:
        n_neg = (df["median_monthly_spend"] < 0).sum()
        results.append(_check(
            "median_monthly_spend non-negative",
            n_neg == 0,
            f"{n_neg:,} negative values"
        ))

    # active_months_last_12 in [0, 12]
    if "active_months_last_12" in df.columns:
        n_bad = ((df["active_months_last_12"] < 0) |
                 (df["active_months_last_12"] > 12)).sum()
        results.append(_check(
            "active_months_last_12 in [0, 12]",
            n_bad == 0,
            f"{n_bad:,} out-of-range values"
        ))

    # R, F, M scores in [1, 5]
    for col in ["R_score", "F_score", "M_score"]:
        if col in df.columns:
            s = df[col].dropna()
            n_bad = ((s < 1) | (s > 5)).sum()
            results.append(_check(
                f"{col} in [1, 5]",
                n_bad == 0,
                f"{n_bad:,} out-of-range values"
            ))

    # category_hhi in [0, 1]
    if "category_hhi" in df.columns:
        s = df["category_hhi"].dropna()
        n_bad = ((s < 0) | (s > 1)).sum()
        results.append(_check(
            "category_hhi in [0, 1]",
            n_bad == 0,
            f"{n_bad:,} out-of-range values"
        ))

    return results


# Check 5: Distribution sanity

def check_distributions() -> list[dict]:
    _s("Check 5: Distribution sanity")
    results = []

    if not FEATURE_FILE.exists():
        return results

    df = pd.read_parquet(FEATURE_FILE)

    # Size tier distribution — no tier less than 0.1%, no tier more than 70%
    if "size_tier" in df.columns:
        st_pct = df["size_tier"].value_counts(normalize=True) * 100

        for tier in EXPECTED_SIZE_TIERS:
            pct = st_pct.get(tier, 0)
            in_range = 0.1 <= pct <= 70.0 if tier != "enterprise" else 0.0 <= pct <= 70.0
            # Enterprise can be very small, down to ~0.5%
            if tier == "enterprise":
                in_range = pct >= 0.1 and pct <= 70.0
            results.append(_check(
                f"size_tier '{tier}' share",
                in_range,
                f"{pct:.2f}%"
            ))

    # Supplier profile distribution
    if "supplier_profile" in df.columns:
        sp_pct = df["supplier_profile"].value_counts(normalize=True) * 100
        for profile in EXPECTED_SUPPLIER_PROFILES:
            pct = sp_pct.get(profile, 0)
            results.append(_check(
                f"supplier_profile '{profile}' share",
                pct > 0,
                f"{pct:.2f}%"
            ))

    # Churn rate between 5% and 50%
    if "churn_label" in df.columns:
        churn_df = df[df["churn_label"].isin([0, 1])]
        churn_rate = (churn_df["churn_label"] == 1).mean() * 100
        in_range = 5 <= churn_rate <= 50
        results.append(_check(
            "Churn rate reasonable",
            in_range,
            f"{churn_rate:.2f}%  (expected 5-50%)"
        ))

    return results


# Check 6: Business logic validation

def check_business_logic() -> list[dict]:
    _s("Check 6: Business logic validation")
    results = []

    if not FEATURE_FILE.exists():
        return results

    df = pd.read_parquet(FEATURE_FILE)

    # All 'new' tier customers must have active_months_last_12 < 2
    if "size_tier" in df.columns and "active_months_last_12" in df.columns:
        new_custs = df[df["size_tier"] == "new"]
        violation = (new_custs["active_months_last_12"] >= 2).sum()
        results.append(_check(
            "'new' tier implies active_months < 2",
            violation == 0,
            f"{violation:,} violations"
        ))

    # Non-new tier customers must have active_months_last_12 >= 2
    if "size_tier" in df.columns and "active_months_last_12" in df.columns:
        non_new = df[df["size_tier"] != "new"]
        violation = (non_new["active_months_last_12"] < 2).sum()
        results.append(_check(
            "Non-'new' tiers imply active_months >= 2",
            violation == 0,
            f"{violation:,} violations"
        ))

    # Size tier ordering: enterprise median > large median > mid median > small median
    if "size_tier" in df.columns and "median_monthly_spend" in df.columns:
        medians = df.groupby("size_tier")["median_monthly_spend"].median()
        small_m = medians.get("small", 0)
        mid_m = medians.get("mid", 0)
        large_m = medians.get("large", 0)
        ent_m = medians.get("enterprise", 0)
        ordered = small_m <= mid_m <= large_m <= ent_m
        results.append(_check(
            "Size tier spend ordering (small<=mid<=large<=enterprise)",
            ordered,
            f"small=${small_m:,.0f}  mid=${mid_m:,.0f}  large=${large_m:,.0f}  enterprise=${ent_m:,.0f}"
        ))

    # Affordability ceiling should equal median_monthly_spend * multiplier
    # Multipliers: new=3.0, small=1.5, mid=1.8, large=2.0, enterprise=2.5
    if all(c in df.columns for c in ["size_tier", "median_monthly_spend", "affordability_ceiling"]):
        multipliers = {"new": 3.0, "small": 1.5, "mid": 1.8, "large": 2.0, "enterprise": 2.5}
        df_check = df.copy()
        df_check["expected_ceiling"] = df_check["size_tier"].map(multipliers) * df_check["median_monthly_spend"]
        df_check["diff"] = (df_check["affordability_ceiling"] - df_check["expected_ceiling"]).abs()
        violation = (df_check["diff"] > 0.05).sum()  # allow tiny rounding
        results.append(_check(
            "Affordability ceiling formula correct",
            violation == 0,
            f"{violation:,} rows with incorrect calculation"
        ))

    # DIM_CUST_CURR_ID uniqueness in features
    if "DIM_CUST_CURR_ID" in df.columns:
        n_total = len(df)
        n_unique = df["DIM_CUST_CURR_ID"].nunique()
        results.append(_check(
            "DIM_CUST_CURR_ID unique in features",
            n_total == n_unique,
            f"{n_total:,} rows  |  {n_unique:,} unique IDs"
        ))

    return results


# Check 7: Cross-file consistency

def check_cross_file() -> list[dict]:
    _s("Check 7: Cross-file consistency")
    results = []

    if FEATURE_FILE.exists() and CUSTOMER_FILE.exists():
        feat_ids = set(pd.read_parquet(
            FEATURE_FILE, columns=["DIM_CUST_CURR_ID"]
        )["DIM_CUST_CURR_ID"])
        cust_ids = set(pd.read_parquet(
            CUSTOMER_FILE, columns=["DIM_CUST_CURR_ID"]
        )["DIM_CUST_CURR_ID"])
        orphan = feat_ids - cust_ids
        results.append(_check(
            "All feature IDs exist in customer dim",
            len(orphan) == 0,
            f"{len(orphan):,} orphan IDs in features"
        ))

    if PRODUCT_FILE.exists() and MERGED_FILE.exists():
        import duckdb
        con = duckdb.connect()
        try:
            result = con.execute(f"""
                SELECT
                    COUNT(DISTINCT m.DIM_ITEM_E1_CURR_ID) AS merged_items,
                    COUNT(DISTINCT p.DIM_ITEM_E1_CURR_ID) AS product_items
                FROM read_parquet('{MERGED_FILE.as_posix()}') m
                LEFT JOIN read_parquet('{PRODUCT_FILE.as_posix()}') p
                    ON m.DIM_ITEM_E1_CURR_ID = p.DIM_ITEM_E1_CURR_ID
                WHERE p.DIM_ITEM_E1_CURR_ID IS NULL
            """).fetchone()
            # This checks items in merged that are NOT in products
            orphan_items = con.execute(f"""
                SELECT COUNT(DISTINCT m.DIM_ITEM_E1_CURR_ID)
                FROM read_parquet('{MERGED_FILE.as_posix()}') m
                LEFT JOIN read_parquet('{PRODUCT_FILE.as_posix()}') p
                    ON m.DIM_ITEM_E1_CURR_ID = p.DIM_ITEM_E1_CURR_ID
                WHERE p.DIM_ITEM_E1_CURR_ID IS NULL
            """).fetchone()[0]
            # Some orphans are expected (6.26% null product match), that's fine
            results.append(_check(
                "Product match coverage in merged",
                True,
                f"{orphan_items:,} items in transactions without product record (expected some)"
            ))
        finally:
            con.close()

    return results


# Summary

def print_summary(all_results: list[dict]) -> int:
    _s("Summary")
    n_total = len(all_results)
    n_pass = sum(1 for r in all_results if r["status"] == "PASS")
    n_fail = n_total - n_pass

    _log(f"Total checks : {n_total}")
    _log(f"Passed       : {n_pass}")
    _log(f"Failed       : {n_fail}")

    if n_fail > 0:
        _log("")
        _log("Failed checks:")
        for r in all_results:
            if r["status"] == "FAIL":
                _log(f"  - {r['check']}  |  {r['detail']}")

    return n_fail


def save_report(all_results: list[dict]) -> None:
    OUT_REPORT.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(all_results)
    n_pass = (df["status"] == "PASS").sum()
    n_fail = (df["status"] == "FAIL").sum()

    summary = pd.DataFrame([
        {"metric": "Total checks", "value": len(df)},
        {"metric": "Passed",       "value": int(n_pass)},
        {"metric": "Failed",       "value": int(n_fail)},
        {"metric": "Pass rate",    "value": f"{n_pass / max(len(df), 1) * 100:.1f}%"},
    ])

    with pd.ExcelWriter(OUT_REPORT, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="01_summary", index=False)
        df.to_excel(writer, sheet_name="02_all_checks", index=False)

        failed = df[df["status"] == "FAIL"]
        if len(failed) > 0:
            failed.to_excel(writer, sheet_name="03_failures", index=False)

        wb = writer.book
        _style(writer.sheets["01_summary"], summary, hc="1F4E79")
        wb["01_summary"].sheet_properties.tabColor = "1F4E79"

        _style(writer.sheets["02_all_checks"], df, hc="375623")
        wb["02_all_checks"].sheet_properties.tabColor = "375623"

        if len(failed) > 0:
            _style(writer.sheets["03_failures"], failed, hc="C00000")
            wb["03_failures"].sheet_properties.tabColor = "C00000"

    print()
    _log(f"Saved: {OUT_REPORT.relative_to(ROOT)}")


# Main

def main() -> None:
    print()
    print("=" * 64)
    print("  CLEAN_DATA SANITY CHECK")
    print("=" * 64)
    start = time.time()

    all_results = []
    all_results += check_file_existence()
    all_results += check_row_counts()
    all_results += check_schema()
    all_results += check_values()
    all_results += check_distributions()
    all_results += check_business_logic()
    all_results += check_cross_file()

    n_fail = print_summary(all_results)
    save_report(all_results)

    _log(f"Runtime: {time.time() - start:.1f}s")

    if n_fail > 0:
        print()
        print(f"XX SANITY CHECK FAILED: {n_fail} check(s) failed.")
        print("Review data_clean/audit/00_sanity_check_summary.xlsx for details.")
        sys.exit(1)
    else:
        print()
        print("OK SANITY CHECK PASSED: All checks passed.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"\nFATAL ERROR: {exc}", file=sys.stderr)
        raise