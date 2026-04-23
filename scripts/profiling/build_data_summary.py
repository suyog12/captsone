from __future__ import annotations

import sys
from pathlib import Path
from typing import Dict, List

import duckdb
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Configuration
OUTPUT_FILE = "data_summary.xlsx"

TAB_COLORS = {
    "00_README":             "1F4E79",
    "01_data_overview":      "2E75B6",
    "02_cust_by_type":       "375623",
    "02_cust_by_specialty":  "375623",
    "02_cust_by_market":     "375623",
    "02_cust_by_mms":        "375623",
    "03_prod_by_family":     "7030A0",
    "03_prod_by_category":   "7030A0",
    "03_prod_by_supplier":   "7030A0",
    "03_prod_brand_split":   "7030A0",
    "04_sales_by_month":     "C00000",
    "05_sales_cust_type":    "ED7D31",
    "06_sales_specialty":    "1F6B75",
    "07_geo_state":          "833C00",
    "07_geo_zip":            "833C00",
    "08_sales_product":      "375623",
    "09_brand_overall":      "7030A0",
    "09_brand_cust_type":    "7030A0",
    "09_brand_specialty":    "7030A0",
    "10_yoy_overall":        "C00000",
    "10_yoy_by_month":       "C00000",
    "10_yoy_product_family": "C00000",
    "10_yoy_customers":      "C00000",
    "11_returns_summary":    "FF0000",
    "11_revenue_outliers":   "FF0000",
}

# Building helper funnctions
def ql(v: str) -> str:
    return "'" + v.replace("'", "''") + "'"

def find_data_folder(base: Path) -> Path:
    for c in [base / "data", base / "Data"]:
        if c.is_dir():
            return c
    raise FileNotFoundError("No 'data' or 'Data' folder found.")

def find_parquet_files(data_dir: Path) -> List[Path]:
    files = sorted(data_dir.rglob("*.parquet"))
    if not files:
        raise FileNotFoundError(f"No parquet files under {data_dir}")
    return files

def table_group(p: Path) -> str:
    return p.parent.name.strip()

def q_df(con, sql: str) -> pd.DataFrame:
    return con.execute(sql).df()

#Excel formatting for ease of reading
THIN = Side(style="thin", color="D0D0D0")
BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT  = "F2F7FF"

def hdr(cell, bg="1F4E79", fc="FFFFFF", bold=True, size=10):
    cell.font      = Font(name="Arial", bold=bold, color=fc, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border    = BDR

def dat(cell, bg="FFFFFF", bold=False, center=False):
    cell.font      = Font(name="Arial", bold=bold, size=9, color="000000")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left", vertical="center"
    )
    cell.border    = BDR

def write_sheet(ws, df: pd.DataFrame,
                hdr_color: str = "1F4E79", tab_color: str = None):
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    if df.empty:
        ws.cell(1, 1, "No data")
        return
    for ci, col in enumerate(df.columns, 1):
        hdr(ws.cell(1, ci, col), bg=hdr_color)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = ALT if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            v    = val if not (isinstance(val, float) and pd.isna(val)) else ""
            cell = ws.cell(ri, ci, v)
            dat(cell, bg=bg, center=isinstance(val, (int, float)))
    ws.freeze_panes = ws.cell(2, 1)
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max(w + 2, 8), 42)

# Registering Views
def register_views(con: duckdb.DuckDBPyConnection, data_dir: Path) -> None:
    files  = find_parquet_files(data_dir)
    groups: Dict[str, List[str]] = {}
    for fp in files:
        groups.setdefault(table_group(fp), []).append(str(fp))

    print("Registering views (data stays on disk)...")

    cust_paths = [p for g, ps in groups.items()
                  if "cust" in g and "dim" in g for p in ps]
    item_paths = [p for g, ps in groups.items()
                  if "item" in g and "dim" in g for p in ps]
    fact_parts: List[str] = []
    for g, ps in groups.items():
        if ("fct" in g or "sales" in g.lower()) and "dim" not in g:
            fy = "FY2425" if "2425" in g else "FY2526"
            for p in ps:
                fact_parts.append(
                    f"SELECT *, '{fy}' AS fiscal_year FROM read_parquet({ql(p)})"
                )

    if not cust_paths:
        raise RuntimeError("Customer dimension files not found.")
    if not item_paths:
        raise RuntimeError("Product dimension files not found.")
    if not fact_parts:
        raise RuntimeError("Sales fact files not found.")

    con.execute("CREATE VIEW dim_cust AS " +
                " UNION ALL ".join(
                    f"SELECT * FROM read_parquet({ql(p)})"
                    for p in cust_paths))

    con.execute("CREATE VIEW dim_item AS " +
                " UNION ALL ".join(
                    f"SELECT * FROM read_parquet({ql(p)})"
                    for p in item_paths))

    # Dedup sales via ROW_NUMBER in a view
    con.execute(f"""
        CREATE VIEW fct_sales AS
        SELECT * EXCLUDE (_rn) FROM (
            SELECT *,
                ROW_NUMBER() OVER (
                    PARTITION BY ORDR_NUM, ORDR_LINE_NUM, fiscal_year
                    ORDER BY (SELECT NULL)
                ) AS _rn
            FROM ({" UNION ALL ".join(fact_parts)})
        ) t WHERE _rn = 1
    """)

    # Merged view — only columns needed for summaries
    con.execute("""
        CREATE VIEW merged AS
        SELECT
            s.fiscal_year,
            s.DIM_CUST_CURR_ID,
            s.DIM_ITEM_E1_CURR_ID,
            s.ORDR_NUM,
            s.ORDR_LINE_NUM,
            s.ORDR_QTY,
            s.UNIT_SLS_AMT,
            s.ORDR_MTHD_DSC,
            s.ORDR_SRC_DSC,
            CAST(s.DIM_ORDR_DT_ID / 10000 AS INTEGER)          AS order_year,
            CAST((s.DIM_ORDR_DT_ID % 10000) / 100 AS INTEGER)  AS order_month,
            TRY_CAST(
                CONCAT(
                    CAST(s.DIM_ORDR_DT_ID / 10000 AS VARCHAR), '-',
                    LPAD(CAST((s.DIM_ORDR_DT_ID % 10000) / 100 AS VARCHAR), 2, '0'), '-',
                    LPAD(CAST(s.DIM_ORDR_DT_ID % 100 AS VARCHAR), 2, '0')
                ) AS DATE
            )                                                   AS order_date,
            CASE WHEN s.ORDR_QTY < 0 THEN true ELSE false END  AS is_return,
            c.CUST_TYPE_CD,
            c.CUST_TYPE_DSC,
            c.SPCLTY_CD,
            c.SPCLTY_DSC,
            c.MKT_CD,
            c.MMS_CLASS_CD,
            c.MMS_CLASS_DSC,
            c.ACTV_FLG,
            c.ZIP,
            c.STATE,
            c.CITY,
            i.ITEM_DSC,
            i.PRVT_BRND_FLG,
            i.PROD_FMLY_LVL1_CD,
            i.PROD_FMLY_LVL1_DSC,
            i.PROD_CTGRY_LVL2_CD,
            i.PROD_CTGRY_LVL2_DSC,
            i.SUPLR_DSC,
            i.SUPLR_ROLLUP_DSC,
            i.GNRC_FLG
        FROM fct_sales s
        LEFT JOIN dim_cust c ON s.DIM_CUST_CURR_ID  = c.DIM_CUST_CURR_ID
        LEFT JOIN dim_item i ON s.DIM_ITEM_E1_CURR_ID = i.DIM_ITEM_E1_CURR_ID
    """)

    print(f"  dim_cust  → {len(cust_paths)} files")
    print(f"  dim_item  → {len(item_paths)} files")
    print(f"  fct_sales → {len(fact_parts)} files (deduped in view)")
    print("  merged view ready — all queries run directly against parquet")

# Query Summary
def build_data_overview(con) -> pd.DataFrame:
    rows = []
    def add(tbl, metric, val):
        rows.append({"table": tbl, "metric": metric, "value": val})

    for view, label in [("dim_cust","Customer dimension"),
                         ("dim_item","Product dimension"),
                         ("fct_sales","Sales facts (deduped)")]:
        try:
            n = con.execute(f"SELECT COUNT(*) FROM {view}").fetchone()[0]
            add(label, "total rows", f"{n:,}")
        except Exception as e:
            add(label, "total rows", f"ERROR: {e}")

    for fy in ["FY2425","FY2526"]:
        try:
            r = con.execute(f"""
                SELECT MIN(order_date), MAX(order_date)
                FROM merged WHERE fiscal_year='{fy}' AND order_date IS NOT NULL
            """).fetchone()
            if r and r[0]:
                add(f"Sales {fy}", "date range", f"{r[0]} to {r[1]}")
        except Exception:
            pass

    try:
        n = con.execute("SELECT COUNT(*) FROM merged WHERE is_return=true").fetchone()[0]
        add("merged","return rows (ORDR_QTY < 0)", f"{n:,}")
    except Exception:
        pass

    for col, label in [
        ("DIM_CUST_CURR_ID",   "unique customers"),
        ("DIM_ITEM_E1_CURR_ID","unique products"),
        ("ORDR_NUM",           "unique order numbers"),
        ("SPCLTY_CD",          "unique specialties"),
        ("CUST_TYPE_CD",       "unique customer types"),
        ("STATE",              "unique states"),
        ("ZIP",                "unique ZIP codes"),
        ("PROD_FMLY_LVL1_CD",  "unique product families"),
    ]:
        try:
            n = con.execute(
                f'SELECT COUNT(DISTINCT "{col}") FROM merged '
                f'WHERE "{col}" IS NOT NULL'
            ).fetchone()[0]
            add("merged", label, f"{n:,}")
        except Exception:
            pass

    try:
        r = con.execute("""
            SELECT SUM(UNIT_SLS_AMT) FROM merged
            WHERE is_return=false AND UNIT_SLS_AMT IS NOT NULL
        """).fetchone()[0]
        if r:
            add("merged","total revenue (excl returns)", f"${r:,.2f}")
    except Exception:
        pass

    return pd.DataFrame(rows)


def build_cust_by_type(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            COALESCE(CUST_TYPE_CD,  'Unknown') AS cust_type_cd,
            COALESCE(CUST_TYPE_DSC, 'Unknown') AS cust_type_dsc,
            COUNT(DISTINCT DIM_CUST_CURR_ID)   AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)           AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)        AS total_revenue,
            ROUND(COUNT(DISTINCT DIM_CUST_CURR_ID)*100.0 /
                SUM(COUNT(DISTINCT DIM_CUST_CURR_ID)) OVER(), 2) AS pct_customers,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER(), 2)                AS pct_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2 ORDER BY 3 DESC
    """)


def build_cust_by_specialty(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            COALESCE(SPCLTY_CD,  'Unknown') AS spclty_cd,
            COALESCE(SPCLTY_DSC, 'Unknown') AS spclty_dsc,
            COUNT(DISTINCT DIM_CUST_CURR_ID) AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)         AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)      AS total_revenue,
            ROUND(AVG(UNIT_SLS_AMT), 2)      AS avg_order_line_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2 ORDER BY 3 DESC
    """)


def build_cust_by_market(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            COALESCE(MKT_CD, 'Unknown')      AS mkt_cd,
            COUNT(DISTINCT DIM_CUST_CURR_ID) AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)         AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)      AS total_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1 ORDER BY 2 DESC
    """)


def build_cust_by_mms(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            COALESCE(MMS_CLASS_CD,  'Unknown') AS mms_class_cd,
            COALESCE(MMS_CLASS_DSC, 'Unknown') AS mms_class_dsc,
            COUNT(DISTINCT DIM_CUST_CURR_ID)    AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)            AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)         AS total_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2 ORDER BY 3 DESC
    """)


def build_prod_by_family(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            COALESCE(PROD_FMLY_LVL1_CD,  'Unknown') AS family_cd,
            COALESCE(PROD_FMLY_LVL1_DSC, 'Unknown') AS family_dsc,
            COUNT(DISTINCT DIM_ITEM_E1_CURR_ID)      AS unique_products,
            COUNT(DISTINCT DIM_CUST_CURR_ID)         AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)                 AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)              AS total_revenue,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER(), 2)    AS pct_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2 ORDER BY 6 DESC
    """)


def build_prod_by_category(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            COALESCE(PROD_FMLY_LVL1_DSC,  'Unknown') AS family_dsc,
            COALESCE(PROD_CTGRY_LVL2_CD,  'Unknown') AS category_cd,
            COALESCE(PROD_CTGRY_LVL2_DSC, 'Unknown') AS category_dsc,
            COUNT(DISTINCT DIM_ITEM_E1_CURR_ID)       AS unique_products,
            COUNT(DISTINCT DIM_CUST_CURR_ID)          AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)                  AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)               AS total_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2,3 ORDER BY 7 DESC
    """)


def build_prod_by_supplier(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            COALESCE(SUPLR_ROLLUP_DSC, SUPLR_DSC, 'Unknown') AS supplier,
            COUNT(DISTINCT DIM_ITEM_E1_CURR_ID)               AS unique_products,
            COUNT(DISTINCT DIM_CUST_CURR_ID)                  AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)                          AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)                       AS total_revenue,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER(), 2)             AS pct_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1 ORDER BY 5 DESC
    """)


def build_prod_brand_split(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            COALESCE(PRVT_BRND_FLG, 'Unknown')        AS prvt_brnd_flg,
            COUNT(DISTINCT DIM_ITEM_E1_CURR_ID)        AS unique_products,
            COUNT(DISTINCT DIM_CUST_CURR_ID)           AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)                   AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)                AS total_revenue,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER(), 2)      AS pct_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1 ORDER BY 5 DESC
    """)


def build_sales_by_month(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            order_year,
            order_month,
            COUNT(DISTINCT ORDR_NUM)            AS unique_orders,
            COUNT(DISTINCT DIM_CUST_CURR_ID)    AS unique_customers,
            COUNT(DISTINCT DIM_ITEM_E1_CURR_ID) AS unique_products,
            COUNT(*)                            AS total_order_lines,
            ROUND(SUM(UNIT_SLS_AMT), 2)         AS total_revenue,
            ROUND(AVG(UNIT_SLS_AMT), 2)         AS avg_line_revenue,
            COUNT(CASE WHEN is_return THEN 1 END) AS return_lines
        FROM merged
        WHERE order_date IS NOT NULL
        GROUP BY 1,2,3 ORDER BY 2,3
    """)


def build_sales_cust_type(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(CUST_TYPE_CD,  'Unknown') AS cust_type_cd,
            COALESCE(CUST_TYPE_DSC, 'Unknown') AS cust_type_dsc,
            COUNT(DISTINCT DIM_CUST_CURR_ID)   AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)           AS unique_orders,
            COUNT(*)                           AS total_order_lines,
            ROUND(SUM(UNIT_SLS_AMT), 2)        AS total_revenue,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER (PARTITION BY fiscal_year), 2
            )                                  AS pct_of_fy_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2,3 ORDER BY 1, 7 DESC
    """)


def build_sales_specialty(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(SPCLTY_CD,  'Unknown') AS spclty_cd,
            COALESCE(SPCLTY_DSC, 'Unknown') AS spclty_dsc,
            COUNT(DISTINCT DIM_CUST_CURR_ID) AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)         AS unique_orders,
            COUNT(*)                         AS total_order_lines,
            ROUND(SUM(UNIT_SLS_AMT), 2)      AS total_revenue,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER (PARTITION BY fiscal_year), 2
            )                                AS pct_of_fy_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2,3 ORDER BY 1, 7 DESC
    """)


def build_geo_state(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(STATE, 'Unknown')         AS state,
            COUNT(DISTINCT DIM_CUST_CURR_ID)   AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)           AS unique_orders,
            COUNT(*)                           AS total_order_lines,
            ROUND(SUM(UNIT_SLS_AMT), 2)        AS total_revenue,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER (PARTITION BY fiscal_year), 2
            )                                  AS pct_of_fy_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2 ORDER BY 1, 6 DESC
    """)


def build_geo_zip(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(STATE, 'Unknown') AS state,
            COALESCE(ZIP,   'Unknown') AS zip,
            COALESCE(CITY,  'Unknown') AS city,
            COUNT(DISTINCT DIM_CUST_CURR_ID) AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)         AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)      AS total_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2,3,4 ORDER BY 1, 7 DESC
    """)


def build_sales_product(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(PROD_FMLY_LVL1_CD,  'Unknown') AS family_cd,
            COALESCE(PROD_FMLY_LVL1_DSC, 'Unknown') AS family_dsc,
            COALESCE(PROD_CTGRY_LVL2_CD,  'Unknown') AS category_cd,
            COALESCE(PROD_CTGRY_LVL2_DSC, 'Unknown') AS category_dsc,
            COALESCE(PRVT_BRND_FLG, 'Unknown')        AS prvt_brnd_flg,
            COUNT(DISTINCT DIM_ITEM_E1_CURR_ID)       AS unique_products,
            COUNT(DISTINCT DIM_CUST_CURR_ID)          AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)                  AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)               AS total_revenue,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER (PARTITION BY fiscal_year), 2
            )                                         AS pct_of_fy_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2,3,4,5,6 ORDER BY 1, 10 DESC
    """)


def build_brand_overall(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(PRVT_BRND_FLG, 'Unknown')        AS prvt_brnd_flg,
            COUNT(DISTINCT DIM_CUST_CURR_ID)           AS unique_customers,
            COUNT(DISTINCT DIM_ITEM_E1_CURR_ID)        AS unique_products,
            COUNT(DISTINCT ORDR_NUM)                   AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)                AS total_revenue,
            ROUND(SUM(UNIT_SLS_AMT)*100.0 /
                SUM(SUM(UNIT_SLS_AMT)) OVER (PARTITION BY fiscal_year), 2
            )                                          AS pct_of_fy_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2 ORDER BY 1, 6 DESC
    """)


def build_brand_cust_type(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(CUST_TYPE_DSC, 'Unknown') AS cust_type_dsc,
            COALESCE(PRVT_BRND_FLG, 'Unknown') AS prvt_brnd_flg,
            COUNT(DISTINCT DIM_CUST_CURR_ID)   AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)           AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)        AS total_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2,3 ORDER BY 1,2, 6 DESC
    """)


def build_brand_specialty(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(SPCLTY_DSC, 'Unknown')    AS spclty_dsc,
            COALESCE(PRVT_BRND_FLG, 'Unknown') AS prvt_brnd_flg,
            COUNT(DISTINCT DIM_CUST_CURR_ID)   AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)           AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)        AS total_revenue
        FROM merged WHERE is_return=false
        GROUP BY 1,2,3 ORDER BY 1,2, 6 DESC
    """)


def build_yoy_overall(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COUNT(DISTINCT DIM_CUST_CURR_ID)          AS unique_customers,
            COUNT(DISTINCT DIM_ITEM_E1_CURR_ID)        AS unique_products,
            COUNT(DISTINCT ORDR_NUM)                   AS unique_orders,
            COUNT(*)                                   AS total_order_lines,
            ROUND(SUM(UNIT_SLS_AMT), 2)                AS total_revenue,
            ROUND(AVG(UNIT_SLS_AMT), 2)                AS avg_line_revenue,
            COUNT(CASE WHEN is_return THEN 1 END)      AS return_lines,
            ROUND(COUNT(CASE WHEN is_return THEN 1 END)*100.0 /
                COUNT(*), 2)                           AS return_pct
        FROM merged GROUP BY 1 ORDER BY 1
    """)


def build_yoy_by_month(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year, order_year, order_month,
            COUNT(DISTINCT DIM_CUST_CURR_ID) AS unique_customers,
            COUNT(DISTINCT ORDR_NUM)         AS unique_orders,
            ROUND(SUM(UNIT_SLS_AMT), 2)      AS total_revenue
        FROM merged
        WHERE is_return=false AND order_date IS NOT NULL
        GROUP BY 1,2,3 ORDER BY 2,3
    """)


def build_yoy_product_family(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            family_dsc,
            SUM(CASE WHEN fiscal_year='FY2425' THEN rev END) AS revenue_fy2425,
            SUM(CASE WHEN fiscal_year='FY2526' THEN rev END) AS revenue_fy2526
        FROM (
            SELECT
                COALESCE(PROD_FMLY_LVL1_DSC,'Unknown') AS family_dsc,
                fiscal_year,
                ROUND(SUM(UNIT_SLS_AMT), 2) AS rev
            FROM merged WHERE is_return=false
            GROUP BY 1,2
        ) sub
        GROUP BY 1
        ORDER BY COALESCE(revenue_fy2526, 0) DESC
    """)


def build_yoy_customers(con) -> pd.DataFrame:
    return q_df(con, """
        WITH fy25 AS (
            SELECT DISTINCT DIM_CUST_CURR_ID FROM merged
            WHERE fiscal_year='FY2425' AND is_return=false
        ),
        fy26 AS (
            SELECT DISTINCT DIM_CUST_CURR_ID FROM merged
            WHERE fiscal_year='FY2526' AND is_return=false
        )
        SELECT 'New in FY25-26 (not in FY24-25)' AS customer_status,
               COUNT(*) AS customer_count
        FROM fy26 WHERE DIM_CUST_CURR_ID NOT IN (SELECT DIM_CUST_CURR_ID FROM fy25)
        UNION ALL
        SELECT 'Retained (in both fiscal years)', COUNT(*)
        FROM fy26 WHERE DIM_CUST_CURR_ID IN (SELECT DIM_CUST_CURR_ID FROM fy25)
        UNION ALL
        SELECT 'Lost in FY25-26 (was in FY24-25)', COUNT(*)
        FROM fy25 WHERE DIM_CUST_CURR_ID NOT IN (SELECT DIM_CUST_CURR_ID FROM fy26)
    """)


def build_returns_summary(con) -> pd.DataFrame:
    return q_df(con, """
        SELECT
            fiscal_year,
            COALESCE(CUST_TYPE_DSC,      'Unknown') AS cust_type_dsc,
            COALESCE(PROD_FMLY_LVL1_DSC, 'Unknown') AS family_dsc,
            COUNT(*)                                 AS return_lines,
            ROUND(SUM(ABS(ORDR_QTY)), 0)             AS total_return_qty,
            ROUND(SUM(ABS(UNIT_SLS_AMT)), 2)         AS total_return_value
        FROM merged WHERE is_return=true
        GROUP BY 1,2,3 ORDER BY 6 DESC
    """)


def build_revenue_outliers(con) -> pd.DataFrame:
    return q_df(con, """
        WITH p999 AS (
            SELECT PERCENTILE_CONT(0.999) WITHIN GROUP
                (ORDER BY UNIT_SLS_AMT) AS threshold
            FROM merged WHERE is_return=false AND UNIT_SLS_AMT IS NOT NULL
        )
        SELECT
            m.fiscal_year,
            m.ORDR_NUM,
            COALESCE(m.ITEM_DSC,      'Unknown') AS item_dsc,
            COALESCE(m.CUST_TYPE_DSC, 'Unknown') AS cust_type_dsc,
            COALESCE(m.SPCLTY_DSC,    'Unknown') AS spclty_dsc,
            m.ORDR_QTY,
            m.UNIT_SLS_AMT,
            m.order_date
        FROM merged m, p999
        WHERE m.UNIT_SLS_AMT > p999.threshold AND m.is_return=false
        ORDER BY m.UNIT_SLS_AMT DESC
        LIMIT 500
    """)


# Readme Sheet
def write_readme(wb):
    ws = wb.create_sheet("00_README", 0)
    ws.sheet_properties.tabColor = TAB_COLORS["00_README"]
    rows = [
        ("Sheet", "Contents", "Use for"),
        ("00_README",             "This navigation guide",                                   "Start here"),
        ("01_data_overview",      "Row counts, date ranges, unique counts, total revenue",   "Confirm data loaded correctly"),
        ("02_cust_by_type",       "Customers by CUST_TYPE_CD — count, orders, revenue",      "Customer type distribution"),
        ("02_cust_by_specialty",  "Customers by SPCLTY_CD — count, orders, revenue",         "Specialty distribution"),
        ("02_cust_by_market",     "Customers by MKT_CD",                                     "Market segment distribution"),
        ("02_cust_by_mms",        "Customers by MMS_CLASS_CD",                               "MMS class distribution"),
        ("03_prod_by_family",     "Products by product family (level 1)",                    "Product portfolio by family"),
        ("03_prod_by_category",   "Products by category (level 2)",                          "Category breakdown"),
        ("03_prod_by_supplier",   "Products by supplier",                                    "Supplier revenue share"),
        ("03_prod_brand_split",   "Private brand vs non-brand: count, orders, revenue",      "Overall brand penetration"),
        ("04_sales_by_month",     "Orders and revenue by month, both fiscal years",          "Sales volume over time"),
        ("05_sales_cust_type",    "Revenue and orders by customer type per fiscal year",     "Which types drive revenue"),
        ("06_sales_specialty",    "Revenue and orders by specialty per fiscal year",         "Which specialties drive revenue"),
        ("07_geo_state",          "Revenue by state per fiscal year",                        "Geographic revenue distribution"),
        ("07_geo_zip",            "Revenue by ZIP code per fiscal year",                     "ZIP-level geography"),
        ("08_sales_product",      "Revenue by family + category + brand flag",               "Product revenue breakdown"),
        ("09_brand_overall",      "Private brand vs non-brand split by fiscal year",         "Brand revenue trend"),
        ("09_brand_cust_type",    "Brand split broken down by customer type",                "Which types buy private brand"),
        ("09_brand_specialty",    "Brand split broken down by specialty",                    "Which specialties buy private brand"),
        ("10_yoy_overall",        "FY24-25 vs FY25-26 headline numbers",                    "Top-level YoY comparison"),
        ("10_yoy_by_month",       "Monthly revenue both fiscal years",                       "Seasonal patterns"),
        ("10_yoy_product_family", "Product family revenue FY24-25 vs FY25-26",              "Which families grew or declined"),
        ("10_yoy_customers",      "New / retained / lost customers between years",           "Customer retention overview"),
        ("11_returns_summary",    "Returns by customer type and product family",             "Return patterns"),
        ("11_revenue_outliers",   "Top 500 order lines by revenue (top 0.1%)",              "Anomalies and large orders"),
    ]
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 38
    for ri, row in enumerate(rows, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            if ri == 1:
                hdr(cell, bg=TAB_COLORS["00_README"])
            else:
                dat(cell, bg=ALT if ri % 2 == 0 else "FFFFFF")
    ws.freeze_panes = "A2"


# Main
def main():
    script_dir = Path(__file__).resolve().parent
    data_dir   = find_data_folder(script_dir)
    output     = script_dir / OUTPUT_FILE

    print(f"Data folder : {data_dir}")
    print(f"Output      : {output}\n")

    # Persistent DuckDB file lets DuckDB spill to disk automatically
    db_path = script_dir / "_summary_temp.duckdb"
    con = duckdb.connect(str(db_path))

    con.execute("SET threads = 4")
    con.execute("SET memory_limit = '6GB'")
    con.execute("SET preserve_insertion_order = false")
    con.execute(f"SET temp_directory = '{str(script_dir).replace(chr(92), '/')}'")

    try:
        register_views(con, data_dir)

        steps = [
            ("01_data_overview",      build_data_overview),
            ("02_cust_by_type",       build_cust_by_type),
            ("02_cust_by_specialty",  build_cust_by_specialty),
            ("02_cust_by_market",     build_cust_by_market),
            ("02_cust_by_mms",        build_cust_by_mms),
            ("03_prod_by_family",     build_prod_by_family),
            ("03_prod_by_category",   build_prod_by_category),
            ("03_prod_by_supplier",   build_prod_by_supplier),
            ("03_prod_brand_split",   build_prod_brand_split),
            ("04_sales_by_month",     build_sales_by_month),
            ("05_sales_cust_type",    build_sales_cust_type),
            ("06_sales_specialty",    build_sales_specialty),
            ("07_geo_state",          build_geo_state),
            ("07_geo_zip",            build_geo_zip),
            ("08_sales_product",      build_sales_product),
            ("09_brand_overall",      build_brand_overall),
            ("09_brand_cust_type",    build_brand_cust_type),
            ("09_brand_specialty",    build_brand_specialty),
            ("10_yoy_overall",        build_yoy_overall),
            ("10_yoy_by_month",       build_yoy_by_month),
            ("10_yoy_product_family", build_yoy_product_family),
            ("10_yoy_customers",      build_yoy_customers),
            ("11_returns_summary",    build_returns_summary),
            ("11_revenue_outliers",   build_revenue_outliers),
        ]

        print("\nRunning summary queries...")
        summaries = {}
        for name, fn in steps:
            print(f"  {name}...", end=" ", flush=True)
            try:
                summaries[name] = fn(con)
                print(f"{len(summaries[name])} rows")
            except Exception as e:
                print(f"WARN — {e}")
                summaries[name] = pd.DataFrame()

        print("\nWriting Excel...")
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            wb = writer.book
            for name, df in summaries.items():
                sname = name[:31]
                if df.empty:
                    ws = wb.create_sheet(sname)
                    ws.cell(1, 1, "No data")
                    ws.sheet_properties.tabColor = TAB_COLORS.get(name, "595959")
                else:
                    df.to_excel(writer, sheet_name=sname, index=False)
                    ws = writer.sheets[sname]
                    write_sheet(ws, df,
                                hdr_color=TAB_COLORS.get(name, "1F4E79"),
                                tab_color=TAB_COLORS.get(name))
            write_readme(wb)

        print(f"\nDone. Output: {output}")

    finally:
        con.close()
        for f in [db_path,
                  script_dir / "_summary_temp.duckdb.wal"]:
            try:
                f.unlink()
            except Exception:
                pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        raise